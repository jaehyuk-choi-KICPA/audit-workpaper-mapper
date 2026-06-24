"""잔액형 총괄표 매핑 품질 평가 하니스 (객관 3기준 평균점수).

기준(각 0~100):
  1) FORM 양식무결성  : validate_grid+check_column_fit+check_cells_present issue 수 (0=100, 건당 -20)
  2) FIT  시각잘림     : 숫자열 내용 길이 > 열너비 인 셀 수(####### 위험). 0=100, 건당 -10
  3) TIE  수치정합     : 본문 기초/기말 합 == 정산표 라우팅 대분류 합(허용오차 1원). 일치=100, 불일치=0

조서별 평균 → 전체 평균. 4회 루프는 외부에서 코드 교정 후 재실행해 점수 추이를 본다.
"""
import sys, re
from pathlib import Path
from collections import defaultdict

sys.path.insert(0, str(Path(__file__).resolve().parent))
import yaml
import openpyxl
from openpyxl.utils import column_index_from_string
from extractors import parse_trial_balance
from pipeline import build_a0, _check_a0_form

REGISTRY = [
    ("A-0", "a0.yaml", "A-0_4000_A000_template.xlsx"),
    ("C",   "c.yaml",  "C_4000_template.xlsx"),
    ("AA",  "aa.yaml", "AA_4000_template.xlsx"),
    ("CC",  "cc.yaml", "CC_4000_template.xlsx"),
    ("D",   "d.yaml",  "D_4000_template.xlsx"),
    ("EE",  "ee.yaml", "EE_4000_template.xlsx"),
    ("R",   "r.yaml",  "R_4000_template.xlsx"),
]
_norm = lambda s: re.sub(r"\s+", "", str(s)) if s is not None else ""


def _disp_len(v):
    if isinstance(v, (int, float)):
        return len(f"{round(v):,}") + (2 if v < 0 else 0)
    return len(str(v)) if v is not None else 0


def score_one(code, cfgf, tmpl, settlement, cfgdir, tmpldir, outdir, tb):
    cfg = yaml.safe_load((Path(cfgdir) / cfgf).read_text(encoding="utf-8"))
    out = str(Path(outdir) / f"{code}.xlsx")
    rep = build_a0(settlement=settlement, template=str(Path(tmpldir) / tmpl),
                   config_dir=cfgdir, output=out, parsed_dir=str(Path(outdir) / "변환자료"),
                   config_file=cfgf)
    info = next((dict(zip(("s",), [None])) for _ in [0]), None)
    # build_a0 doesn't return info; recompute body range from report-free: re-open
    b = cfg["body"]
    sheet = cfg["sheet"]
    ds = b["data_start_row"]
    wb = openpyxl.load_workbook(out)
    ws = wb[sheet]
    name_ci = column_index_from_string(b["columns"]["계정명"])
    marker_ci = column_index_from_string(b["marker_col"])
    end_anchor = _norm(b.get("body_end_anchor", "기초"))
    de = ds
    for r in range(ds, ws.max_row + 1):
        if any(_norm(ws.cell(r, ci).value).startswith(end_anchor) for ci in {name_ci, marker_ci}):
            de = r - 1
            break
    # FORM
    form_issues = _check_a0_form(out, {"body_data_start": ds, "body_data_end": de}, cfgdir, cfgf)
    form = max(0, 100 - 20 * len(form_issues))
    # FIT
    numc = [column_index_from_string(c) for c in b.get("num_cols", [])]
    overflow = 0
    for r in range(ds, de + 1):
        for ci in numc:
            col = openpyxl.utils.get_column_letter(ci)
            w = ws.column_dimensions[col].width or 8
            v = ws.cell(r, ci).value
            if not isinstance(v, str) or not v.startswith("="):   # 수식 제외(계산값 모름)
                if _disp_len(v) > w:
                    overflow += 1
    fit = max(0, 100 - 10 * overflow)
    # TIE: 본문 기초/기말 합 vs 정산표 라우팅 합
    owned = {_norm(g) for sec in b["sections"] for g in sec["groups"]}
    exp_b = sum(r["기초"] or 0 for r in tb if _norm(r["대분류"]) in owned)
    exp_e = sum(r["기말"] or 0 for r in tb if _norm(r["대분류"]) in owned)
    bcol = column_index_from_string(b["columns"]["기초"])
    ecol = column_index_from_string(b["columns"]["기말"])
    got_b = got_e = 0
    label = _norm(b.get("subtotal", {}).get("label", "합계계소계총계"))
    for r in range(ds, de + 1):
        nm = _norm(ws.cell(r, name_ci).value)
        if not nm or nm in ("합계", "소계", "계", "총계", "Total".lower(), _norm(label)):
            continue
        vb = ws.cell(r, bcol).value; ve = ws.cell(r, ecol).value
        if isinstance(vb, (int, float)): got_b += vb
        if isinstance(ve, (int, float)): got_e += ve
    tie = 100 if (abs(got_b - exp_b) <= 1 and abs(got_e - exp_e) <= 1) else 0
    avg = round((form + fit + tie) / 3, 1)
    return {"code": code, "FORM": form, "FIT": fit, "TIE": tie, "avg": avg,
            "detail": f"form_issues={len(form_issues)} overflow={overflow} "
                      f"기초 {got_b:,}/{exp_b:,} 기말 {got_e:,}/{exp_e:,}"}


def main(settlement=None, outdir=None):
    base = Path(__file__).resolve().parent.parent
    S = settlement or str(next(p for p in base.glob("_internal/data/**/*정산표*.xls*")
                                if "~$" not in str(p)))
    cfgdir = str(base / "_internal/config")
    tmpldir = str(base / "_internal/양식")
    outdir = outdir or str(base / "_internal/output/평가")
    tb = parse_trial_balance(S)
    rows = [score_one(c, f, t, S, cfgdir, tmpldir, outdir, tb) for c, f, t in REGISTRY]
    print(f"{'조서':6}{'FORM':>6}{'FIT':>6}{'TIE':>6}{'평균':>7}  진단")
    for r in rows:
        print(f"{r['code']:6}{r['FORM']:>6}{r['FIT']:>6}{r['TIE']:>6}{r['avg']:>7}  {r['detail']}")
    total = round(sum(r["avg"] for r in rows) / len(rows), 1)
    print(f"\n전체 평균: {total}")
    return total


if __name__ == "__main__":
    main()
