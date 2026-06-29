# -*- coding: utf-8 -*-
"""BBDD(장단기차입금) 완성본 조서 생성 — 다중시트 연쇄 graft 파이프라인.

완성본(보조시트·양식컨트롤·매크로 보존)을 템플릿으로, 5개 시트를 채워 무손실 이식한다:
  100_총괄표(1.총괄표 leaf refill + 2.Recap + 이자비용) / 200_차입금 Test(조회박스+좌측+담보) /
  300_미지급이자 / 400_주석 / 조회서_기타조회서류.

데이터 소스(이상적 중간다리 원칙):
  · 전자조회서(confirm)  → parse_bank/loans/collateral → 조회 매핑(200·300·400·조회서기타)
  · 거래처원장(ledger)   → cached_ideal_ledger → borrowings_by_inst → 잔액(200 좌측·100 RECAP)
  · 정산표(settlement)   → parse_trial_balance → 1.총괄표 leaf(refill)
  · 분개장(journal)      → cached_ideal_journal → 이자비용

g_pipeline.py 미러. 부분 실패 허용(_safe): 한 시트가 실패해도 나머지는 생성.
"""

import warnings
from pathlib import Path

import openpyxl
import yaml

warnings.simplefilter("ignore")


def _load_cfg(config_dir):
    cd = Path(config_dir)
    detail = yaml.safe_load((cd / "bbdd_detail.yaml").read_text(encoding="utf-8"))
    from generators.base import apply_col_offset
    leaf = apply_col_offset(yaml.safe_load((cd / "bbdd.yaml").read_text(encoding="utf-8")))
    return detail, leaf


def build_bbdd(template, *, confirm=None, ledger=None, settlement=None, journal=None,
               config_dir, parsed_dir, output, params=None):
    """BBDD 완성본 조서를 생성한다. 가용한 소스만으로 채우고 나머지는 비운다(부분 실패 허용).

    Returns: {"output": 경로, "sheets": {시트키: 결과/에러}, "warnings": [...]}
    """
    detail, leaf = _load_cfg(config_dir)
    warns, results = [], {}

    def _safe(name, fn, default=None):
        try:
            return fn()
        except Exception as e:                       # noqa: BLE001 (견고성 — 절대 전체실패 금지)
            warns.append(f"{name}: {type(e).__name__}: {e}")
            return default

    # ── 소스 파싱(이상적 중간다리) ──
    from extractors.bank import parse_bank, parse_bank_loans, parse_bank_collateral
    from extractors.borrowings import borrowings_by_inst
    from extractors.trial_balance import parse_trial_balance
    from generators.a300 import build_a300_data
    from cache import cached_ideal_ledger, cached_ideal_journal

    loans = _safe("대출(2-2)", lambda: parse_bank_loans(confirm), []) if confirm else []
    coll = _safe("담보(9)", lambda: parse_bank_collateral(confirm), []) if confirm else []
    a300 = _safe("조회기타", lambda: build_a300_data(parse_bank(confirm), loans, coll),
                 {"pension": [], "loans": [], "collateral": []}) if confirm else None
    borrow = _safe("잔액(원장)", lambda: borrowings_by_inst(
        cached_ideal_ledger(ledger, parsed_dir, config_dir)), []) if ledger else []
    tb = _safe("정산표", lambda: parse_trial_balance(settlement), []) if settlement else []
    jr = _safe("분개장(이상)", lambda: cached_ideal_journal(journal, parsed_dir), []) if journal else []

    # ── 시트별 1회용 추출·채움 ──
    from generators.sheet_surgery import extract_light, graft_sheet
    from generators.refill import fill_refill
    from generators import bbdd_detail as B

    tmp = Path(parsed_dir) / "_bbdd_tmp"
    tmp.mkdir(parents=True, exist_ok=True)
    names = openpyxl.load_workbook(template, read_only=True).sheetnames
    SHEETS = {"100": names[8], "200": names[9], "300": names[10],
              "400": names[11], "iq": names[12]}

    def _light(key, fn):
        sn = SHEETS[key]
        p = str(tmp / f"{key}.xlsx")
        _safe(f"{sn} 추출", lambda: extract_light(template, sn, p))
        if not Path(p).exists():
            return None
        wb = openpyxl.load_workbook(p)
        ws = wb[sn]
        results[key] = _safe(f"{sn} 채움", lambda: fn(ws))
        wb.save(p)
        return p

    # 100은 leaf refill(파일경로 기반) + 후처리(ws) 2단계 → 별도 처리
    p100 = str(tmp / "100.xlsx")
    _safe("100 추출", lambda: extract_light(template, SHEETS["100"], p100))
    if Path(p100).exists():
        if tb:
            _safe("100 leaf refill", lambda: fill_refill(p100, tb, p100, leaf))
        wb = openpyxl.load_workbook(p100)
        ws = wb[SHEETS["100"]]
        results["100_recap"] = _safe("100 RECAP", lambda: B.fill_bbdd100_recap(ws, detail, borrow))
        results["100_interest"] = _safe("100 이자비용", lambda: B.fill_bbdd100_interest(ws, detail, jr))
        wb.save(p100)
        if params:                                   # 상단 헤더(회사명/작성자/검토자/날짜)+결론
            from pipeline import _apply_header_conclusion
            _safe("100 헤더", lambda: _apply_header_conclusion(
                p100, SHEETS["100"], leaf, params, has_adjust=False))
    else:
        p100 = None

    p200 = _light("200", lambda ws: B.fill_bbdd200(ws, detail, loans, collateral=coll, balances=borrow))
    p300 = _light("300", lambda ws: B.fill_bbdd300(ws, detail, loans))
    p400 = _light("400", lambda ws: B.fill_bbdd400(ws, detail, loans))
    piq = _light("iq", lambda ws: B.fill_bbdd_inquiry(ws, detail, a300 or {})) if a300 is not None else None

    # ── 연쇄 graft(존재 시트만) ──
    seq = [("100", p100), ("200", p200), ("300", p300), ("400", p400), ("iq", piq)]
    seq = [(k, p) for k, p in seq if p]
    cur = str(template)
    for i, (k, p) in enumerate(seq):
        nxt = str(output) if i == len(seq) - 1 else str(tmp / f"chain{i}.xlsx")
        if _safe(f"{SHEETS[k]} graft", lambda c=cur, pp=p, kk=k, nn=nxt:
                 graft_sheet(c, pp, SHEETS[kk], nn) or nn):
            cur = nxt
    return {"output": cur, "sheets": results, "warnings": warns}
