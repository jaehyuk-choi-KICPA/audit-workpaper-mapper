"""A-1 조서 일괄 생성 파이프라인.

입력 3종(발송 Control Sheet · 거래처원장 · 조회서 취합엑셀)을 파싱해
하나의 작업용 템플릿에 A-100/A-200/A-300 세 시트를 모두 채워 A-1을 완성한다.

체이닝 방식: 템플릿을 복사한 작업 파일에 세 생성기를 순차 적용한다.
각 생성기는 자기 시트만 수정하므로 다른 시트는 보존된다.

무시(감사인 수행영역): FN·월보 시트, 은행연합회 자료 열, 우편조회 회신금액.
"""

import os
import re
import shutil
import warnings
from pathlib import Path

from cache import cached_ideal_ledger, load_with_cache
from extractors import (
    parse_cs,
    parse_bank, parse_bank_loans, parse_bank_collateral,
    parse_insurance, parse_invest, parse_invest_eval,
)
from generators import (
    A100Generator, A200Generator, build_a200_data,
    A300Generator, build_a300_data,
    ControlSheetGenerator, build_ref_map,
)

# 별도정산표/수정사항 파서 버전 — 파서 코드가 바뀌면 올려 해시캐시를 무효화한다(소스 해시만으론
# 못 잡음). build_a0·build_special·후처리 모두 이 상수를 써야 한다(하드코딩 금지 — 버전 드리프트로
# 본문은 신캐시·후처리는 구캐시를 읽어 매칭이 어긋난 사건).
_PARSER_VER = "v15"


class RunReport:
    """실행 진단 리포트. 입력/생성/출력/확인필요 단계별 항목을 모은다.

    엑셀을 못 보는 사용자가 '무엇이 입력값 문제이고 무엇이 출력 양식 문제인지' 알도록,
    그리고 일부가 실패해도 어디가 비었는지 알도록 한다 (절대 전체실패 금지 + 진단).
    """

    def __init__(self):
        self.entries: list[tuple] = []   # (stage, level, msg)  level: ok|warn|error

    def add(self, stage, level, msg):
        self.entries.append((stage, level, str(msg)))

    @property
    def has_error(self):
        return any(lv == "error" for _, lv, _ in self.entries)

    @property
    def has_warn(self):
        return any(lv in ("warn", "error") for _, lv, _ in self.entries)

    def render(self) -> str:
        icon = {"ok": "✓", "warn": "△", "error": "✗"}
        if self.has_error:
            summary = "부분 출력 — 일부 자료가 비었습니다([입력]의 ✗ 확인). 나머지는 정상 생성됨"
        elif self.has_warn:
            summary = "정상 출력 — 단, [확인필요]/빈 자료(△) 항목은 감사인 검토 필요"
        else:
            summary = "정상 출력"
        lines = ["=" * 60, f"  실행 리포트 — {summary}", "=" * 60]
        for stage in ("입력", "생성", "출력", "확인필요"):
            rows = [(lv, m) for st, lv, m in self.entries if st == stage]
            if not rows:
                continue
            lines.append(f"[{stage}]")
            for lv, m in rows:
                lines.append(f"  {icon.get(lv, '·')} {m}")
        return "\n".join(lines)


def build_a1(*, control_sheet, ledger_src, confirm_xlsx, template, config_dir,
             output, params, parsed_dir=None, ref_dir=None):
    """A-1 조서(A-100/200/300)를 한 파일로 생성한다.

    Args:
        control_sheet: 발송 Control Sheet 경로
        ledger_src:    잔액 소스 — 거래처원장(.xls/.xlsx) 또는 결산보고서(.xlsx).
                       파일명에 '결산보고서'가 있으면 결산보고서 변환 경로로 자동 라우팅.
        confirm_xlsx:  조회서 취합엑셀 경로 (BANK/INSURANCE/INVESTMENT)
        template:      작업용 템플릿(.xlsx, A100/A200/A300 시트 포함)
        config_dir:    셀 매핑 YAML 디렉터리 (a100/a200/a300/settlement_report.yaml)
        output:        결과 파일 경로
        params:        공통 헤더 {회사명, 날짜, preparer, reviewer} + 선택 항목
        parsed_dir:    변환자료 폴더 (이상적 양식 캐시 저장/재사용; None이면 출력 폴더 옆)

    Returns:
        RunReport (단계별 진단). 결과 파일은 output 경로에 생성됨(부분 실패해도 항상 생성).
    """
    cfg = Path(config_dir)
    out = Path(output)
    parsed = parsed_dir or str(out.parent / "변환자료")
    report = RunReport()

    # C-2차 학습 동의어 주입(오프라인·무료): 과거 해소된 헤더 매핑을 이번 실행의 헤더 인식에
    # 자동 적용한다. 캐시가 비었으면 무영향(기존 동작과 동일). LLM/키 불필요.
    try:
        from format_adapt import LearnedHeaders
        from extractors._headers import set_learned
        _lh = LearnedHeaders(parsed)
        for _pk in ("ledger", "insurance"):
            syn = _lh.synonyms_for(_pk)
            if syn:
                set_learned(_pk, syn)
                report.add("입력", "ok", f"학습 동의어 적용({_pk}): {len(syn)}건")
    except Exception:
        pass  # 학습 캐시는 보조 — 실패해도 본 흐름 막지 않음

    # 회복력: 회사마다 입력 구조가 달라 일부 파서가 실패할 수 있다. 한 파서 실패가
    # 전체 A-1 생성을 막지 않도록 감싸서, 해당 부분만 비우고 진단을 모은다(부분 출력 보장).
    def _safe_parse(label, fn, *args, default=None, **kwargs):
        try:
            with warnings.catch_warnings(record=True) as caught:
                warnings.simplefilter("always")
                res = fn(*args, **kwargs)
            for w in caught:                       # 파서 내부 경고(예: 잔액불일치) 포착
                m = str(w.message)
                if "default style" in m:
                    continue
                report.add("입력", "warn", f"{label}: {m.splitlines()[0][:160]}")
            n = len(res) if hasattr(res, "__len__") else "?"
            report.add("입력", "ok" if res else "warn",
                       f"{label}: {n}건" + ("" if res else " (빈 값 — 해당 부분 비움)"))
            return res
        except Exception as e:
            report.add("입력", "error", f"{label} 처리 실패({type(e).__name__}: {e}) → 해당 부분 비움")
            return [] if default is None else default

    # 소스 파일 해시 기반 파싱 캐시(변환자료에 저장 → 동일 입력 재실행 시 재파싱 생략).
    # 취합엑셀 하나를 6개 파서가 읽으므로 tag로 키 충돌을 막는다. 거래처원장은
    # cached_ideal_ledger가 이미 변환자료에 캐시하므로 여기서는 confirm/CS만 캐시.
    def _cached(parse_fn, src, tag):
        return load_with_cache(src, parse_fn, cache_dir=parsed, tag=tag)

    # ---- STEP 1. 데이터 추출 (전부 안전 래핑) ----
    cs_rows = _safe_parse("발송 Control Sheet", _cached, parse_cs, control_sheet, "cs")
    ledger_rows = _safe_parse("거래처원장/결산보고서", cached_ideal_ledger, ledger_src, parsed, str(cfg))
    bank_rows = _safe_parse("BANK(현금성/퇴직)", _cached, parse_bank, confirm_xlsx, "bank")
    ins_groups = _safe_parse("INSURANCE(장기금융)", _cached, parse_insurance, confirm_xlsx, "insurance")
    invest_rows = _safe_parse("INVESTMENT(단기금융)", _cached, parse_invest, confirm_xlsx, "invest")
    invest_eval = _safe_parse("INVESTMENT 2번(평가액)", _cached, parse_invest_eval, confirm_xlsx, "invest_eval")
    loan_rows = _safe_parse("대출(BANK 2-2)", _cached, parse_bank_loans, confirm_xlsx, "loans")
    coll_rows = _safe_parse("담보(BANK 9)", _cached, parse_bank_collateral, confirm_xlsx, "collateral")
    ref_map = _safe_parse("취합 요약(조서번호)", _cached, build_ref_map, confirm_xlsx, "refmap", default={})

    # 외화 환율 메모(참고자료) 로드 — 없으면 빈 dict(=환산 비활성, KRW 그대로)
    fx_rates = {}
    if ref_dir:
        from extractors import load_fx_rates
        fx_rates = _safe_parse("기말환율 메모(참고자료)", load_fx_rates, ref_dir, default={}) or {}

    # ---- STEP 2~4. 변환 ----
    a200_data = build_a200_data(ledger_rows, bank_rows, ins_groups, invest_rows,
                                fx_rates=fx_rates, invest_eval=invest_eval)
    a300_data = build_a300_data(bank_rows, loan_rows, coll_rows)

    # 외화 기말환산 불일치 집계(진단)
    fx_diff = sum(1 for r in a200_data["cash"] if r.get("_기말환산불일치"))
    if fx_diff:
        report.add("확인필요", "warn",
                   f"외화 기말환산 불일치 {fx_diff}건(환율×조회서외화값 ≠ 명세서 원화) — A-0에서 조정")

    # 확인필요(감사인 영역) 집계
    cash_review = sum(1 for r in a200_data["cash"] if r["_확인필요"])
    lt_review = sum(1 for r in a200_data["longterm"] if r["_확인필요"])
    if cash_review or lt_review:
        report.add("확인필요", "warn",
                   f"A200 금액 불일치/우편 등 확인필요: 현금성 {cash_review}건, 장기금융 {lt_review}건")

    # ---- STEP 3. 조서 생성 (시트별 순차, 한 시트 실패가 나머지를 막지 않음) ----
    work = str(out)
    shutil.copy(template, work)

    def _safe_gen(sheet, gen_cls, yaml_name, fill_args):
        try:
            g = gen_cls(work, str(cfg / yaml_name))
            g.fill(*fill_args)
            g.save(work)
            report.add("생성", "ok", f"{sheet} 생성")
        except Exception as e:
            report.add("생성", "error", f"{sheet} 생성 실패({type(e).__name__}: {e}) → 이 시트는 양식 기본값")

    _safe_gen("A-100", A100Generator, "a100.yaml", (cs_rows, params))
    _safe_gen("A-200", A200Generator, "a200.yaml", (a200_data, params))
    _safe_gen("A-300", A300Generator, "a300.yaml", (a300_data, params))
    _safe_gen("Control Sheet", ControlSheetGenerator, "control_sheet.yaml",
              (control_sheet, ref_map, params))

    # ---- 양식무결성 HOOK(시각 확인 대체): 생성 후 격자·컬럼·제목 검사 ----
    try:
        for issue in _check_form_integrity(work, n_cs=len(cs_rows), config_dir=config_dir):
            report.add("출력", "warn", issue)
    except Exception as e:
        report.add("출력", "warn", f"양식무결성 검사 자체 실패({type(e).__name__}: {e})")

    return report


def _check_form_integrity(path, *, n_cs: int, config_dir: str) -> list:
    """생성된 A-1의 양식 무결성(격자·컬럼잘림·제목/헤더)을 검사하는 표준 게이트.

    양식 붕괴는 조서 신뢰성 훼손 → 데이터 채우기 후 반드시 이 검사를 거친다.
    config_dir은 build_a1가 받은 값을 그대로 사용(.exe frozen 모드 경로 안전).
    """
    from validator import (validate_grid, validate_a200_form,
                           check_column_fit, check_cells_present,
                           check_outside_col_fill)
    import yaml

    cfg = Path(config_dir)
    out = []
    # Control Sheet: 격자 + 컬럼 잘림 + 제목·헤더 존재 (시각 확인 대체)
    cs_cfg = yaml.safe_load((cfg / "control_sheet.yaml").read_text(encoding="utf-8"))
    cs_sheet = cs_cfg["sheet"]
    cs_start, cs_end = cs_cfg["data_start_row"], cs_cfg["data_start_row"] + n_cs - 1
    out += validate_grid(path, cs_sheet, header_row=3, data_start=cs_cfg["data_start_row"],
                         n_rows=n_cs, cols=list(range(1, 11)))
    out += check_column_fit(path, cs_sheet, cols=list(range(1, 11)),
                            data_start=cs_start, data_end=cs_end, header_row=3)
    out += check_cells_present(path, cs_sheet, [cs_cfg.get("title_cell", "A1"), "B3", "E3", "G3"])
    # A100 바깥열 회색 음영 보존 검사
    a1_cfg = yaml.safe_load((cfg / "a100.yaml").read_text(encoding="utf-8"))
    outside_a1 = a1_cfg.get("layout", {}).get("outside_gray_cols", [])
    if outside_a1:
        out += check_outside_col_fill(path, a1_cfg["sheet"], outside_gray_cols=outside_a1)
    # A200 전용(소계·수식·테두리) + 바깥열 회색 음영 보존 검사
    a2 = yaml.safe_load((cfg / "a200.yaml").read_text(encoding="utf-8"))
    out += validate_a200_form(path, a2["sheet"], a2)
    outside_a2 = a2.get("layout", {}).get("outside_gray_cols", [])
    if outside_a2:
        out += check_outside_col_fill(path, a2["sheet"], outside_gray_cols=outside_a2)
    return out


def build_a0(*, settlement, template, config_dir, output, params=None, parsed_dir=None,
             config_file="a0.yaml"):
    """A-0 총괄표(본문 + 수정사항)를 생성한다.

    정산표(별도정산표·수정사항집계)를 단일 소스로 총괄표 본문(기초/기말/수정사항/수정후)과
    '4.수정사항' 분개표를 채운다. 나머지 섹션은 손대지 않는다(감사인/후속).

    Args:
        settlement: 정산표 파일 경로(별도정산표·수정사항집계 시트 포함)
        template:   A-0 작업용 템플릿(.xlsx, '4000_A000 총괄표' 시트)
        config_dir: 셀 매핑 YAML 디렉터리(a0.yaml)
        output:     결과 파일 경로
        params:     공통 헤더(선택; 현재 본문/수정사항만 채우므로 미사용 가능)
        parsed_dir: 변환자료 폴더(파싱 해시캐시; None이면 출력 폴더 옆)

    Returns:
        RunReport (부분 실패 허용 — 항상 출력 파일 생성).
    """
    from extractors import parse_trial_balance, parse_adjustments
    from generators import A0Generator

    cfg = Path(config_dir)
    out = Path(output)
    parsed = parsed_dir or str(out.parent / "변환자료")
    report = RunReport()

    # 캐시 태그에 파서 버전을 포함 → 파서 코드가 바뀌면 캐시가 자동 무효화(소스 해시만으론 못 잡음).


    def _cached(fn, src, tag):
        return load_with_cache(src, fn, cache_dir=parsed, tag=f"{tag}_{_PARSER_VER}")

    # ---- STEP 1. 정산표 추출 (안전 래핑 + 해시캐시) ----
    try:
        tb_rows = _cached(parse_trial_balance, settlement, "trial_balance")
        report.add("입력", "ok" if tb_rows else "warn",
                   f"별도정산표: {len(tb_rows)}행" + ("" if tb_rows else " (빈 값 — 헤더 미인식?)"))
    except Exception as e:
        tb_rows = []
        report.add("입력", "error", f"별도정산표 처리 실패({type(e).__name__}: {e})")
    try:
        adj = _cached(parse_adjustments, settlement, "adjustments")
        report.add("입력", "ok" if adj else "warn",
                   f"수정사항집계: {len(adj)} entries" + ("" if adj else " (없음)"))
    except Exception as e:
        adj = []
        report.add("입력", "error", f"수정사항집계 처리 실패({type(e).__name__}: {e})")

    # ---- STEP 2~3. 생성 ----
    work = str(out)
    shutil.copy(template, work)
    info = {}
    try:
        g = A0Generator(work, str(cfg / config_file))
        info = g.fill(tb_rows, adj, params or {})
        g.save(work)
        n_body = info["body_data_end"] - info["body_data_start"] + 1
        report.add("생성", "ok",
                   f"총괄표 생성 (본문 {n_body}행, 수정사항 entry {info['n_adj_entries']}건)")
    except Exception as e:
        report.add("생성", "error", f"총괄표 생성 실패({type(e).__name__}: {e}) → 양식 기본값")

    # ---- 양식 무결성 HOOK ----
    if info:
        try:
            for issue in _check_a0_form(work, info, str(cfg), config_file, template=template):
                report.add("출력", "warn", issue)
        except Exception as e:
            report.add("출력", "warn", f"양식무결성 검사 자체 실패({type(e).__name__}: {e})")

    return report


def _check_a0_form(path, info: dict, config_dir: str, config_file: str = "a0.yaml",
                   template: str = None) -> list:
    """총괄표 양식 무결성 게이트(격자·열잘림·헤더존재·헤더채움색) — config 주도로 조서 무관."""
    from validator import (validate_grid, check_column_fit, check_cells_present,
                           check_header_fill)
    from openpyxl.utils import column_index_from_string
    import yaml

    from generators.base import apply_col_offset
    cfg = apply_col_offset(yaml.safe_load((Path(config_dir) / config_file).read_text(encoding="utf-8")))
    sheet = cfg["sheet"]
    b = cfg["body"]
    cols = b["columns"]
    ds, de = info["body_data_start"], info["body_data_end"]
    hr = b["header_row"]
    # 리터럴 값 열(계정명/기초/기말/수정사항)만 너비검사(수식열은 텍스트길이 오탐 제외)
    lit = [cols[k] for k in ("계정명", "기초", "기말", "수정사항") if k in cols]
    lit_ci = [column_index_from_string(c) for c in lit]
    # 격자검사 대상 = 채우는 열(num_cols) − grid_skip(템플릿상 의도적으로 비고 테두리 없는 이동열,
    #   예: G 유형자산의 취득·처분·감가상각 컬럼). skip은 빈 이동컬럼 오탐 방지.
    skip = set(b.get("grid_skip", []))
    num = [c for c in b.get("num_cols", []) if c not in skip]
    grid_ci = sorted(set(lit_ci + [column_index_from_string(c) for c in num]))
    out = []
    out += validate_grid(path, sheet, header_row=hr, data_start=ds,
                         n_rows=de - ds + 1, cols=grid_ci)
    out += check_column_fit(path, sheet, cols=lit_ci, data_start=ds, data_end=de, header_row=hr)
    out += check_cells_present(path, sheet,
                               [f"{cols['계정명']}{hr}", f"{cols['기초']}{hr}"])
    # 헤더 채움색 보존 검사(소실·파란색 변질 방지) — 템플릿과 대조
    if template:
        out += check_header_fill(path, template, sheet, header_row=hr, cols=grid_ci)
    return out


def _account_nature(rec: dict) -> str:
    """계정 성격(정산표 섹션 flag 기반). 미매핑 리포트 표시·귀속제안용."""
    for flag, label in (("부채", "부채"), ("자본", "자본"), ("매출", "수익(매출)"),
                        ("매출원가", "매출원가"), ("판관비", "판관비"),
                        ("영업외수익", "영업외수익"), ("영업외비용", "영업외비용"),
                        ("재고", "재고자산"), ("유형자산", "유형자산"), ("무형자산", "무형자산")):
        if rec.get(flag):
            return label
    return "자산/기타"


# 귀속제안 = 규칙(키워드) 기반 — 오프라인·결정적(LLM/임베딩 불필요; 프로젝트 원칙 일치).
_SUGGEST_RULES = [
    (("차입금", "사채", "유동성장기", "전환사채"), "BBDD(차입금)"),
    (("매입채무", "외상매입금", "지급어음"), "AA(매입채무)"),
    (("퇴직", "연금"), "EE(퇴직급여)"),
    (("법인세",), "CC/법인세"),
    (("충당", "미지급", "예수", "선수", "보증", "이연"), "CC(기타부채)"),
    (("매출채권", "외상매출", "받을어음", "공사미수"), "C(매출채권)"),
    (("미수", "선급", "대급금", "선납", "예치", "회원권"), "D(기타자산)"),
    (("현금", "예금", "금융상품", "대여금"), "A-0(현금·금융)"),
    (("매도가능", "지분법"), "B(유가증권)"),
    (("상품", "제품", "원재료", "재공", "재고"), "E(재고자산)"),
    (("자본금", "주식발행", "자기주식", "잉여금", "준비금", "평가이익", "평가손실", "재평가"), "GG(자본)"),
    (("감가상각", "상각", "무형"), "G(유무형자산)"),
]
_NATURE_FALLBACK = {"부채": "CC(기타부채)", "자본": "GG(자본)", "수익(매출)": "P(매출)",
                    "매출원가": "Q(매출원가)", "판관비": "R(판관비)", "영업외수익": "S(기타손익)",
                    "영업외비용": "S(기타손익)", "재고자산": "E(재고자산)",
                    "유형자산": "G(유무형자산)", "무형자산": "G(유무형자산)"}


def _suggest_owner(name: str, nature: str) -> str:
    n = re.sub(r"\s+", "", str(name or ""))
    for kws, code in _SUGGEST_RULES:
        if any(k in n for k in kws):
            return code
    return _NATURE_FALLBACK.get(nature, "수동확인")


def routing_completeness(settlement, config_dir, config_files, *, parsed_dir=None):
    """정산표 전 대분류가 조서 config들에 1:1로 소유되는지 검사(중복·누락 안전망).

    Returns: {"unmapped":[대분류...], "duplicate":{대분류:[조서...]}, "owned":N}
    """
    from extractors import parse_trial_balance
    from generators.adjust import owned_tokens, _norm
    import yaml

    tb = parse_trial_balance(settlement)
    # 조서별 소유 토큰을 adjust.owned_tokens로 통일 산출(엔진 groups/groups_flag+exclude_kw,
    # refill/sales name_map 규칙·키워드, owns_flag(s), bs_rows, rows, owns 전부 반영).
    # → 규칙·flag로 실제 렌더/흡수되는 계정이 '미매핑'으로 오표시되던 가짜 양성 제거.
    cfgs = []
    for cf in config_files:
        cfg = yaml.safe_load((Path(config_dir) / cf).read_text(encoding="utf-8"))
        cfgs.append((cfg.get("sheet", cf), owned_tokens(cfg, tb)))

    def _owners(cls: str) -> list:
        cn = _norm(cls)
        out = []
        for sheet, toks in cfgs:
            if cn in toks["exact"] or any(len(k) >= 2 and k in cn for k in toks["kw"]):
                out.append(sheet)
        return sorted(set(out))

    owned = {r["대분류"]: _owners(r["대분류"]) for r in tb}
    present = {r["대분류"] for r in tb
              if any(r.get(k) not in (None, 0) for k in ("기초", "기말", "수정후"))}
    # 제조원가명세서 전용 대분류(그 대분류의 모든 행이 제조원가=True)는 완전성 체크 제외 —
    # Q가 '당기총공사비용' 총계로 흡수하므로 개별 매핑 불필요(사용자 지시: 원가명세서쪽 미매핑 무시).
    mfg_only = {c for c in present if all(r.get("제조원가") for r in tb if r["대분류"] == c)}
    present -= mfg_only
    unmapped = sorted(c for c in present if not owned.get(c))
    # 미매핑 상세: 금액(수정후 우선, 없으면 기말 합)·성격·귀속제안(규칙기반). 리포트에 그대로 실음.
    first = {}
    amt = {}
    for r in tb:
        c = r["대분류"]
        if c not in unmapped:
            continue
        first.setdefault(c, r)
        amt[c] = amt.get(c, 0) + (r.get("수정후") if r.get("수정후") is not None else (r.get("기말") or 0))
    detail = []
    for c in unmapped:
        nat = _account_nature(first.get(c, {}))
        detail.append({"계정": c, "금액": amt.get(c, 0), "성격": nat,
                       "제안": _suggest_owner(c, nat)})
    return {
        "unmapped": unmapped,
        "unmapped_detail": detail,
        "duplicate": {c: owned[c] for c in present if len(owned.get(c, [])) > 1},
        "owned": sum(1 for c in present if owned.get(c)),
    }


def build_special(*, kind, settlement, template, config_dir, config_file, output,
                  parsed_dir=None):
    """특수 총괄표(고정격자 리필·자본·매출원가·매출)를 생성한다.

    엔진(a0.py 섹션 렌더)으로 안 맞는 조서용. kind별 fill 함수로 분기한다:
      refill  → generators.refill.fill_refill   (R-2 인건비 등 고정 계산행)
      capital → generators.capital.fill_capital (GG 자본)
      cogs    → generators.sales.fill_cogs       (Q 매출원가, 공사원가 롤포워드+상품매출원가)
      sales   → generators.sales.fill_sales      (P 매출, IS 동적+BS 교차참조)
    정산표 파싱은 build_a0과 같은 해시캐시를 공유한다(재파싱 0). 생성 후 경량 양식게이트.
    """
    import yaml
    from extractors import parse_trial_balance, parse_adjustments

    cfg_dir = Path(config_dir)
    out = Path(output)
    parsed = parsed_dir or str(out.parent / "변환자료")
    report = RunReport()


    def _cached(fn, src, tag):
        return load_with_cache(src, fn, cache_dir=parsed, tag=f"{tag}_{_PARSER_VER}")

    try:
        tb_rows = _cached(parse_trial_balance, settlement, "trial_balance")
        report.add("입력", "ok" if tb_rows else "warn", f"별도정산표: {len(tb_rows)}행")
    except Exception as e:
        tb_rows = []
        report.add("입력", "error", f"별도정산표 처리 실패({type(e).__name__}: {e})")

    from generators.base import apply_col_offset
    cfg = apply_col_offset(yaml.safe_load((cfg_dir / config_file).read_text(encoding="utf-8")))
    work = str(out)
    shutil.copy(template, work)
    try:
        if kind == "refill":
            from generators.refill import fill_refill
            res = fill_refill(work, tb_rows, work, cfg)
        elif kind == "capital":
            from generators.capital import fill_capital
            adj = _cached(parse_adjustments, settlement, "adjustments")
            res = fill_capital(work, tb_rows, adj, work, cfg)
        elif kind == "cogs":
            from generators.sales import fill_cogs
            res = fill_cogs(work, tb_rows, work, cfg)
        elif kind == "sales":
            from generators.sales import fill_sales
            res = fill_sales(work, tb_rows, work, cfg)
        elif kind == "retirement":
            from generators.sales import fill_retirement
            res = fill_retirement(work, tb_rows, work, cfg)
        elif kind == "rows":
            from generators.refill import fill_rows
            res = fill_rows(work, tb_rows, work, cfg)
        else:
            raise ValueError(f"알 수 없는 kind: {kind}")
        report.add("생성", "ok", f"{cfg.get('sheet')} 생성 {res}")
    except Exception as e:
        report.add("생성", "error", f"생성 실패({type(e).__name__}: {e}) → 양식 기본값")

    # 경량 양식게이트(고정격자라 테두리는 보존됨 → 열 잘림 위주). 값열을 너비검사.
    try:
        for issue in _check_special_form(work, cfg):
            report.add("출력", "warn", issue)
    except Exception as e:
        report.add("출력", "warn", f"양식무결성 검사 실패({type(e).__name__}: {e})")
    return report


def _check_special_form(path, cfg: dict) -> list:
    """특수조서 경량 양식게이트: 숫자 값열(기초·기말·수정)의 ####### 위험만 검사한다.

    고정격자 템플릿이라 테두리·헤더는 in-place로 보존된다. 유일한 실위험은 큰 금액이
    열 너비를 넘어 #######로 잘리는 것. 수식/대사설명 텍스트는 표시값이 다르므로 제외
    (그 길이로 너비를 재면 오탐 → 사용자 불필요 크로스체크). 실제 숫자 셀만 잰다.
    """
    import openpyxl
    from openpyxl.utils import column_index_from_string

    sheet = cfg["sheet"]
    val_cols = [cfg[k] for k in ("base_col", "end_col", "adj_col") if cfg.get(k)]
    for spec in cfg.get("rows", []):            # 고정 다중행(B): 행별 컬럼
        val_cols += [spec[k] for k in ("base_col", "end_col", "adj_col") if spec.get(k)]
    val_cols = list(dict.fromkeys(val_cols))
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet]
    issues = []
    for col in val_cols:
        ci = column_index_from_string(col)
        maxlen = max((len(f"{v:,.0f}") for r in range(1, ws.max_row + 1)
                      if isinstance((v := ws.cell(r, ci).value), (int, float))
                      and not isinstance(v, bool)), default=0)
        if maxlen == 0:
            continue
        width = ws.column_dimensions[col].width or 8.43
        if width + 1 < maxlen:
            issues.append(f"[너비] {sheet} {col}열: 너비 {width:.0f} < 숫자 {maxlen}자 (####### 위험)")
    wb.close()
    return issues


def _coerce_date(v):
    """입력 날짜 문자열을 datetime으로(셀 날짜서식 유지). 파싱 실패 시 원문 문자열."""
    if hasattr(v, "year"):           # 이미 date/datetime
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            from datetime import datetime
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return s


def _apply_header_conclusion(path, sheet, cfg, params, has_adjust):
    """생성본(gen, 1회용)에 상단 헤더(회사명/Preparer/Reviewer/날짜)와 결론 문구를 쓴다.

    회사명은 'label: 값' 결합형이라 라벨을 보존하며 값만 교체. 결론은 수정사항 유무로 하드코딩
    ('수정사항 제외하고 특이사항 없음' / '특이사항 없음'). openpyxl로 gen(1회용)에만 적용 →
    완성본 컨트롤엔 영향 없음(이식으로 총괄표만 옮겨감)."""
    import openpyxl
    hc = cfg.get("header_cells") or {}
    concl = cfg.get("conclusion") or {}
    if not hc and not concl:
        return
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.active
    company = (params or {}).get("회사명")
    if hc.get("회사명") and company:
        cell = ws[hc["회사명"]]
        cur = str(cell.value or "")
        prefix = cur.split(":")[0] if ":" in cur else "회사명"
        cell.value = f"{prefix}: {company}"
    if hc.get("preparer") and (params or {}).get("preparer"):
        ws[hc["preparer"]] = params["preparer"]
    if hc.get("reviewer") and (params or {}).get("reviewer"):
        ws[hc["reviewer"]] = params["reviewer"]
    if hc.get("날짜") and (params or {}).get("날짜"):
        ws[hc["날짜"]] = _coerce_date(params["날짜"])
    # 결론: 하드코딩 셀은 분개 행삽입으로 위치가 밀려 어긋난다(각주와 충돌). → **마지막 '결론' 헤더를
    # 동적 탐지**해 그 2행 아래(텍스트 칸)에 쓴다. conclusion 키가 있는 조서만(특수구조 조서 제외 가능).
    if concl:
        text = "수정사항 제외하고 특이사항 없음" if has_adjust else "특이사항 없음"
        last = None
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and "결론" in v and len(v.strip()) <= 10:
                    last = (cell.row, cell.column)      # 마지막(최하단) '결론' 헤더 = 메인 결론
        if last:
            ws.cell(last[0] + 2, last[1]).value = text
    wb.save(path)
    wb.close()


def _light_cached(full, sheet, parsed, code):
    """완성본에서 총괄표만 뽑은 가벼운 생성용 템플릿(변환자료 캐시, 소스 변경 시 자동 갱신)."""
    import hashlib
    from generators.sheet_surgery import extract_light
    st = Path(full).stat()
    h = hashlib.sha1(f"{int(st.st_mtime)}_{st.st_size}_{_PARSER_VER}".encode()).hexdigest()[:10]
    light = str(Path(parsed) / f"__light_{code}_{h}.xlsx")
    if not os.path.exists(light):
        for old in Path(parsed).glob(f"__light_{code}_*.xlsx"):   # 구버전 정리
            try:
                old.unlink()
            except OSError:
                pass
        extract_light(full, sheet, light)
    return light


def build_lead_all(*, settlement, registry, config_dir, template_root, output_dir,
                   params=None, parsed_dir=None, progress=None):
    """레지스트리의 모든 총괄표를 완성본(보조시트·컨트롤 포함)에 채워 일괄 생성한다.

    progress: 조서 1개 끝날 때마다 progress(code, ok:bool) 호출(실시간 진행표시용, 선택).

    조서별 흐름: **완성본에서 총괄표만 가벼운 템플릿으로 추출(_light_cached) → 그 위에 정산표
    데이터 생성(build_a0/build_special) → 수정사항 분개 렌더 → 헤더·결론 작성 → 완성본에
    총괄표 시트만 zip 이식(graft_sheet)**. 완성본은 openpyxl로 절대 열지 않아 보조시트의
    양식컨트롤·매크로·도형이 100% 보존된다.

    정산표/수정사항 파싱은 변환자료 캐시로 공유. 라우팅 완전성·미매핑 분개도 함께 점검.
    Returns: ({code: RunReport}, completeness_dict)
    """
    import yaml
    from extractors import parse_trial_balance, parse_adjustments
    from generators import adjust, sheet_surgery
    from generators.base import apply_col_offset

    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    parsed = parsed_dir or str(out_dir.parent / "변환자료")
    Path(parsed).mkdir(parents=True, exist_ok=True)
    reports = {}
    engine_configs = []

    def _cfg(item):
        return apply_col_offset(yaml.safe_load((Path(config_dir) / item["config"]).read_text(encoding="utf-8")))

    # ---- 분개 매핑 선계산(생성과 무관, 계정 기반) ----
    adj = []
    rel_by_code = {}
    entry_codes = {}
    mapped = set()
    try:
        tb = load_with_cache(settlement, parse_trial_balance, cache_dir=parsed, tag=f"trial_balance_{_PARSER_VER}")
        adj = load_with_cache(settlement, parse_adjustments, cache_dir=parsed, tag=f"adjustments_{_PARSER_VER}")
        for item in registry:
            rel = adjust.related_entries(adj, adjust.owned_tokens(_cfg(item), tb), tb)
            rel_by_code[item["code"]] = rel
            mapped.update(e["no"] for e in rel)
            for e in rel:
                entry_codes.setdefault(e["no"], []).append(item["code"])
    except Exception as e:
        warnings.warn(f"[분개 선계산] 실패: {type(e).__name__}: {e}")

    # ---- 조서별: 추출 → 생성 → 분개 → 헤더/결론 → 이식 ----
    for item in registry:
        code = item["code"]
        kind = item.get("kind", "engine")
        full = str(Path(template_root) / item["template"])
        # 출력 파일명 = 템플릿명에서 'template' 토큰만 제거(예 C_4000_template_매출채권.xlsm → C_4000_매출채권.xlsm)
        out_name = item["template"].replace("_template_", "_").replace("_template", "").replace("template_", "")
        outp = str(out_dir / out_name)                          # 완성본 확장자 유지(.xlsm 매크로 보존)
        engine_configs.append(item["config"])
        cfg = _cfg(item)
        sheet = cfg["sheet"]
        report = RunReport()
        reports[code] = report
        gen = str(out_dir / f"__gen_{code}.xlsx")
        try:
            light = _light_cached(full, sheet, parsed, code)
            if kind == "engine":
                report = build_a0(settlement=settlement, template=light, config_dir=config_dir,
                                  output=gen, params=params, parsed_dir=parsed, config_file=item["config"])
            else:
                report = build_special(kind=kind, settlement=settlement, template=light,
                                       config_dir=config_dir, config_file=item["config"],
                                       output=gen, parsed_dir=parsed)
            reports[code] = report
            # 수정사항 분개(gen에 렌더 — gen은 1회용이라 openpyxl 무방)
            try:
                if code == "GG":
                    ec = {no: [c for c in cs if c != "GG"] for no, cs in entry_codes.items()}
                    n = adjust.render(gen, sheet, adj, mode="refer", entry_codes=ec)
                else:
                    n = adjust.render(gen, sheet, rel_by_code.get(code, []))
                if n:
                    report.add("생성", "ok", f"수정사항 분개 {n}건 반영")
            except Exception as e:
                report.add("출력", "warn", f"수정사항 렌더 실패({type(e).__name__}: {e})")
            # 헤더 + 결론
            try:
                # GG는 전체 분개를 refer로 표시하므로 결론도 '분개 존재 여부'로 판단.
                _has_adj = bool(adj) if code == "GG" else bool(rel_by_code.get(code))
                _apply_header_conclusion(gen, sheet, cfg, params, has_adjust=_has_adj)
            except Exception as e:
                report.add("출력", "warn", f"헤더/결론 작성 실패({type(e).__name__}: {e})")
            # 이식: 완성본에 총괄표 시트만 옮김(컨트롤·매크로 보존)
            sheet_surgery.graft_sheet(full, gen, sheet, outp)
            report.add("출력", "ok", f"완성본 이식 → {Path(outp).name}")
        except Exception as e:
            report.add("생성", "error", f"{code} 생성 실패({type(e).__name__}: {e})")
        finally:
            if os.path.exists(gen):
                try:
                    os.remove(gen)
                except OSError:
                    pass
        if progress:
            try:
                progress(code, not reports[code].has_error)
            except Exception:
                pass

    # 미매핑 분개(어느 조서에도 안 붙음) → 경고
    comp_adj = [(e["no"], [l["계정"] for l in e["lines"]]) for e in adj if e["no"] not in mapped]
    comp = routing_completeness(settlement, config_dir, engine_configs, parsed_dir=parsed)
    comp["unmapped_adjustments"] = comp_adj
    return reports, comp
