"""수정사항(감사 수정분개) 통합 렌더러 — 조서 무관.

수정사항집계 분개를 **그 분개가 건드린 계정을 소유한 조서**의 수정사항 섹션에 그린다.
한 분개가 여러 조서 계정을 건드리면 그 조서들 모두에 등장한다(중복 정상 — 총괄표 본문에
수정사항이 있으면 그 조서엔 당연히 해당 분개가 있다).

매칭 원리(사용자 지적 반영): 별도정산표 수정사항은 수정사항집계 SUMIF로 걸려 있으므로,
분개의 각 계정을 그 계정의 대분류로 환원해 '이 조서가 그 대분류를 소유하는가'로 연결한다.
소유 판정은 config(엔진 groups/groups_flag, refill name_map, sales is_groups/bs_rows, 규칙 키워드)
에서 모은 소유 토큰으로 한다.
"""

import re
from copy import copy

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter

_NUMF = '#,##0;[Red]\\(#,##0\\);"-"'
_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HDR_FILL = PatternFill("solid", fgColor="D9D9D9")


def _norm(s) -> str:
    return re.sub(r"\s+", "", str(s)) if s is not None else ""


def owned_tokens(cfg: dict, tb_rows: list) -> dict:
    """이 조서가 소유한 토큰. {'exact': 대분류/계정명 정확매칭, 'kw': 부분일치 키워드}.
    정확/키워드를 분리해 '광고매출'(매출)이 '서비스및광고매출원가'(매출원가)에 잘못 걸리는 등의
    교차오염을 막는다(정확 토큰은 ==, 키워드는 포함)."""
    exact, kw = set(), set()
    for sec in cfg.get("body", {}).get("sections", []):
        exact.update(_norm(g) for g in sec.get("groups", []))
        flag = sec.get("groups_flag")
        if flag:
            excl = [_norm(x) for x in sec.get("exclude_kw", [])]   # 타 조서 소유분 제외(CC 잔여부채 등)
            for r in tb_rows:
                if not r.get(flag):
                    continue
                kn = _norm(r["대분류"])
                if any(e and e in kn for e in excl):
                    continue
                exact.add(kn)
    # owns_flag(단일)/owns_flags(복수): 그 flag 섹션의 대분류+계정명을 정확 토큰으로(매출·매출원가·원가상세·제조 등)
    ofs = list(cfg.get("owns_flags", []))
    if cfg.get("owns_flag"):
        ofs.append(cfg["owns_flag"])
    for of in ofs:
        for r in tb_rows:
            if r.get(of):
                exact.add(_norm(r["대분류"])); exact.add(_norm(r["계정명"]))
    for v in cfg.get("name_map", {}).values():
        if isinstance(v, str):
            exact.add(_norm(v))
        elif isinstance(v, dict):
            kw.update(_norm(k) for k in v.get("포함", []))
    for g in cfg.get("is_groups", []):
        exact.add(_norm(g))
    for spec in cfg.get("bs_rows", []):
        exact.add(_norm(spec["name"]))
    for spec in cfg.get("rows", []):            # 고정 다중행(B 등)
        exact.add(_norm(spec["name"]))
    for o in cfg.get("owns", []):               # 명시 소유(템플릿 행 없어도 소관 — 지분법 등)
        exact.add(_norm(o))
    for spec in cfg.get("expense_rows", []):
        kw.update(_norm(k) for k in spec.get("rule", {}).get("포함", []))
    for key in ("commodity_rule", "mfg_total_rule"):
        kw.update(_norm(k) for k in cfg.get(key, {}).get("포함", []))
    exact.discard(""); kw.discard("")
    return {"exact": exact, "kw": kw}


def related_entries(entries: list, toks: dict, tb_rows: list) -> list:
    """분개 각 라인의 계정을 대분류로 환원해 소유 토큰과 매칭(정확 토큰은 ==, 키워드는 포함)."""
    name_to_class = {_norm(r["계정명"]): _norm(r["대분류"]) for r in tb_rows}
    exact, kw = toks["exact"], toks["kw"]

    def line_hit(acc: str) -> bool:
        n = _norm(acc)
        cls = name_to_class.get(n, n)
        if cls in exact or n in exact:
            return True
        return any(len(k) >= 2 and (k in cls or k in n) for k in kw)

    return [e for e in entries if any(line_hit(l["계정"]) for l in e["lines"])]


def _find_anchor(ws, header_kw: str, col_hint: int = None):
    """수정사항 '섹션 헤더' 행/열을 찾는다. 반드시 숫자/로마자 + '.' 섹션번호 접두가 있는 헤더만
    잡는다(예 '4. 수정사항', '5. 회사제시 수정사항'). 본문 표의 컬럼 헤더 '수정사항'은 접두가
    없어 제외된다(과거 그걸 잘못 잡아 분개표를 본문 한가운데 삽입한 버그 방지)."""
    for r in range(1, ws.max_row + 1):
        for c in range(1, 9):
            v = _norm(ws.cell(r, c).value)
            if not v or "수정사항" not in v:
                continue
            if re.match(r"^[0-9０-９Ⅰ-Ⅻ]+\.", v):     # 'N.'/로마자 섹션번호 접두 필수
                return r, c
    return None, None


def _entry_desc(e) -> str:
    """entry의 대표 설명(수정사항집계 Description) — 라인 설명 중 첫 비어있지 않은 것 + notes."""
    for ln in e["lines"]:
        if ln.get("설명"):
            return str(ln["설명"]).strip()
    return "; ".join(str(n).strip() for n in e.get("notes", []) if str(n).strip())


def render(path: str, sheet: str, entries: list, *, header_kw="수정사항",
           mode="normal", entry_codes=None):
    """수정사항 섹션에 분개표를 그린다.

    레이아웃(전 조서 공통): 섹션 헤더 아래 한 행 띄고 분개표. **적요란엔 (*1)(*2) 마커만** 달고,
    표 아래 한 행 띄운 뒤 **(*N) [수정사항집계 설명 원문]** 을 각주로 적는다.

    mode="refer"(GG 미처분이익잉여금): 전체 분개를 긁어오고, 적요엔 **'refer to [대응 조서]'**를
    빨간 기울임 글씨로 쓴다(각주 없음). entry_codes={entry_no: [조서코드...]} 필요.
    """
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.active
    hr, hc = _find_anchor(ws, header_kw)
    if hr is None:
        wb.close()
        return 0

    refer = (mode == "refer")
    seq_of = {e["no"]: i for i, e in enumerate(entries, 1)}     # 분개 → (*N) 순번
    foots = [] if refer else [(seq_of[e["no"]], _entry_desc(e)) for e in entries if _entry_desc(e)]
    n_lines = sum(len(e["lines"]) for e in entries)
    n_rows = 1 + n_lines + (2 + len(foots) if foots else 0) if entries else 0  # 헤더+라인+(빈행+각주)

    # ── 다음 섹션 경계(next_sec) robust 탐지 ──────────────────────────────
    # 근본문제: 템플릿이 분개용으로 예약한 고정공간에 의존하면, 분개가 작을 땐 잉여행의 격자(테두리)가
    # 남아 표가 커 보이고 아래 '주석/증감분석' 위에 그려진 듯 보인다. → 예약공간 의존을 버리고
    # [분개헤더, 다음섹션) 영역을 '데이터 크기에 정확히' 맞춘다(양방향). 먼저 경계를 정확히 잡는다.
    #   종료 인식: 'N.'(섹션) · 'N)'(서브헤더) · 키워드(주석/증감분석/결론/Recap/Test/Nature/검토).
    # 경계 = 다음 **메인 섹션**('N.' 점접두)만. 'N)' 괄호접두는 수정사항 내부 서브섹션
    # (예 '2) 검토 결과')이라 경계로 보면 그 아래 placeholder를 못 지운다 → 제외.
    # 키워드(결론/주석 등)는 'N.'가 없는 비표준 헤더 대비, 짧은 헤더형 셀에서만 보조 인정.
    _SEC_KW = ("주석", "증감분석", "결론", "recap", "nature")

    def _is_next_section(rr):
        for c in range(1, 10):
            raw = ws.cell(rr, c).value
            v = _norm(raw)
            if not v:
                continue
            if re.match(r"^[0-9０-９Ⅰ-Ⅻ]+\)", v):       # 'N)' 서브섹션('2) 결론' 등) → 경계 아님(스킵)
                continue
            if re.match(r"^[0-9０-９Ⅰ-Ⅻ]+\.", v):       # 'N.' 메인 섹션 → 경계
                return True
            if len(str(raw).strip()) <= 14 and any(k in v.lower() for k in _SEC_KW):
                return True                              # 'N.' 없는 메인 헤더(예 P의 '결론') 보조 인식
        return False

    next_sec = ws.max_row + 1
    for rr in range(hr + 2, ws.max_row + 1):       # hr+1은 보통 빈 줄, hr+2부터 탐색
        if _is_next_section(rr):
            next_sec = rr
            break

    # ── 영역 완전 정규화: [hr+1, next_sec) 의 값·테두리·채움·병합 전부 초기화 ──
    # (값만 지우던 과거 버그 → 예약 격자 잔존. 이제 테두리/채움까지 비워 잉여행이 빈칸이 되게 한다.)
    from openpyxl.worksheet.cell_range import CellRange
    from openpyxl.styles import Border as _B, PatternFill as _PF
    for ref in [str(m) for m in ws.merged_cells.ranges]:
        rng = CellRange(ref)
        if rng.min_row < next_sec and rng.max_row >= hr + 1:
            ws.unmerge_cells(ref)
    _blank_b, _blank_f = _B(), _PF()

    def _normalize(r0, r1):
        for rr in range(r0, r1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(rr, c)
                cell.value = None
                cell.border = _blank_b
                cell.fill = _blank_f

    _normalize(hr + 1, next_sec)

    # 관련 분개 없음 → '해당사항 없음'만(영역은 위에서 완전 정규화됨 → 잔존 격자 0).
    if not entries:
        ws.cell(hr + 2, hc).value = "해당사항 없음"
        wb.save(path)
        return 0

    start = hr + 2                                              # 섹션 헤더 아래 한 행 띄고
    avail = max(0, next_sec - start)                           # 다음 섹션 전까지 빈 공간
    if n_rows > avail:                                          # 데이터가 더 크면 그만큼 삽입(아래로 밀어냄)
        ws.insert_rows(start, n_rows - avail)
    # (데이터가 더 작으면: 잉여행은 위에서 정규화돼 빈칸 → 표가 정확히 entries 크기만큼만 보임)

    c0 = hc
    cols = {"no": c0, "acct": c0 + 1, "dr": c0 + 2, "cr": c0 + 3, "desc": c0 + 4}

    def put(r, c, val, *, num=False, hdr=False, bold=False, font=None, border=True, align=None):
        cell = ws.cell(r, c)
        cell.value = val
        if border:
            cell.border = _BORDER
        cell.alignment = align or Alignment(
            horizontal="center" if (hdr or c in (cols["no"], cols["dr"], cols["cr"])) else "left",
            vertical="center")
        if num:
            cell.number_format = _NUMF
            cell.alignment = Alignment(horizontal="right", vertical="center")
        if hdr:
            cell.fill = _HDR_FILL
        if font is not None:
            cell.font = font
        elif hdr or bold:
            cell.font = Font(bold=True)

    r = start
    for txt, key in (("#", "no"), ("계정과목", "acct"), ("차변", "dr"), ("대변", "cr"), ("적요", "desc")):
        put(r, cols[key], txt, hdr=True)
    r += 1
    for e in entries:
        seq = seq_of[e["no"]]
        first = True
        for ln in e["lines"]:
            put(r, cols["no"], f"#{e['no']}" if first else None)
            put(r, cols["acct"], ln["계정"])
            dr = ln.get("금액") if ln["side"] == "차변" else None
            crv = ln.get("금액") if ln["side"] == "대변" else None
            put(r, cols["dr"], dr, num=True)
            put(r, cols["cr"], crv, num=True)
            if first:
                if refer:                                       # GG: refer to [조서] 빨강 기울임
                    codes = (entry_codes or {}).get(e["no"], [])
                    txt = ("refer to " + ", ".join(codes)) if codes else ""
                    put(r, cols["desc"], txt, font=Font(color="FFFF0000", italic=True))
                else:
                    put(r, cols["desc"], f"(*{seq})")           # 적요엔 마커만
            else:
                put(r, cols["desc"], None)
            first = False
            r += 1

    # 각주: 표 아래 한 행 띄고 (*N) 설명원문
    if foots:
        r += 1
        for seq, desc in foots:
            put(r, cols["no"], f"(*{seq})", border=False, align=Alignment(horizontal="left", vertical="top"))
            cell = ws.cell(r, cols["acct"])
            cell.value = desc
            # 자동줄바꿈 OFF(사용자 지정) — 매핑된 Description은 한 줄로 표시.
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
            r += 1

    for c, w in ((cols["no"], 6), (cols["acct"], 18), (cols["dr"], 15), (cols["cr"], 15), (cols["desc"], 16)):
        cur = ws.column_dimensions[get_column_letter(c)].width or 0
        ws.column_dimensions[get_column_letter(c)].width = max(cur, w)

    wb.save(path)
    return len(entries)
