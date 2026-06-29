# -*- coding: utf-8 -*-
"""G200 취득·처분 Test — 샘플선정 근거 + TARGET TEST(분개장 매핑).

(3) 샘플선정 근거 = 총괄표에 취득(또는 처분) 금액이 있는 계정만 '1.~, 2.~'로 동적 나열
(4개 초과 시 행 추가). TARGET TEST 표 = 그 계정들의 분개장 거래를 계정당 1~2건 매핑
(계정과목·날짜·적요·거래처·취득액). **증빙확인(K) = 취득액(H) 동일금액 하드코딩(수식 금지).**
취득/처분이 분개장에 없으면 표본 절차 생략 문구만.
"""

import re
from copy import copy

from openpyxl.utils import column_index_from_string

from .base import shift_formula_rows


def _norm(s) -> str:
    return re.sub(r"\s+", "", str(s)) if s is not None else ""


_NUMF = '_(* #,##0_);[Red]_(* \\(#,##0\\);_(* "-"_);_(@_)'
_TBL_C0, _TBL_C1 = 3, 13   # 표 열범위 C..M


def fill_g200(ws, cfg: dict, movements: dict, intangible=None) -> dict:
    """G200 취득·처분 TARGET TEST를 모두 채운다(새 양식: 처분 폼은 템플릿에 선구성됨).

    표를 초과하는 표본은 **테두리째 행삽입**으로 표를 확장(테두리 없는 빈행에 흘리지 않음).
    처분은 **유형만**(무형 대변=직접법 상각이라 처분 아님). 증빙확인(K)=금액(H) 하드코딩, 금액일치(L)=K-H.

    Args:
        movements: {계정: {취득:[line], 처분:[line], 취득합, 처분합}}  line={날짜,적요,거래처,금액}
        intangible: 무형 계정명(처분 표본에서 제외)
    """
    g = cfg["g200"]
    cols = g["cols"]
    per = g.get("per_account", 2)
    sci = column_index_from_string(g.get("sample_col", "C"))
    intang = {_norm(x) for x in (intangible or [])}

    # 취득/처분 표본 라인 (계정당 per)
    acq = [(a, m) for a, m in movements.items() if (m.get("취득합") or 0) > 0]
    disp = [(a, m) for a, m in movements.items()
            if (m.get("처분합") or 0) > 0 and _norm(a) not in intang]
    acq_lines = [(a, ln) for a, m in acq for ln in (m.get("취득") or [])[:per]]
    disp_lines = [(a, ln) for a, m in disp for ln in (m.get("처분") or [])[:per]]

    # 샘플선정 근거 2개(취득=첫째, 처분=둘째)
    sa_rows = _find_all(ws, g.get("sample_anchor", "샘플선정"))
    if sa_rows:
        _fill_samples(ws, sa_rows[0], [a for a, _ in acq], sci)
    if len(sa_rows) > 1:
        _fill_samples(ws, sa_rows[1], [a for a, _ in disp], sci)

    out = {}
    h_acq = _find_row(ws, g.get("target_amount_label", "취득액"))
    if h_acq:
        out["acq"] = _fill_target(ws, cols, h_acq, acq_lines)
    h_disp = _find_row(ws, g.get("disp_amount_label", "처분액"))
    if h_disp:
        out["disp"] = _fill_target(ws, cols, h_disp, disp_lines)
    return out


def _fill_samples(ws, anchor_row, accounts, sci):
    """(3) 샘플선정 근거를 '1.~,2.~'로 채운다(앵커행 C열부터, 초과 시 행삽입)."""
    start = anchor_row if _norm(ws.cell(anchor_row, sci).value) else anchor_row + 1
    cur = 0
    for r in range(start, start + 30):
        v = ws.cell(r, sci).value
        if v and re.match(r"\d+\.", str(v).strip()):
            cur += 1
        else:
            break
    need = len(accounts)
    if need > cur:
        _insert(ws, start + cur, need - cur, sci)
    for r in range(start, start + max(cur, need)):
        ws.cell(r, sci).value = None
    for i, a in enumerate(accounts):
        ws.cell(start + i, sci).value = f"{i + 1}.{a}"


def _fill_target(ws, cols, header_row, lines):
    """TARGET TEST 표(header_row 아래)를 채운다. 표(테두리) 초과 시 **테두리째 행삽입**으로 확장."""
    ds = header_row + 1
    # 테두리 있는 표 행 수(좌측 테두리 기준)
    ext = 0
    for r in range(ds, ds + 40):
        b = ws.cell(r, _TBL_C0).border
        if b and b.left and b.left.style:
            ext += 1
        else:
            break
    donor = {c: (copy(ws.cell(ds, c).border), copy(ws.cell(ds, c).font),
                 copy(ws.cell(ds, c).alignment)) for c in range(_TBL_C0, _TBL_C1 + 1)}
    needed = len(lines)
    if needed > ext and ext > 0:                   # 표 초과 → 테두리째 행삽입
        n = needed - ext
        saved = None
        for m in list(ws.merged_cells.ranges):     # 표 안 병합(증빙 I열) 떼었다 확장 재병합
            if ds <= m.min_row and m.max_row <= ds + ext - 1:
                saved = (m.min_col, m.max_col, m.min_row)
                ws.unmerge_cells(str(m))
        ws.insert_rows(ds + ext, n)
        shift_formula_rows(ws, ds + ext, n)
        for r in range(ds + ext, ds + ext + n):
            for c, (bd, ft, al) in donor.items():
                cell = ws.cell(r, c)
                cell.border, cell.font, cell.alignment = copy(bd), copy(ft), copy(al)
        if saved:
            ws.merge_cells(start_row=saved[2], start_column=saved[0],
                           end_row=ds + ext + n - 1, end_column=saved[1])
    # 표 전 영역(표본 + 빈행)에 테두리 보장 + L수식(=K-H) 행별 보정
    span = max(needed, ext)
    for r in range(ds, ds + span):
        for c, (bd, _f, _a) in donor.items():
            ws.cell(r, c).border = copy(bd)
        ws[f"L{r}"] = f"=K{r}-H{r}"
    for i, (a, ln) in enumerate(lines):
        r = ds + i
        ws[f"{cols['계정과목']}{r}"] = a
        if ln.get("날짜") is not None:
            ws[f"{cols['날짜']}{r}"] = ln["날짜"]
            ws[f"{cols['날짜']}{r}"].number_format = "yyyy-mm-dd"
        ws[f"{cols['적요']}{r}"] = ln.get("적요")
        ws[f"{cols['거래처']}{r}"] = ln.get("거래처")
        amt = ln.get("금액")
        ws[f"{cols['금액']}{r}"] = amt
        ws[f"{cols['금액']}{r}"].number_format = _NUMF
        ws[f"{cols['증빙확인']}{r}"] = amt          # K=H 동일금액 하드코딩(수식 X)
        ws[f"{cols['증빙확인']}{r}"].number_format = _NUMF
        if not ws[f"J{r}"].value:
            ws[f"J{r}"] = "O"
        if not ws[f"M{r}"].value:
            ws[f"M{r}"] = "O"
    return needed


def _find_all(ws, anchor):
    a = _norm(anchor)
    return [r for r in range(1, ws.max_row + 1)
            if any(a in _norm(ws.cell(r, c).value) for c in range(2, 14))]


def build_disposal_form(ws, cfg: dict, movements: dict, intangible=None) -> dict:
    """'2. 처분 Test' 아래에 취득 TEST와 동일한 1)~4) 폼을 신설하고 처분 데이터를 채운다.

    취득 블록(1) Test 목적 ~ 4) Test 결론)을 캡처해 처분 헤더 아래에 찍고 '취득'→'처분' 치환.
    유형 처분(처분합>0)만 TARGET TEST에 분개장 매핑(무형은 대변=직접법 상각이라 처분 아님 → 제외).
    실제 처분 없으면 결론에 처분 없음 문구.
    """
    intangible = {_norm(x) for x in (intangible or [])}
    from copy import copy
    g = cfg["g200"]
    cols = g["cols"]
    per = g.get("per_account", 2)
    # 취득 블록 범위: '1) Test'~'4) Test 결론'
    b_start = _find_row(ws, "1) Test")
    b_end = _find_row(ws, "4) Test 결론")
    disp = _find_row(ws, g.get("disp_title", "처분 Test"))
    if not (b_start and b_end and disp and b_end > b_start):
        return {"built": False}
    nrows = b_end - b_start + 1

    # 블록 캡처(값+스타일) + 내부 병합(상대 오프셋)
    block = []
    for r in range(b_start, b_end + 1):
        row = {}
        for c in range(1, 16):
            cell = ws.cell(r, c)
            row[c] = {"v": cell.value, "font": copy(cell.font), "border": copy(cell.border),
                      "fill": copy(cell.fill), "align": copy(cell.alignment), "nf": cell.number_format}
        block.append(row)
    merges = []
    for m in list(ws.merged_cells.ranges):
        if b_start <= m.min_row and m.max_row <= b_end:
            merges.append((m.min_row - b_start, m.max_row - b_start, m.min_col, m.max_col))

    # 처분 헤더 아래 삽입
    at = disp + 1
    ws.insert_rows(at, nrows)
    shift_formula_rows(ws, at, nrows)
    for i, row in enumerate(block):
        rr = at + i
        for c, st in row.items():
            cell = ws.cell(rr, c)
            v = st["v"]
            if isinstance(v, str):
                v = v.replace("취득", "처분").replace("처분결의서", "처분품의서")
            cell.value = v
            cell.font, cell.border = copy(st["font"]), copy(st["border"])
            cell.fill, cell.alignment = copy(st["fill"]), copy(st["align"])
            cell.number_format = st["nf"]
    for (r0, r1, c0, c1) in merges:
        try:
            ws.merge_cells(start_row=at + r0, start_column=c0, end_row=at + r1, end_column=c1)
        except Exception:
            pass

    disp_acc = [(a, m) for a, m in movements.items()
                if (m.get("처분합") or 0) > 0 and _norm(a) not in intangible]
    # 처분 샘플선정 근거(블록 내)
    sa = _find_row_from(ws, g.get("sample_anchor", "샘플선정"), at)
    if sa is not None and disp_acc:
        sci = column_index_from_string(g.get("sample_col", "C"))
        ss = sa if _norm(ws.cell(sa, sci).value) else sa + 1
        for r in range(ss, ss + 12):
            v = ws.cell(r, sci).value
            if v and re.match(r"\d+\.", str(v).strip()):
                ws.cell(r, sci).value = None
            else:
                break
        for i, (a, _) in enumerate(disp_acc):
            ws.cell(ss + i, sci).value = f"{i + 1}.{a}"

    # 처분 TARGET TEST 채움 (새 블록 내 헤더 '처분액' 찾아 그 아래)
    hdr = _find_row_from(ws, "처분액", at)
    if hdr is None:
        hdr = _find_row_from(ws, "취득액", at)   # 치환 안 된 경우 대비
    n_filled = 0
    if hdr is not None and disp_acc:
        lines = []
        for a, m in disp_acc:
            for ln in (m.get("처분") or [])[:per]:
                lines.append((a, ln))
        numf = '_(* #,##0_);[Red]_(* \\(#,##0\\);_(* "-"_);_(@_)'
        for i, (a, ln) in enumerate(lines):
            r = hdr + 1 + i
            ws[f"{cols['계정과목']}{r}"] = a
            if ln.get("날짜") is not None:
                ws[f"{cols['날짜']}{r}"] = ln["날짜"]
                ws[f"{cols['날짜']}{r}"].number_format = "yyyy-mm-dd"
            ws[f"{cols['적요']}{r}"] = ln.get("적요")
            ws[f"{cols['거래처']}{r}"] = ln.get("거래처")
            ws[f"{cols['금액']}{r}"] = ln.get("금액")
            ws[f"{cols['금액']}{r}"].number_format = numf
            ws[f"{cols['증빙확인']}{r}"] = ln.get("금액")
            ws[f"{cols['증빙확인']}{r}"].number_format = numf
        n_filled = len(lines)
    elif not disp_acc:
        # 처분 없음 — 4) 결론에 문구
        concl = _find_row_from(ws, "처분 내역이", at) or _find_row_from(ws, "4) Test 결론", at)
        if concl is not None:
            ws[f"{cols['계정과목']}{concl}"] = g.get("no_disp_text", "당기 중 처분 내역이 없음.")
    return {"built": True, "disp_rows": n_filled, "inserted": nrows}


def _find_row(ws, anchor):
    return _find_row_from(ws, anchor, 1)


def _find_row_from(ws, anchor, start):
    a = _norm(anchor)
    for r in range(start, ws.max_row + 1):
        for c in range(2, 14):
            if a in _norm(ws.cell(r, c).value):
                return r
    return None


def _insert(ws, at, n, donor_col):
    ws.insert_rows(at, n)
    shift_formula_rows(ws, at, n)
    for nr in range(at, at + n):                # 도너 스타일(at 위 행)
        for c in range(2, 14):
            src, dst = ws.cell(at - 1, c), ws.cell(nr, c)
            dst.border = copy(src.border)
            dst.font = copy(src.font)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
