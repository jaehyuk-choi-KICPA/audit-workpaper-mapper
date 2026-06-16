"""A-200 (조회서 대사) 생성기.

회사제시 명세서(거래처원장)를 base로 두고, 조회서 취합엑셀(BANK/INSURANCE/
INVESTMENT)을 **역방향 매핑**해 STEP5 데이터 표를 구성한다.

핵심 원칙 (회계사 조언 반영):
  - 회사제시 금액(col8) = 거래처원장 잔액 (명세서가 모집단의 기준).
  - 거래처원장엔 계좌번호가 불완전하므로, **금액으로 조회서 회신금액과 1:1 매칭**해
    일치 시 구분/계좌번호/통화/이자율/만기/조회서번호를 조회서에서 역으로 채운다.
  - 금액 불일치분은 감사인 영역 → 결론(col22)에 (*n) 각주를 달고, 표 아래에 각주 문구를 둔다.
  - 차이금액(col16)=수식, 소계 H/O/P=SUM수식, STEP2 모집단=소계 합 수식.
  - 일치여부 5개열(col17~21)=전부 공란(감사인 확인영역), 사용제한=n/a.

분류:
  - 현금성자산   : 거래처원장 보통예금        ↔ BANK 1번(현금성)
  - 장기금융상품 : 거래처원장 장기금융(잔액>0) ↔ INSURANCE 1번 그룹(해약환급금>0)
  - 단기금융상품 : 거래처원장 단기매매/단기금융 ↔ INVESTMENT 1번 (단기매매주식 등)

전자조회분만 취합엑셀에 존재한다. 우편조회분(명세서엔 있으나 취합엑셀에 없음)은
회신금액을 채울 수 없으므로 각주(*n)로 감사인 확인을 표시한다.

★ 양식 보존: 템플릿은 계정별 고정 블록(소계 포함)이므로, 행수가 다르면 양식이
  무너진다. 생성기는 도너 행의 스타일을 복사해 동적으로 블록을 재구성하고,
  소계/STEP2/각주/Step6 위치를 최종 위치로 재계산한다.
"""

import re
from collections import defaultdict

from .base import WorkpaperGenerator


_CASH_LEDGER_ACCOUNTS = ("보통예금",)
_LONGTERM_LEDGER_KEYWORD = "장기금융"
_SHORTTERM_LEDGER_KEYWORDS = ("단기매매", "단기금융", "단기투자")

_CONFIRM_ELECTRONIC = "전자"
_CONFIRM_POSTAL = "우편"
_V = "V"
_NA = "n/a"


def _amt(v) -> int:
    if v is None or isinstance(v, bool):
        return 0
    if isinstance(v, (int, float)):
        return round(v)
    s = re.sub(r"[^\d.\-]", "", str(v))
    try:
        return round(float(s)) if s else 0
    except ValueError:
        return 0


def _clean_inst_name(s) -> str:
    if not s:
        return ""
    s = re.sub(r"\(.*?\)", "", str(s))
    return re.sub(r"\s+", "", s).strip()


# ---------------------------------------------------------------------------
# 현금성자산: 거래처원장 보통예금 ↔ BANK 현금성 (금액 1:1 역매핑)
# ---------------------------------------------------------------------------

def _norm_acc(s) -> str:
    """계좌번호 정규화: 숫자만. 'NNN-NNNNNN-NN-NNNNN' → 'NNNNNNNNNNNNNNNN'(하이픈 제거)."""
    return re.sub(r"[^\d]", "", str(s or ""))


def _ledger_foreign_hint(L: dict) -> bool:
    txt = f"{L.get('적요') or ''}{L.get('거래처명') or ''}"
    return "외화" in txt or "외환" in txt


def _build_cash_rows(ledger_cash: list[dict], bank_cash: list[dict],
                     fx_rates: "dict | None" = None) -> tuple[list[dict], list[dict]]:
    # 외화 조회서 행은 KRW가 아니라 외화로 회신되므로 금액 매칭이 불가 →
    # 계좌번호(정규화)로, 그래도 안 되면 '외화' 표기 명세서 행에 한해 이름으로 매칭.
    fx_banks = [b for b in bank_cash if _is_foreign(b.get("통화"))]
    krw_banks = [b for b in bank_cash if not _is_foreign(b.get("통화"))]
    used_fx = [False] * len(fx_banks)

    by_amt: dict[int, list[dict]] = defaultdict(list)
    for b in krw_banks:
        by_amt[_amt(b["금액"])].append(b)

    rows = []
    for L in ledger_cash:
        ledger_name = _clean_inst_name(L["거래처명"])
        lacc = _norm_acc(L.get("비고"))           # 현금성 명세서는 계좌번호를 비고에 둔다
        chosen = None

        # 1) 외화 우선: 계좌번호(강) → '외화' 표기 행 한정 이름(약)
        for i, b in enumerate(fx_banks):
            if used_fx[i]:
                continue
            bacc = _norm_acc(b.get("계좌번호"))
            if lacc and bacc and (lacc == bacc or lacc in bacc or bacc in lacc):
                chosen = b; used_fx[i] = True; break
        if chosen is None and _ledger_foreign_hint(L):
            for i, b in enumerate(fx_banks):
                if used_fx[i]:
                    continue
                bname = _clean_inst_name(b.get("금융기관명"))
                if bname and (bname in ledger_name or ledger_name in bname):
                    chosen = b; used_fx[i] = True; break

        # 2) KRW: 금액 1:1 (기존) — 동일 금액 내 이름 우선
        if chosen is None:
            cands = by_amt.get(_amt(L["잔액"]))
            if cands:
                for i, b in enumerate(cands):
                    bname = _clean_inst_name(b["금융기관명"])
                    if bname and (bname in ledger_name or ledger_name in bname):
                        chosen = cands.pop(i); break
                if chosen is None:
                    chosen = cands.pop(0)

        rows.append(_cash_row(L, chosen, fx_rates))

    leftover = [b for i, b in enumerate(fx_banks) if not used_fx[i]]
    leftover += [b for lst in by_amt.values() for b in lst]
    return rows, leftover


def _is_foreign(cur) -> bool:
    """통화가 외화인가 (KRW/원/None이 아니면 외화)."""
    if cur is None:
        return False
    c = str(cur).strip().upper()
    return c not in ("", "KRW", "WON", "원", "￦", "\\")


def _cash_row(ledger: dict, bank: "dict | None", fx_rates: "dict | None" = None) -> dict:
    matched = bank is not None
    company_krw = _amt(ledger["잔액"])          # 명세서 금액(H) = 명세서상 원화

    cur = bank.get("통화") if matched else None
    confirm_krw = _amt(bank["금액"]) if matched else None   # 조회서 회신 KRW (기본)

    # 외화 계좌 + 환율 메모 + 조회서 외화값이 있으면, 회신금액(O) = 환율 × 외화값으로 환산.
    fx_amount = _to_float(bank.get("외화금액")) if matched else None
    rate = (fx_rates or {}).get(str(cur).strip().upper()) if (matched and _is_foreign(cur)) else None
    fx_translated = bool(rate and fx_amount)
    if fx_translated:
        confirm_krw = round(rate * fx_amount)

    return {
        "계정분류":   "현금성자산",
        "금융기관":   (bank["금융기관명"] if matched else ledger["거래처명"]),
        "구분":       (bank["금융상품종류"] if matched else None),
        "계좌번호":   (bank["계좌번호"] if matched else None),
        "통화":       (cur if matched else None),
        "금액":       company_krw,
        "이자율":     (bank["연이자율"] if matched else None),
        "만기":       (bank["만기"] if matched else None),
        "사용제한":   _NA,
        "회신여부":   (_V if matched else None),
        "회수방법":   (_CONFIRM_ELECTRONIC if matched else None),
        "조회서번호": (bank["조서번호"] if matched else None),
        "회신금액":   confirm_krw,
        # 외화 환산 메타 (생성기: 외화값 열·환율박스·기말환산 불일치 각주)
        "_외화금액":  (fx_amount if fx_translated else None),
        "_환율":      (rate if fx_translated else None),
        "_기말환산불일치": bool(fx_translated and confirm_krw != company_krw),
        "_확인필요":  (not matched),
    }


def _to_float(v):
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[^\d.\-]", "", str(v))
    try:
        return float(s) if s else None
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# 장기금융상품: 거래처원장 장기금융(잔액>0) ↔ INSURANCE 그룹(해약환급금>0)
# ---------------------------------------------------------------------------

def _build_longterm_rows(ledger_lt: list[dict], ins_groups: list[dict]) -> list[dict]:
    rows = []
    for L in ledger_lt:
        company_amt = _amt(L["잔액"])
        if company_amt <= 0:
            continue
        ledger_name = _clean_inst_name(L["거래처명"])
        match = None
        for g in ins_groups:
            gname = _clean_inst_name(g["금융기관명"])
            if gname and (ledger_name in gname or gname in ledger_name):
                match = g
                break
        matched = match is not None
        rows.append({
            "계정분류":   "장기금융상품",
            "금융기관":   L["거래처명"],
            "구분":       (match["보험의종류"] if matched else None),
            "계좌번호":   (match["증권번호"] if matched else None),
            "통화":       ("KRW" if matched else None),
            "금액":       company_amt,
            "이자율":     "-",
            "만기":       "-",
            "사용제한":   _NA,
            "회신여부":   (_V if matched else None),
            "회수방법":   (_CONFIRM_ELECTRONIC if matched else _CONFIRM_POSTAL),
            "조회서번호": (match.get("조서번호") if matched else None),
            "회신금액":   (_amt(match["조회서금액"]) if matched else None),
            "_확인필요":  (not matched),
        })
    return rows


# ---------------------------------------------------------------------------
# 단기금융상품: 거래처원장 단기매매/단기금융 ↔ INVESTMENT (금액 1:1 역매핑)
# ---------------------------------------------------------------------------

def _build_shortterm_rows(ledger_st: list[dict], invest_rows: list[dict]) -> list[dict]:
    by_amt: dict[int, list[dict]] = defaultdict(list)
    for v in invest_rows:
        by_amt[_amt(v["금액"])].append(v)

    rows = []
    for L in ledger_st:
        company_amt = _amt(L["잔액"])
        if company_amt <= 0:
            continue
        ledger_name = _clean_inst_name(L["거래처명"])
        chosen = None
        cands = by_amt.get(company_amt)
        if cands:
            for i, v in enumerate(cands):
                vname = _clean_inst_name(v["금융기관명"])
                if vname and (vname in ledger_name or ledger_name in vname):
                    chosen = cands.pop(i)
                    break
            if chosen is None:
                chosen = cands.pop(0)
        matched = chosen is not None
        rows.append({
            "계정분류":   "단기금융상품",
            "금융기관":   (chosen["금융기관명"] if matched else L["거래처명"]),
            "구분":       (chosen["금융상품종류"] if matched else None),
            "계좌번호":   (chosen["계좌번호"] if matched else None),
            "통화":       ("KRW" if matched else None),
            "금액":       company_amt,
            "이자율":     "-",
            "만기":       "-",
            "사용제한":   _NA,
            "회신여부":   (_V if matched else None),
            "회수방법":   (_CONFIRM_ELECTRONIC if matched else None),
            "조회서번호": (chosen["조서번호"] if matched else None),
            "회신금액":   (_amt(chosen["금액"]) if matched else None),
            "_확인필요":  (not matched),
        })
    return rows


# ---------------------------------------------------------------------------
# 공개 변환 API
# ---------------------------------------------------------------------------

def build_a200_data(ledger_rows: list[dict], bank_rows: list[dict],
                    insurance_groups: list[dict],
                    invest_rows: "list[dict] | None" = None,
                    fx_rates: "dict | None" = None) -> dict:
    """A-200 STEP5 데이터(현금성/장기금융/단기금융) + 미매칭 회신을 계산해 반환한다.

    fx_rates: {통화: 기말환율}. 외화 계좌의 회신금액을 환율×조회서외화값으로 환산한다.
    """
    invest_rows = invest_rows or []
    ledger_cash = [r for r in ledger_rows if r.get("계정과목") in _CASH_LEDGER_ACCOUNTS]
    ledger_lt = [r for r in ledger_rows
                 if _LONGTERM_LEDGER_KEYWORD in (r.get("계정과목") or "")]
    ledger_st = [r for r in ledger_rows
                 if any(k in (r.get("계정과목") or "") for k in _SHORTTERM_LEDGER_KEYWORDS)]
    bank_cash = [b for b in bank_rows if b.get("계정분류") == "현금성자산"]

    cash, leftover = _build_cash_rows(ledger_cash, bank_cash, fx_rates)
    longterm = _build_longterm_rows(ledger_lt, insurance_groups)
    shortterm = _build_shortterm_rows(ledger_st, invest_rows)

    # 실제 사용된 외화·환율만 모아 환율박스용으로 반환 (없으면 빈 dict → 박스 생략)
    fx_used = {}
    for r in cash:
        if r.get("_환율"):
            fx_used[str(r["통화"]).strip().upper()] = r["_환율"]

    return {
        "cash": cash,
        "longterm": longterm,
        "shortterm": shortterm,
        "cash_leftover_bank": leftover,
        "fx_used": fx_used,
    }


# ---------------------------------------------------------------------------
# 엑셀 출력 (템플릿 = 회계사가 손본 작업용 양식; 스타일 보존 재구성)
# ---------------------------------------------------------------------------

class A200Generator(WorkpaperGenerator):
    """A-200의 STEP2/STEP5/STEP6을 채운다. 회사명/날짜는 템플릿 수식(A100 참조)을 보존."""

    def fill(self, data: dict, params: dict) -> None:
        self.open_template()
        cfg = self.config
        self._fill_header(cfg.get("header", {}), params)
        self._fill_step1_step2(cfg, params)
        self._fill_step5_step6(cfg, data)

    # --- 공통 항목 (회사명/날짜는 템플릿 수식 보존, Preparer/Reviewer만 선택 기입) ---
    def _fill_header(self, h: dict, params: dict) -> None:
        if h.get("preparer_cell") and params.get("preparer"):
            self.write_cell(h["preparer_cell"], params["preparer"])
        if h.get("reviewer_cell") and params.get("reviewer"):
            self.write_cell(h["reviewer_cell"], params["reviewer"])

    # --- Step2 모집단 (총금액은 소계 합 수식으로 — STEP5 작성 후 채움) ---
    def _fill_step1_step2(self, cfg: dict, params: dict) -> None:
        s2 = cfg["step2"]
        defaults = cfg.get("defaults", {})
        if s2.get("source_name_cell"):
            self.write_cell(s2["source_name_cell"],
                            params.get("source_name") or defaults.get("source_name"))
        if s2.get("completeness_cell"):
            self.write_cell(s2["completeness_cell"],
                            params.get("completeness") or defaults.get("completeness"))
        # population_cell 은 STEP5에서 소계 행번호 확정 후 수식으로 기입

    # --- Step5 표 + Step6 (공용 도구 render_subtotal_sections 사용) ---
    def _fill_step5_step6(self, cfg: dict, data: dict) -> None:
        s5 = cfg["step5"]
        cols = s5["columns"]
        start = s5["data_start_row"]
        diff_c = s5["diff_col"]
        conc_c = s5["conclusion_col"]
        match_cs = s5["match_cols"]
        hcol = s5["amount_company_col"]
        ocol = s5["amount_confirm_col"]
        label_col = cols["No."]

        # 0) 도너 스타일 캡처 (초기화 전) → 1) 영역 초기화 (base 공용 도구)
        #    Step6/각주도 캡처해야 빨강·파랑 등 양식이 보존됨 (안 하면 무서식으로 날아감)
        s6 = cfg.get("step6", {})
        data_style = self.capture_row_style(s5["style_donor_data_row"])
        sub_style = self.capture_row_style(s5["style_donor_subtotal_row"])
        foot_style = self.capture_row_style(s5["style_donor_footnote_row"])
        title_style = self.capture_row_style(s6["style_donor_title_row"])
        body_style = self.capture_row_style(s6["style_donor_body_row"])
        self.clear_region(start, s5["clear_until_row"])

        # 2) 각주 번호 부여 (현금성→장기→단기 순서) — A200 고유
        footnotes = []  # (마커, 문구)
        fx_col = s5.get("fx_value_col")          # 외화값 열 (Conclusion 옆)
        fx_used = data.get("fx_used") or {}

        def decorate(rec, r):
            from copy import copy
            for mc in match_cs:           # 일치여부 5개열 공란
                self.ws[f"{mc}{r}"] = None
            # 외화값(조회서) — 외화 계좌만, Conclusion 열 테두리 복사
            if fx_col and rec.get("_외화금액"):
                c = self.ws[f"{fx_col}{r}"]
                c.value = rec["_외화금액"]
                c.number_format = "#,##0.00"
                c.border = copy(self.ws[f"{conc_c}{r}"].border)

            # 각주: 미회신(_확인필요) 또는 외화 기말환산 불일치
            note = None
            if rec.get("_확인필요"):
                note = ("우편조회분으로 전자취합에 미포함됨 → 회신금액 확인 필요"
                        if rec["계정분류"] == "장기금융상품"
                        else "조회서상 금액과 명세서 금액이 불일치함에 따라 확인 필요")
            elif rec.get("_기말환산불일치"):
                note = (f"기말환산 불일치 / {rec.get('통화') or '외화'} "
                        f"(환율×조회서외화값 ≠ 명세서 원화) — A-0에서 조정")
            if note is None:
                self.ws[f"{conc_c}{r}"] = None
                return
            marker = f"(*{len(footnotes) + 1})"
            footnotes.append((marker, f"{marker} {note}"))
            self.ws[f"{conc_c}{r}"] = marker

        # 외화값 열 헤더 (외화 있을 때만) — Conclusion 헤더 테두리 복사
        if fx_col and fx_used:
            from copy import copy
            self.ws[f"{fx_col}43"] = s5.get("fx_value_header", "조회서 외화")
            for hr in (43, 44):
                self.ws[f"{fx_col}{hr}"].border = copy(self.ws[f"{conc_c}{hr}"].border)

        sections = [
            ("현금성자산 합계", data.get("cash", [])),
            ("장기금융상품 합계", data.get("longterm", [])),
            ("단기금융상품 합계", data.get("shortterm", [])),
        ]
        subtotal_rows, r = self.render_subtotal_sections(
            start_row=start, columns=cols, sections=sections,
            data_style=data_style, subtotal_style=sub_style,
            subtotal_label_col=label_col,
            subtotal_label_merge=s5["subtotal_label_merge"],
            sum_cols=(hcol, ocol, diff_c),
            formula_cols={diff_c: f"={ocol}{{r}}-{hcol}{{r}}"},
            row_decorator=decorate, 
        )

        # 2.5) 환율박스 (외화 있을 때만, STEP5 제목 옆 빈 영역)
        if fx_used and cfg.get("fx_box"):
            self._render_fx_box(cfg["fx_box"], fx_used)

        # 3) STEP2 모집단 총금액 = 소계 합 수식
        pop_cell = cfg["step2"].get("population_cell")
        if pop_cell:
            self.write_cell(pop_cell, "=" + "+".join(f"{hcol}{s}" for s in subtotal_rows))

        # 4) 각주 문구 (표 아래) — 도너 스타일 적용
        r += cfg.get("step5", {}).get("gap_before_footnote", 1)
        for _, note in footnotes:
            self.apply_row_style(r, foot_style)
            self.ws[f"{label_col}{r}"] = note
            r += 1

        # 5) Step6 결론 — 제목 바/본문 스타일 적용
        r += s6.get("gap_after_footnote", 1)
        self._write_conclusion(r, data, label_col, title_style, body_style)

    def _render_fx_box(self, box_cfg: dict, fx_used: dict) -> None:
        """STEP5 제목 옆 빈 영역에 기말환율 박스를 그린다 (실제 사용된 외화만).

        (제목) / (통화 | 환율) / (USD | 1,434.90) ... 크기는 외화 수에 따라 (N+2)x2.
        """
        from openpyxl.styles import Border, Side, Font, Alignment
        from openpyxl.utils import coordinate_to_tuple, get_column_letter

        r0, c0 = coordinate_to_tuple(box_cfg["anchor_cell"])   # (row, col)
        thin = Side(style="thin")
        box = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center")

        def cell(rr, cc):
            return self.ws.cell(row=rr, column=cc)

        # 제목 (2칸 병합)
        t = cell(r0, c0); t.value = box_cfg.get("title", "기말환율")
        t.font = Font(bold=True); t.alignment = center
        self.safe_merge(f"{get_column_letter(c0)}{r0}:{get_column_letter(c0 + 1)}{r0}")
        # 헤더
        h1 = cell(r0 + 1, c0); h1.value = box_cfg.get("col_currency_header", "통화")
        h2 = cell(r0 + 1, c0 + 1); h2.value = box_cfg.get("col_rate_header", "환율(₩)")
        h1.font = h2.font = Font(bold=True)
        # 통화별 환율 (정렬 = 통화코드)
        rr = r0 + 2
        for cur in sorted(fx_used):
            cell(rr, c0).value = cur
            v = cell(rr, c0 + 1); v.value = fx_used[cur]; v.number_format = "#,##0.00"
            rr += 1
        # 박스 테두리·정렬 일괄
        for row_i in range(r0, rr):
            for col_i in (c0, c0 + 1):
                cc = cell(row_i, col_i)
                cc.border = box
                if cc.alignment is None or cc.alignment.horizontal is None:
                    cc.alignment = center

    def _write_conclusion(self, row: int, data: dict, label_col: str,
                          title_style: dict, body_style: dict) -> None:
        cash_review = sum(1 for x in data["cash"] if x["_확인필요"])
        lt_review = sum(1 for x in data.get("longterm", []) if x["_확인필요"])
        st_review = sum(1 for x in data.get("shortterm", []) if x["_확인필요"])
        leftover = len(data.get("cash_leftover_bank", []))

        lines = ["Step 6. 결론 (초안 — 감사인 검토 필요)"]
        lines.append(
            f"현금성자산: 명세서 {len(data['cash'])}건을 조회서 회신금액과 1:1 대사함. "
            f"금액 불일치 {cash_review}건은 확인 필요(장부·회신 차이, 우편조회 등)."
        )
        if leftover:
            lines.append(
                f"조회서에는 회신되었으나 명세서 금액과 일치하지 않는 {leftover}건이 존재함 → 감사인 확인."
            )
        lt_n = len(data.get("longterm", []))
        if lt_n:
            lines.append(
                f"장기금융상품: 명세서 {lt_n}건. 우편조회분 {lt_review}건은 회신금액 별도 입력 필요."
            )
        st_n = len(data.get("shortterm", []))
        if st_n:
            lines.append(f"단기금융상품: 명세서 {st_n}건 (미매칭 {st_review}건 확인 필요).")
        # A-0 단서 (회계사 요청)
        lines.append("상기 차이금액은 4000-A-0(현금및현금성자산·장단기금융상품)에서 "
                     "소명 및 수정분개를 통해 조정함.")
        for i, text in enumerate(lines):
            rr = row + i
            self.apply_row_style(rr, title_style if i == 0 else body_style)
            self.ws[f"{label_col}{rr}"] = text
