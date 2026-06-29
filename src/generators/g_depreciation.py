# -*- coding: utf-8 -*-
"""G300 감가상각비 검토 — (1)모집단표·(3)재계산표 동적 구성.

두 표 모두 총괄표 존재 감가상각비 계정(유형 자산별 + 무형)을 따라 행을 동적 구성한다.
  · (1)모집단: D고정자산관리대장(=회사제시 감가비 하드코딩) | E총괄표상 금액(=−G100!J{감가누계행})
               | F일치(=D=E).  유형 계(감가상각비 계) + 무형 계(무형고정자산상각 계).
  · (3)재계산: D고정자산관리대장(동일 하드코딩) | E총 재계산(=관리대장!M{소계행}) | F차이(=D−E).

각 자산의 G100 감가누계행·관리대장 재계산 소계행은 G100/관리대장 생성 결과에서 받는다
(integration: 세 시트 연쇄 graft에서 좌표가 확정됨).
"""

import re
from copy import copy

from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string

from .base import shift_formula_rows


def _norm(s) -> str:
    return re.sub(r"\s+", "", str(s)) if s is not None else ""


def fill_g300(ws, cfg: dict, tangible: list, intangible: list) -> dict:
    """G300 두 표를 채운다.

    Args:
        tangible:   [{계정, g100_row, register_row, 회사감가비}] (유형, 총괄표 순서)
        intangible: 동일 구조 (무형)
    """
    g = cfg["g300"]
    nc = g.get("name_col", "C")
    dc = g.get("ledger_col", "D")
    ec = g.get("amount_col", "E")
    fc = g.get("match_col", "F")
    nci = column_index_from_string(nc)
    g100 = g.get("g100_sheet", "G100_총괄표")
    g100j = g.get("g100_dep_col", "J")
    reg = g.get("register_sheet", "고정자산 관리대장_25")
    regm = g.get("register_recalc_col", "M")
    t_total = g.get("tangible_total", "감가상각비 계")
    i_total = g.get("intangible_total", "무형고정자산상각 계")

    def find_row(anchor, start=1):
        a = _norm(anchor)
        for r in range(start, ws.max_row + 1):
            for c in range(2, 8):
                if a in _norm(ws.cell(r, c).value):
                    return r
        return None

    pop_hdr = find_row(g.get("pop_anchor", "총괄표상"))
    rec_hdr = find_row(g.get("recalc_anchor", "총 재계산"))

    # 재계산 표 먼저(아래) → 모집단(위) 순으로 처리(행삽입이 위 표 안 밀게)
    out = {}
    if rec_hdr:
        out["recalc"] = self_n = _rebuild(
            ws, rec_hdr + 1, nci, nc, dc, ec, fc, t_total, i_total,
            tangible, intangible,
            e_formula=lambda it: f"='{reg}'!{regm}{it['register_row']}",
            f_formula=lambda r: f"={dc}{r}-{ec}{r}")    # 재계산 F=차이(D−E)
    if pop_hdr:
        out["pop"] = _rebuild(
            ws, pop_hdr + 1, nci, nc, dc, ec, fc, t_total, i_total,
            tangible, intangible,
            e_formula=lambda it: f"=-'{g100}'!{g100j}{it['g100_row']}",
            f_formula=lambda r: f"={dc}{r}={ec}{r}")     # 모집단 F=일치(D=E)
    return out


def _rebuild(ws, start, nci, nc, dc, ec, fc, t_total, i_total,
             tangible, intangible, *, e_formula, f_formula):
    """한 표(유형 rows + 유형계 + 무형 rows + 무형계)를 재구성한다. Returns 채운 자산 수."""
    # 현재 유형계 / 무형계 행 위치
    t_row = _find_label(ws, nci, start, t_total)
    i_row = _find_label(ws, nci, start, i_total)
    if t_row is None or i_row is None:
        return 0
    cur_t = t_row - start            # 유형 데이터 행 수
    cur_i = i_row - (t_row + 1)       # 무형 데이터 행 수
    numf = ws.cell(start, column_index_from_string(ec)).number_format

    # ── 무형 영역 행수 맞춤(아래 먼저) ──
    i_start = t_row + 1
    _adjust_rows(ws, i_start, cur_i, len(intangible), i_row)   # 무형 영역(아래) 먼저
    _adjust_rows(ws, start, cur_t, len(tangible), t_row)       # 유형 영역
    # 위치 재계산
    t_row = start + len(tangible)
    i_start = t_row + 1
    i_row = i_start + len(intangible)

    def write_block(r0, items):
        for i, it in enumerate(items):
            r = r0 + i
            ws[f"{nc}{r}"] = it["계정"]
            ws[f"{dc}{r}"] = it.get("회사감가비") or 0
            ws[f"{dc}{r}"].number_format = numf
            ws[f"{ec}{r}"] = e_formula(it)
            ws[f"{ec}{r}"].number_format = numf
            ws[f"{fc}{r}"] = f_formula(r)

    write_block(start, tangible)
    # 유형계
    if tangible:
        ws[f"{ec}{t_row}"] = f"=SUM({ec}{start}:{ec}{t_row - 1})"
        ws[f"{dc}{t_row}"] = f"=SUM({dc}{start}:{dc}{t_row - 1})"
        ws[f"{fc}{t_row}"] = f"=SUM({fc}{start}:{fc}{t_row - 1})"
        for col in (dc, ec, fc):
            ws[f"{col}{t_row}"].number_format = numf
    write_block(i_start, intangible)
    if intangible:
        ws[f"{ec}{i_row}"] = f"=SUM({ec}{i_start}:{ec}{i_row - 1})"
        ws[f"{dc}{i_row}"] = f"=SUM({dc}{i_start}:{dc}{i_row - 1})"
        ws[f"{fc}{i_row}"] = f"=SUM({fc}{i_start}:{fc}{i_row - 1})"
        for col in (dc, ec, fc):
            ws[f"{col}{i_row}"].number_format = numf
    return len(tangible) + len(intangible)


def _find_label(ws, nci, start, label):
    # 템플릿 오타 '무형고장자산상각'(고장) ↔ '무형고정자산상각'(고정) 정규화
    a = _norm(label).replace("고장", "고정")
    for r in range(start, min(start + 60, ws.max_row + 1)):
        if _norm(ws.cell(r, nci).value).replace("고장", "고정") == a:
            return r
    return None


def _adjust_rows(ws, block_start, cur, need, total_row):
    """블록(block_start~) 행수를 cur→need로 insert/delete. total_row(계 행)는 그 직후. 새 계행 반환."""
    if need > cur:
        n = need - cur
        ws.insert_rows(block_start + cur, n)
        shift_formula_rows(ws, block_start + cur, n)
        for nr in range(block_start + cur, block_start + cur + n):  # 도너 스타일(블록 첫 행)
            for c in range(2, 8):
                src, dst = ws.cell(block_start, c), ws.cell(nr, c)
                dst.border = copy(src.border)
                dst.font = copy(src.font)
                dst.alignment = copy(src.alignment)
                dst.number_format = src.number_format
        return total_row + n
    if need < cur:
        n = cur - need
        ws.delete_rows(block_start + need, n)
        return total_row - n
    return total_row
