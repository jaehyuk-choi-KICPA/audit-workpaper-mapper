"""GG 자본 총괄표 전용 생성기.

자본 총괄표는 일반 잔액형(대분류별/섹션별 소계)과 구조가 다르다:
  - 로마자 구분(Ⅰ.자본금~Ⅴ.이익잉여금)이 **상위 합계 행**이고 그 아래 leaf 계정이 온다.
    (완성본은 로마/자본총계가 이미 하위 참조 SUM 수식 → leaf만 갱신하면 자동 재계산)
  - 수정사항: 이익잉여금(미처분이익잉여금)에 **수정분개의 이익잉여금 Effect 합계**가 집계된다
    (개별 계정 K-L이 아니라 손익→이익잉여금 전체 효과).

따라서 엔진 렌더 대신 **완성본 격자 보존 in-place 리필**을 한다: leaf 행의 계정명을 정산표
대분류에 매칭(최장 공통 접두사)해 기초/기말을 채우고, 미처분이익잉여금 행 수정사항에 집계액을 쓴다.
"""

import re
from collections import defaultdict
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

from .base import shift_formula_rows

_ROMAN = ("Ⅰ", "Ⅱ", "Ⅲ", "Ⅳ", "Ⅴ", "Ⅵ")
_NUMF = '#,##0;[Red]\\(#,##0\\);"-"'


def _norm(s) -> str:
    return re.sub(r"\s+", "", str(s)) if s is not None else ""


def _copy_style(ws, src_r, dst_r, ncols):
    for c in range(1, ncols + 1):
        sc = ws.cell(src_r, c)
        if sc.has_style:
            dc = ws.cell(dst_r, c)
            dc.font = copy(sc.font); dc.border = copy(sc.border)
            dc.fill = copy(sc.fill); dc.alignment = copy(sc.alignment)
            dc.number_format = sc.number_format


def _term_text(blk: dict) -> "str | None":
    """CE 블록(기수/기간) → '제 19 기 2025년 1월 1일부터 \\n 2025년 12월 31일까지' (템플릿 양식)."""
    n, rng = blk.get("기수"), blk.get("기간")
    if not (n and rng):
        return None
    return f"제 {n} 기 {rng[0]}부터 \n {rng[1]}까지"


def _fill_appropriation(ws, ce: "dict | None", bcol="D", ecol="E", acol="H") -> int:
    """이익잉여금처분계산서: 제N기 연도헤더 + **당기 칼럼의 미처분 흐름**만 자동 채운다.

    설계(사용자 확정):
      · 연도/기수 헤더(제N기 YYYY) ← CE(자본변동표) B4/B5. 당기·전기 두 라벨 교체(잔재 교정).
      · 당기 전기이월미처분 = 총괄표 **미처분이익잉여금 leaf 기초**(천원). 당기순이익 = 미처분 leaf
        **증감(기말−기초+수정)**(천원). → 차기이월(=합계 수식) 이 자동으로 총괄표 미처분 수정후와 **대사**.
        (CE '이익잉여금' 열은 미처분+적립금 **총액**이라 미처분-only와 안 맞음 → CE는 헤더에만 사용.)
      · 전기 칼럼(제N-1기)은 **자동 채우지 않고 비운다**(수기 — 전기조서 영역). 잔재 방지로 값만 클리어.
      · 처분일자(처분예정일/확정일)는 보존(제N기 정규식이 안 잡음).
    행은 절대행이 아니라 **과목 라벨 앵커**로 찾는다(로마자 행삽입에도 견고).

    A열 제외(필수): A열은 [수치]/[멘트] 가이드 마커 규약 전용이라 '차기이월미처분이익잉여금'·'이익준비금
    적립' 같은 계정어를 담아, 포함하면 라벨 매칭·섹션앵커를 가로채 오염시킨다 → 라벨 스캔은 B열(=2)부터.
    """
    if not ce:
        return 0
    LBL0 = 2

    # ── 0) 섹션 앵커 ──
    start_row = None
    for r in range(1, ws.max_row + 1):
        if "이익잉여금처분계산서" in _norm("".join(str(ws.cell(r, c).value or "") for c in range(LBL0, 6))):
            start_row = r
            break
    if start_row is None:
        return 0

    # ── 1) 연도/기수 헤더(제 N 기 YYYY) 두 셀: 당기=첫째, 전기=둘째 ──
    term_cells = []
    for r in range(start_row, ws.max_row + 1):
        for c in range(LBL0, min(ws.max_column, 10) + 1):
            v = ws.cell(r, c).value
            if v and re.search(r"제\s*\d+\s*기.*\d{4}", str(v)):
                term_cells.append((r, c))
    cur_txt = _term_text(ce.get("당기") or {})
    prev_txt = _term_text(ce.get("전기") or {})
    if len(term_cells) >= 1 and cur_txt:
        ws.cell(*term_cells[0]).value = cur_txt
    if len(term_cells) >= 2 and prev_txt:
        ws.cell(*term_cells[1]).value = prev_txt
    n = 1 if term_cells else 0

    # ── 2) 당기/전기 칼럼 위치(과목 헤더행의 '당기'·'전기') ──
    cur = pv_sub = hdr_row = None
    for r in range(start_row, min(start_row + 8, ws.max_row + 1)):
        for c in range(LBL0, min(ws.max_column, 14) + 1):
            t = _norm(ws.cell(r, c).value)
            if t == "당기" and cur is None:
                cur, hdr_row = c, r
            elif t == "전기" and pv_sub is None:
                pv_sub, hdr_row = c, r
    if not cur:
        return n

    # ── 3) 총괄표 미처분이익잉여금 leaf 행(섹션 위) ──
    mip = None
    for r in range(1, start_row):                      # 이름 열만(B·C) — 값 열 숫자 제외
        lab = _norm("".join(str(ws.cell(r, c).value or "") for c in range(LBL0, 4)))
        if lab.endswith("미처분이익잉여금") and "이월" not in lab and lab[:1] not in _ROMAN:
            mip = r                                    # 마지막(가장 아래) 매칭 = 본문 leaf

    # ── 4) 본문 행 순회: 당기 전기이월·당기순이익만 채우고, 전기 칼럼 값은 클리어(수기) ──
    end_row = ws.max_row
    for r in range(start_row, ws.max_row + 1):
        rl = _norm("".join(str(ws.cell(r, c).value or "") for c in range(LBL0, cur)))
        if not rl:
            continue
        if mip and "전기이월미처분이익잉여금" in rl:
            ws.cell(r, cur).value = f"=ROUND({bcol}{mip}/1000,0)"               # 미처분 leaf 기초
            ws.cell(r, cur).number_format = _NUMF
            n += 1
        elif mip and "당기순이익" in rl:
            ws.cell(r, cur).value = f"=ROUND((({ecol}{mip}-{bcol}{mip})+{acol}{mip})/1000,0)"  # 미처분 증감
            ws.cell(r, cur).number_format = _NUMF
            n += 1
        if "차기이월미처분이익잉여금" in rl:
            end_row = r                                # 전기 클리어 종료 = 차기이월 행
            break

    # 전기 칼럼(sub+roman) 값 클리어 — 타사/전기 잔재 제거, 수기 입력용 빈칸(테두리 보존).
    # 과목 헤더행(='전기' 라벨)은 보존: 헤더 아래(hdr_row+1)부터 차기이월행까지만 비운다.
    if pv_sub and hdr_row:
        for r in range(hdr_row + 1, end_row + 1):
            for c in (pv_sub, pv_sub + 1):
                if ws.cell(r, c).value is not None:
                    ws.cell(r, c).value = None
    return n


def _fill_capital_roman(template, tb_rows, adj_entries, output, cfg, roman_groups, ce=None):
    """로마자 분류 그룹핑 방식(사용자 지정): 자본 flag 계정을 로마자(Ⅰ~Ⅴ)에 매칭해 그 아래
    leaf로 동적 렌더(누락 계정 자동 편입). leaf 라벨에 의존하지 않는다. 로마자 SUM은 멤버로 재작성,
    자본총계 등 절대참조는 행삽입 시 shift_formula_rows로 보정. 로마자 처리는 **아래→위**(삽입이
    위쪽 로마자 위치에 영향 없게)."""
    sheet = cfg["sheet"]
    nci = cfg.get("name_col", 2)
    bcol = cfg.get("base_col", "C")
    ecol = cfg.get("end_col", "D")
    acol = cfg.get("adj_col", "G")
    hr = cfg.get("header_row", 9)
    end_anchor = _norm(cfg.get("data_end_anchor", "자본총계"))

    # 자본 flag 대분류 집계(정산표 순서 유지)
    order, agg = [], {}
    for rec in tb_rows:
        if not rec.get("자본"):
            continue
        kn = _norm(rec["대분류"])
        if kn not in agg:
            agg[kn] = {"name": rec["대분류"], "기초": 0, "기말": 0}
            order.append(kn)
        agg[kn]["기초"] += rec["기초"] or 0
        agg[kn]["기말"] += rec["기말"] or 0

    def cat_of(kn):
        for roman_kw, kws in roman_groups.items():     # 위에서부터 먼저 매칭(특수→일반)
            if any(_norm(x) in kn for x in kws):
                return roman_kw
        return None

    members = {rk: [] for rk in roman_groups}
    for kn in order:
        c = cat_of(kn)
        if c:
            members[c].append(agg[kn])

    re_adj = sum(l["이익잉여금"] for e in adj_entries for l in e["lines"] if l.get("이익잉여금"))

    wb = openpyxl.load_workbook(template)
    ws = wb[sheet]
    ncols = ws.max_column

    # 로마자 행과 각 leaf 영역 [로마자+1, 다음 로마자) 파악
    roman_rows = []
    for r in range(hr + 1, ws.max_row + 1):
        c = _norm(ws.cell(r, nci).value)
        if not c:
            continue
        if c.startswith(end_anchor):
            break
        if c[:1] in _ROMAN:
            cat = next((rk for rk in roman_groups if _norm(rk) in c), None)
            roman_rows.append((r, cat))
    bounds = []
    for i, (rr, cat) in enumerate(roman_rows):
        nxt = roman_rows[i + 1][0] if i + 1 < len(roman_rows) else None
        if nxt is None:
            te = rr + 1
            while te <= ws.max_row and not _norm(ws.cell(te, nci).value).startswith(end_anchor):
                te += 1
            nxt = te
        bounds.append((rr, cat, rr + 1, nxt))      # leaf 영역 [rr+1, nxt)

    n_filled = 0
    for rr, cat, ls, le in reversed(bounds):        # 아래→위
        mem = members.get(cat, []) if cat else []
        avail = le - ls
        need = max(len(mem), 1)                      # 최소 1행(빈 분류도 leaf 1행 유지)
        if need > avail:
            ins = need - avail
            ws.insert_rows(le, ins)
            shift_formula_rows(ws, le, ins)         # 자본총계 등 절대참조 보정
            for k in range(ins):
                _copy_style(ws, ls, le + k, ncols)  # 도너=기존 첫 leaf
            le += ins
        r = ls
        for m in mem:
            ws.cell(r, nci).value = m["name"]
            ws[f"{bcol}{r}"] = m["기초"]; ws[f"{bcol}{r}"].number_format = _NUMF
            ws[f"{ecol}{r}"] = m["기말"]; ws[f"{ecol}{r}"].number_format = _NUMF
            if "미처분이익잉여금" in _norm(m["name"]) and re_adj:
                ws[f"{acol}{r}"] = re_adj; ws[f"{acol}{r}"].number_format = _NUMF
            else:
                ws[f"{acol}{r}"] = None
            n_filled += 1
            r += 1
        for rb in range(r, le):                     # 잉여 leaf 비움(값·라벨)
            ws.cell(rb, nci).value = None
            ws[f"{bcol}{rb}"] = None; ws[f"{ecol}{rb}"] = None; ws[f"{acol}{rb}"] = None
        last = r - 1 if mem else ls                  # SUM 범위 끝
        if mem:
            ws[f"{bcol}{rr}"] = f"=SUM({bcol}{ls}:{bcol}{last})"
            ws[f"{ecol}{rr}"] = f"=SUM({ecol}{ls}:{ecol}{last})"
        else:
            ws[f"{bcol}{rr}"] = 0; ws[f"{ecol}{rr}"] = 0
        ws[f"{bcol}{rr}"].number_format = _NUMF
        ws[f"{ecol}{rr}"].number_format = _NUMF

    for col in (bcol, ecol, "E", acol, "H"):
        cur = ws.column_dimensions[col].width or 0
        ws.column_dimensions[col].width = max(cur, 16)
    n_appr = _fill_appropriation(ws, ce, bcol=bcol, ecol=ecol, acol=acol)  # 이잉처분계산서 전기칼럼·헤더
    Path(output).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return {"n_filled": n_filled, "re_adj": re_adj, "mode": "roman", "n_appr": n_appr}


def _common_prefix(a: str, b: str) -> int:
    n = 0
    for x, y in zip(a, b):
        if x == y:
            n += 1
        else:
            break
    return n


def fill_capital(template: str, tb_rows: list, adj_entries: list, output: str, cfg: dict,
                 settlement: "str | None" = None) -> dict:
    """자본 총괄표를 정산표로 in-place 리필한다.

    cfg: {sheet, header_row, data_end_anchor, name_col, base_col, end_col, adj_col}
    roman_groups가 있으면 로마자 분류 그룹핑 방식(_fill_capital_roman)으로 동적 렌더(누락 계정 편입).
    settlement: 정산표 경로(주어지면 CE 시트=자본변동표에서 이익잉여금처분계산서 전기칼럼·연도헤더를 채움).
    """
    ce = None
    if settlement:
        try:
            from extractors.equity_changes import parse_equity_changes
            ce = parse_equity_changes(settlement, cfg.get("ce_sheet", "CE"))
        except Exception as e:                          # 부분 실패 허용 — 자본 총괄표는 정상 생성
            import warnings
            warnings.warn(f"[자본 이잉처분계산서] CE 파싱 실패: {type(e).__name__}: {e}")

    roman_groups = cfg.get("roman_groups")
    if roman_groups:
        return _fill_capital_roman(template, tb_rows, adj_entries, output, cfg, roman_groups, ce=ce)

    sheet = cfg["sheet"]
    nci = cfg.get("name_col", 2)        # 계정명 열(1-indexed)
    bcol = cfg.get("base_col", "C")     # 기초
    ecol = cfg.get("end_col", "D")      # 기말(회사제시)
    acol = cfg.get("adj_col", "G")      # 수정사항
    hr = cfg.get("header_row", 9)
    end_anchor = _norm(cfg.get("data_end_anchor", "자본총계"))

    # 정산표 대분류별 기초/기말 합산
    by_class = defaultdict(lambda: {"기초": 0, "기말": 0})
    for rec in tb_rows:
        g = by_class[_norm(rec["대분류"])]
        g["기초"] += rec["기초"] or 0
        g["기말"] += rec["기말"] or 0

    # 명시적 매핑(자본 leaf 라벨 → 정산표 대분류) — 자본 계정은 명칭이 비슷해 fuzzy 오매칭 위험
    # (예: '매도가능증권평가익'(자본)이 '매도가능증권'(자산)에 걸림). 명시 매핑으로 정확히 잡는다.
    name_map = {_norm(k): _norm(v) for k, v in cfg.get("name_map", {}).items()}

    def find(name):
        n = _norm(name)
        if name_map:
            tgt = name_map.get(n)
            return by_class.get(tgt) if tgt else None
        if n in by_class:
            return by_class[n]
        best, bl = None, 3
        for k in by_class:
            c = _common_prefix(k, n)
            if c > bl:
                bl, best = c, k
        return by_class[best] if best else None

    # 이익잉여금 수정사항 = 수정분개의 이익잉여금 Effect 합계
    re_adj = sum(l["이익잉여금"] for e in adj_entries for l in e["lines"] if l.get("이익잉여금"))

    wb = openpyxl.load_workbook(template)
    ws = wb[sheet]
    n_filled = 0
    for r in range(hr + 1, ws.max_row + 1):
        b = ws.cell(r, nci).value
        nb = _norm(b)
        if not nb:
            continue
        if nb.startswith(end_anchor):
            break
        if nb.startswith(_ROMAN) or "자본총계" in nb:
            continue                       # 로마/총계는 수식 자동 재계산
        rec = find(b)
        if rec:
            ws[f"{bcol}{r}"] = rec["기초"]; ws[f"{bcol}{r}"].number_format = _NUMF
            ws[f"{ecol}{r}"] = rec["기말"]; ws[f"{ecol}{r}"].number_format = _NUMF
            if "미처분이익잉여금" in nb and re_adj:
                ws[f"{acol}{r}"] = re_adj; ws[f"{acol}{r}"].number_format = _NUMF
            n_filled += 1
        else:
            # ★ 블랭크화: 정산표에 없는 leaf(타 회사 잔재)는 값을 비운다(0). 라벨/구조는 보존.
            ws[f"{bcol}{r}"] = 0; ws[f"{ecol}{r}"] = 0
            ws[f"{acol}{r}"] = None

    for col in (bcol, ecol, "E", acol, "H"):
        cur = ws.column_dimensions[col].width or 0
        ws.column_dimensions[col].width = max(cur, 16)

    n_appr = _fill_appropriation(ws, ce, bcol=bcol, ecol=ecol, acol=acol)  # 이잉처분계산서 전기칼럼·헤더
    Path = __import__("pathlib").Path
    Path(output).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return {"n_filled": n_filled, "re_adj": re_adj, "n_appr": n_appr}
