# -*- coding: utf-8 -*-
"""G100 총괄표 이동 컬럼(취득·처분·감가상각비) 채우기 — 분개장 + CF 기반.

총괄표 본문은 엔진(A0Generator)이 기초·기말·수정후만 채우고 이동 컬럼은 비운다(grid_skip).
이 후처리가 **분개장 취득/처분**과 **CF 자산별 감가상각비**로 그 컬럼을 채운다(사용자 지시:
취득/처분은 분개장 분석만, 감가상각비는 CF, 나머지[역분개·대체·손상] 칸은 무시).

부호 규칙(처분 시 감가상각누계액이 −부호로 빠지는 것 포함):
  · 유형 본체행(건물·차량운반구…): E취득 = +분개장 차변, F처분 = −분개장 대변(실제 처분원가).
  · 감가상각누계액행("X 감가상각누계액"): J감가상각비 = −CF[X], F처분 = +누계감소(분개장 차변).
  · 무형 본체행(누계행 없음, 예 소프트웨어, 직접법): J감가상각비 = −CF[X],
    F처분 = −(분개장 대변 − CF[X])  (무형 대변엔 상각이 섞여 있어 CF만큼 빼야 실제 처분).
"""

import re

from openpyxl.utils import column_index_from_string


def _norm(s) -> str:
    return re.sub(r"\s+", "", str(s)) if s is not None else ""


def fill_note(ws, cfg: dict, base_row: dict, accum_row: dict) -> dict:
    """주석 4.변동표 — 헬퍼 테이블(U~AF)을 G100 참조 수식으로 미러링.

    헬퍼 계정명(U열)을 G100 본체/누계 행에 매칭해 V~AC를 '=G100!{col}{row}'로 채운다.
    보이는 변동표(ROUND(헬퍼/1000))는 자동 계산된다. 헬퍼에 매칭 G100 행이 없으면 그 행 비움.
    """
    note = cfg.get("note")
    if not note:
        return {"filled": 0}
    g100 = note.get("g100_sheet", "G100_총괄표")
    nci = column_index_from_string(note.get("mirror_name_col", "U"))
    scan = note.get("mirror_scan_rows", [6, 26])
    m2g = note.get("mirror_to_g100", {})
    accum_kw = "감가상각누계액"

    def g100_row_of(name):
        nn = _norm(name)
        if accum_kw in nn:
            base = _norm(re.sub(accum_kw, "", str(name)).strip(" _"))
            return accum_row.get(base)
        return base_row.get(nn)

    filled = 0
    for r in range(scan[0], scan[1] + 30):
        nm = ws.cell(r, nci).value
        if not nm or not str(nm).strip():
            continue
        n = _norm(nm)
        if n in ("유형자산", "무형자산", "합계") or n.startswith("계정"):
            continue
        gr = g100_row_of(nm)
        for hcol, gcol in m2g.items():
            cell = ws[f"{hcol}{r}"]
            cell.value = f"='{g100}'!{gcol}{gr}" if gr else None
        if gr:
            filled += 1
    return {"filled": filled}


def _group_memo(entries):
    """메모 entries를 (내용연수, 상각방법)으로 묶어 [(계정과목들, 연수, 방법들)] 반환."""
    merged, order = {}, []
    for m in entries:
        key = (m.get("내용연수"), tuple(m.get("상각방법") or []))
        if key not in merged:
            merged[key] = []
            order.append(key)
        merged[key].append(m["계정과목"])
    return [(merged[k], k[0], k[1]) for k in order]


def fill_g100_policy(ws, cfg: dict, memo: list) -> dict:
    """회계정책 박스(2.8 유형자산표·2.9 무형자산표)를 메모 기반으로 채운다.

    '회계정책' 라벨 아래 첫 '계정과목' 표머리=유형, 둘째=무형. 각 표 데이터행을 비우고
    메모(구분별, (내용연수,방법) 묶음)로 다시 쓴다. 항목이 기존 공간보다 많으면 insert_rows.
    """
    from .base import shift_formula_rows
    pol = cfg.get("policy")
    if not pol or not memo:
        return {"유형": 0, "무형": 0}
    label = _norm(pol.get("header_label", "계정과목"))
    sect = _norm(pol.get("section_anchor", "회계정책"))
    cc, cd, ce = pol.get("계정과목", "C"), pol.get("내용연수", "D"), pol.get("방법", "E")
    cci = column_index_from_string(cc)

    # '회계정책' 라벨 행 찾기
    sect_row = None
    for r in range(1, ws.max_row + 1):
        if sect in _norm(ws.cell(r, cci).value):
            sect_row = r
            break
    if sect_row is None:
        return {"유형": 0, "무형": 0}
    # 그 아래 '계정과목' 표머리 2개(유형/무형)
    headers = [r for r in range(sect_row, ws.max_row + 1)
               if _norm(ws.cell(r, cci).value) == label][:2]
    by_kind = {"유형": [], "무형": []}
    for m in memo:
        by_kind.get(m.get("구분", "유형"), by_kind["유형"]).append(m)

    out = {"유형": 0, "무형": 0}
    kinds = ["유형", "무형"][: len(headers)]
    # 아래→위 처리(아래 표 행삽입이 위 표 머리 위치를 밀지 않도록)
    for hdr, kind in reversed(list(zip(headers, kinds))):
        groups = _group_memo(by_kind.get(kind, []))
        start = hdr + 1
        end = start                            # 데이터 영역 = 표머리+1 ~ 다음 빈 줄/섹션 전
        for r in range(start, ws.max_row + 1):
            v = ws.cell(r, cci).value
            if v in (None, "") or label in _norm(v) or re.match(r"\d+[.)]", str(v or "").strip()):
                break
            end = r + 1
        avail = end - start
        need = len(groups)
        if need > avail and need > 0:           # 공간 부족 → 행 삽입(도너=표머리+1 스타일 복사)
            n = need - avail
            ws.insert_rows(start + avail, n)
            shift_formula_rows(ws, start + avail, n)
            from copy import copy
            for nr in range(start + avail, start + avail + n):
                for col in (cci, cci + 1, cci + 2):
                    src, dst = ws.cell(start, col), ws.cell(nr, col)
                    dst.border, dst.font = copy(src.border), copy(src.font)
                    dst.alignment = copy(src.alignment)
        for r in range(start, start + max(avail, need)):   # 기존 데이터 비움
            for col in (cc, cd, ce):
                ws[f"{col}{r}"] = None
        for i, (accts, years, methods) in enumerate(groups):
            r = start + i
            ws[f"{cc}{r}"] = ", ".join(accts)
            ws[f"{cd}{r}"] = f"{years}년" if years else ""
            ws[f"{ce}{r}"] = ", ".join(f"{x}법" for x in methods)
        out[kind] = need
    return out


def fill_g100_movements(ws, cfg: dict, movements: dict, cf_dep: dict) -> dict:
    """G100 시트 본문 이동 컬럼을 채운다.

    Args:
        ws: openpyxl worksheet (G100_총괄표).
        cfg: g.yaml (movement_cols / movement_name_col / accum_suffix / body 포함).
        movements: {자산계정: {취득합, 처분합, 누계감소, ...}} (parse_fixed_asset_movements).
        cf_dep: {자산계정: 당기감가상각비} (parse_cf_depreciation).
    Returns:
        {filled: 채운 행 수}
    """
    mc = cfg["movement_cols"]
    name_col = column_index_from_string(cfg.get("movement_name_col", "C"))
    accum_kw = _norm(cfg.get("accum_suffix", "감가상각누계액"))
    numf = cfg.get("body", {}).get("num_format", "#,##0")
    E, F, J = mc["취득"], mc["처분"], mc["감가상각비"]
    Dc = cfg.get("movement_base_col", "D")     # 기초
    Kc = cfg.get("movement_end_col", "K")      # 기말

    ds = cfg.get("body", {}).get("data_start_row", 11)
    end_anchor = _norm(cfg.get("body", {}).get("body_end_anchor", "계정과목"))

    def _base(name):
        """'감가상각누계액_건물' / '건물 감가상각누계액' → '건물'(누계 키워드·구분자 제거)."""
        t = re.sub(r"감가상각누계액|상각누계액", "", str(name))
        return t.strip(" _\t").strip()

    # 본문 행 범위 + 계정명 수집(누계행 존재 판정용)
    names = {}
    for r in range(ds, ws.max_row + 1):
        v = ws.cell(r, name_col).value
        nm = str(v).strip() if v not in (None, "") else ""
        if nm and _norm(nm).startswith(end_anchor):   # 감가상각비 명세 표머리 → 본문 끝
            break
        if nm:
            names[r] = nm
    # 누계행이 존재하는 자산(base) 집합 — 유형/무형(직접법) 구분
    accum_bases = {_norm(_base(n)) for n in names.values() if accum_kw in _norm(n)}

    def put(r, col, val):
        if val in (None, 0):
            return
        c = ws[f"{col}{r}"]
        c.value = val
        c.number_format = numf

    def mv(name):
        return movements.get(name) or movements.get(_norm(name)) or {}

    def cf(name):
        if name in cf_dep:
            return cf_dep[name]
        nn = _norm(name)
        for k, v in cf_dep.items():
            if _norm(k) == nn:
                return v
        return 0

    filled = 0
    for r, nm in names.items():
        nn = _norm(nm)
        if accum_kw in nn:                        # 감가상각누계액 행
            base = _base(nm)
            dep = cf(base)
            jval = -dep if dep else 0
            put(r, J, jval)
            # 처분 시 감가상각누계액 제거(−부호로 빠짐) = 기말−기초−감가비 (누계는 상각+처분제거로만 변동)
            기초 = ws[f"{Dc}{r}"].value or 0
            기말 = ws[f"{Kc}{r}"].value or 0
            if isinstance(기초, (int, float)) and isinstance(기말, (int, float)):
                처분제거 = 기말 - 기초 - jval
                put(r, F, 처분제거 if 처분제거 else 0)
            if dep:
                filled += 1
            continue
        # 본체 자산행
        m = mv(nm)
        취득 = m.get("취득합") or 0
        처분 = m.get("처분합") or 0
        has_accum = nn in accum_bases
        put(r, E, 취득)
        if has_accum:                            # 유형(간접법): 대변=실제 처분
            put(r, F, -처분 if 처분 else 0)
        else:                                    # 무형(직접법): 본체에 상각, 대변−CF=실제 처분
            dep = cf(nm)
            put(r, J, -dep if dep else 0)
            real = 처분 - dep
            put(r, F, -real if real else 0)
        if 취득 or 처분 or (not has_accum and cf(nm)):
            filled += 1
    return {"filled": filled}
