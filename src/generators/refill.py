"""고정 계산행 총괄표용 범용 in-place 리필 생성기.

Q(매출원가) 처럼 본문이 계정 리스트가 아니라 **고정 계산행**(기초미완성·당기총공사비용·계·
공사매출원가 등, 일부는 수식)인 조서에 쓴다. 완성본 격자를 그대로 두고, 계정명이 정산표
(제조원가명세서 포함)에 매칭되는 행만 기초/기말/수정사항을 채운다(나머지 수식/라벨 보존).
"""

import re
from collections import defaultdict
from pathlib import Path

import openpyxl

_NUMF = '#,##0;[Red]\\(#,##0\\);"-"'


def fill_rows(template, tb_rows, output, cfg):
    """고정 다중행 리필 — 행마다 컬럼이 다른 총괄표(B 매도가능증권 등)용.

    각 행(label)을 정산표(name=대분류/계정명)의 기초/기말/수정사항으로 채운다. 행별로 컬럼이
    다를 수 있다(B: BS행 기초C/기말H/수정I, IS행 기초C/기말D/수정G). 중간 명세(취득·평가·처분·
    대체 등)는 비워둔다(개별 명세서 후속) — 수정후/증감 등은 템플릿 수식이 자동 계산.
    정산표에 없는 행은 0으로 비운다(타 회사 완성본 잔재 제거).

    cfg: {sheet, name_col, rows:[{label, name, base_col, end_col, adj_col}, ...]}
    """
    sheet = cfg["sheet"]
    nci = cfg.get("name_col", 2)
    by_class = defaultdict(lambda: {"기초": 0, "기말": 0, "수정사항": 0})
    by_name = {}
    for rec in tb_rows:
        g = by_class[_norm(rec["대분류"])]
        for k in ("기초", "기말", "수정사항"):
            g[k] += rec[k] or 0
        by_name[_norm(rec["계정명"])] = {"기초": rec["기초"] or 0, "기말": rec["기말"] or 0,
                                        "수정사항": rec["수정사항"] or 0}
    wb = openpyxl.load_workbook(template)
    ws = wb[sheet]
    n = 0
    cols = set()
    for spec in cfg.get("rows", []):
        lbl, nm = _norm(spec["label"]), _norm(spec["name"])
        rec = by_name.get(nm) or by_class.get(nm)
        bcol, ecol, acol = spec["base_col"], spec["end_col"], spec["adj_col"]
        cols.update([bcol, ecol, acol])
        for r in range(1, ws.max_row + 1):
            if _norm(ws.cell(r, nci).value) == lbl:
                if rec and any(rec[k] for k in ("기초", "기말", "수정사항")):
                    ws[f"{bcol}{r}"] = rec["기초"]; ws[f"{bcol}{r}"].number_format = _NUMF
                    ws[f"{ecol}{r}"] = rec["기말"]; ws[f"{ecol}{r}"].number_format = _NUMF
                    if rec["수정사항"]:
                        ws[f"{acol}{r}"] = rec["수정사항"]; ws[f"{acol}{r}"].number_format = _NUMF
                    n += 1
                else:                       # 정산표에 없음 → 비움(잔재 제거)
                    ws[f"{bcol}{r}"] = 0; ws[f"{ecol}{r}"] = 0; ws[f"{acol}{r}"] = None
                break
    for c in cols:
        cur = ws.column_dimensions[c].width or 0
        ws.column_dimensions[c].width = max(cur, 16)
    Path(output).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return {"n_filled": n}


def _norm(s):
    return re.sub(r"\s+", "", str(s)) if s is not None else ""


def _cp(a, b):
    n = 0
    for x, y in zip(a, b):
        if x == y:
            n += 1
        else:
            break
    return n


def fill_refill(template, tb_rows, output, cfg):
    """cfg: {sheet, header_row, name_col, base_col, end_col, adj_col, formulas{col:tmpl}, data_end_anchor}"""
    sheet = cfg["sheet"]
    nci = cfg.get("name_col", 2)
    bcol, ecol, acol = cfg["base_col"], cfg["end_col"], cfg["adj_col"]
    hr = cfg["header_row"]
    end_anchor = _norm(cfg.get("data_end_anchor", "기초"))
    formulas = cfg.get("formulas", {})

    by_class = defaultdict(lambda: {"기초": 0, "기말": 0, "수정사항": 0})
    by_name = {}
    for rec in tb_rows:
        g = by_class[_norm(rec["대분류"])]
        g["기초"] += rec["기초"] or 0
        g["기말"] += rec["기말"] or 0
        g["수정사항"] += rec["수정사항"] or 0
        by_name[_norm(rec["계정명"])] = {"기초": rec["기초"] or 0, "기말": rec["기말"] or 0,
                                        "수정사항": rec["수정사항"] or 0}

    # 명시적 매핑(구분 라벨 → 정산표 계정명/규칙) — R-2 인건비처럼 판관비/노무비 구분이 필요한 교차조서용.
    #   값이 문자열이면: 정산표 계정명(없으면 대분류) 매칭.
    #   값이 dict(규칙)이면: {제조: bool?, 포함: [kw...], 제외: [kw...]}로 tb_rows를 직접 집계.
    #     제조 인건비는 제조원가명세서 급여/상여금이 판관비와 대분류가 같을 수 있어(예: 대분류='급여')
    #     단순 계정명/대분류 매칭으론 못 가르므로, 제조 flag + 계정명 키워드로 합산한다.
    name_map = {_norm(k): v for k, v in cfg.get("name_map", {}).items()}

    def _match_rule(rule):
        want_mfg = rule.get("제조")
        inc = [_norm(x) for x in rule.get("포함", [])]
        exc = [_norm(x) for x in rule.get("제외", [])]
        agg = {"기초": 0, "기말": 0, "수정사항": 0}
        hit = False
        for rec in tb_rows:
            if want_mfg is not None and bool(rec.get("제조원가")) != bool(want_mfg):
                continue
            text = _norm(rec["대분류"]) + _norm(rec["계정명"])
            if inc and not any(k in text for k in inc):
                continue
            if exc and any(k in text for k in exc):
                continue
            hit = True
            agg["기초"] += rec["기초"] or 0
            agg["기말"] += rec["기말"] or 0
            agg["수정사항"] += rec["수정사항"] or 0
        return agg if hit else {"기초": 0, "기말": 0, "수정사항": 0}

    def find(name):
        n = _norm(name)
        if name_map:
            tgt = name_map.get(n)
            if tgt is None:
                return None
            if isinstance(tgt, dict):
                return _match_rule(tgt)            # 규칙 기반 집계(제조 flag+키워드)
            t = _norm(tgt)
            return by_name.get(t) or by_class.get(t)   # 계정명 우선, 없으면 대분류 합산
        if n in by_class:
            return by_class[n]
        best, bl = None, 4
        for k in by_class:
            c = _cp(k, n)
            if c > bl:
                bl, best = c, k
        return by_class[best] if best else None

    wb = openpyxl.load_workbook(template)
    ws = wb[sheet]
    n = 0
    for r in range(hr + 1, ws.max_row + 1):
        nm = ws.cell(r, nci).value
        if _norm(nm).startswith(end_anchor):
            break
        rec = find(nm)
        if not rec:
            # ★ 블랭크화: name_map에 있는 데이터 행인데 정산표 매칭 없음(타 회사 잔재) → 비움(0).
            #   라벨/수식행(name_map에 없음)은 건드리지 않는다.
            if name_map and _norm(nm) in name_map:
                ws[f"{bcol}{r}"] = 0; ws[f"{ecol}{r}"] = 0; ws[f"{acol}{r}"] = None
            continue
        ws[f"{bcol}{r}"] = rec["기초"]; ws[f"{bcol}{r}"].number_format = _NUMF
        ws[f"{ecol}{r}"] = rec["기말"]; ws[f"{ecol}{r}"].number_format = _NUMF
        if rec["수정사항"]:
            ws[f"{acol}{r}"] = rec["수정사항"]; ws[f"{acol}{r}"].number_format = _NUMF
        for col, tmpl in formulas.items():
            ws[f"{col}{r}"] = tmpl.format(r=r); ws[f"{col}{r}"].number_format = _NUMF
        n += 1
    # 너비 보정: 데이터 블록 전체(기초~증감율 등, 수식 계산열 포함)를 넉넉히 — 큰 금액 ####### 방지.
    from openpyxl.utils import column_index_from_string, get_column_letter
    left = min(column_index_from_string(c) for c in (bcol, ecol, acol))
    for ci in range(left, left + 7):
        col = get_column_letter(ci)
        cur = ws.column_dimensions[col].width or 0
        ws.column_dimensions[col].width = max(cur, 16)
    import pathlib
    pathlib.Path(output).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return {"n_filled": n}
