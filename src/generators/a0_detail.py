# -*- coding: utf-8 -*-
"""A-0(현금및현금성자산) 개별시트 매핑 — 총괄표 연동.

총괄표(정산표 매핑)가 채워진 뒤, 그 값에 **연동**해 개별시트를 동적 구성한다:
  - A050 주석   : 총괄표 계정별 수정후(J)/기초(E)를 =ROUND(!/1000,0)로 참조(이름 매칭, 없는 계정 skip).
  - A020 실사   : 기말환율표(참고자료) 배치 + 외화 원금×환율=원화환산액 가이드.
  - 보험가입현황 : 조회서 INSURANCE 해약환급금>0 per-policy 매핑(거래처/보험명/가입/만기/잔액).
  - 총괄표 정리  : 잔재 <유동> 마커 제거.

순수 ws 편집 함수 — 호출자(easy_run_a0)가 extract_one+graft로 완성본 보존하며 적용한다.
"""
import re
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.cell.cell import MergedCell

_thin = Side(style="thin")
_BOX = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_CEN = Alignment(horizontal="center", vertical="center")
_YEL = PatternFill("solid", fgColor="FFFF00")
_NOFILL = PatternFill(fill_type=None)


def _amt(v):
    if v is None:
        return 0
    s = re.sub(r"[^0-9.\-]", "", str(v))
    try:
        return float(s) if s else 0
    except ValueError:
        return 0


def _fmt_date(v):
    s = re.sub(r"[^0-9]", "", str(v or ""))
    if len(s) == 8:
        y = int(s[:4])
        return "종신" if y > 2100 else "%s-%s-%s" % (s[:4], s[4:6], s[6:8])
    return str(v or "")


# ── 총괄표 잔재 마커 정리 ───────────────────────────────────────────────
def cleanup_total_marker(ws):
    """헤더 바로 아래 잔재 <유동> 마커(생성된 마커와 중복) 제거."""
    for r in range(10, 13):
        for c in range(2, 7):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "<유동>" in v and ws.cell(r + 1, 3).value in ("<유동>", "<비유동>"):
                ws.cell(r, c).value = None


# ── A050 주석 ← 총괄표 ROUND 연동 ───────────────────────────────────────
def _scan_total(total_ws):
    """총괄표에서 {계정명: 행}, {대분류: 합계행} 추출(연동 참조용)."""
    acct, sub, last = {}, {}, None
    for r in range(12, 40):
        C = total_ws.cell(r, 3).value
        D = total_ws.cell(r, 4).value
        cs = str(C).strip() if C else ""
        if cs and cs not in ("합계", "<유동>", "<비유동>"):
            last = cs
            if D:
                acct[str(D).strip()] = r
        elif cs == "합계" and last:
            sub.setdefault(last, r)
    return acct, sub


def fill_a050_notes(ws, total_ws, total_sheet="4000_A000 총괄표"):
    """A050 라벨(B열)을 총괄표 계정과 이름매칭 → 당기말/전기말 ROUND 수식. 매칭 안 되면 skip(연동 없음)."""
    acct, sub = _scan_total(total_ws)
    R = lambda tr, col: "=ROUND('%s'!%s%d/1000,0)" % (total_sheet, col, tr)

    def tot_row(label):
        """A050 라벨 → 총괄표 행. 계정명 직접 매칭 → 대분류 합계('계'/'합계' 라벨) 매칭."""
        l = label.replace(" ", "")
        if label in acct:
            return acct[label], "acct"
        for name, r in acct.items():
            if name.replace(" ", "") == l:
                return r, "acct"
        if "현금및현금성" in l:
            return sub.get("현금및현금성자산"), "sub"
        if "장기금융" in l or "저축성보험" in l:
            return acct.get("장기금융상품") or (list(sub.get("장기금융상품", [None]) for _ in [0])[0]), "acct"
        return None, None

    for r in range(6, 14):
        label = ws.cell(r, 2).value
        if not isinstance(label, str) or not label.strip():
            continue
        lab = label.strip()
        # 계 / 합계 행은 A050 내부 합산 수식으로(연동은 sub-account가 담당)
        if lab.endswith("계") or lab == "합  계" or lab == "합계":
            continue
        tr, kind = tot_row(lab)
        if not tr:
            continue
        for cc, dc in (("C", "D"), ("J", "K")):
            ws["%s%d" % (cc, r)] = R(tr, "J")   # 당기말 = 수정후
            ws["%s%d" % (dc, r)] = R(tr, "E")   # 전기말 = 기초
            ws["%s%d" % (cc, r)].number_format = "#,##0"
            ws["%s%d" % (dc, r)].number_format = "#,##0"
    # A열 잔재 마커 제거
    for r in range(4, 14):
        a = ws.cell(r, 1)
        if not isinstance(a, MergedCell) and isinstance(a.value, str) and "수치" in a.value:
            a.value = None
            a.fill = _NOFILL


# ── A020 금융자산 실사 ← 기말환율표 ─────────────────────────────────────
def fill_a020_fx(ws, fx_rates, anchor=(12, 10)):
    """기말환율표(참고) 배치 + 외화 원금×환율=원화환산액 가이드. fx_rates={통화:환율}."""
    if not fx_rates:
        return
    r0, c0 = anchor
    ws.cell(r0, c0).value = "기말환율(참고)"
    ws.cell(r0, c0).font = Font(bold=True); ws.cell(r0, c0).alignment = _CEN
    ws.cell(r0 + 1, c0).value = "통화"; ws.cell(r0 + 1, c0 + 1).value = "환율(₩)"
    ws.cell(r0 + 1, c0).font = ws.cell(r0 + 1, c0 + 1).font = Font(bold=True)
    rr = r0 + 2
    for cur in sorted(fx_rates):
        ws.cell(rr, c0).value = cur
        v = ws.cell(rr, c0 + 1); v.value = fx_rates[cur]; v.number_format = "#,##0.00"
        rr += 1
    for ri in range(r0, rr):
        for ci in (c0, c0 + 1):
            ws.cell(ri, ci).border = _BOX
            if ws.cell(ri, ci).alignment.horizontal is None:
                ws.cell(ri, ci).alignment = _CEN
    a = ws["A14"]
    a.value = "[수치] 외화 계좌가 있으면 원금(외화)×기말환율 = 원화환산액(F열). 환율 우측 표 참조."
    a.fill = _YEL; a.font = Font(italic=True, size=9, color="806000")


# ── 장기금융상품 보험가입현황 ← 조회서 INSURANCE ───────────────────────
def parse_insurance_holdings(confirm_xlsx):
    """조회서 INSURANCE 1번 섹션에서 해약환급금>0 per-policy 추출(보험가입현황 매핑용)."""
    import openpyxl
    from extractors._sections import find_section_bounds
    from extractors._headers import map_headers
    # 해약환급금은 **숫자 전용열(_금액)** 우선, 없으면 표시열. (BANK와 동일 원칙)
    SYN = {"금융기관명": "거래처", "보험의 종류": "보험명", "보험의종류": "보험명", "증권번호": "증권번호",
           "부보기간_시작": "가입일", "부보기간_종료": "만기",
           "해약환급금_금액": "해약환급금_금액", "해약환급금": "해약환급금_disp"}
    out = []
    try:
        wb = openpyxl.load_workbook(confirm_xlsx, read_only=True, data_only=True)
        rows = (list(wb["INSURANCE"].iter_rows(values_only=True))
                if "INSURANCE" in wb.sheetnames else [])
        wb.close()
    except Exception:
        rows = []
    b = find_section_bounds(rows, "1.") if rows else None
    if b:
        si, se = b
        hi = cm = None
        for i in range(si + 1, min(si + 6, se)):
            m = map_headers(rows[i], synonyms=SYN, required={"거래처", "보험명"})
            if m and ("해약환급금_금액" in m or "해약환급금_disp" in m):
                hi, cm = i, m; break
        if cm:
            for i in range(hi + 1, se):
                r = rows[i]

                def g(k):
                    idx = cm.get(k)
                    return r[idx] if idx is not None and idx < len(r) else None

                def surr():
                    v = g("해약환급금_금액")
                    return v if v not in (None, "") else g("해약환급금_disp")
                if not g("거래처") or _amt(surr()) <= 0:
                    continue
                out.append({"거래처": str(g("거래처")).strip(), "보험명": str(g("보험명") or "").strip(),
                            "증권번호": str(g("증권번호") or "").strip(), "가입일": _fmt_date(g("가입일")),
                            "만기": _fmt_date(g("만기")), "잔액": int(_amt(surr()))})
    # 기타 시트(GUARANTEE 등)의 해약환급금도 장기금융 보험가입현황에 포함(위치 불문 스캔)
    try:
        from extractors.surrender import parse_surrender_value
        for s in parse_surrender_value(confirm_xlsx, exclude_sheets=("INSURANCE",)):
            out.append({"거래처": s["금융기관명"], "보험명": s["보험의종류"],
                        "증권번호": s["증권번호"], "가입일": "", "만기": "",
                        "잔액": int(s["해약환급금"])})
    except Exception:
        pass
    return out


def fill_insurance_holdings(ws, policies):
    """보험가입현황 표(헤더+행+합계+가이드) 렌더. policies=parse_insurance_holdings 결과."""
    ws["A2"] = "보험가입현황"; ws["A2"].font = Font(bold=True, size=12)
    ws["A3"] = "기말 기준"; ws["I4"] = "(단위 : 원)"
    hdr = ["거래처", "보험명", "피보험자", "가입일자", "만기일자", "적요", "월 납입금액", "잔액"]
    cols = ["A", "B", "C", "D", "E", "F", "H", "I"]
    for col, h in zip(cols, hdr):
        c = ws["%s5" % col]; c.value = h; c.font = Font(bold=True); c.alignment = _CEN
        c.fill = PatternFill("solid", fgColor="D9E1F2"); c.border = _BOX
    n = max(len(policies), 8)
    for i in range(n):
        r = 6 + i
        for col in cols + ["G"]:
            ws["%s%d" % (col, r)].border = _BOX
        if i < len(policies):
            p = policies[i]
            ws["A%d" % r] = p["거래처"]; ws["B%d" % r] = p["보험명"]; ws["D%d" % r] = p["가입일"]
            ws["E%d" % r] = p["만기"]; ws["F%d" % r] = p["증권번호"]
            ws["I%d" % r] = p["잔액"]; ws["I%d" % r].number_format = "#,##0"
    tr = 6 + n
    ws["A%d" % tr] = "합  계"; ws["A%d" % tr].font = Font(bold=True)
    for col in cols + ["G"]:
        ws["%s%d" % (col, tr)].border = _BOX
    ws["I%d" % tr] = "=SUM(I6:I%d)" % (tr - 1); ws["I%d" % tr].font = Font(bold=True)
    ws["I%d" % tr].number_format = "#,##0"
    g = ws["A1"]
    g.value = ("[수치] 조회서 INSURANCE 회신 장기금융(보험 해약환급금) 매핑. 잔액 합계는 조회서 회신분 — "
               "미회신/회사 추가 보험은 회사 보험명세로 보완(총괄표 장기금융 기말과 대사).")
    g.fill = _YEL; g.font = Font(italic=True, size=9, color="806000")


# 보험가입현황 열너비(graft가 빈 템플릿 cols를 지우므로 XML로 재주입할 폭)
HOLDINGS_COL_WIDTHS = {1: 16, 2: 26, 3: 10, 4: 13, 5: 13, 6: 22, 8: 13, 9: 18}


# ── 총괄표 "5) 장기금융상품" 평가대사 표 ← A-1/조회서 기반 ───────────────
def _clean_name(s):
    return re.sub(r"\(.*?\)", "", str(s or "")).replace(" ", "").strip()


def build_longterm_recon(confirm_xlsx, ledger_path):
    """장기금융 평가대사: 조회서 보험사(해약환급금) ↔ 거래처원장 잔액(이름매칭) → 차액.

    계정명이 '보통예금'으로 뭉쳐 있어도 **조회서 보험사 ↔ 거래처 거래처명 매칭**으로 장기금융을
    식별한다(계정명 키워드 아님). 매칭은 가장 구체적인(긴) 거래처명 우선.
    Returns: [{거래처, 잔액, 해약환급금, 조서번호}] (차액=잔액-해약환급금은 시트에서 수식).
    """
    from collections import defaultdict
    from extractors.surrender import parse_longterm_groups
    from extractors.ledger import parse_ledger
    try:
        ins = parse_longterm_groups(confirm_xlsx)   # INSURANCE + 해약환급금(기타시트) 합산
    except Exception:
        ins = []
    by = {}
    for g in ins:
        k = _clean_name(g["금융기관명"])
        if not k:
            continue
        if k not in by:
            by[k] = {"name": g["금융기관명"], "해약환급금": 0, "조서번호": g.get("조서번호")}
        by[k]["해약환급금"] += _amt(g.get("조회서금액"))
    bal = defaultdict(float)
    try:
        for r in parse_ledger(ledger_path):
            bal[_clean_name(r.get("거래처명"))] += (r.get("잔액") or 0)
    except Exception:
        pass
    rows = []
    for k, v in by.items():
        cands = [(lk, lv) for lk, lv in bal.items() if lk and (lk in k or k in lk)]
        cands.sort(key=lambda x: -len(x[0]))      # 가장 구체적(긴) 이름 우선(부분일치 오매칭 방지)
        잔액 = cands[0][1] if cands else 0
        rows.append({"거래처": v["name"], "잔액": int(잔액),
                     "해약환급금": int(v["해약환급금"]), "조서번호": v["조서번호"]})
    return rows


def build_shortterm_recon(confirm_xlsx, ledger_path):
    """단기금융 평가대사: 조회서 INVESTMENT 2번 평가액 ↔ 거래처원장 잔액(이름매칭) → 차액.

    **단기금융 존재 판단** = 이 결과가 비면 단기금융 없음(4)는 prune). 평가액(시가)이 회사계상과
    다르면 미평가 의심. Returns: [{거래처, 잔액, 평가액, 조서번호}].
    """
    from collections import defaultdict
    from extractors.invest import parse_invest_eval
    from extractors.ledger import parse_ledger
    try:
        ev = parse_invest_eval(confirm_xlsx)
    except Exception:
        ev = []
    if not ev:
        return []                              # 단기금융 없음
    bal = defaultdict(float)
    try:
        for r in parse_ledger(ledger_path):
            bal[_clean_name(r.get("거래처명"))] += (r.get("잔액") or 0)
    except Exception:
        pass
    rows = []
    for e in ev:
        k = _clean_name(e.get("금융기관명"))
        cands = [(lk, lv) for lk, lv in bal.items() if lk and (lk in k or k in lk)]
        cands.sort(key=lambda x: -len(x[0]))
        잔액 = cands[0][1] if cands else 0
        rows.append({"거래처": e.get("금융기관명"), "잔액": int(잔액),
                     "평가액": int(_amt(e.get("평가액"))), "조서번호": None})
    return rows


def fill_total_longterm(ws, recon):
    """총괄표의 '5) 장기금융상품' 평가대사 표(거래처/잔액/해약환급금/차액/조서번호)를 채운다.

    헤더행을 동적으로 찾는다(해약환급금+거래처 포함). 데이터 행에 recon 매핑. 차액=수식.
    """
    from openpyxl.utils import get_column_letter as L
    def nz(c, r):  # 셀값 공백 제거(헤더가 '거 래 처 명'처럼 띄어쓰기 있음)
        return str(ws.cell(r, c).value or "").replace(" ", "")
    hdr = None
    for r in range(40, 90):
        vals = [nz(c, r) for c in range(3, 12)]
        if any("해약환급금" in v for v in vals) and any("거래처" in v for v in vals):
            hdr = r
            break
    if not hdr:
        return 0
    col = {}
    for c in range(3, 12):
        v = nz(c, hdr)
        if "거래처" in v:
            col["거래처"] = c
        elif "잔액" in v:
            col["잔액"] = c
        elif "해약" in v:
            col["해약"] = c
        elif "차액" in v:
            col["차액"] = c
        elif "조서" in v:
            col["조서"] = c
        elif "계정과목" in v:
            col["계정"] = c
    if "거래처" not in col or "해약" not in col:
        return 0
    r0 = hdr + 1
    for i, x in enumerate(recon):
        r = r0 + i
        if col.get("계정"):
            ws.cell(r, col["계정"]).value = "장기금융상품"
        ws.cell(r, col["거래처"]).value = x["거래처"]
        if col.get("잔액"):
            c = ws.cell(r, col["잔액"]); c.value = x["잔액"]; c.number_format = "#,##0"
        ch = ws.cell(r, col["해약"]); ch.value = x["해약환급금"]; ch.number_format = "#,##0"
        if col.get("차액") and col.get("잔액"):
            cd = ws.cell(r, col["차액"])
            cd.value = "=%s%d-%s%d" % (L(col["잔액"]), r, L(col["해약"]), r)
            cd.number_format = "#,##0"
        if col.get("조서"):
            ws.cell(r, col["조서"]).value = x["조서번호"]
    _fix_total_diff_sum(ws, col, r0)
    return len(recon)


def fill_total_shortterm(ws, recon):
    """총괄표 '4) 단기금융상품' 평가대사 표(거래처/잔액/평가액/차액)를 채운다(존재 시만).

    헤더에 '평가액'+'거래처'가 있는 단기 표를 동적으로 찾는다. 없으면(=양식에 4)표 없음) no-op.
    recon이 비면(단기금융 없음) 아무것도 안 함(prune).
    """
    from openpyxl.utils import get_column_letter as L
    if not recon:
        return 0

    def nz(c, r):
        return str(ws.cell(r, c).value or "").replace(" ", "")
    hdr = None
    for r in range(40, 90):
        vals = [nz(c, r) for c in range(3, 12)]
        if any("평가액" in v for v in vals) and any("거래처" in v for v in vals):
            hdr = r
            break
    if not hdr:
        return 0
    col = {}
    for c in range(3, 12):
        v = nz(c, hdr)
        if "거래처" in v:
            col["거래처"] = c
        elif "잔액" in v:
            col["잔액"] = c
        elif "평가액" in v:
            col["평가"] = c
        elif "차액" in v:
            col["차액"] = c
        elif "계정과목" in v:
            col["계정"] = c
    if "거래처" not in col or "평가" not in col:
        return 0
    r0 = hdr + 1
    for i, x in enumerate(recon):
        r = r0 + i
        if col.get("계정"):
            ws.cell(r, col["계정"]).value = "단기금융상품"
        ws.cell(r, col["거래처"]).value = x["거래처"]
        if col.get("잔액"):
            c = ws.cell(r, col["잔액"]); c.value = x["잔액"]; c.number_format = "#,##0"
        cp = ws.cell(r, col["평가"]); cp.value = x["평가액"]; cp.number_format = "#,##0"
        if col.get("차액") and col.get("잔액"):
            cd = ws.cell(r, col["차액"])
            cd.value = "=%s%d-%s%d" % (L(col["평가"]), r, L(col["잔액"]), r)
            cd.number_format = "#,##0"
    _fix_total_diff_sum(ws, col, r0)
    return len(recon)


def _fix_total_diff_sum(ws, col, r0):
    """총계행 차액 SUM 범위를 잔액 SUM과 동일하게 보정(템플릿이 첫행 누락하는 경우)."""
    from openpyxl.utils import get_column_letter as L
    if not (col.get("차액") and col.get("잔액")):
        return
    for rr in range(r0, r0 + 15):
        gv = ws.cell(rr, col["잔액"]).value
        if isinstance(gv, str) and gv.upper().startswith("=SUM"):
            m = re.search(r"(\d+)\s*:\s*[A-Z]+(\d+)", gv)
            if m:
                ws.cell(rr, col["차액"]).value = "=SUM(%s%s:%s%s)" % (
                    L(col["차액"]), m.group(1), L(col["차액"]), m.group(2))
            break
