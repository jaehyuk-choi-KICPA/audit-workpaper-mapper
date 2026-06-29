"""조서 생성기 공통 기반 클래스.

맵핑 후 양식 보존(스타일 복사 + 가변 섹션 동적 재구성)은 조서 종류와 무관한
공통 문제이므로 이 기반 클래스에 **재사용 도구**로 둔다. A-200/A-300/A-0 등이
상속해 동일하게 활용한다.
"""

import re
from copy import copy, deepcopy
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
import yaml


# 셀참조: 선택적 $ + 열(1~3자) + 선택적 $ + 행번호. 앞은 영숫자/_/!(시트참조) 아님,
# 뒤는 영숫자/_/'(' 아님 → 함수명(LOG10()·정의된이름(Q1Total) 오인 방지.
_REF_RE = re.compile(r"(?<![A-Za-z0-9_!])(\$?)([A-Z]{1,3})(\$?)(\d+)(?![A-Za-z0-9_(])")


def shift_formula_rows(ws, at_row: int, n: int) -> int:
    """ws의 모든 수식에서 at_row 이상을 가리키는 **행참조를 +n** 이동(insert_rows 직후 호출).

    openpyxl `insert_rows`는 셀 값·스타일만 내리고 **수식 문자열의 행참조는 그대로** 두므로,
    삽입 지점 아래로 밀린 보존영역 수식(예: EE PL표 `=E35-D35`, 체크행)이 엉뚱한 행(빈셀·텍스트
    헤더)을 가리켜 #VALUE!/#DIV/0!/#REF!가 난다. Excel처럼 at_row 이상 행참조를 일괄 +n 한다.
    같은 시트 단순 셀참조(A1·$A$1·범위 양끝)만 대상 — 시트간 참조('시트'!A1)는 `!` 뒤라 제외.
    이후 생성기가 명시적으로 덮어쓰는 셀(본문·합계)은 정상 수식으로 다시 써지므로 무해하다.
    n이 음수면 행 삭제(delete_rows) 보정: at_row 이상 행참조를 -|n| 이동(삭제구간 아래 행이
    위로 당겨진 것 반영). 삭제구간 내부를 가리키던 참조는 그대로 둔다(빈 placeholder 삭제 전제).
    Returns: 바뀐 수식 셀 수.
    """
    if n == 0:
        return 0

    def bump(m):
        dc, col, dr, row = m.groups()
        rr = int(row)
        return f"{dc}{col}{dr}{rr + n if rr >= at_row else rr}"

    changed = 0
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                nv = _REF_RE.sub(bump, v)
                if nv != v:
                    cell.value = nv
                    changed += 1
    return changed


def apply_col_offset(cfg: dict) -> dict:
    """config의 `col_offset`만큼 본문 컬럼 참조를 일괄 +이동(로드 시 1회).

    사용자가 총괄표 본문 앞에 빈 열을 추가하면 모든 본문 컬럼이 우측으로 밀린다. config 값
    자체를 다시 쓰지 않고(주석 보존), 로드 시 컬럼류 필드만 offset만큼 옮긴다. **행번호·셀주소
    (header_cells/conclusion)·계정명/앵커 텍스트는 건드리지 않는다** — 순수 본문 컬럼만.
    """
    off = cfg.get("col_offset", 0)
    if not off:
        return cfg
    cfg = deepcopy(cfg)

    def sl(L):                                  # 컬럼 문자 +off
        return get_column_letter(column_index_from_string(L) + off)

    def sf(s):                                  # 수식 내 'X{r}' 컬럼 참조 +off
        return re.sub(r"([A-Z]+)(?=\{r\})", lambda m: sl(m.group(1)), s)

    for k in ("base_col", "end_col", "adj_col", "pct_col"):
        if isinstance(cfg.get(k), str):
            cfg[k] = sl(cfg[k])
    if isinstance(cfg.get("name_col"), int):
        cfg["name_col"] += off
    if isinstance(cfg.get("formulas"), dict):
        cfg["formulas"] = {sl(k): sf(v) for k, v in cfg["formulas"].items()}
    for row in cfg.get("rows", []) or []:       # 고정 다중행(B) — 행별 컬럼
        for k in ("base_col", "end_col", "adj_col"):
            if isinstance(row.get(k), str):
                row[k] = sl(row[k])

    b = cfg.get("body")
    if isinstance(b, dict):
        if isinstance(b.get("marker_col"), str):
            b["marker_col"] = sl(b["marker_col"])
        if isinstance(b.get("columns"), dict):
            b["columns"] = {k: (sl(v) if isinstance(v, str) else v) for k, v in b["columns"].items()}
        if isinstance(b.get("detail_formulas"), dict):
            b["detail_formulas"] = {sl(k): sf(v) for k, v in b["detail_formulas"].items()}
        for k in ("num_cols", "pct_cols", "grid_skip"):
            if isinstance(b.get(k), list):
                b[k] = [sl(x) for x in b[k]]
        st = b.get("subtotal")
        if isinstance(st, dict):
            if isinstance(st.get("label_col"), str):
                st["label_col"] = sl(st["label_col"])
            if isinstance(st.get("label_merge"), str):
                st["label_merge"] = sf(st["label_merge"])
            if isinstance(st.get("sum_cols"), list):
                st["sum_cols"] = [sl(x) for x in st["sum_cols"]]
            if isinstance(st.get("formulas"), dict):
                st["formulas"] = {sl(k): sf(v) for k, v in st["formulas"].items()}
    return cfg


class WorkpaperGenerator:
    """양식 템플릿을 복사해 데이터를 채우는 기반 클래스.

    하위 클래스는 fill() 메서드를 구현해 계정별 로직을 정의한다.
    셀 주소는 모두 config YAML에서 읽어 소스코드에 하드코딩하지 않는다.
    """

    def __init__(self, template_path: str, config_path: str):
        self.template_path = Path(template_path)
        self.config = self._load_config(config_path)
        self.wb = None
        self.ws = None

    def _load_config(self, config_path: str) -> dict:
        with open(config_path, encoding="utf-8") as f:
            return apply_col_offset(yaml.safe_load(f))

    def open_template(self):
        self.wb = openpyxl.load_workbook(self.template_path)
        sheet_name = self.config.get("sheet")
        self.ws = self.wb[sheet_name] if sheet_name else self.wb.active

    def write_cell(self, cell_addr: str, value):
        """고정 셀에 값을 쓴다."""
        self.ws[cell_addr] = value

    def insert_data_rows(self, section: dict, rows: list[dict]):
        """데이터 단락에 행을 삽입하고 채운다.

        section: config에서 읽은 단락 정의 (start_row, columns)
        rows: 삽입할 데이터 목록
        """
        start_row = section["start_row"]
        columns = section["columns"]
        n = len(rows)

        if n == 0:
            return

        # 템플릿의 시작 행 아래에 (n-1)개 행 삽입 (첫 행은 기존 행 활용)
        if n > 1:
            self.ws.insert_rows(start_row + 1, n - 1)

        for i, row_data in enumerate(rows):
            r = start_row + i
            for col_letter, key in columns.items():
                self.ws[f"{col_letter}{r}"] = row_data.get(key)

    def inject_narrative(self, narrative: dict):
        """서술(목적·방법·결론)을 지정 셀에 쓴다."""
        narr_cfg = self.config.get("narrative", {})
        for key, cell_addr in narr_cfg.items():
            if key in narrative:
                self.write_cell(cell_addr, narrative[key])

    def save(self, output_path: str):
        """결과를 새 파일로 저장한다. 템플릿 원본은 보존된다."""
        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(out)

    def fill(self, *args, **kwargs):
        raise NotImplementedError("하위 클래스에서 fill()을 구현해야 합니다.")

    # =======================================================================
    # 양식 보존 공용 도구 (조서 종류 무관 — 모든 생성기가 상속해 사용)
    # =======================================================================

    def capture_row_style(self, row: int) -> dict:
        """한 행의 열별 스타일(폰트·테두리·채움·정렬·서식)을 캡처한다.

        초기화(clear_region) 전에 도너 행에서 호출해 두고, 새로 쓰는 행에
        apply_row_style로 입혀 테두리 등 양식을 보존한다.
        """
        styles = {"_height": self.ws.row_dimensions[row].height}   # 행 높이도 보존(삽입행 높이 불일치 방지)
        for col in range(1, self.ws.max_column + 1):
            c = self.ws.cell(row=row, column=col)
            styles[col] = {
                "font": copy(c.font),
                "border": copy(c.border),
                "fill": copy(c.fill),
                "alignment": copy(c.alignment),
                "number_format": c.number_format,
            }
        return styles

    def apply_row_style(self, row: int, styles: dict, skip_fill: bool = False) -> None:
        """capture_row_style로 캡처한 스타일(+행 높이)을 지정 행에 입힌다.

        skip_fill=True면 채움색(fill)은 건드리지 않는다 — 템플릿에 이미 있는 행(회색 banding 등)에
        도너의 단일 fill을 덮어써 banding이 뭉개지는 것을 막기 위함(폰트·테두리·서식만 통일).
        """
        h = styles.get("_height")
        if h is not None:
            self.ws.row_dimensions[row].height = h
        for col, st in styles.items():
            if col == "_height":
                continue
            c = self.ws.cell(row=row, column=col)
            c.font = copy(st["font"])
            c.border = copy(st["border"])
            if not skip_fill:
                c.fill = copy(st["fill"])
            c.alignment = copy(st["alignment"])
            c.number_format = st["number_format"]

    def clear_values(self, start: int, end: int) -> None:
        """행 범위의 **값만** 비운다. 스타일·테두리·병합은 보존.

        템플릿에 격자가 미리 그려진 고정표(예: Control Sheet)에 데이터를 채울 때 사용.
        (clear_region은 스타일까지 리셋하므로 미리 그려진 격자를 파괴 → 고정표엔 금지)

        병합 영역의 비-앵커 셀(MergedCell)은 읽기전용이라 건너뛴다(앵커만 비우면 값이 사라짐).
        """
        from openpyxl.cell.cell import MergedCell
        for row in range(start, end + 1):
            for col in range(1, self.ws.max_column + 1):
                c = self.ws.cell(row=row, column=col)
                if isinstance(c, MergedCell):
                    continue
                c.value = None

    def clear_region(self, start: int, end: int) -> None:
        """행 범위의 값·스타일·병합을 초기화한다(동적 재구성 전제).

        주의: 미리 격자가 그려진 고정표에는 쓰지 말 것(격자 파괴). 그 경우 clear_values 사용.
        """
        for rng in list(self.ws.merged_cells.ranges):
            if rng.min_row >= start:
                self.ws.unmerge_cells(str(rng))
        plain = (Font(), Border(), PatternFill(), Alignment())
        for row in range(start, end + 1):
            for col in range(1, self.ws.max_column + 1):
                c = self.ws.cell(row=row, column=col)
                c.value = None
                c.font, c.border, c.fill, c.alignment = (copy(x) for x in plain)
                c.number_format = "General"

    def safe_merge(self, rng: str) -> None:
        try:
            self.ws.merge_cells(rng)
        except Exception:
            pass

    def fit_column_widths(self, cols, data_start: int, data_end: int, *,
                          min_width: float = 8, max_width: float = 100,
                          header_row: "int | None" = None, grow_only: bool = True) -> None:
        """열 너비를 내용 길이에 맞게 보정한다 (긴 주소 등이 잘리지 않게).

        한글 등 전각 문자는 약 2배 폭으로 계산. grow_only면 템플릿 너비보다 좁히지 않는다.
        양식 보존 도구의 하나 — 데이터 채우기 후 호출해 열이 내용에 맞도록.
        """
        from openpyxl.utils import get_column_letter
        for c in cols:
            rows = list(range(data_start, data_end + 1))
            if header_row:
                rows.append(header_row)
            maxw = 0
            for r in rows:
                v = self.ws.cell(row=r, column=c).value
                if v is None:
                    continue
                w = sum(2 if ord(ch) > 0x1100 else 1 for ch in str(v))
                maxw = max(maxw, w)
            if maxw == 0:
                continue
            width = min(max_width, max(min_width, maxw + 2))
            letter = get_column_letter(c)
            if grow_only:
                cur = self.ws.column_dimensions[letter].width or 0
                width = max(width, cur)
            self.ws.column_dimensions[letter].width = width

    def capture_row(self, row: int) -> dict:
        """행의 열별 값+스타일을 함께 캡처한다 (제목·헤더 행 재배치용).

        clear_region 전에 캡처해 두고 stamp_row로 다른 위치에 그대로 찍는다.
        """
        out = {}
        for col in range(1, self.ws.max_column + 1):
            c = self.ws.cell(row=row, column=col)
            out[col] = {
                "value": c.value,
                "font": copy(c.font),
                "border": copy(c.border),
                "fill": copy(c.fill),
                "alignment": copy(c.alignment),
                "number_format": c.number_format,
            }
        return out

    def stamp_row(self, dst_row: int, captured: dict, overrides: dict = None) -> None:
        """capture_row 결과(값+스타일)를 지정 행에 찍는다. overrides={col: value}로 값 교체."""
        overrides = overrides or {}
        for col, item in captured.items():
            c = self.ws.cell(row=dst_row, column=col)
            c.value = overrides.get(col, item["value"])
            c.font = copy(item["font"])
            c.border = copy(item["border"])
            c.fill = copy(item["fill"])
            c.alignment = copy(item["alignment"])
            c.number_format = item["number_format"]

    def render_rows(self, start_row, columns, rows, style, *,
                    formula_cols=None, row_decorator=None, start_no=1):
        """데이터 행들을 양식 보존(스타일 복사)하며 쓴다. (next_row, next_no) 반환.

        columns: {표준키: 열문자} ('No.' 키는 start_no부터 자동 채번)
        formula_cols: {열문자: '수식템플릿'} ({r} 치환)
        row_decorator: fn(rec, row) optional
        """
        formula_cols = formula_cols or {}
        r, no = start_row, start_no
        for rec in rows:
            for key, col in columns.items():
                self.ws[f"{col}{r}"] = no if key == "No." else rec.get(key)
            for col, tmpl in formula_cols.items():
                self.ws[f"{col}{r}"] = tmpl.format(r=r)
            self.apply_row_style(r, style)
            if row_decorator:
                row_decorator(rec, r)
            r += 1
            no += 1
        return r, no

    def render_subtotal_sections(self, *, start_row, columns, sections,
                                 data_style, subtotal_style,
                                 subtotal_label_col, subtotal_label_merge=None,
                                 sum_cols=(), formula_cols=None, row_decorator=None):
        """가변 길이 섹션들을 (데이터 행 + 소계 행)으로 양식 보존하며 렌더한다.

        조서 맵핑 공통 패턴: 계정/구분별 가변 행 + 각 섹션 끝 소계. 행수가
        템플릿과 달라도 도너 스타일을 복사하고 소계/수식 위치를 재계산하므로
        양식이 무너지지 않는다.

        Args:
            start_row: 첫 데이터 행
            columns: {표준키: 열문자} 데이터 행 매핑 ('No.' 키는 1부터 자동 채번)
            sections: [(소계라벨, [row_dict, ...]), ...]
            data_style/subtotal_style: capture_row_style 결과
            subtotal_label_col: 소계 라벨을 쓸 열문자
            subtotal_label_merge: 소계 라벨 병합 패턴(예 'B{r}:F{r}') 또는 None
            sum_cols: 소계에서 =SUM(범위) 걸 열문자들
            formula_cols: {열문자: '수식템플릿'} 데이터 행마다 ({r} 치환)
            row_decorator: fn(rec, row) 행별 추가 처리(각주 등) optional

        Returns:
            (subtotal_rows, next_row): 소계 행번호 리스트, 다음 빈 행번호
        """
        r = start_row
        no = 1
        subtotal_rows = []

        for label, rows in sections:
            first = r
            r, no = self.render_rows(r, columns, rows, data_style,
                                     formula_cols=formula_cols,
                                     row_decorator=row_decorator, start_no=no)
            last = r - 1

            # 소계 행
            self.apply_row_style(r, subtotal_style)
            self.ws[f"{subtotal_label_col}{r}"] = label
            if subtotal_label_merge:
                self.safe_merge(subtotal_label_merge.format(r=r))
            for col in sum_cols:
                self.ws[f"{col}{r}"] = (f"=SUM({col}{first}:{col}{last})"
                                        if last >= first else 0)
            subtotal_rows.append(r)
            r += 1

        return subtotal_rows, r
