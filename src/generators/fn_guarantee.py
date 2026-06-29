# -*- coding: utf-8 -*-
"""A-1 FN 시트에 GUARANTEE(보증) 1번 박스를 그대로 렌더.

기존 주석(B~F열)과 겹치지 않게 우측(I열~)·지정 행에 붙인다. 값은 회신값 그대로.
"""
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from openpyxl.cell.cell import MergedCell

_thin = Side(style="thin")
_BOX = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_HDR_FILL = PatternFill("solid", fgColor="D9E1F2")


def render_fn_guarantee(ws, data, start_row=15, start_col=9):
    """GUARANTEE 1번 박스(title+header+rows)를 FN!(start_row, start_col=I)에 렌더.

    data = {title, header:[...], rows:[[...],...]}. 회신값 그대로 기입(재계산 없음).
    """
    if not data or not data.get("rows"):
        return 0
    header = data["header"]
    rows = data["rows"]
    ncol = max(len(header), max((len(r) for r in rows), default=0))
    r0, c0 = start_row, start_col

    # 제목
    t = ws.cell(r0, c0)
    if not isinstance(t, MergedCell):
        t.value = data.get("title") or "보증(보험) 조회 내역"
        t.font = Font(bold=True)
        t.alignment = Alignment(vertical="center")

    # 헤더
    hr = r0 + 1
    for j in range(ncol):
        c = ws.cell(hr, c0 + j)
        if isinstance(c, MergedCell):
            continue
        c.value = header[j] if j < len(header) else None
        c.font = Font(bold=True, size=9)
        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BOX

    # 데이터(선금/지급/합계 등) — 그대로
    for i, row in enumerate(rows):
        rr = hr + 1 + i
        for j in range(ncol):
            c = ws.cell(rr, c0 + j)
            if isinstance(c, MergedCell):
                continue
            v = row[j] if j < len(row) else None
            c.value = v
            c.border = _BOX
            c.font = Font(size=9)
            # 금액류(쉼표 숫자 문자열/숫자)는 우측정렬
            if isinstance(v, (int, float)) or (isinstance(v, str) and v.replace(",", "").replace("-", "").isdigit() and len(v) > 3):
                c.alignment = Alignment(horizontal="right")

    # 열 너비(내용 맞춤, 과하지 않게)
    from openpyxl.utils import get_column_letter as L
    for j in range(ncol):
        col = L(c0 + j)
        cur = ws.column_dimensions[col].width or 8.43
        ws.column_dimensions[col].width = max(cur, 12)
    return len(rows)
