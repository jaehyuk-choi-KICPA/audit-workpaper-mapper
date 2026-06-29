# -*- coding: utf-8 -*-
"""고정자산 관리대장_25 생성기 — 회사제시 입력열 설치 + 감사인 재계산.

레이아웃(사용자 지정): **회사제시 입력=왼쪽(B~J)**, **감사인재계산=오른쪽(L~P)**,
내용연수/상각방법 **참조표=감사인재계산 위(L1:N5)**. 생성기가 헤더영역(행1~7)과
데이터(행8+)를 전부 새로 그린다(A열 마커는 보존).

  1. 메모(내용연수·상각방법) 기반 참조표(L1:N5) 작성,
  2. 회사제시 입력열(B~J)에 관리대장 파서 결과를 자산별로 채우고,
  3. 감사인 재계산 수식(L~P)을 **행별로** 설치한다(내용연수·정률율 치환, 1000 비망·월할),
  4. 자산분류별 소계 + 총계 — G300 재계산표가 각 소계행(M열)을 참조.

재계산 로직(완성본 수식 일반화):
  정액 = MIN(INT(기말/내용연수×월할), MAX(0, 기말-전기누계-비망))
  정률 = MIN(INT((기말-전기누계)×정률율×월할), MAX(0, 기말-전기누계-비망))
  월할 = 당기취득(YEAR(취득일)=결산연도)이면 (13-월)/12, 아니면 1
  처분(당기감소>0 & 기말=0)이면 회사제시 상각비 사용.
상각방법이 애매하면(메모에 정률·정액 병기) **회사제시 감가상각비를 정액 추정치와 비교해
추론**(가까우면 정액, 아니면 정률). 메모에 그 계정이 없으면 **관리대장 상각율**로 추론.
"""

import re

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter

from .base import WorkpaperGenerator

_THIN = Side(style="thin")
_MEDIUM = Side(style="medium")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HDR_FILL = PatternFill("solid", fgColor="D9D9D9")   # 헤더 회색


def _norm(s) -> str:
    return re.sub(r"\s+", "", str(s)) if s is not None else ""


def _ci(letter: str) -> int:
    return column_index_from_string(letter)


def _cl(idx: int) -> str:
    return get_column_letter(idx)


class FixedAssetRegisterGenerator(WorkpaperGenerator):

    def fill(self, ledger_rows: list, memo: list, params: dict = None) -> dict:
        params = params or {}
        cfg = self.config
        self.open_template()
        ye = self._year_end(params, cfg)
        inp, rc, lg = cfg["inputs"], cfg["recalc"], cfg.get("legend", {})
        self.numf = cfg.get("num_format", "#,##0")
        self.memo_val = cfg.get("memo_value", 1000)
        self.rates = {int(k): float(v) for k, v in (cfg.get("declining_rates") or {}).items()}
        self.hr = cfg.get("header_row", 7)
        self.ghr = cfg.get("group_header_row", 6)
        ds = cfg.get("data_start_row", 8)
        memo_map = {_norm(m["계정과목"]): m for m in (memo or [])}

        # 헤더+데이터 영역을 B열~재계산 마지막열까지 새로 그림(A열 마커 보존)
        c1, c2 = _ci(inp["계정과목"]), _ci(rc["차액2"])
        self._clear_block(1, max(self.ws.max_row, ds), c1, c2)

        self._write_legend(memo, lg)
        self._write_headers(inp, rc)

        groups = self._group_by_class(ledger_rows)
        r = ds
        subtotal_rows, all_subtotals = {}, []
        first_asset = r
        for cls, items in groups:
            grp_start = r
            for rec in items:
                self._write_asset_row(r, rec, cls, inp, rc, ye, memo_map)
                r += 1
            self._write_subtotal(r, grp_start, r - 1, f"{cls} 소계", inp, rc)
            subtotal_rows[cls] = r
            all_subtotals.append(r)
            r += 1
        total_row = None
        if all_subtotals:
            self._write_total(r, all_subtotals, inp, rc)
            total_row = r
            r += 1

        self.fit_column_widths([_ci(inp["계정과목"]), _ci(inp["자산명"])],
                               first_asset, max(first_asset, r - 1), min_width=12, max_width=40)
        return {"subtotal_rows": subtotal_rows, "total_row": total_row, "data_end": r - 1}

    # ------------------------------------------------------------------ helpers
    def _clear_block(self, r1, r2, c1, c2):
        """사각 영역 값·스타일·병합 초기화(A열 등 영역 밖은 보존)."""
        from copy import copy
        from openpyxl.cell.cell import MergedCell
        for rng in list(self.ws.merged_cells.ranges):
            if rng.min_col >= c1 and rng.max_col <= c2 and rng.min_row >= r1 and rng.max_row <= r2:
                self.ws.unmerge_cells(str(rng))
        plain = (Font(), Border(), PatternFill(), Alignment())
        for row in range(r1, r2 + 1):
            for col in range(c1, c2 + 1):
                c = self.ws.cell(row, col)
                if isinstance(c, MergedCell):
                    continue
                c.value = None
                c.font, c.border, c.fill, c.alignment = (copy(x) for x in plain)
                c.number_format = "General"

    def _year_end(self, params, cfg) -> int:
        d = params.get("날짜") or params.get("기준일")
        if d:
            m = re.search(r"(\d{4})", str(d))
            if m:
                return int(m.group(1))
        return int(cfg.get("year_end", 2025))

    def _group_by_class(self, rows):
        order, groups = [], {}
        for rec in rows or []:
            cls = rec.get("계정과목") or "미분류"
            if cls not in groups:
                groups[cls] = []
                order.append(cls)
            groups[cls].append(rec)
        return [(c, groups[c]) for c in order]

    def _hdr_cell(self, addr, value, *, bold=True, fill=True, wrap=True):
        c = self.ws[addr]
        c.value = value
        c.font = Font(bold=bold)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
        c.border = _BORDER
        if fill:
            c.fill = _HDR_FILL
        return c

    def _write_legend(self, memo, lg):
        if not lg:
            return
        sr = lg.get("start_row", 1)
        cc, ci_, cm = lg.get("계정과목", "L"), lg.get("내용연수", "M"), lg.get("상각방법", "N")
        self._hdr_cell(f"{cc}{sr}", "계 정 과 목")
        self._hdr_cell(f"{ci_}{sr}", "내 용 연 수")
        self._hdr_cell(f"{cm}{sr}", "감가상각방법")
        if not memo:
            return
        merged, order = {}, []
        for m in memo:
            key = (m.get("내용연수"), tuple(m.get("상각방법") or []))
            if key not in merged:
                merged[key] = []
                order.append(key)
            merged[key].append(m["계정과목"])
        rr = sr + 1
        for years, methods in order:
            self.ws[f"{cc}{rr}"] = ", ".join(merged[(years, methods)])
            self.ws[f"{ci_}{rr}"] = (f"{years}년" if years else "")
            self.ws[f"{cm}{rr}"] = ", ".join(f"{x}법" for x in methods)
            for col in (cc, ci_, cm):
                self.ws[f"{col}{rr}"].border = _BORDER
                self.ws[f"{col}{rr}"].alignment = Alignment(horizontal="center")
            rr += 1

    def _write_headers(self, inp, rc):
        labels = {
            "계정과목": "계정과목(자산분류)", "자산명": "자산명", "취득일자": "취득일자",
            "기초가액": "기초가액", "당기취득": "당기취득", "당기감소": "당기감소",
            "기말잔액": "기말잔액", "전기말상각누계": "전기말상각누계액", "당기감가상각비": "당기감가상각비",
        }
        rc_labels = {"방법": "감가상각방법", "상각비": "당기감가상각비",
                     "차액1": "차액", "누계": "감가상각누계액", "차액2": "차액"}
        # 그룹 헤더(행6)
        ic1, ic2 = _ci(inp["계정과목"]), _ci(inp["당기감가상각비"])
        rcc1, rcc2 = _ci(rc["방법"]), _ci(rc["차액2"])
        gh = self._hdr_cell(f"{_cl(ic1)}{self.ghr}", "회사제시 (고정자산관리대장)")
        self.safe_merge(f"{_cl(ic1)}{self.ghr}:{_cl(ic2)}{self.ghr}")
        self._hdr_cell(f"{_cl(rcc1)}{self.ghr}", "감사인 재계산")
        self.safe_merge(f"{_cl(rcc1)}{self.ghr}:{_cl(rcc2)}{self.ghr}")
        # 컬럼 헤더(행7)
        for key, col in inp.items():
            self._hdr_cell(f"{col}{self.hr}", labels.get(key, key))
        for key, col in rc.items():
            self._hdr_cell(f"{col}{self.hr}", rc_labels.get(key, key))

    def _border_row(self, r, c_from, c_to, *, bold=False, top=None):
        for c in range(c_from, c_to + 1):
            cell = self.ws.cell(r, c)
            cell.border = (Border(left=_THIN, right=_THIN, top=_MEDIUM, bottom=_THIN)
                           if top == "medium" else _BORDER)
            if bold:
                cell.font = Font(bold=True)

    def _write_asset_row(self, r, rec, cls, inp, rc, ye, memo_map):
        ws = self.ws
        vals = {
            "계정과목": cls, "자산명": rec.get("자산명"), "취득일자": rec.get("취득일자"),
            "기초가액": rec.get("기초가액"), "당기취득": rec.get("당기취득"),
            "당기감소": rec.get("당기감소"), "기말잔액": rec.get("기말잔액"),
            "전기말상각누계": rec.get("전기말상각누계"), "당기감가상각비": rec.get("당기감가상각비"),
        }
        for key, col in inp.items():
            cell = ws[f"{col}{r}"]
            cell.value = vals.get(key)
            if key not in ("계정과목", "자산명", "취득일자"):
                cell.number_format = self.numf
        years, methods = self._method_of(cls, rec, memo_map)
        rate = self.rates.get(years, 0.0) if years else 0.0
        self._write_recalc(r, inp, rc, ye, years, methods, rate)
        self._border_row(r, _ci(inp["계정과목"]), _ci(rc["차액2"]))

    def _method_of(self, cls, rec, memo_map):
        """(내용연수:int|None, methods:[정액|정률]) — 메모 우선, 없으면 관리대장 상각율로 추론."""
        m = memo_map.get(_norm(cls))
        if m:
            return m.get("내용연수"), (m.get("상각방법") or ["정액"])
        years = int(rec["내용연수"]) if rec.get("내용연수") else None
        rate, methods = rec.get("상각율"), ["정액"]
        if years and rate:
            methods = ["정액"] if abs(rate - 1.0 / years) < 0.02 else ["정률"]
        return years, methods

    def _write_recalc(self, r, inp, rc, ye, years, methods, rate):
        ws = self.ws
        N, O, P = inp["기말잔액"], inp["전기말상각누계"], inp["당기감가상각비"]
        J, M = inp["취득일자"], inp["당기감소"]
        Bc, Cc, Dc, Ec, Fc = rc["방법"], rc["상각비"], rc["차액1"], rc["누계"], rc["차액2"]
        frac = f'IF({J}{r}="",1,IF(YEAR({J}{r})={ye},(13-MONTH({J}{r}))/12,1))'
        base = f"MAX(0,{N}{r}-{O}{r}-{self.memo_val})"
        sl = f"MIN(INT({N}{r}/{years}*{frac}),{base})" if years else "0"
        dl = f"MIN(INT(({N}{r}-{O}{r})*{rate}*{frac}),{base})" if rate else "0"

        if len(methods) >= 2:   # 정률/정액 병기 → 회사제시 금액 보고 추론
            ws[f"{Bc}{r}"] = (f'=IF(OR({P}{r}=0,{N}{r}=0),"정률법",'
                              f'IF(ABS({P}{r}-INT({N}{r}/{years}*{frac}))<=1,"정액법","정률법"))') \
                if years else '="정액법"'
            dep = f'IF({Bc}{r}="정액법",{sl},{dl})'
        else:
            ws[f"{Bc}{r}"] = "정액법" if methods[0] == "정액" else "정률법"
            dep = sl if methods[0] == "정액" else dl
        ws[f"{Cc}{r}"] = f"=IF(AND({M}{r}>0,{N}{r}=0),{P}{r},IF({N}{r}=0,0,{dep}))"
        ws[f"{Dc}{r}"] = f"={Cc}{r}-{P}{r}"
        ws[f"{Ec}{r}"] = f"=IF({N}{r}=0,0,{O}{r}+{Cc}{r})"
        ws[f"{Fc}{r}"] = f"={Ec}{r}-({O}{r}+{P}{r})"
        for col in (Cc, Dc, Ec, Fc):
            ws[f"{col}{r}"].number_format = self.numf

    def _sum_cols(self, inp, rc):
        return [inp["기초가액"], inp["당기취득"], inp["당기감소"], inp["기말잔액"],
                inp["전기말상각누계"], inp["당기감가상각비"], rc["상각비"], rc["차액1"],
                rc["누계"], rc["차액2"]]

    def _write_subtotal(self, r, a, b, label, inp, rc):
        ws = self.ws
        ws[f"{inp['자산명']}{r}"] = label
        for col in self._sum_cols(inp, rc):
            ws[f"{col}{r}"] = f"=SUM({col}{a}:{col}{b})"
            ws[f"{col}{r}"].number_format = self.numf
        self._border_row(r, _ci(inp["계정과목"]), _ci(rc["차액2"]), bold=True, top="medium")

    def _write_total(self, r, subtotal_rows, inp, rc):
        ws = self.ws
        ws[f"{inp['자산명']}{r}"] = "총   계"
        for col in self._sum_cols(inp, rc):
            ws[f"{col}{r}"] = "=" + "+".join(f"{col}{sr}" for sr in subtotal_rows)
            ws[f"{col}{r}"].number_format = self.numf
        self._border_row(r, _ci(inp["계정과목"]), _ci(rc["차액2"]), bold=True, top="medium")
