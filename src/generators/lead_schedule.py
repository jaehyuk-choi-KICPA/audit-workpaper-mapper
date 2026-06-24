"""잔액형 총괄표 통일 생성기.

A-0(현금·장단기금융)에서 검증한 '총괄표 본문 + 수정사항' 패턴을 **표준양식 잔액형 조서**
(B/C/D/AA/CC/BBDD/EE …)에 일반화한다. 조서별 차이는 config(YAML)로만 주입한다.

처리:
  1. 1.총괄표 본문 — account_routing으로 이 조서에 속한 별도정산표 계정을 모아
     계정명/기초/기말/수정사항(K-L)/수정후를 채운다(증가·증감율·수정후=수식). 음수=빨간 괄호.
     템플릿에 미리 그려진 도너 행 스타일을 복사해 행수가 달라도 격자를 보존·확장한다.
  2. 7.수정사항 — 표준양식엔 분개표 격자가 없으므로 **A-0(4000_A000)과 동일한 분개표**
     (계정과목[차변/대변]|금액[차변/대변]|Effect[손익/이익잉여금]|Description)를
     '1) 회사 제시 수정분개' 아래에 테두리째 그려 넣고, 이 조서 계정과 관련된 분개만 재현한다.

유형자산(G)·무형자산(H)은 변동분해형(롤포워드)이라 이 엔진 대상이 아니다(별도 — 후속).
"""

import re

from openpyxl.styles import Border, Side

from .base import WorkpaperGenerator

_THIN = Side(style="thin")
_BOX = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _norm(s) -> str:
    return re.sub(r"\s+", "", str(s)) if s is not None else ""


class LeadScheduleGenerator(WorkpaperGenerator):

    def fill(self, tb_rows, adj_entries, routing: dict, params: dict = None) -> dict:
        self.open_template()
        cfg = self.config
        code = cfg["routing_code"]
        mine = [r for r in tb_rows if routing.get(r["대분류"]) == code]
        body_end = self._fill_body(mine, cfg["body"])
        adj_end, n_adj = self._fill_adjustments(mine, adj_entries, cfg["adjustments"])
        return {"routing_code": code, "n_accounts": len(mine),
                "body_end": body_end, "adj_end": adj_end, "n_adj_entries": n_adj}

    # ---- 1. 총괄표 본문 ----
    def _fill_body(self, rows, b):
        donor = self.capture_row_style(b["donor_detail_row"])   # 삽입 전 캡처
        # 계정 수만큼 행을 삽입해 아래 섹션(대사 설명·2.Recap 등, 병합 영역)을 밀어낸다.
        needed = len(rows) + (1 if (b.get("subtotal") and rows) else 0)
        if needed:
            self.ws.insert_rows(b["data_start_row"], needed)
        cols = b["columns"]
        detail_f = b.get("detail_formulas", {})
        numf, pctf = b["num_format"], b["pct_format"]
        numc, pctc = set(b["num_cols"]), set(b["pct_cols"])

        def setfmt(col, row):
            if col in numc:
                self.ws[f"{col}{row}"].number_format = numf
            elif col in pctc:
                self.ws[f"{col}{row}"].number_format = pctf

        r = b["data_start_row"]
        first = r
        for rec in rows:
            self.apply_row_style(r, donor)          # 도너 격자 복사(행 확장)
            for key, col in cols.items():
                val = rec.get(key)
                if key == "수정사항" and val in (0, None):
                    val = None
                self.ws[f"{col}{r}"] = val
                setfmt(col, r)
            for col, tmpl in detail_f.items():
                self.ws[f"{col}{r}"] = tmpl.format(r=r)
                setfmt(col, r)
            r += 1
        last = r - 1

        sub = b.get("subtotal")
        if sub and rows:
            self.apply_row_style(r, donor)
            self.ws[f'{sub["label_col"]}{r}'] = sub["label"]
            if sub.get("label_merge"):
                self.safe_merge(sub["label_merge"].format(r=r))
            for col in sub.get("sum_cols", []):
                self.ws[f"{col}{r}"] = f"=SUM({col}{first}:{col}{last})" if last >= first else 0
                setfmt(col, r)
            for col, tmpl in sub.get("formulas", {}).items():
                self.ws[f"{col}{r}"] = tmpl.format(r=r)
                setfmt(col, r)
            r += 1
        return r - 1

    # ---- 7. 수정사항(분개표 그려서 재현) ----
    def _fill_adjustments(self, mine, adj_entries, a):
        names = {_norm(r["계정명"]) for r in mine} | {_norm(r["대분류"]) for r in mine}

        def related(e):
            return any(_norm(l["계정"]) in names for l in e["lines"])

        entries = [e for e in adj_entries if related(e)]

        # 앵커('1) 회사 제시 수정분개') 행 탐색 → 그 아래에 표 그림
        anchor = self._find_text_row(a["anchor_text"])
        if anchor is None:
            return None, len(entries)
        c = a["cols"]
        r = anchor + a.get("gap_after_anchor", 1)
        numf, numc = a["num_format"], set(a["num_cols"])

        # 표가 들어갈 만큼 행을 삽입해 아래 섹션('2) 검토 결과' 등)을 밀어낸다(덮어쓰기 방지).
        n_lines = sum(len(e["lines"]) + len(e.get("notes", [])) for e in entries)
        needed = (2 + n_lines) if entries else 1
        self.ws.insert_rows(r, needed)

        if not entries:
            self.ws[f'{c["debit_acct"]}{r}'] = "해당사항 없음"
            self.ws[f'{c["debit_acct"]}{r}'].border = _BOX
            return r, 0

        # 헤더 2행 (A-0 분개표 양식)
        h1, h2 = r, r + 1
        self.ws[f'{c["debit_acct"]}{h1}'] = "계정과목"
        self.ws[f'{c["debit_amt"]}{h1}'] = "금액"
        self.ws[f'{c["pl"]}{h1}'] = "Effect"
        self.ws[f'{c["desc"]}{h1}'] = "Description"
        self.safe_merge(f'{c["debit_acct"]}{h1}:{c["credit_acct"]}{h1}')
        self.safe_merge(f'{c["debit_amt"]}{h1}:{c["credit_amt"]}{h1}')
        self.safe_merge(f'{c["pl"]}{h1}:{c["re"]}{h1}')
        self.safe_merge(f'{c["desc"]}{h1}:{c["desc"]}{h2}')
        self.ws[f'{c["debit_acct"]}{h2}'] = "차변"
        self.ws[f'{c["credit_acct"]}{h2}'] = "대변"
        self.ws[f'{c["debit_amt"]}{h2}'] = "차변"
        self.ws[f'{c["credit_amt"]}{h2}'] = "대변"
        self.ws[f'{c["pl"]}{h2}'] = "손익"
        self.ws[f'{c["re"]}{h2}'] = "이익잉여금"
        order = [c["no"], c["debit_acct"], c["credit_acct"], c["debit_amt"],
                 c["credit_amt"], c["pl"], c["re"], c["desc"]]
        for hr in (h1, h2):
            for col in order:
                self.ws[f"{col}{hr}"].border = _BOX
        r = h2 + 1

        def setamt(col, row, val):
            self.ws[f"{col}{row}"] = val
            if col in numc:
                self.ws[f"{col}{row}"].number_format = numf

        for e in entries:
            first = True
            estart = r
            for ln in e["lines"]:
                if first:
                    self.ws[f'{c["no"]}{r}'] = f'#{e["no"]}'
                    first = False
                if ln["side"] == "차변":
                    self.ws[f'{c["debit_acct"]}{r}'] = ln["계정"]
                    if ln.get("금액") is not None:
                        setamt(c["debit_amt"], r, ln["금액"])
                else:
                    self.ws[f'{c["credit_acct"]}{r}'] = ln["계정"]
                    if ln.get("금액") is not None:
                        setamt(c["credit_amt"], r, ln["금액"])
                if ln.get("손익") is not None:
                    setamt(c["pl"], r, ln["손익"])
                if ln.get("이익잉여금") is not None:
                    setamt(c["re"], r, ln["이익잉여금"])
                if ln.get("설명"):
                    self.ws[f'{c["desc"]}{r}'] = ln["설명"]
                for col in order:
                    self.ws[f"{col}{r}"].border = _BOX
                r += 1
            for note in e.get("notes", []):
                self.ws[f'{c["desc"]}{r}'] = note
                for col in order:
                    self.ws[f"{col}{r}"].border = _BOX
                r += 1
        return r - 1, len(entries)

    # ---- helper ----
    def _find_text_row(self, text: str) -> "int | None":
        t = _norm(text)
        for row in range(1, self.ws.max_row + 1):
            for col in range(1, 12):
                if t and t in _norm(self.ws.cell(row, col).value):
                    return row
        return None
