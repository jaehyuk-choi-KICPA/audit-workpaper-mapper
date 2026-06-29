"""zip 외과수술 — 생성된 총괄표 한 시트를 완성본(.xlsx/.xlsm)에 무손실 이식.

배경: 회계법인 완성본 조서는 보조시트(Test·주석)에 양식컨트롤·VML·도형·매크로가
가득하다(실측: 조서당 컨트롤 ~94개 등). openpyxl로 열고 저장하면 이 요소가 **전부 소실**된다
(실측 검증). 그래서 완성본은 **openpyxl로 절대 열지 않고**, 가벼운 단일시트 템플릿에서 생성한
총괄표 시트만 **zip/XML 레벨로 이식**한다.

핵심 난점 = 스타일 인덱스. 완성본의 다른 시트들은 원본 styles.xml 인덱스를 그대로 참조하므로
(원본 바이트 보존), 총괄표가 쓰는 스타일을 **원본 styles.xml에 덧붙이고(append)** 총괄표의
`s=` 인덱스를 그 새 위치로 **재매핑**한다(원본 인덱스는 절대 건드리지 않아 다른 시트 안전).
문자열은 openpyxl이 인라인(`t="inlineStr"`)으로 저장하므로 sharedStrings 의존이 없다
(혹시 `t="s"`면 light sharedStrings로 인라인화).

수정되는 파트 = `xl/styles.xml`(병합) + 대상 시트 XML(재매핑) 뿐. 나머지(컨트롤·도형·매크로·
다른 시트 전부)는 원본 zip 바이트 그대로 복사 → 무손실.
"""

import os
import re
import shutil
import zipfile
from pathlib import Path
import xml.etree.ElementTree as ET

_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
ET.register_namespace("", _MAIN)
ET.register_namespace("r", _R)

_XMLDECL = b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'


def _q(tag: str) -> str:
    return f"{{{_MAIN}}}{tag}"


# ---------------------------------------------------------------------------
# 시트명 → zip 내 파트 경로
# ---------------------------------------------------------------------------

def _sheet_part(zf: zipfile.ZipFile, sheet_name: str) -> "str | None":
    wbx = ET.fromstring(zf.read("xl/workbook.xml"))
    sheets = wbx.find(_q("sheets"))
    rid = None
    for s in (sheets if sheets is not None else []):
        if s.get("name") == sheet_name:
            rid = s.get(f"{{{_R}}}id")
            break
    if rid is None:
        return None
    relsx = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    for rel in relsx:
        if rel.get("Id") == rid:
            t = rel.get("Target")
            if t.startswith("/"):
                return t.lstrip("/")
            return "xl/" + t if not t.startswith("xl/") else t
    return None


def trim_blank_tail(full_path: str, sheet_name: str, output_path: str = None,
                    margin: int = 40) -> int:
    """대상 시트의 **내용 없는 꼬리 행**(빈 styled 행)을 zip/XML 레벨로 잘라낸다(스타일·다른 시트 무손실).

    회계법인 완성본은 일부 시트가 dimension이 수십만 행으로 부풀어(빈 styled 행 누적) 100MB+가 되어
    openpyxl 로드/저장이 분 단위로 느려진다. 실제 값(`<v>`/`<f>`/`<is>`)이 있는 마지막 행 + margin
    까지만 남기고 그 아래 빈 행을 제거(+dimension 축소)한다. 클린 시트엔 사실상 무영향(전 행 보존).
    openpyxl로 열지 않아 컨트롤·매크로·도형 보존. output_path 생략 시 제자리. Returns: 제거한 행 수.
    """
    output_path = output_path or full_path
    with zipfile.ZipFile(full_path) as zf:
        part = _sheet_part(zf, sheet_name)
        if part is None:
            raise ValueError(f"[trim] 시트 '{sheet_name}' 없음")
        data = {n: zf.read(n) for n in zf.namelist()}

    sheet = data[part]
    head, sd_open, rest = sheet.partition(b"<sheetData>")
    if not sd_open:                                   # <sheetData/> 빈 시트 등
        if output_path != full_path:
            shutil.copyfile(full_path, output_path)
        return 0
    body, sd_close, tail = rest.partition(b"</sheetData>")

    chunks = body.split(b"<row ")                     # chunks[0]=선두공백, 이후 각 행(접두 제외)
    last_content = 0
    parsed = []                                       # (rnum, has_content, raw_chunk)
    for ch in chunks[1:]:
        try:
            rnum = int(ch.split(b'"', 2)[1])          # r="N"
        except (IndexError, ValueError):
            parsed.append((None, True, ch))           # 파싱 불가 → 보존
            continue
        has = (b"<v" in ch) or (b"<f" in ch) or (b"<is" in ch)
        if has:
            last_content = max(last_content, rnum)
        parsed.append((rnum, has, ch))

    keep = last_content + margin
    kept = [chunks[0]]
    removed = 0
    for rnum, _has, ch in parsed:
        if rnum is not None and rnum > keep:
            removed += 1
            continue
        kept.append(b"<row " + ch)
    new_body = b"".join(kept)

    sheet = head + sd_open + new_body + sd_close + tail
    # dimension 축소(열 범위는 보존, 행만 keep로)
    sheet = re.sub(rb'(<dimension ref="[A-Z]+\d+:[A-Z]+)\d+"',
                   lambda m: m.group(1) + str(keep).encode() + b'"', sheet, count=1)
    data[part] = sheet

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    tmp = str(output_path) + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as out:
        for n, b in data.items():
            out.writestr(n, b)
    os.replace(tmp, output_path)
    return removed


# ---------------------------------------------------------------------------
# styles.xml 병합 (문자열 조작 — 네임스페이스 mc/x14 등 원형 보존; ET 재직렬화는
# 이 특수 네임스페이스를 ns0/ns1로 변질시켜 Excel이 파일을 거부함 → 실측 확인)
# ---------------------------------------------------------------------------

_SING = {"numFmts": "numFmt", "fonts": "font", "fills": "fill",
         "borders": "border", "cellXfs": "xf"}


def _find_table(xml: str, name: str):
    """<name ...>...</name> 또는 빈 <name .../>. (start, end, inner) 또는 None."""
    m = re.search(r"<%s\b[^>]*?>.*?</%s>" % (name, name), xml, re.DOTALL)
    if m:
        inner = re.search(r"<%s\b[^>]*?>(.*)</%s>" % (name, name), m.group(0), re.DOTALL).group(1)
        return m.start(), m.end(), inner
    m = re.search(r"<%s\b[^>]*?/>" % name, xml)
    if m:
        return m.start(), m.end(), ""
    return None


def _elements(inner: str, tag: str) -> list:
    """테이블 내부 문자열에서 최상위 <tag> 요소들을 원문 그대로 추출(같은 태그 비중첩 가정)."""
    return re.findall(r"<%s\b(?:[^>]*?/>|[^>]*?>.*?</%s>)" % (tag, tag), inner, re.DOTALL)


def _merge_styles(base_bytes: bytes, light_bytes: bytes, used_xf: set) -> "tuple[bytes, dict]":
    """light styles의 used_xf(셀이 쓰는 cellXf 인덱스)를 base styles에 **문자열로** 병합.

    base의 기존 요소·인덱스는 절대 바뀌지 않는다(append만) → 다른 시트(원본 바이트) 안전.
    Returns: (병합 styles.xml bytes, {light_xf_idx: base_xf_idx}).
    """
    base = base_bytes.decode("utf-8")
    light = light_bytes.decode("utf-8")

    def table_elems(xml: str, name: str) -> list:
        t = _find_table(xml, name)
        return _elements(t[2], _SING[name]) if t else []

    base_tab = {n: table_elems(base, n) for n in _SING}
    light_tab = {n: table_elems(light, n) for n in _SING}

    def nf_codes(elems: list) -> dict:
        d = {}
        for e in elems:
            mid = re.search(r'numFmtId="(\d+)"', e)
            code = re.search(r'formatCode="((?:[^"\\]|\\.)*)"', e)
            if mid and code:
                d[int(mid.group(1))] = code.group(1)
        return d

    base_codes = nf_codes(base_tab["numFmts"])
    base_code_to_id = {c: i for i, c in base_codes.items()}
    light_codes = nf_codes(light_tab["numFmts"])
    next_id = max([163] + list(base_codes.keys())) + 1

    add = {n: [] for n in _SING}        # 테이블별 새로 덧붙일 요소
    sub_cache = {"fonts": {}, "fills": {}, "borders": {}}
    nf_map: dict = {}

    def map_sub(name: str, lidx: int) -> int:
        if lidx in sub_cache[name]:
            return sub_cache[name][lidx]
        new_idx = len(base_tab[name]) + len(add[name])
        add[name].append(light_tab[name][lidx])
        sub_cache[name][lidx] = new_idx
        return new_idx

    def map_nf(lid: int) -> int:
        if lid < 164:                   # builtin — 공통
            return lid
        if lid in nf_map:
            return nf_map[lid]
        code = light_codes.get(lid)
        if code is None:
            return lid
        if code in base_code_to_id:
            nf_map[lid] = base_code_to_id[code]
            return nf_map[lid]
        nonlocal next_id
        nid = next_id
        next_id += 1
        add["numFmts"].append('<numFmt numFmtId="%d" formatCode="%s"/>' % (nid, code))
        base_code_to_id[code] = nid
        nf_map[lid] = nid
        return nid

    xf_map: dict = {}
    light_xfs = light_tab["cellXfs"]
    for li in sorted(used_xf):
        if li >= len(light_xfs):
            xf_map[li] = 0
            continue
        xf = light_xfs[li]

        def a(name):
            m = re.search(r'%s="(\d+)"' % name, xf)
            return int(m.group(1)) if m else None

        new = xf
        if a("fontId") is not None:
            new = re.sub(r'fontId="\d+"', 'fontId="%d"' % map_sub("fonts", a("fontId")), new, 1)
        if a("fillId") is not None:
            new = re.sub(r'fillId="\d+"', 'fillId="%d"' % map_sub("fills", a("fillId")), new, 1)
        if a("borderId") is not None:
            new = re.sub(r'borderId="\d+"', 'borderId="%d"' % map_sub("borders", a("borderId")), new, 1)
        if a("numFmtId") is not None:
            new = re.sub(r'numFmtId="\d+"', 'numFmtId="%d"' % map_nf(a("numFmtId")), new, 1)
        xf_map[li] = len(base_tab["cellXfs"]) + len(add["cellXfs"])
        add["cellXfs"].append(new)

    # 각 테이블을 (기존+추가)로 재작성하고 count 갱신 — 문자열 치환(네임스페이스 보존)
    out = base
    for name in _SING:
        if not add[name] and _find_table(out, name) is not None:
            # 변경 없음 → count만 정확하면 그대로 둠(기존 count 신뢰)
            continue
        elems = base_tab[name] + add[name]
        new_table = '<%s count="%d">%s</%s>' % (name, len(elems), "".join(elems), name)
        t = _find_table(out, name)
        if t is None:                   # 없던 테이블(numFmts) → styleSheet 여는 태그 직후 삽입
            mm = re.search(r"<styleSheet\b[^>]*>", out)
            out = out[:mm.end()] + new_table + out[mm.end():]
        else:
            out = out[:t[0]] + new_table + out[t[1]:]

    return out.encode("utf-8"), xf_map


# ---------------------------------------------------------------------------
# 대상 시트 XML: 문자열 인라인화 + s= 인덱스 재매핑
# ---------------------------------------------------------------------------

def _inline_shared_strings(sheet_xml: str, sst_bytes: "bytes | None") -> str:
    """t="s" 셀(sharedStrings 참조)을 인라인 문자열로 변환. sst 없으면 그대로."""
    if 't="s"' not in sheet_xml or sst_bytes is None:
        return sheet_xml
    sst = ET.fromstring(sst_bytes)
    strings = []
    for si in sst:
        # si의 모든 t 텍스트 이어붙임(서식분할 r 요소 대비)
        txt = "".join(t.text or "" for t in si.iter(_q("t")))
        strings.append(txt)

    def repl(m):
        body = m.group(0)
        vm = re.search(r"<v>(\d+)</v>", body)
        if not vm:
            return body
        idx = int(vm.group(1))
        s = strings[idx] if idx < len(strings) else ""
        s = (s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))
        # t="s" → t="inlineStr", <v>..</v> → <is><t xml:space="preserve">..</t></is>
        new = body.replace('t="s"', 't="inlineStr"')
        new = re.sub(r"<v>\d+</v>", f'<is><t xml:space="preserve">{s}</t></is>', new)
        return new

    # 셀 단위 치환 (t="s" 포함하는 <c ...>...</c>)
    return re.sub(r'<c\b[^>]*\bt="s"[^>]*>.*?</c>', repl, sheet_xml, flags=re.DOTALL)


# 스타일 인덱스 속성은 셀/행/열의 ' s="N"'(앞에 공백). **단어경계(\b)** 로 한정해야
# showGridLines="0"/showZeros="0" 같은 's'로 끝나는 다른 속성의 's="0"'을 오인·훼손하지 않는다.
# (과거: s="(\d+)"가 showGridLines="0"의 s="0"을 잡아 격자선 설정을 깨뜨림 → 빈칸 박스선 사고.)
_S_ATTR = re.compile(r'\bs="(\d+)"')


def _remap_style_indices(sheet_xml: str, xf_map: dict) -> str:
    """워크시트 XML의 셀·행·열 스타일 s="N"을 xf_map으로 치환(다른 속성의 s= 는 제외)."""
    def repl(m):
        old = int(m.group(1))
        return f's="{xf_map.get(old, old)}"'
    return _S_ATTR.sub(repl, sheet_xml)


def _used_style_indices(sheet_xml: str) -> set:
    return {int(x) for x in _S_ATTR.findall(sheet_xml)}


# 열 정의(<cols>) — openpyxl(gen)이 catch-all 열(min=42..16384)에 customWidth="1"을 덧붙여
# Excel이 인쇄범위를 16384열까지 "사용자지정폭"으로 잡고 FitToPagesWide=1이 전부 축소 → 우측
# 빈칸이 격자처럼 보이는 사고. 열 너비는 행 삽입과 무관하므로 완성본 원본 <cols>로 통째 교체.
# (원본 cols의 스타일 인덱스는 merged styles의 base이므로 그대로 유효 — 재매핑 불필요.)
_COLS_RE = re.compile(r"<cols>.*?</cols>", re.DOTALL)


def _restore_template_cols(gen_xml: str, full_sheet_xml: str) -> str:
    m = _COLS_RE.search(full_sheet_xml)
    tmpl_cols = m.group(0) if m else None
    if tmpl_cols is None:                       # 템플릿에 cols 없음 → gen cols 제거
        return _COLS_RE.sub("", gen_xml, count=1)
    if _COLS_RE.search(gen_xml):
        return _COLS_RE.sub(lambda _: tmpl_cols, gen_xml, count=1)
    return gen_xml.replace("<sheetData", tmpl_cols + "<sheetData", 1)  # gen에 cols 없음 → 삽입


# ---------------------------------------------------------------------------
# 공개 API
# ---------------------------------------------------------------------------

def extract_light(full_path: str, sheet_name: str, out_path: str) -> str:
    """완성본에서 sheet_name(총괄표) 한 장만 담은 **가벼운 단독 xlsx**를 생성(생성용).

    생성기는 가벼운 템플릿에서 빨리 채운 뒤 graft_sheet로 완성본에 되돌린다. 이 light 파일은
    1회용이므로 openpyxl 저장으로 컨트롤이 사라져도 무방(완성본 원본은 절대 openpyxl로 안 엶).
    A-0처럼 깨진 도형 참조(image Target="NULL")로 openpyxl 로드가 실패하면, 그 NULL 관계만
    제거한 임시본으로 재시도(총괄표 시트는 도형이 없어 영향 없음). Returns: out_path.
    """
    import openpyxl as _ox
    src = full_path
    tmp = None
    try:
        wb = _ox.load_workbook(full_path)
    except Exception:
        # 깨진 도형/이미지(예 Target="NULL")로 로드 실패 → 도형·이미지 파트와 참조를 통째 제거한
        # 임시본으로 재시도. 총괄표 시트는 도형이 없으므로 추출 결과엔 영향 없음.
        tmp = str(out_path) + ".fix.xlsx"
        with zipfile.ZipFile(full_path) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for n in zin.namelist():
                if n.startswith("xl/drawings/") or n.startswith("xl/media/"):
                    continue                                   # 도형·이미지 파트 제거
                b = zin.read(n)
                if n.startswith("xl/worksheets/") and n.endswith(".xml"):
                    b = re.sub(rb'<drawing\b[^>]*/>', b'', b)
                    b = re.sub(rb'<legacyDrawing\b[^>]*/>', b'', b)
                elif n.endswith(".rels"):
                    b = re.sub(rb'<Relationship[^>]*relationships/(?:drawing|image|vmlDrawing)"[^>]*/>',
                               b'', b)
                elif n == "[Content_Types].xml":
                    b = re.sub(rb'<Override[^>]*(?:drawing|vmlDrawing)[^>]*/>', b'', b)
                zout.writestr(n, b)
        src = tmp
        wb = _ox.load_workbook(src)
    for s in list(wb.sheetnames):
        if s != sheet_name:
            del wb[s]
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    wb.close()
    if tmp and os.path.exists(tmp):
        os.remove(tmp)
    return out_path


def blank_numeric_constants(full_path: str, sheet_name: str, rows, output_path: str = None) -> int:
    """완성본의 sheet_name 시트에서 지정 행들의 **숫자 상수 셀만 값을 비운다**(스타일 보존).

    숫자 상수 = `<v>`가 있고 `<f>`(수식)도 `t=`(문자/불린/inlineStr)도 없는 셀. 따라서 **수식·텍스트는
    절대 비우지 않는다**(회사고유 숫자만 제거). openpyxl로 열지 않아 보조시트 컨트롤·매크로·도형 100% 보존.
    셀 요소를 `<c r=.. s=..>...</c>` → `<c r=.. s=../>`로 바꿔 값만 제거하고 스타일 인덱스(s)는 유지.

    rows: 비울 행 번호 집합(1-indexed). output_path 생략 시 제자리 갱신. Returns: 비운 셀 수.
    """
    output_path = output_path or full_path
    rows = set(int(r) for r in rows)
    with zipfile.ZipFile(full_path) as zf:
        part = _sheet_part(zf, sheet_name)
        if part is None:
            raise ValueError(f"[blank] 시트 '{sheet_name}' 없음")
        data = {n: zf.read(n) for n in zf.namelist()}

    sheet = data[part].decode("utf-8")
    n_done = [0]

    def repl(m):
        full = m.group(0)
        col, row, attrs = m.group("col"), int(m.group("row")), m.group("attrs")
        if row not in rows:
            return full
        if "<f" in full:                       # 수식 보존
            return full
        if re.search(r'\bt="', attrs):         # 문자/불린/inlineStr 보존
            return full
        if "<v>" not in full and "<v " not in full:   # 값 없음 → 그대로
            return full
        sm = re.search(r'\bs="\d+"', attrs)
        s = " " + sm.group(0) if sm else ""
        n_done[0] += 1
        return '<c r="%s%d"%s/>' % (col, row, s)

    sheet = re.sub(r'<c r="(?P<col>[A-Z]+)(?P<row>\d+)"(?P<attrs>[^>]*?)(?:/>|>.*?</c>)',
                   repl, sheet, flags=re.DOTALL)
    data[part] = sheet.encode("utf-8")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    tmp = str(output_path) + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as out:
        for n, b in data.items():
            out.writestr(n, b)
    os.replace(tmp, output_path)
    return n_done[0]


def graft_sheet(full_path: str, generated_path: str, sheet_name: str,
                output_path: str) -> None:
    """완성본(full_path)에 generated_path의 sheet_name 시트를 무손실 이식해 output_path 생성.

    full_path  : 보조시트·컨트롤·매크로가 든 완성본(.xlsx/.xlsm). openpyxl로 열지 않음.
    generated_path: 가벼운 템플릿에서 총괄표를 채워 생성한 파일(openpyxl 산출, 인라인 문자열).
    sheet_name : 이식할 총괄표 시트명(두 파일에서 동일해야 함).
    output_path: 결과(완성본 확장자 유지 권장 — .xlsm이면 매크로 보존).

    완성본의 다른 모든 파트는 바이트 그대로 복사된다. styles.xml만 병합되고 대상 시트만 교체.
    """
    full_path, generated_path, output_path = str(full_path), str(generated_path), str(output_path)

    with zipfile.ZipFile(full_path) as zf_full:
        full_part = _sheet_part(zf_full, sheet_name)
        if full_part is None:
            raise ValueError(f"[graft] 완성본에 시트 '{sheet_name}' 없음")
        base_styles = zf_full.read("xl/styles.xml")
        full_names = zf_full.namelist()
        full_data = {n: zf_full.read(n) for n in full_names}

    with zipfile.ZipFile(generated_path) as zf_gen:
        gen_part = _sheet_part(zf_gen, sheet_name)
        if gen_part is None:
            raise ValueError(f"[graft] 생성본에 시트 '{sheet_name}' 없음")
        gen_sheet = zf_gen.read(gen_part).decode("utf-8")
        gen_styles = zf_gen.read("xl/styles.xml")
        try:
            gen_sst = zf_gen.read("xl/sharedStrings.xml")
        except KeyError:
            gen_sst = None

    # ① 문자열 인라인화 → ② 사용 스타일 수집 → ③ styles 병합 → ④ s= 재매핑
    sheet_xml = _inline_shared_strings(gen_sheet, gen_sst)
    used = _used_style_indices(sheet_xml)
    merged_styles, xf_map = _merge_styles(base_styles, gen_styles, used)
    sheet_xml = _remap_style_indices(sheet_xml, xf_map)
    sheet_xml = _restore_template_cols(sheet_xml, full_data[full_part].decode("utf-8"))
    sheet_bytes = (_XMLDECL + sheet_xml.encode("utf-8")) if not sheet_xml.startswith("<?xml") \
        else sheet_xml.encode("utf-8")

    # 대상 시트의 _rels(원본 총괄표의 도형·하이퍼링크 참조)는 새 시트가 참조 안 하므로 드롭.
    target_rels = re.sub(r"(.*/)?([^/]+)$", r"\1_rels/\2.rels", full_part) \
        if "/" in full_part else f"xl/worksheets/_rels/{full_part}.rels"
    # full_part 형태: 'xl/worksheets/sheetN.xml' → rels: 'xl/worksheets/_rels/sheetN.xml.rels'
    p = Path(full_part)
    target_rels = str(p.parent / "_rels" / (p.name + ".rels")).replace("\\", "/")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as out:
        for name in full_names:
            if name == "xl/styles.xml":
                out.writestr(name, merged_styles)
            elif name == full_part:
                out.writestr(name, sheet_bytes)
            elif name == target_rels:
                continue  # 드롭(새 총괄표는 도형 참조 안 함)
            elif name == "xl/calcChain.xml":
                continue  # 수식 재계산 체인 — 시트 교체로 무효화 → 제거(Excel이 자동 재생성)
            elif name == "[Content_Types].xml":
                # 제거한 calcChain의 Override 항목도 제거(없는 파트 참조 방지)
                b = full_data[name]
                b = re.sub(rb'<Override[^>]*calcChain[^>]*/>', b'', b)
                out.writestr(name, b)
            elif name == "xl/workbook.xml":
                # calcChain을 떼고 openpyxl 수식엔 캐시값이 없으므로, Excel이 열 때 전체 재계산하도록
                # calcPr에 fullCalcOnLoad="1" 주입(없으면). 안 하면 수식 결과가 빈칸으로 보임.
                b = full_data[name].decode("utf-8")
                if "fullCalcOnLoad" not in b:
                    if "<calcPr" in b:
                        b = re.sub(r"<calcPr\b([^>]*?)/>", r'<calcPr\1 fullCalcOnLoad="1"/>', b, count=1)
                    else:  # calcPr 없으면 sheets 앞에 추가
                        b = b.replace("<sheets>", '<calcPr calcId="191029" fullCalcOnLoad="1"/><sheets>', 1)
                out.writestr(name, b.encode("utf-8"))
            else:
                out.writestr(name, full_data[name])
