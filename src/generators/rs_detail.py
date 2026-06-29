# -*- coding: utf-8 -*-
"""R(판매비와관리비)·S(기타손익) 개별시트 매핑 — 총괄표 연동.

총괄표(정산표 매핑)가 채워진 뒤, 그 값에 **연동**해 보조시트를 동적 구성한다:
  - R110 증감분석 : 1)전기대비증감 표를 총괄표 판관비 계정과 동기화(이름매칭).
                    C(전기)=총괄표 D(기초)·D(당기)=총괄표 I(수정후) 직접 링크, E/F 수식 재작성.
                    Threshold C12 = PM×0.5. ABOVE 계정 비고에 X3,X4…(빨강) + 2.증감분석 블록.
                    (급여/퇴직/감가상각 등 Refer-to 계정은 참조 유지·X마커 생략. X1·X2는 템플릿 예시.)
  - R 주석검토   : "15.부가가치 계산" 고정 6계정을 총괄표에 =ROUND(!/1000,0) 연동(천원). a0_detail 패턴.
  - S200 증감분석: 섹션형(영업외수익/비용/법인세) 표를 총괄표와 동기화(동적 길이). F열 수식 신규.
                    Threshold C29 = PM×0.5. ABOVE 비고 빨강 X1,X2… + 2.결론~멘트 사이 블록(없으면 특이사항 없음).

순수 ws 편집 함수 — 호출자(easy_run_rs)가 extract_one+graft로 완성본 보존하며 적용한다.
총괄표 컬럼: 계정명=C, 전기(기초)=D, 당기(수정후)=I  (R/S 공통; r.yaml col_offset 반영).
"""
import re
from copy import copy

from openpyxl.styles import Font

from generators.base import shift_formula_rows

# 총괄표 컬럼(R/S 공통)
_LEAD_NAME_COL = 3     # C 계정명
_LEAD_PRIOR_COL = "D"  # 전기(기초)
_LEAD_CURR_COL = "I"   # 당기(수정후)

_RED = Font(color="FFFF0000")
ANALYSIS_PLACEHOLDER = ("당 계정의 전기 대비 증감에 대하여 감사인은 증감 원인을 파악하고 "
                        "거래내역·증빙과 대사하여 합리성을 검토한다.")

# 주석검토 고정 계정(B8:B13)
_NOTES_ACCOUNTS = ["급여", "퇴직급여", "복리후생비", "세금과공과", "감가상각비", "지급임차료"]


# ── 공용 헬퍼 ───────────────────────────────────────────────────────────
def _norm(s):
    return re.sub(r"\s+", "", str(s or ""))


def _is_subtotal_label(s):
    n = _norm(s)
    return (not n) or n.endswith("계") or n in ("합계", "총계", "소계")


def read_pm(settlement_path):
    """별도정산표에서 'Performance Materiality' 라벨을 1~8행에서 찾아 **우측 셀**(숫자)을 반환.

    회사마다 PM 셀 위치가 달라(M3/N3, N3/O3, J3/K3, P3/Q3…) 라벨 검색으로 찾는다.
    못 찾으면 None(=threshold 수동입력). data_only로 계산값을 읽는다.
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(settlement_path, read_only=True, data_only=True)
    except Exception:
        return None
    if "별도정산표" not in wb.sheetnames:
        wb.close()
        return None
    ws = wb["별도정산표"]
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True)):
        rows.append(row)
    wb.close()
    for row in rows:
        for c, v in enumerate(row):
            if isinstance(v, str) and "performance materiality" in v.lower():
                # 라벨 우측에서 첫 숫자 셀
                for cc in range(c + 1, min(c + 4, len(row))):
                    nv = row[cc]
                    if isinstance(nv, (int, float)) and not isinstance(nv, bool):
                        return float(nv)
    return None


def lead_amounts(settlement_path):
    """별도정산표 파싱 → {정규화 계정명: (전기, 당기)} (전기=기초, 당기=수정후).

    ABOVE 판정(엑셀 미평가 수식 대신 파이썬)과 이름매칭 보조용. 빈 결과면 {}.
    """
    try:
        from extractors.trial_balance import parse_trial_balance
        recs = parse_trial_balance(settlement_path)
    except Exception:
        return {}
    out = {}
    for r in recs:
        nm = r.get("계정명")
        if not nm:
            continue
        prior = r.get("기초") or 0
        curr = r.get("수정후")
        if curr is None:
            curr = r.get("기말") or 0
        out[_norm(nm)] = (float(prior or 0), float(curr or 0))
    return out


def _scan_lead_accounts(lead_ws, first_row, last_scan=300):
    """총괄표 본문에서 [(계정명, 행)] 순서대로 수집.

    본문 끝(소계 '총계'/'기초:' 범례) 도달 시 중단 — 그 아래 감사조서요약·수정사항 등 비계정 행을
    계정으로 오인하지 않게 한다(R 본문은 단일 섹션이라 첫 소계가 곧 본문 끝).
    """
    out = []
    blank = 0
    for r in range(first_row, last_scan):
        v = lead_ws.cell(r, _LEAD_NAME_COL).value
        s = str(v).strip() if v is not None else ""
        if not s:
            blank += 1
            if blank >= 3 and out:
                break
            continue
        blank = 0
        if s.startswith("기초") or s.startswith("기말") or s.startswith("수정후"):
            break                                # 본문 끝 범례 → 중단
        if s in ("<유동>", "<비유동>"):
            continue
        if _is_subtotal_label(s):
            if out:
                break                            # 첫 소계('총계') = 본문 끝
            continue
        out.append((s, r))
    return out


def _scan_lead_sections(lead_ws, first_row, last_scan=300):
    """총괄표 섹션 스캔 → [(소계라벨, [(계정명, 행), ...]), ...] (… '계' 행으로 섹션 닫음)."""
    sections = []
    cur = []
    blank = 0
    for r in range(first_row, last_scan):
        v = lead_ws.cell(r, _LEAD_NAME_COL).value
        s = str(v).strip() if v is not None else ""
        if not s:
            blank += 1
            if blank >= 3 and sections:
                break
            continue
        blank = 0
        if s.startswith("기초") or s.startswith("기말") or s.startswith("수정후"):
            break                                # 본문 끝 범례 → 중단
        if s in ("<유동>", "<비유동>"):
            continue
        n = _norm(s)
        if n.endswith("계") and cur:           # 섹션 소계 → 섹션 닫기
            sections.append((s, cur))
            cur = []
            continue
        if _is_subtotal_label(s):               # 그 외 합계/총계 등은 무시
            continue
        cur.append((s, r))
    if cur:                                     # 소계행 없이 끝난 잔여
        sections.append(("계", cur))
    return sections


def _capture_style(ws, row, cols):
    st = {}
    for c in cols:
        cell = ws.cell(row, c)
        st[c] = dict(font=copy(cell.font), border=copy(cell.border),
                     fill=copy(cell.fill), alignment=copy(cell.alignment),
                     number_format=cell.number_format)
    return st


def _apply_style(ws, row, st):
    for c, item in st.items():
        cell = ws.cell(row, c)
        cell.font = copy(item["font"]); cell.border = copy(item["border"])
        cell.fill = copy(item["fill"]); cell.alignment = copy(item["alignment"])
        cell.number_format = item["number_format"]


def _link(sheet, col, lead_row, thousands=False):
    if thousands:
        return "=ROUND('%s'!%s%d/1000,0)" % (sheet, col, lead_row)
    return "='%s'!%s%d" % (sheet, col, lead_row)


def _is_above(name, amounts, threshold):
    """ABOVE 판정(파이썬): |수정후-기초| > threshold. 금액·threshold 없으면 False."""
    if threshold is None:
        return False
    a = amounts.get(_norm(name))
    if not a:
        return False
    return abs(a[1] - a[0]) > threshold


# ── R110 증감분석 비고 분류(Refer-to / X1·X2) ───────────────────────────
def _refer_text(name):
    """Refer-to 분류(타조서에서 분석되는 계정). 표준 cross-ref이라 규칙 인코딩."""
    n = _norm(name)
    if "퇴직" in n:
        return "Refer to 4000-EE"
    if ("급여" in n or "상여" in n or "임금" in n) and "퇴직" not in n:
        return "Refer to 4000-R-2"
    if "감가상각" in n or "무형자산상각" in n:
        return "Refer to 4000-G-1"
    return None


def _preset_xmark(name):
    """템플릿 예시 X1(복리후생비)·X2(세금과공과) 고정 매핑."""
    n = _norm(name)
    if "복리후생" in n:
        return "X1"
    if "세금과공과" in n:
        return "X2"
    return None


# ── R110 ────────────────────────────────────────────────────────────────
def fill_r110(ws, lead_ws, lead_sheet, threshold, amounts):
    """R110 증감분석 시트를 총괄표 판관비 계정과 동기화."""
    HDR, FIRST, SUB = 14, 15, 40              # 헤더14·데이터15~·소계40(템플릿)
    accts = _scan_lead_accounts(lead_ws, 12)
    n = len(accts)
    data_style = _capture_style(ws, FIRST, range(2, 8))   # B..G
    # X1 헤더/설명 스타일(증감분석 블록 도너)
    x_hdr_style = _capture_style(ws, 48, [2])
    x_desc_style = _capture_style(ws, 50, [2])

    # 1) 표 영역 크기 보정 (n>25면 소계행 위로 행삽입)
    cap = SUB - FIRST                          # 25
    sub_row = SUB
    if n > cap:
        add = n - cap
        ws.insert_rows(SUB, add)
        shift_formula_rows(ws, SUB, add)
        for r in range(FIRST + cap, FIRST + n):   # 새 행 스타일
            _apply_style(ws, r, data_style)
        sub_row = SUB + add

    # 2) 계정행 채우기
    new_x, x_no = [], 3                          # 신규 X마커는 X3부터
    for i, (name, lr) in enumerate(accts):
        r = FIRST + i
        ws.cell(r, 2).value = name                                  # B 계정명
        ce = ws.cell(r, 3); ce.value = _link(lead_sheet, _LEAD_PRIOR_COL, lr)  # C 전기
        cd = ws.cell(r, 4); cd.value = _link(lead_sheet, _LEAD_CURR_COL, lr)   # D 당기
        ws.cell(r, 5).value = "=D%d-C%d" % (r, r)                   # E 증감
        ws.cell(r, 6).value = '=IF(ABS(E%d)>C$12,"Above","below")' % r  # F
        # G 비고
        g = ws.cell(r, 7)
        refer = _refer_text(name)
        xpre = _preset_xmark(name)
        if refer:
            g.value = refer; g.font = copy(data_style[7]["font"])
        elif xpre:
            g.value = xpre; g.font = copy(_RED)
        elif _is_above(name, amounts, threshold):
            g.value = "X%d" % x_no; g.font = copy(_RED)
            new_x.append((x_no, name)); x_no += 1
        else:
            g.value = None; g.font = copy(data_style[7]["font"])
    # 잔여 템플릿 행 값 비우기(테두리 보존)
    for r in range(FIRST + n, sub_row):
        for c in range(2, 8):
            ws.cell(r, c).value = None

    # 3) 소계 SUM 범위 재지정
    ws.cell(sub_row, 3).value = "=SUM(C%d:C%d)" % (FIRST, sub_row - 1)
    ws.cell(sub_row, 4).value = "=SUM(D%d:D%d)" % (FIRST, sub_row - 1)

    # 4) Threshold
    if threshold is not None:
        c12 = ws["C12"]; c12.value = int(round(threshold)); c12.number_format = "#,##0"

    # 5) 2.증감분석 블록: 기존 X1/X2 다음에 신규 X3+ append
    _append_analysis_blocks(ws, new_x, x_hdr_style, x_desc_style)
    return {"accounts": n, "above": len(new_x), "threshold": threshold}


def _append_analysis_blocks(ws, new_x, hdr_style, desc_style):
    """B열에서 마지막 'X\\d ' 블록을 찾아 그 아래로 신규 분석블록을 붙인다."""
    if not new_x:
        return
    last = None
    for r in range(45, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and re.match(r"^X\d+\s", v.strip()):
            last = r
    base = (last + 3) if last else 47          # 헤더→설명 +2, 다음 블록 +1 갭
    r = base
    for no, name in new_x:
        ws.cell(r, 2).value = "X%d %s" % (no, name)
        _apply_style(ws, r, hdr_style)
        ws.cell(r + 2, 2).value = ANALYSIS_PLACEHOLDER
        _apply_style(ws, r + 2, desc_style)
        r += 4


# ── R 주석검토 (천원, 고정 6계정) ───────────────────────────────────────
def fill_r_notes(ws, lead_ws, lead_sheet):
    """주석검토 "15.부가가치 계산" 고정계정을 총괄표에 =ROUND(!/1000,0) 연동."""
    accts = {_norm(name): row for name, row in _scan_lead_accounts(lead_ws, 12)}
    matched = 0
    for i, label in enumerate(_NOTES_ACCOUNTS):
        r = 8 + i                              # B8..B13
        lr = _match_account(label, accts)
        if lr is None:
            ws.cell(r, 3).value = None; ws.cell(r, 4).value = None
            continue
        c = ws.cell(r, 3); c.value = _link(lead_sheet, _LEAD_CURR_COL, lr, thousands=True)  # C 당기
        c.number_format = "#,##0"
        d = ws.cell(r, 4); d.value = _link(lead_sheet, _LEAD_PRIOR_COL, lr, thousands=True)  # D 전기
        d.number_format = "#,##0"
        matched += 1
    # 합계행(B14) — 현재 빈칸
    ws.cell(14, 3).value = "=SUM(C8:C13)"; ws.cell(14, 3).number_format = "#,##0"
    ws.cell(14, 4).value = "=SUM(D8:D13)"; ws.cell(14, 4).number_format = "#,##0"
    return {"matched": matched, "total": len(_NOTES_ACCOUNTS)}


def _match_account(label, accts):
    """이름매칭: 정확 → 부분일치(퇴직 가드). accts={정규화명: 행}."""
    n = _norm(label)
    if n in accts:
        return accts[n]
    cands = []
    for k, row in accts.items():
        if ("퇴직" in k) != ("퇴직" in n):
            continue
        if k in n or n in k:
            cands.append((len(k), row))
    return max(cands)[1] if cands else None


# ── S200 ─────────────────────────────────────────────────────────────────
def fill_s200(ws, lead_ws, lead_sheet, threshold, amounts):
    """S200 증감분석 섹션형 표를 총괄표와 동기화(동적 길이)."""
    HDR, FIRST, OLD_END = 12, 13, 27          # 헤더12·데이터13~·표끝27(템플릿)
    sub_donor = 20                             # 소계 도너행
    sections = _scan_lead_sections(lead_ws, 11)
    data_style = _capture_style(ws, FIRST, range(2, 8))
    sub_style = _capture_style(ws, sub_donor, range(2, 8))

    needed = sum(len(a) for _, a in sections) + len(sections)
    old_rows = OLD_END - FIRST + 1
    delta = needed - old_rows
    if delta > 0:
        ws.insert_rows(OLD_END + 1, delta)
        shift_formula_rows(ws, OLD_END + 1, delta)
    elif delta < 0:
        ws.delete_rows(FIRST + needed, -delta)
        shift_formula_rows(ws, OLD_END + 1, delta)

    # 표 리사이즈 후 Threshold 셀 위치 재탐지(표 밑이라 행이 밀림) → F 수식이 이 셀을 참조
    thr_row = _find_label_cell(ws, "Threshold", col=2) or (OLD_END + 2)

    # 표 재구성
    above_accts = []
    r = FIRST
    for label, rows in sections:
        first = r
        for name, lr in rows:
            _apply_style(ws, r, data_style)
            ws.cell(r, 2).value = name
            ws.cell(r, 3).value = _link(lead_sheet, _LEAD_PRIOR_COL, lr)   # C 전기
            ws.cell(r, 4).value = _link(lead_sheet, _LEAD_CURR_COL, lr)    # D 당기
            ws.cell(r, 5).value = "=D%d-C%d" % (r, r)                      # E
            ws.cell(r, 6).value = '=IF(ABS(E%d)>C$%d,"Above","below")' % (r, thr_row)  # F
            if _is_above(name, amounts, threshold):
                above_accts.append((r, name))
            r += 1
        last = r - 1
        _apply_style(ws, r, sub_style)
        ws.cell(r, 2).value = label
        for col, ltr in ((3, "C"), (4, "D"), (5, "E")):
            ws.cell(r, col).value = ("=SUM(%s%d:%s%d)" % (ltr, first, ltr, last)
                                     if last >= first else 0)
        r += 1

    # ABOVE 비고 빨강 X1,X2…
    for i, (rr, name) in enumerate(above_accts, start=1):
        g = ws.cell(rr, 7); g.value = "X%d" % i; g.font = copy(_RED)

    # Threshold 값
    if threshold is not None:
        c = ws.cell(thr_row, 3); c.value = int(round(threshold)); c.number_format = "#,##0"

    # 2.결론 ~ (*)멘트 사이 블록
    concl = _find_label_cell(ws, "2. 결론", col=2)
    ment = _find_ment_row(ws, concl)
    _fill_s_conclusion(ws, concl, ment, above_accts, data_style)
    return {"sections": len(sections), "above": len(above_accts), "threshold": threshold}


def _find_label_cell(ws, text, col=2):
    t = _norm(text)
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, col).value
        if isinstance(v, str) and _norm(v) == t:
            return r
    # 부분일치 폴백
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, col).value
        if isinstance(v, str) and t in _norm(v):
            return r
    return None


def _find_ment_row(ws, after_row):
    """(*) 분석대상/멘트 행(A열 [멘트]) 탐색 — 결론 헤더 다음."""
    start = (after_row or 1) + 1
    for r in range(start, ws.max_row + 2):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and ("멘트" in v or "분석대상" in v):
            return r
    return None


def _fill_s_conclusion(ws, concl, ment, above_accts, data_style):
    """결론 헤더와 멘트 사이에 분석블록 기재(없으면 특이사항 없음)."""
    if concl is None:
        return
    if not above_accts:
        ws.cell(concl + 1, 2).value = "특이사항 없음"
        return
    # 필요한 행수 = 블록당 1행 + 헤더 갭, 멘트 직전에 공간 확보
    need = len(above_accts)
    gap = (ment - concl - 1) if ment else 99
    if ment and gap < need + 1:
        ins = (need + 1) - gap
        ws.insert_rows(ment, ins)
        shift_formula_rows(ws, ment, ins)
    r = concl + 2                              # 헤더 아래 1행 갭
    for i, (_, name) in enumerate(above_accts, start=1):
        cell = ws.cell(r, 2)
        cell.value = "X%d %s — %s" % (i, name, ANALYSIS_PLACEHOLDER)
        cell.font = Font(bold=False)
        r += 1
