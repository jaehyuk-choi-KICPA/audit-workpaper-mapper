"""A-100 (완전성 체크) 시트 생성기.

CS Control Sheet(발송엑셀) 파싱 결과로 3.Test 내역 표와 4.Test 결론을 채운다.

체크 컬럼 규칙 (완성본 V로 확정):
  - 당기 회사제시 CS / 금융자산·부채 원장 / 관련 손익 원장 : 무조건 V
  - 은행연합회 자료 : 공란 (외부 대조 결과라 입력 3파일로 판단 불가 → 감사인 영역)
  - 발송여부 : V

공통 헤더(회사·날짜·Preparer·Reviewer)는 사용자가 주는 런타임 값을 그대로 얹는다.
회사명/날짜는 이 시트에 기입하며, A-200/A-300은 이 셀을 참조하는 수식으로 가져온다.

양식 보존: base 공용 도구(capture_row_style/clear_region/render_rows) 사용.
"""

from .base import WorkpaperGenerator

_V = "V"
_BLANK = ""


class A100Generator(WorkpaperGenerator):

    def fill(self, cs_rows: list[dict], params: dict) -> None:
        self.open_template()
        cfg = self.config
        self._fill_header(cfg.get("header", {}), params)
        last_row = self._fill_test3_conclusion(cfg["test3"], cfg.get("conclusion", {}), cs_rows)
        self._clean_outside_content(cfg.get("layout", {}), last_row)

    def _fill_header(self, h: dict, params: dict) -> None:
        if h.get("company_cell") and params.get("회사명"):
            self.write_cell(h["company_cell"],
                            f"{h.get('company_prefix', '')}{params['회사명']}")
        if h.get("date_cell") and params.get("날짜"):
            self.write_cell(h["date_cell"], params["날짜"])
        if h.get("preparer_cell") and params.get("preparer"):
            self.write_cell(h["preparer_cell"], params["preparer"])
        if h.get("reviewer_cell") and params.get("reviewer"):
            self.write_cell(h["reviewer_cell"], params["reviewer"])

    def _fill_test3_conclusion(self, t3: dict, c: dict, cs_rows: list[dict]) -> int:
        """3.Test 내역 표 + 4.Test 결론을 채운다. 결론은 표 바로 아래로 동적 배치.

        (기존 버전은 결론이 R58 고정 → 기관 수가 적으면 표와 결론 사이 빈칸이 과도)
        Returns: 마지막으로 쓴 행 번호(인쇄영역·정리 경계 계산용).
        """
        data_start = t3["data_start_row"]
        b_start, b_end = c["block_start_row"], c["block_end_row"]
        gap = c.get("gap_after_data", 1)
        token = c.get("count_token", "N개")

        # 0) 도너 스타일 + 결론 블록(값+스타일+병합) 캡처 (초기화 전)
        style = self.capture_row_style(t3["style_donor_row"])
        block = [self.capture_row(r) for r in range(b_start, b_end + 1)]
        block_merges = [(m.min_row, m.min_col, m.max_row, m.max_col)
                        for m in self.ws.merged_cells.ranges if b_start <= m.min_row <= b_end]

        # 1) 표~옛 결론 영역 일괄 초기화 (값·스타일·병합)
        self.clear_region(data_start, b_end)

        # 2) 데이터 행 렌더
        rows = [{
            "구분":           r.get("구분"),
            "일련번호":        r.get("일련번호"),
            "금융기관":        r.get("금융기관명") or "",
            "당기회사제시CS":   _V,
            "은행연합회자료":   _BLANK,   # 감사인 영역
            "금융자산부채원장": _V,
            "관련손익원장":     _V,
            "발송여부":        _V,
        } for r in cs_rows]
        self.render_rows(data_start, t3["columns"], rows, style)
        data_end = data_start + len(rows) - 1

        # 3) 결론 블록을 표 바로 아래로 stamp + 병합 재현 + 'N개' 치환
        new_start = data_end + gap + 1
        delta = new_start - b_start
        for i, cap in enumerate(block):
            dst = new_start + i
            self.stamp_row(dst, cap)
            for col in list(cap):
                v = self.ws.cell(row=dst, column=col).value
                if isinstance(v, str) and token in v:
                    self.ws.cell(row=dst, column=col).value = v.replace(token, f"{len(rows)}개")
        for (r1, c1, r2, c2) in block_merges:
            self.safe_merge_idx(r1 + delta, c1, r2 + delta, c2)

        return new_start + (b_end - b_start)

    def safe_merge_idx(self, r1, c1, r2, c2):
        from openpyxl.utils import get_column_letter
        self.safe_merge(f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}")

    def _clean_outside_content(self, layout: dict, last_row: int) -> None:
        """인쇄영역을 내용 범위로 한정한다.

        바깥 열(A·B·M+)의 회색 음영은 양식 원본 디자인이므로 건드리지 않는다.
        (validator.check_outside_col_fill이 이 음영 보존 여부를 게이트로 검사)
        """
        from openpyxl.utils import get_column_letter, column_index_from_string
        last_col = column_index_from_string(layout.get("content_last_col", "L"))
        self.ws.print_area = f"A1:{get_column_letter(last_col)}{last_row}"
