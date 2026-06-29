# -*- coding: utf-8 -*-
"""BBDD(장단기차입금) 상세 시트 채우기 — 후처리(extract_light 산출 시트에 직접 작업).

100_총괄표 상단 1.총괄표는 refill.fill_refill(bbdd.yaml)이 담당. 이 모듈은 상세 영역을 채운다:
  · 200_차입금 Test  : 1.3) 금융거래조회서 확인 표(우=전자조회서 BANK 2-2) + 담보표(BANK 9)
  · 300_차입금명세표  : 명세(2-2) + 변동(거래처원장)
  · 400_주석사항      : 장기차입금 종류(대출종류)
  · 100 RECAP/이자비용 : 거래처원장 잔액 + 분개장 이자비용

동적변환 원칙(위반 금지): 행수는 데이터에 따라 가변. insert/delete 직후 shift_formula_rows로
보존영역 수식참조 보정, 도너행 스타일 복사로 테두리 보존, 합계/check SUM 범위는 실제 데이터행으로
재작성, 존재 안 하는 계정 섹션은 삭제. 위치는 절대행이 아니라 라벨/앵커로 찾는다.

BBDD 특유 규칙: 이자율이 다른 차입금은 같은 금융기관명이라도 각각 별도 행(파서가 행단위 → 그대로).
"""

import re
from copy import copy

from openpyxl.utils import column_index_from_string

from .base import shift_formula_rows


_NONE_TOKENS = {"", "'", "해당없음", "해당없슴", "해당사항없음", "담보보증내역없음", "내역없음", "-"}
_RATE_FMT = "0.00%"   # 이자율 통일: 퍼센트·소수점 둘째자리 반올림 (사용자 지정)


def _norm(s) -> str:
    return re.sub(r"\s+", "", str(s)) if s is not None else ""


def _clean(v):
    """'해당없음'·"' "·내역없음 등 빈 토큰을 None으로."""
    if v is None:
        return None
    return None if _norm(v) in _NONE_TOKENS else v


def _amt(v) -> float:
    """문자/숫자 → float. 비수치는 0."""
    if v is None or isinstance(v, bool):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[^\d.\-]", "", str(v))
    try:
        return float(s) if s not in ("", "-", ".") else 0.0
    except ValueError:
        return 0.0


def _fmt_date(v):
    """YYYYMMDD → YYYY-MM-DD. sentinel(00000000/00010101)·무효는 빈칸."""
    s = str(v).strip() if v is not None else ""
    if re.fullmatch(r"\d{8}", s):
        if s in ("00000000", "00010101", "00001010"):
            return ""
        return f"{s[:4]}-{s[4:6]}-{s[6:]}"
    return v


def _find_label_row(ws, col_idx, label, start=1, end=None):
    """col_idx 열에서 label(정규화 부분일치)을 가진 행을 동적 탐색."""
    target = _norm(label)
    end = end or ws.max_row
    for r in range(start, end + 1):
        if target and target in _norm(ws.cell(r, col_idx).value):
            return r
    return None


def _copy_row_style(ws, src_row, dst_row, c0, c1):
    """src_row의 c0~c1 열 스타일(테두리·폰트·정렬·서식·채움)을 dst_row에 복사."""
    for c in range(c0, c1 + 1):
        s, d = ws.cell(src_row, c), ws.cell(dst_row, c)
        d.border = copy(s.border)
        d.font = copy(s.font)
        d.alignment = copy(s.alignment)
        d.fill = copy(s.fill)
        d.number_format = s.number_format
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def _resize_block(ws, data_start, cur, need, donor_row, c0, c1):
    """데이터 블록 행수를 cur→need로 insert/delete. 삽입행은 donor 스타일 복사 + shift_formula_rows.

    반환: 행수 변화 delta(아래 보존행 이동량). 합계/check 등 호출자가 delta로 위치 재계산.
    """
    if need > cur:
        n = need - cur
        at = data_start + cur                 # 마지막 데이터행 다음(=합계 위)
        ws.insert_rows(at, n)
        shift_formula_rows(ws, at, n)
        for r in range(at, at + n):
            _copy_row_style(ws, donor_row, r, c0, c1)
        return n
    if need < cur:
        n = cur - need
        ws.delete_rows(data_start + need, n)
        shift_formula_rows(ws, data_start + need + n, -n)  # 삭제 아래 보존 수식 행참조 -n 보정(check 등)
        return -n
    return 0


def _fill_table(ws, sec: dict, rows: list) -> dict:
    """헤더행 아래 placeholder 데이터 영역을 rows로 동적 렌더(insert/delete + 도너스타일).

    sec: {header_row, placeholder_rows, cols{열문자:키|@리터럴}, donor_cols, date_cols, rate_cols, rate_div}
    합계 없는 단순 표 전용(담보·명세 등). 데이터 0건이면 placeholder 1행 비움.
    """
    hr = sec["header_row"]
    ds = hr + 1
    c0, c1 = sec.get("donor_cols", [2, 8])
    _resize_block(ws, ds, sec.get("placeholder_rows", 1), max(len(rows), 1), ds, c0, c1)
    _render_cols(ws, sec, rows, ds)
    return {"filled": len(rows)}


def _render_cols(ws, sec, rows, ds):
    """ds부터 rows를 cols 맵으로 렌더(날짜/이자율/금액 포맷 적용). 0건이면 placeholder 1행 비움."""
    cols = sec["cols"]
    date_cols = set(sec.get("date_cols", []))
    rate_cols = set(sec.get("rate_cols", []))
    rate_div = sec.get("rate_div", 1) or 1
    num_cols = set(sec.get("num_cols", []))
    num_fmt = sec.get("num_format", "#,##0")
    rate_fmt = sec.get("rate_format", _RATE_FMT)
    for i, rec in enumerate(rows):
        r = ds + i
        for col, key in cols.items():
            cell = ws[f"{col}{r}"]
            if isinstance(key, str) and key.startswith("@"):
                cell.value = key[1:]
            elif col in date_cols:
                cell.value = _fmt_date(rec.get(key))
            elif col in rate_cols:
                cell.value = _amt(rec.get(key)) / rate_div
                cell.number_format = rate_fmt
            elif col in num_cols:
                cell.value = _amt(rec.get(key))
                cell.number_format = num_fmt
            else:
                cell.value = _clean(rec.get(key))
    if not rows:
        for col in cols:
            ws[f"{col}{ds}"] = None


def _fill_with_total(ws, sec: dict, rows: list) -> dict:
    """헤더+데이터+합계 표(동적). 합계행은 라벨로 탐지, SUM 범위는 실데이터행으로 재작성.

    sec: _fill_table 키 + {total_label, sum_cols}. (300 미지급이자 표1/표2.)
    """
    hr = sec["header_row"]
    ds = hr + 1
    c0, c1 = sec.get("donor_cols", [2, 9])
    tot0 = _find_label_row(ws, 2, sec["total_label"], ds)
    if tot0 is None:
        return {"filled": 0, "error": "합계행 미탐지"}
    cur = tot0 - ds
    need = max(len(rows), 1)
    _resize_block(ws, ds, cur, need, ds, c0, c1)
    _render_cols(ws, sec, rows, ds)
    tot = ds + need
    de = ds + len(rows) - 1
    if rows:
        for col in sec.get("sum_cols", []):
            ws[f"{col}{tot}"] = f"=SUM({col}{ds}:{col}{de})"
    return {"filled": len(rows), "total_row": tot}


def fill_bbdd300(ws, cfg: dict, loans: list) -> dict:
    """300 미지급이자 — 표1(완전성)·표2(재계산). 2-2 활성대출(대출금액>0) 처음 N열.

    하단(표2 R32)→상단(표1 R11) 순으로 처리(상단 행삽입이 하단 위치를 밀어도 안전).
    """
    active = [r for r in loans if _amt(r.get("대출금액")) > 0]
    return {
        "t2": _fill_with_total(ws, cfg["spec_accrued2"], active),
        "t1": _fill_with_total(ws, cfg["spec_accrued1"], active),
    }


def fill_bbdd_inquiry(ws, cfg: dict, a300: dict) -> dict:
    """조회서_기타조회서류 — 퇴직연금(1번)/대출(2-2)/담보(9번) 3표. build_a300_data 결과 소비.

    하단(담보 R13)→상단(대출 R8 → 퇴직연금 R4) 순 처리.
    """
    return {
        "collateral": _fill_table(ws, cfg["inquiry_collateral"], a300.get("collateral", [])),
        "loans": _fill_table(ws, cfg["inquiry_loans"], a300.get("loans", [])),
        "pension": _fill_table(ws, cfg["inquiry_pension"], a300.get("pension", [])),
    }


def fill_bbdd200(ws, cfg: dict, loans: list, collateral: list | None = None,
                 balances: list | None = None) -> dict:
    """200_차입금 Test — 조회확인 표(우=BANK 2-2) + 담보표(3.2=BANK 9).

    하단(담보)→상단(조회박스) 순 처리: 상단 행삽입이 하단 위치를 밀어도 안전(절대행 가정 유지).
    loans: parse_bank_loans(대출금액>0). collateral: parse_bank_collateral. balances: Phase 2(좌측).
    """
    out = {}
    if collateral is not None and cfg.get("test_collateral"):
        out["collateral"] = _fill_table(ws, cfg["test_collateral"], collateral)
    ti = cfg["test_inquiry"]
    ds = ti["data_start"]
    nci = 2  # B열(거래처명/합계 라벨)
    c0, c1 = ti.get("donor_cols", [2, 20])

    active = [r for r in loans if _amt(r.get(ti["loan_amount_key"])) > 0]
    n = len(active)

    tot0 = _find_label_row(ws, nci, ti["total_label"], ds)
    if tot0 is None:
        return {"filled": 0, "error": "합계행 미탐지"}
    cur = tot0 - ds

    delta = _resize_block(ws, ds, cur, max(n, 1), ds, c0, c1)
    tot = tot0 + delta
    de = ds + n - 1                            # 마지막 데이터행

    # 우측(전자조회서) 렌더
    right = ti["right_cols"]
    right_formulas = ti.get("right_formulas", {})
    date_cols = set(ti.get("date_cols", []))
    rate_cols = set(ti.get("rate_cols", []))
    rate_div = ti.get("rate_div", 1) or 1
    clear_cols = ti.get("clear_cols", [])
    for i, ln in enumerate(active):
        r = ds + i
        for col, key in right.items():
            cell = ws[f"{col}{r}"]
            if isinstance(key, str) and key.startswith("@"):
                cell.value = key[1:]
            elif col in date_cols:
                cell.value = _fmt_date(ln.get(key))     # native mm-dd-yy는 문자열에 무영향
            elif col in rate_cols:
                cell.value = _amt(ln.get(key)) / rate_div  # 분수
                cell.number_format = ti.get("rate_format", _RATE_FMT)  # 0.00% 통일
            else:
                cell.value = _clean(ln.get(key))
        for col, tmpl in right_formulas.items():
            ws[f"{col}{r}"] = tmpl.format(r=r)
        for col in clear_cols:
            ws[f"{col}{r}"] = None

    # 좌측(회사잔액 split) — Phase 2
    if balances:
        _fill_left_balances(ws, ti, active, balances, ds)

    # 합계 SUM 범위 재작성(실제 데이터행)
    for col in ti.get("sum_cols", []):
        ws[f"{col}{tot}"] = f"=SUM({col}{ds}:{col}{de})"
    # check 행 재작성
    chk = tot + ti.get("check_offset", 1)
    for col, tmpl in ti.get("check", {}).items():
        ws[f"{col}{chk}"] = tmpl.format(tot=tot)

    return {"filled": n, "total_row": tot, "data_end": de}


def _fill_left_balances(ws, ti, active, balances, ds):
    """좌측 회사잔액 split — 조회 대출금액과 동일 금액(기말) 라인정렬. 불일치 시 (*1) 마커.

    balances: borrowings_by_inst 결과 [{금융기관, 대출종류, 계정분류, 기초, 기말}].
    active:   200 우측에 렌더된 대출 목록(라인 순서 동일). 매칭=기말==대출금액(중복은 소비).
    """
    left = ti["left_cols"]
    rowf = ti.get("row_total_formula", {})
    mcol = ti.get("mismatch_col")
    mtext = ti.get("mismatch_mark", "(*1)")
    by_amt = {}
    for b in balances or []:
        by_amt.setdefault(round(_amt(b.get("기말"))), []).append(b)
    for i, ln in enumerate(active):
        r = ds + i
        cands = by_amt.get(round(_amt(ln.get("대출금액"))))
        b = cands.pop(0) if cands else None
        if b:
            ws[f"{left['거래처명']}{r}"] = b.get("금융기관")
            cls = b.get("계정분류")
            if cls in left:
                ws[f"{left[cls]}{r}"] = _amt(b.get("기말"))
        elif mcol:                              # 잔액명세서에 매칭 없음 → 확인필요
            ws[f"{mcol}{r}"] = mtext
        for col, tmpl in rowf.items():
            ws[f"{col}{r}"] = tmpl.format(r=r)


# ===========================================================================
# 100 RECAP (2.Recap) — 거래처원장 차입금 시설별 기초/기말 (장기/유동성장기/단기)
# ===========================================================================

def _find_recap_title(ws, cci, kw, exclude=None, start=1):
    """C열 start행부터 kw 포함(exclude 제외) 타이틀 행 탐색.

    ★start로 'Recap' 섹션 아래로 제한 — 그렇지 않으면 1.총괄표 leaf의 동명 계정(단기차입금/
    유동성장기부채/장기차입금)이 먼저 잡혀 RECAP 대신 총괄표 영역을 파괴한다(반복 사고).
    """
    k = _norm(kw)
    ex = _norm(exclude) if exclude else None
    for r in range(start, ws.max_row + 1):
        v = _norm(ws.cell(r, cci).value)
        if k in v and (not ex or ex not in v):
            return r
    return None


def _fill_recap_table(ws, recap, sec, rows):
    """한 RECAP 표(장기/유동성장기/단기) 채움 — 동적 행수 + 합계 SUM 재작성.

    순번(C 자동) / 금융기관(D) / 대출종류(E) / 기초(F) / 기말(K) / 증감(L=K-F). 차입·상환·대체는 공란.
    """
    cci = recap["num_col"]
    anchor = _find_label_row(ws, cci, recap.get("section_anchor", "Recap"))  # 'Recap' 아래로 제한(매번 재탐색)
    title = _find_recap_title(ws, cci, sec["title_kw"], sec.get("title_exclude"), start=anchor or 1)
    if title is None:
        return {"filled": 0, "error": "타이틀 미탐지"}
    ds = title + recap.get("header_offset", 2) + 1
    tot0 = _find_label_row(ws, cci, recap["total_label"], ds)
    if tot0 is None:
        return {"filled": 0, "error": "합계행 미탐지"}
    cur = tot0 - ds
    need = max(len(rows), 1)
    c0, c1 = recap.get("donor_cols", [3, 12])
    _resize_block(ws, ds, cur, need, ds, c0, c1)
    tot = ds + need
    de = ds + len(rows) - 1
    ic, tc, bc, ec = recap["inst_col"], recap["type_col"], recap["base_col"], recap["end_col"]
    rowf = recap.get("row_formula", {})
    for i, b in enumerate(rows):
        r = ds + i
        ws.cell(r, cci).value = i + 1
        ws[f"{ic}{r}"] = b.get("금융기관")
        ws[f"{tc}{r}"] = b.get("대출종류")
        ws[f"{bc}{r}"] = _amt(b.get("기초"))
        ws[f"{ec}{r}"] = _amt(b.get("기말"))
        for col, tmpl in rowf.items():
            ws[f"{col}{r}"] = tmpl.format(r=r)
    if not rows:                                # 데이터 없으면 placeholder 1행 전체 비움(L/M 수식 #VALUE! 방지)
        for c in range(cci, 14):                # C~M
            ws.cell(ds, c).value = None
    else:
        for col in recap.get("sum_cols", []):
            ws[f"{col}{tot}"] = f"=SUM({col}{ds}:{col}{de})"
    return {"filled": len(rows), "total_row": tot}


def _prune_recap_table(ws, recap, sec) -> int:
    """빈 RECAP 표(타이틀~다음 섹션 직전)를 통째 삭제. 삭제 행수 반환(shift 보정 포함)."""
    cci = recap["num_col"]
    anchor = _find_label_row(ws, cci, recap.get("section_anchor", "Recap"))
    title = _find_recap_title(ws, cci, sec["title_kw"], sec.get("title_exclude"), start=anchor or 1)
    if title is None:
        return 0
    end = None
    for r in range(title + 1, ws.max_row + 1):       # 다음 번호 섹션(차입금/이자비용)
        v = _norm(ws.cell(r, cci).value)
        if re.match(r"\d+\.", v) and ("차입금" in v or "이자비용" in v):
            end = r
            break
    if end is None:
        return 0
    n = end - title
    ws.delete_rows(title, n)
    shift_formula_rows(ws, title, -n)
    return n


def _renumber_recap(ws, recap) -> None:
    """2.Recap 아래 번호 섹션(차입금 표 + 이자비용)을 순차 재번호(prune로 생긴 빈 번호 메움)."""
    cci = recap["num_col"]
    anchor = _find_label_row(ws, cci, recap.get("section_anchor", "Recap"))
    if not anchor:
        return
    end = _find_label_row(ws, cci, "감사조서 요약", anchor) or ws.max_row
    i = 1
    for r in range(anchor + 1, end):
        v = ws.cell(r, cci).value
        if v and re.match(r"\s*\d+\.", str(v)) and ("차입금" in str(v) or "이자비용" in str(v)):
            ws.cell(r, cci).value = re.sub(r"^\s*\d+\.\s*", f"{i}. ", str(v))
            i += 1


def fill_bbdd100_recap(ws, cfg: dict, borrowings: list) -> dict:
    """100 2.Recap — 장기/유동성장기/단기 표를 거래처원장 차입금 시설로 채움(하단→상단).

    데이터 없는 표는 subtractive 삭제 + 번호 재정렬(사용자 대원칙: 존재 안 하는 계정 번호 제거).
    """
    recap = cfg["recap"]
    by_cls = {}
    for b in borrowings or []:
        by_cls.setdefault(b["계정분류"], []).append(b)
    out = {}
    for sec in reversed(recap["tables"]):
        out[sec["class"]] = _fill_recap_table(ws, recap, sec, by_cls.get(sec["class"], []))
    if recap.get("prune_empty", True):
        for sec in reversed(recap["tables"]):        # 하단→상단 삭제(상단 타이틀 위치 보존)
            if out[sec["class"]].get("filled", 0) == 0:
                _prune_recap_table(ws, recap, sec)
        _renumber_recap(ws, recap)
    return out


# ===========================================================================
# 100 4.이자비용 — 이상적 분개장에서 이자비용 필터 → 거래처(은행)별 집계
# ===========================================================================

def _bracket_code(s):
    m = re.match(r"\s*\[([^\]]+)\]", str(s or ""))
    return m.group(1).strip() if m else None


def _strip_bracket(s):
    return re.sub(r"^\s*\[[^\]]*\]\s*", "", str(s or "")).strip()


def _inst_only(s):
    """거래처 정규화: 괄호(계좌/용도) 제거 + 공백 제거. 이미 깨끗하면 그대로."""
    if s is None:
        return None
    s = re.sub(r"\s+", "", re.sub(r"\(.*?\)", "", str(s)))
    return s or None


def fill_bbdd100_interest(ws, cfg: dict, journal_lines: list) -> dict:
    """100 4.이자비용 — 이상적 분개장에서 '이자비용' 계정 필터 → (계정과목,거래처) 집계 렌더.

    전기이월 0 / 증가=Σ차변 / 감소=Σ대변 / 잔액=전기이월+증가-감소(수식). 합계 잔액 = 이자비용 leaf와 대사.
    """
    sec = cfg["interest"]
    cci = 3
    kw = sec.get("filter_kw", "이자비용")
    # 이자비용 = Σ차변(음수 정정분개 포함, 정산표 일치 387M). 대변(월말 역분개·손익대체 마감)은
    # 회계 메커니즘 → 증가에 미반영. 거래처별 net 차변으로 집계, 0-합 그룹(손익대체 마감)은 제외.
    # 미지급이자(발생-지급 차이)는 300 시트가 별도 검토.
    izl = [l for l in (journal_lines or []) if kw in str(l.get("계정과목") or "")]
    groups = {}
    for l in izl:
        acct_raw = str(l.get("계정과목") or "")
        key = (_strip_bracket(acct_raw), _bracket_code(acct_raw), _inst_only(l.get("거래처")))
        groups[key] = groups.get(key, 0.0) + _amt(l.get("차변"))
    rows = [{"계정과목": k[0], "코드": k[1], "거래처": k[2], "전기이월": 0, "증가": v, "감소": 0}
            for k, v in groups.items() if abs(v) > 0.5]
    rows.sort(key=lambda r: -r["증가"])

    ra = _find_label_row(ws, cci, sec.get("section_anchor", "Recap"))  # leaf 동명 회피 + 재번호 무관
    title = _find_label_row(ws, cci, sec["title"], ra or 1)
    if title is None:
        return {"filled": 0, "error": "이자비용 섹션 미탐지"}
    hr = _find_label_row(ws, cci, sec["header_label"], title + 1)
    ds = hr + 1
    tot0 = _find_label_row(ws, cci, sec["total_label"], ds)
    cur = tot0 - ds
    need = max(len(rows), 1)
    c0, c1 = sec.get("donor_cols", [3, 9])
    _resize_block(ws, ds, cur, need, ds, c0, c1)
    tot = ds + need
    de = ds + len(rows) - 1
    cols = sec["cols"]
    num_cols = set(sec.get("num_cols", []))
    num_fmt = sec.get("num_format", "###,###")
    rowf = sec.get("row_formula", {})
    for i, rec in enumerate(rows):
        r = ds + i
        for col, key in cols.items():
            cell = ws[f"{col}{r}"]
            if col in num_cols:
                cell.value = _amt(rec.get(key))
                cell.number_format = num_fmt
            else:
                cell.value = _clean(rec.get(key))
        for col, tmpl in rowf.items():
            ws[f"{col}{r}"] = tmpl.format(r=r)
    if not rows:
        for c in range(cci, 10):
            ws.cell(ds, c).value = None
    else:
        for col in sec.get("sum_cols", ["I"]):
            ws[f"{col}{tot}"] = f"=SUM({col}{ds}:{col}{de})"
    return {"filled": len(rows), "total_row": tot,
            "이자비용합": sum(r["증가"] for r in rows)}


# ===========================================================================
# 400 주석 8.차입금 (우측 감사인) — 차입처/종류/이자율 (금액은 200!F 연동 자동)
# ===========================================================================

def fill_bbdd400(ws, cfg: dict, loans: list) -> dict:
    """400 주석 8.차입금 우측 감사인 표 — 활성대출별 O차입처/P종류/Q이자율 + R금액 수식 재작성.

    금액 R = ROUND('200'!F{24+i}/1000). extract_light가 시트분리 시 일부 교차참조를 #REF!로
    깨뜨리므로, 기존 수식에 의존하지 않고 활성대출 순서대로 직접 채운다(앵커 기반 data_start).
    """
    sec = cfg["note"]
    active = [l for l in loans if _amt(l.get("대출금액")) > 0]
    acol = column_index_from_string(sec.get("anchor_col", "O"))
    hr = _find_label_row(ws, acol, sec.get("anchor", "차입처"))
    if hr is None:
        return {"filled": 0, "error": "차입처 헤더 미탐지"}
    ds = hr + sec.get("data_offset", 2)
    amount_col = sec.get("amount_col", "R")
    ref = sec.get("ref_sheet", "200_차입금 Test")
    refc, ref0, div = sec.get("ref_sheet_col", "F"), sec.get("ref_start_row", 24), sec.get("amount_div", 1000)
    cols = sec["cols"]
    rate_col = sec.get("rate_col")
    # 데이터 영역(헤더~소계) 전체 클리어 — 템플릿 원래 ROUND/#REF! 잔재 제거
    sub = _find_label_row(ws, acol, sec.get("subtotal", "소계"), ds) or (ds + 9)
    clear_cols = [c for c in (list(cols) + [rate_col, amount_col]) if c]
    for r in range(ds, sub):
        for col in clear_cols:
            ws[f"{col}{r}"] = None
    for i, ln in enumerate(active):
        r = ds + i
        for col, key in cols.items():
            ws[f"{col}{r}"] = _clean(ln.get(key))
        if rate_col:
            c = ws[f"{rate_col}{r}"]
            c.value = _amt(ln.get("연이율")) / 100
            c.number_format = _RATE_FMT
        ws[f"{amount_col}{r}"] = f"=ROUND('{ref}'!{refc}{ref0 + i}/{div},0)"
    return {"filled": len(active)}
