"""A-0 총괄표 생성기 — 본문(1.총괄표) + 수정사항(4.수정사항)만 채운다.

1차 범위: 정산표(별도정산표)로 총괄표 본문을, 수정사항집계로 '4.수정사항' 분개표를 채운다.
나머지 섹션(2.Recap·3.증감분석·Test·서술·결론)은 손대지 않는다(감사인/후속).

본문(대분류별 상세계정 + 합계, 유동/비유동 마커):
  대분류 | 계정명 | 기초 | 기말(회사제시) | 증가(=E-D) | 증감율(=F/D) | 수정사항(=K-L 그대로) | 수정후(=E+H) | 비고
  - 기초/기말/수정사항 = 별도정산표 값(수정사항은 부호 그대로). 증가/증감율/수정후 = 수식.

수정사항: 수정사항집계 분개를 **이 총괄표 계정과 관련된 entry만** 그대로 재현(차변/대변·금액·Effect·설명).

★ 양식 보존(고정 격자): A-0 총괄표는 테두리·병합·서식이 미리 그려진 고정표다. 따라서 값만
  비우는 clear_values를 쓰고(테두리/병합/서식 보존), 도너 스타일 재적용은 하지 않는다.
  (clear_region은 시작행 아래 시트 전체의 병합까지 해제하므로 금지 — Test표·수정사항 헤더 병합이 깨짐.)
  음수는 빨간 괄호 서식을 셀에 직접 지정한다(감사조서 표기 규약).
"""

import re

from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string

from .base import WorkpaperGenerator


def _norm(s) -> str:
    return re.sub(r"\s+", "", str(s)) if s is not None else ""


def _nonzero_group(rows: list[dict]) -> bool:
    """그룹에 의미있는(0/None 아닌) 잔액·수정이 하나라도 있으면 True."""
    for r in rows:
        for k in ("기초", "기말", "수정사항", "수정후"):
            if r.get(k) not in (None, 0):
                return True
    return False


class A0Generator(WorkpaperGenerator):

    def fill(self, tb_rows: list[dict], adj_entries: list[dict], params: dict = None) -> dict:
        """A-0 총괄표(본문+수정사항)를 채운다. 렌더 범위(양식검증·진단용)를 반환."""
        self.open_template()
        cfg = self.config
        body_end = self._fill_body(tb_rows, cfg["body"])
        adj_end, n_adj = self._fill_adjustments(tb_rows, adj_entries, cfg["body"], cfg["adjustments"])
        return {
            "body_data_start": cfg["body"]["data_start_row"],
            "body_data_end": body_end,
            "adj_start": cfg.get("adjustments", {}).get("data_start_row"),
            "adj_end": adj_end,
            "n_adj_entries": n_adj,
        }

    # ---- 본문 ----
    def _fill_body(self, tb_rows, b):
        """회사 불문 본문 렌더. 소유맵(sections의 대분류)에 해당하고 정산표에 존재하는
        대분류만 렌더한다. 완성본 격자 행수보다 많으면 도너 스타일로 행을 삽입해 맞춘다."""
        # 제조원가명세서 행(제조원가=True) 제외 옵션: 판관비(R) 등 손익계산서 본문 총괄표는
        # 제조 인건비(급여/상여금/퇴직급여 제조분)와 대분류가 겹쳐 합산되면 안 된다.
        # 제조분은 매출원가(Q)·인건비 노무비(R-2)에서만 쓴다.
        exclude_mfg = b.get("exclude_mfg", False)
        by_class: dict = {}
        for r in tb_rows:
            if exclude_mfg and r.get("제조원가"):
                continue
            by_class.setdefault(r["대분류"], []).append(r)

        cols = b["columns"]
        detail_f = b.get("detail_formulas", {})
        sub = b.get("subtotal")
        numf, pctf = b["num_format"], b["pct_format"]
        numc, pctc = set(b["num_cols"]), set(b["pct_cols"])
        marker_col = b["marker_col"]
        scope = b.get("subtotal_scope", "group") if sub else "none"
        ds_row = b["data_start_row"]

        def wset(addr, val):
            # 렌더 경로에 남은 병합 셀(insert_rows가 안 옮긴 설명/Test 병합 잔재)에 쓰면
            # MergedCell 쓰기 오류 → 그 병합을 해제하고 쓴다(데이터 영역이라 해제가 맞다).
            if isinstance(self.ws[addr], MergedCell):
                for rng in list(self.ws.merged_cells.ranges):
                    if addr in rng:
                        self.ws.unmerge_cells(str(rng))
                        break
            self.ws[addr] = val

        def setfmt(col, row):
            if col in numc:
                self.ws[f"{col}{row}"].number_format = numf
            elif col in pctc:
                self.ws[f"{col}{row}"].number_format = pctf

        def render_detail(r, rec):
            for key, col in cols.items():
                val = rec.get(key)
                if key == "수정사항" and val in (0, None):
                    val = None
                wset(f"{col}{r}", val)
                setfmt(col, r)
            for col, tmpl in detail_f.items():
                wset(f"{col}{r}", tmpl.format(r=r))
                setfmt(col, r)

        def render_subtotal(r, first, last, label=None):
            wset(f'{sub["label_col"]}{r}', label or sub["label"])
            if sub.get("label_merge"):
                self.safe_merge(sub["label_merge"].format(r=r))
            for col in sub.get("sum_cols", []):
                wset(f"{col}{r}", (f"=SUM({col}{first}:{col}{last})"
                                   if last >= first else 0))
                setfmt(col, r)
            for col, tmpl in sub.get("formulas", {}).items():
                wset(f"{col}{r}", tmpl.format(r=r))
                setfmt(col, r)

        # groups_flag: 명시 목록 대신 정산표의 플래그(예: 판관비) True 대분류를 정산표 순서로
        # 수집(회사불문 — 명칭변형·신규계정 자동 포함). 비면 section["groups"]로 폴백(안전망).
        def _section_groups(section):
            flag = section.get("groups_flag")
            if not flag:
                return section.get("groups", [])
            seen, dyn = set(), []
            for r in tb_rows:
                k = r["대분류"]
                if r.get(flag) and k not in seen:
                    seen.add(k); dyn.append(k)
            return dyn or section.get("groups", [])

        # 렌더할 섹션/그룹(있는 대분류만) + 필요한 행수 계산
        live = []
        needed = 0
        for section in b["sections"]:
            groups = [g for g in _section_groups(section)
                      if by_class.get(g) and _nonzero_group(by_class[g])]
            if not groups:
                continue
            live.append((section, groups))
            if section.get("marker"):
                needed += 1
            for g in groups:
                needed += len(by_class[g])
                if scope == "group":
                    needed += 1
            if scope == "section":
                needed += 1

        # 본문 영역 끝(‘기초 :’ 설명 행) 탐색 → 기존 값 비우고, 부족하면 도너로 행 삽입.
        # 앵커는 조서마다 계정명 열 또는 마커 열에 있으므로 둘 다 스캔.
        name_ci = column_index_from_string(cols["계정명"])
        marker_ci = column_index_from_string(marker_col)
        end_anchor = _norm(b.get("body_end_anchor", "기초"))
        note_row = ds_row
        for rr in range(ds_row, self.ws.max_row + 1):
            if any(_norm(self.ws.cell(rr, ci).value).startswith(end_anchor)
                   for ci in {name_ci, marker_ci}):
                note_row = rr
                break
        self.clear_values(ds_row, note_row - 1)            # 기존 본문 값 비움(격자 보존)
        dd = self.capture_row_style(b["donor_detail_row"])  # 삽입 전 도너 캡처
        dm = self.capture_row_style(b.get("donor_marker_row", b["donor_detail_row"]))
        dsub = self.capture_row_style(b["donor_subtotal_row"]) if sub else dd
        delta = needed - (note_row - ds_row)
        if delta > 0:
            # ★ openpyxl insert_rows의 병합 처리는 버전마다 달라(안 옮기거나 metadata만 옮김) 렌더
            #   경로에 병합 잔재가 남으면 MergedCell 쓰기오류·테두리 소실이 난다. 결정적으로:
            #   삽입 '전에' ds_row 이상 병합을 떼고 → 삽입 → delta만큼 내려 다시 병합(설명/Test
            #   영역은 렌더 구간 바로 아래로 정확히 이동, 렌더 구간은 병합 없는 깨끗한 셀).
            from openpyxl.worksheet.cell_range import CellRange
            saved = [str(m) for m in self.ws.merged_cells.ranges if m.min_row >= ds_row]
            for ref in saved:
                self.ws.unmerge_cells(ref)
            self.ws.insert_rows(ds_row, delta)
            for ref in saved:
                rng = CellRange(ref); rng.shift(0, delta)
                self.ws.merge_cells(str(rng))

        # 렌더 구간 병합 무조건 해제(삽입 여부 무관 — delta<=0면 템플릿이 렌더보다 행이 많아
        #   기존 설명/Test 병합이 렌더 경로에 남는다 → apply_row_style가 MergedCell에 테두리를
        #   못 입혀 'R..테두리 소실' 발생. 데이터 영역이라 해제가 맞다).
        from openpyxl.worksheet.cell_range import CellRange as _CR
        last_render = ds_row + needed - 1
        for ref in [str(m) for m in self.ws.merged_cells.ranges]:
            rng = _CR(ref)
            if rng.min_row <= last_render and rng.max_row >= ds_row:
                self.ws.unmerge_cells(ref)

        # 렌더(도너 스타일 적용 → 삽입 행도 테두리 보존)
        r = ds_row
        for section, groups in live:
            if section.get("marker"):
                self.apply_row_style(r, dm)
                wset(f'{marker_col}{r}', section["marker"])
                r += 1
            sec_first = r
            for g in groups:
                grp_first = r
                for rec in by_class[g]:
                    self.apply_row_style(r, dd)
                    render_detail(r, rec)
                    r += 1
                if scope == "group":
                    self.apply_row_style(r, dsub)
                    render_subtotal(r, grp_first, r - 1)
                    r += 1
            if scope == "section":
                self.apply_row_style(r, dsub)
                render_subtotal(r, sec_first, r - 1, label=section.get("subtotal_label"))
                r += 1
        # 계정명 열 너비 보정(긴 계정명) + 숫자/율 열 최소 너비 보장(긴 금액 ####### 방지)
        self.fit_column_widths([name_ci], data_start=ds_row, data_end=r - 1, header_row=b["header_row"])
        for c in b.get("num_cols", []):
            cur = self.ws.column_dimensions[c].width or 0
            self.ws.column_dimensions[c].width = max(cur, 16)
        for c in b.get("pct_cols", []):
            cur = self.ws.column_dimensions[c].width or 0
            self.ws.column_dimensions[c].width = max(cur, 8)
        return r - 1

    # ---- 수정사항(분개 재현) ----
    def _fill_adjustments(self, tb_rows, adj_entries, b, a):
        # 분개 재현 섹션이 없는 조서(PL 총괄표 등 — 본문 수정사항 열로 충분)는 건너뜀.
        if not a.get("enabled", True):
            return None, 0
        # 이 총괄표에 속한 계정명/대분류 집합(관련 entry 판정용)
        a0_classes = {g for sec in b["sections"] for g in sec["groups"]}
        names = {_norm(r["계정명"]) for r in tb_rows if r["대분류"] in a0_classes}
        names |= {_norm(c) for c in a0_classes}

        def related(entry):
            return any(_norm(l["계정"]) in names for l in entry["lines"])

        entries = [e for e in adj_entries if related(e)]

        # 항상 비운다(다른 회사 완성본 템플릿의 분개 잔재 제거 — 관련 분개가 없는데 템플릿을 그대로
        # 두면 타사 분개가 노출됨). 관련 분개가 없으면 '해당사항 없음'만 남긴다.
        self.clear_values(a["data_start_row"], a["clear_until_row"])
        if not entries:
            self.ws[f'{a["debit_acct_col"]}{a["data_start_row"]}'] = "해당사항 없음"
            return None, 0

        numf, numc = a["num_format"], set(a["num_cols"])

        def setamt(col, row, val):
            self.ws[f"{col}{row}"] = val
            if col in numc:
                self.ws[f"{col}{row}"].number_format = numf

        r = a["data_start_row"]
        for e in entries:
            first_line = True
            for ln in e["lines"]:
                if first_line:
                    self.ws[f'{a["no_col"]}{r}'] = f'#{e["no"]}'
                    first_line = False
                if ln["side"] == "차변":
                    self.ws[f'{a["debit_acct_col"]}{r}'] = ln["계정"]
                    if ln.get("금액") is not None:
                        setamt(a["debit_amt_col"], r, ln["금액"])
                else:
                    self.ws[f'{a["credit_acct_col"]}{r}'] = ln["계정"]
                    if ln.get("금액") is not None:
                        setamt(a["credit_amt_col"], r, ln["금액"])
                if ln.get("손익") is not None:
                    setamt(a["pl_col"], r, ln["손익"])
                if ln.get("이익잉여금") is not None:
                    setamt(a["re_col"], r, ln["이익잉여금"])
                if ln.get("설명"):
                    self.ws[f'{a["desc_col"]}{r}'] = ln["설명"]
                r += 1
            for note in e.get("notes", []):
                self.ws[f'{a["desc_col"]}{r}'] = note
                r += 1
        return r - 1, len(entries)
