# -*- coding: utf-8 -*-
"""기타자산(D200)·기타부채(CC200) 두번째 시트 동적 상세 생성기 — 공용.

2단계(별도 EXE): 1단계 산출물의 총괄표(D100/CC100)에 **실제 존재하는 계정 집합이 드라이버**.
계정마다 소표 블록(거래처별 기초/기말 + 증감 + 합계 + 총괄표 tie-out + Check)을 동적 렌더.
회사마다 기타자산/기타부채 계정이 달라지므로 템플릿 미리그림 블록에 의존하지 않고 총괄표 계정
수만큼 블록을 생성한다(신규 계정 동적 추가, 총괄표에 없는 템플릿 블록은 제거).

수치만 규칙기반(거래처원장=이상적 잔액명세서에서 계정명 매핑). Nature 텍스트·증감분석 서술·Test
샘플링은 감사인 영역(빈 양식 보존). 완성본은 절대 openpyxl로 열어 저장하지 않는다 — D100/CC100은
read-only로 읽고(계정→행 추출), D200/CC200은 extract_light→fill→graft_sheet로 무손실 이식한다.

동적변환 원칙(위반 금지): 행수 가변. insert/delete 직후 shift_formula_rows로 보존영역(Section2)
수식참조 보정, 도너 스타일 복사로 테두리 보존, 합계/tie/check SUM은 실데이터행으로 재작성. tie-out은
절대행이 아니라 총괄표 실제 행(scan 결과)으로 재지정.
"""

import re
from copy import copy

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.cell_range import CellRange

from .base import WorkpaperGenerator, shift_formula_rows
from .bbdd_detail import _norm, _amt, _clean


def _col(letter: str) -> int:
    return column_index_from_string(letter)


def _gen(ws) -> WorkpaperGenerator:
    """기존 ws에 base의 양식보존 도구(capture/apply/clear/stamp/fit)를 그대로 붙여 쓰기 위한 래퍼."""
    g = WorkpaperGenerator.__new__(WorkpaperGenerator)
    g.ws = ws
    g.wb = None
    g.config = {}
    g.template_path = None
    return g


def _sheet_ref(name: str) -> str:
    """수식용 시트 참조 접두사. 공백 있으면 작은따옴표로 감싼다."""
    return f"'{name}'!" if re.search(r"\s", name) else f"{name}!"


def resolve_sheet(path: str, prefix: str) -> "str | None":
    """완성본에서 접두(prefix, 예 'D200'/'D100')로 시작하는 실제 시트명을 찾는다(read-only).

    시트명은 법인 템플릿 버전마다 다르므로(예 'D200_명세서…' vs 'D200_계정별…') 접두로 해석한다.
    """
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        pn = _norm(prefix)
        for name in wb.sheetnames:
            if _norm(name).startswith(pn):
                return name
    finally:
        wb.close()
    return None


# ===========================================================================
# 1) 총괄표(D100/CC100) 계정 → 행·기초·기말 스캔 (read-only)
# ===========================================================================

def scan_lead_accounts(lead_path: str, lead_cfg: dict) -> list:
    """완성본 lead sheet를 read-only로 읽어 [{name, norm, row, base, end, section}] 추출.

    row = tie-out 절대참조의 행번호(회사별 실제 위치). 소계/합계행·본문끝(앵커) 스킵,
    섹션 마커로 유동/비유동 부여. 읽기 전용이라 완성본 무손상.
    """
    wb = openpyxl.load_workbook(lead_path, read_only=True, data_only=True)
    try:
        ws = wb[lead_cfg["sheet"]]
        ds = lead_cfg["data_start_row"]
        nci = _col(lead_cfg["name_col"])
        bci = _col(lead_cfg["base_col"])
        eci = _col(lead_cfg["end_col"])
        markers = lead_cfg.get("section_markers", [])
        marker_norms = {_norm(m): m for m in markers}
        subs = {_norm(x) for x in lead_cfg.get("subtotal_labels", [])}
        end_anchor = _norm(lead_cfg.get("body_end_anchor", ""))
        out = []
        section = None
        for r in range(ds, (ws.max_row or ds) + 1):
            v = ws.cell(r, nci).value
            ns = str(v).strip() if v not in (None, "") else ""
            if not ns:
                continue
            nn = _norm(ns)
            if nn in marker_norms:               # 섹션 마커
                section = marker_norms[nn]
                continue
            if end_anchor and nn.startswith(end_anchor):
                break                            # 본문 끝(예 '기초 :' 설명행)
            if nn in subs:
                continue                         # 소계/합계행
            out.append({
                "name": ns, "norm": nn, "row": r,
                "base": _amt(ws.cell(r, bci).value),
                "end": _amt(ws.cell(r, eci).value),
                "section": section,
            })
        return out
    finally:
        wb.close()


# ===========================================================================
# 2) 거래처원장(이상적 잔액명세서) 계정 매핑
# ===========================================================================

def group_ledger(records: list) -> dict:
    """이상적 원장 records를 정규화 계정과목 키로 묶는다 → {norm계정: [rec, ...]}."""
    by = {}
    for r in records or []:
        by.setdefault(_norm(r.get("계정과목")), []).append(r)
    return by


def match_account(by: dict, acct_norm: str, used: set, alias: dict, fuzzy: bool) -> list:
    """총괄표 계정명(norm)에 대응하는 원장 거래처 행 목록을 찾는다(중복배정 금지).

    ① 정확 일치 → ② alias 명시 → ③ fuzzy(부분일치, 긴 키 우선, 기본 off — 차감계정 오매칭 방지).
    못 찾으면 [](부분출력: 데이터행 없이 tie+check만).
    """
    if acct_norm in by and acct_norm not in used:
        used.add(acct_norm)
        return by[acct_norm]
    for cand in alias.get(acct_norm, []) if alias else []:
        cn = _norm(cand)
        if cn in by and cn not in used:
            used.add(cn)
            return by[cn]
    if fuzzy:
        cands = [k for k in by if k not in used and (k in acct_norm or acct_norm in k)]
        cands.sort(key=lambda k: -len(k))
        if cands:
            used.add(cands[0])
            return by[cands[0]]
    return []


# ===========================================================================
# 3) region 모드 동적 렌더 (D200) — Section1 블록 일괄 재구성 + Section2 밀기
# ===========================================================================

def _resize_region(ws, region_start: int, region_end: int, needed: int) -> None:
    """[region_start, region_end] 영역을 needed 행으로 맞춘다(값 비움 + 병합 정리 + insert/delete + shift).

    영역 내 병합은 제거(렌더가 재생성), 영역 아래(Section2) 병합·수식은 delta만큼 이동.
    """
    # 병합 먼저 해제(영역 + 아래). 아래 것은 delta 이동해 재병합.
    below = []
    for m in list(ws.merged_cells.ranges):
        if m.min_row >= region_start:
            ref = str(m)
            if m.min_row > region_end:
                below.append(ref)
            ws.unmerge_cells(ref)
    _gen(ws).clear_values(region_start, region_end)
    old = region_end - region_start + 1
    delta = needed - old
    if delta > 0:
        ws.insert_rows(region_start, delta)
    elif delta < 0:
        ws.delete_rows(region_start, -delta)
    if delta:
        shift_formula_rows(ws, region_start, delta)
    for ref in below:
        rng = CellRange(ref)
        if delta:
            rng.shift(0, delta)
        try:
            ws.merge_cells(str(rng))
        except Exception:
            pass


def _plan_blocks(accounts: list, by_ledger: dict, cfg: dict):
    """렌더 계획 수립: [('marker', text) | ('block', acct, recs)]. 미매칭 계정명 리스트도 반환."""
    alias = {(_norm(k)): v for k, v in (cfg.get("alias") or {}).items()}
    fuzzy = bool(cfg.get("fuzzy", False))
    used = set()
    plan = []
    unmatched = []
    cur_section = object()
    for a in accounts:
        sec = a.get("section") or ""
        if sec != cur_section:
            plan.append(("marker", sec))
            cur_section = sec
        recs = match_account(by_ledger, a["norm"], used, alias, fuzzy)
        if not recs:
            unmatched.append(a["name"])
        plan.append(("block", a, recs))
    return plan, unmatched


def render_detail(ws, cfg: dict, accounts: list, by_ledger: dict) -> dict:
    """모드 분기 진입점. region=D200(Section1 일괄 재구성), unit=CC200(계정 단위 스탬프)."""
    mode = cfg.get("mode", "region")
    if mode == "unit":
        return render_cc200(ws, cfg, accounts, by_ledger)
    return render_d200(ws, cfg, accounts, by_ledger)


def render_d200(ws, cfg: dict, accounts: list, by_ledger: dict) -> dict:
    """D200 Section1을 총괄표 계정 기반으로 동적 재구성. Returns 진단 dict."""
    g = _gen(ws)
    lead = cfg["lead"]
    b = cfg["block"]
    reg = cfg["region"]
    dn = cfg["donor"]
    lead_sheet = lead["sheet"]
    ref = _sheet_ref(lead_sheet)

    # region_end = Section2 앵커행 - 1
    rs = reg["start_row"]
    aci = _col(reg["anchor_col"])
    anchor = _norm(reg["section2_anchor"])
    region_end = None
    for r in range(rs, ws.max_row + 1):
        if anchor and anchor in _norm(ws.cell(r, aci).value):
            region_end = r - 1
            break
    if region_end is None:
        region_end = ws.max_row

    # 도너 캡처(클리어 전): 스타일 + 헤더는 값까지
    st_marker = g.capture_row_style(dn["marker_row"])
    st_title = g.capture_row_style(dn["title_row"])
    st_blank = g.capture_row_style(dn["blank_row"])
    cap_header = g.capture_row(dn["header_row"])
    st_data = g.capture_row_style(dn["data_row"])
    st_sub = g.capture_row_style(dn["subtotal_row"])
    st_tie = g.capture_row_style(dn["tie_row"])
    st_check = g.capture_row_style(dn["check_row"])

    plan, unmatched = _plan_blocks(accounts, by_ledger, cfg)

    # 필요 행수
    needed = 0
    for item in plan:
        if item[0] == "marker":
            needed += 1
        else:
            needed += 8 + max(len(item[2]), 1)   # 고정 8 + 데이터 max(N,1)

    _resize_region(ws, rs, region_end, needed)

    # 렌더(상단→하단)
    r = rs
    idx = 0
    for item in plan:
        if item[0] == "marker":
            g.apply_row_style(r, st_marker)
            ws.cell(r, _col(b["marker_col"])).value = item[1]
            r += 1
        else:
            _, a, recs = item
            idx += 1
            r = _render_block(g, ws, r, idx, a, recs, b, lead, ref,
                              st_title, st_blank, cap_header, st_data, st_sub, st_tie, st_check)

    g.fit_column_widths([_col(b["cols"]["거래처"])], data_start=rs, data_end=max(rs, r - 1))
    return {"blocks": idx, "accounts": len(accounts), "unmatched": unmatched}


# ===========================================================================
# 4) unit 모드 동적 렌더 (CC200) — 계정마다 3서브섹션 유닛 통째 스탬프
# ===========================================================================

def _content_end(ws, start: int) -> int:
    """start행 이하에서 값이 있는 마지막 행(전 영역 clear/재구성 범위)."""
    last = start
    for r in range(start, (ws.max_row or start) + 1):
        for c in range(1, 13):
            if ws.cell(r, c).value not in (None, ""):
                last = r
                break
    return last


def _text_overrides(cap: dict, donor_acct: str, name: str) -> dict:
    """캡처행 문자열 값에서 도너 계정명 토큰을 대상계정명으로 치환한 overrides 생성."""
    ov = {}
    if not donor_acct or donor_acct == name:
        return ov
    for col, item in cap.items():
        v = item.get("value")
        if isinstance(v, str) and donor_acct in v:
            ov[col] = v.replace(donor_acct, name)
    return ov


def _detect_unit_geometry(ws, u: dict) -> "dict | None":
    """도너 유닛(첫 계정 블록) 기하를 라벨로 자동탐지 — 회계법인마다 CC200 행 위치가 달라
    하드코딩 절대행은 깨진다(도너 캡처가 옆 유닛까지 삼켜 tie행 중복 등). 탐지 실패 시 None."""
    tci = _col(u["title_col"])
    maxr = ws.max_row or 1
    title_re = re.compile(r"^\s*\d+\.\s*\S")

    def is_title(r):
        v = ws.cell(r, tci).value
        return isinstance(v, str) and bool(title_re.match(v))

    us = next((r for r in range(1, maxr + 1) if is_title(r)), None)
    if us is None:
        return None
    nxt = next((r for r in range(us + 1, maxr + 1) if is_title(r)), None)

    def find(label, cols, start, end):
        t = _norm(label)
        for r in range(start, (end or maxr) + 1):
            for c in cols:
                v = ws.cell(r, c).value
                if isinstance(v, str) and t and t in _norm(v):
                    return r
        return None

    end_bound = (nxt - 1) if nxt else maxr
    sub = find("합계", (3, 4), us + 1, end_bound)
    if sub is None:
        return None
    tie = find("총괄표", (3, 4), sub + 1, end_bound)
    chk = find("Check", (3, 4), (tie or sub) + 1, end_bound)
    # Nature 헤더: us~합계 사이, 기초·기말(또는 전기·잔액) 라벨이 있는 행
    hdr = None
    for r in range(us + 1, sub):
        labs = [_norm(ws.cell(r, c).value) for c in range(1, (ws.max_column or 1) + 1)]
        if any(("기초" in x or "전기" in x) for x in labs) and \
           any(("기말" in x or "잔액" in x) for x in labs):
            hdr = r
            break
    if hdr is None:
        hdr = max(us + 1, sub - 2)
    ue = (nxt - 1) if nxt else ((chk or tie or sub) + 2)
    return {
        "unit_start": us, "unit_end": ue, "header_row": hdr,
        "data_donor_row": hdr + 1, "data_placeholder_end": sub - 1,
        "subtotal_donor_row": sub,
        "tie_donor_row": tie or (sub + 2),
        "check_donor_row": chk or ((tie or sub) + 1),
    }


def render_cc200(ws, cfg: dict, accounts: list, by_ledger: dict) -> dict:
    """CC200 — 총괄표 계정마다 도너 유닛(3서브섹션)을 통째 스탬프. Returns 진단 dict."""
    g = _gen(ws)
    lead = cfg["lead"]
    # 유닛 기하는 라벨로 자동탐지(템플릿 버전 무관). 탐지 실패 시 config 값 폴백.
    u = dict(cfg["unit"])
    geo = _detect_unit_geometry(ws, u)
    if geo:
        u.update(geo)
    ref = _sheet_ref(lead["sheet"])
    us, ue = u["unit_start"], u["unit_end"]
    hr, dd = u["header_row"], u["data_donor_row"]
    ph_end = u["data_placeholder_end"]
    donor_acct = u.get("donor_account", "")
    title_ci = _col(u["title_col"])
    ccol, ecol, fcol = _col(u["cols"]["거래처"]), _col(u["cols"]["기초"]), _col(u["cols"]["기말"])
    pre_len = hr - us + 1                 # R8..R13
    post_start = ph_end + 1              # R36
    post_rows = list(range(post_start, ue + 1))   # R36..R65

    # 도너 캡처(클리어 전): 유닛 전 행 값+스타일 + 데이터행 스타일
    caps = {r: g.capture_row(r) for r in range(us, ue + 1)}
    st_data = g.capture_row_style(dd)

    plan, unmatched = _plan_blocks(accounts, by_ledger, cfg)
    units = [(it[1], it[2]) for it in plan if it[0] == "block"]   # marker 무시(CC200은 마커 미렌더)

    def unit_height(n):
        return pre_len + max(n, 1) + len(post_rows)

    region_start = us
    region_end = _content_end(ws, us)
    needed = sum(unit_height(len(recs)) for _, recs in units)
    _resize_region(ws, region_start, region_end, needed)

    r = region_start
    idx = 0
    for a, recs in units:
        idx += 1
        r = _render_unit(g, ws, r, idx, a, recs, caps, st_data, u, lead, ref,
                         donor_acct, us, hr, dd, ph_end, post_rows, pre_len,
                         title_ci, ccol, ecol, fcol)
    g.fit_column_widths([ccol], data_start=region_start, data_end=max(region_start, r - 1))
    return {"blocks": idx, "accounts": len(accounts), "unmatched": unmatched}


def _render_unit(g, ws, top, idx, a, recs, caps, st_data, u, lead, ref, donor_acct,
                 us, hr, dd, ph_end, post_rows, pre_len,
                 title_ci, ccol, ecol, fcol) -> int:
    """한 계정 유닛 스탬프: pre(제목/Nature헤더) + Nature데이터×N + post(합계/tie/check/2)/3))."""
    # ── pre 세그먼트(R us..hr) ──
    for i, dr in enumerate(range(us, hr + 1)):
        cur = top + i
        ov = _text_overrides(caps[dr], donor_acct, a["name"])
        if dr == us:
            ov[title_ci] = f"{idx}. {a['name']}"
        g.stamp_row(cur, caps[dr], overrides=ov)
    # ── Nature 데이터 ──
    data_top = top + pre_len
    n = len(recs)
    for i, rec in enumerate(recs):
        rr = data_top + i
        g.apply_row_style(rr, st_data)
        ws.cell(rr, 1).value = None
        ws.cell(rr, ccol).value = _clean(rec.get("거래처명"))
        ws.cell(rr, ecol).value = _amt(rec.get("전기이월"))
        ws.cell(rr, fcol).value = _amt(rec.get("잔액"))
        for cl, tmpl in u.get("row_formulas", {}).items():
            ws[f"{cl}{rr}"] = tmpl.format(r=rr)
    if n == 0:
        g.apply_row_style(data_top, st_data)
        ws.cell(data_top, 1).value = None
        ws.cell(data_top, ccol).value = None
        for cl in u.get("row_formulas", {}):
            ws[f"{cl}{data_top}"] = None
    de = data_top + max(n, 1) - 1
    # ── post 세그먼트(R post_rows) ──
    post_top = data_top + max(n, 1)
    tot = tie = None
    sampling = set(u.get("sampling_data_rows", []))
    for i, dr in enumerate(post_rows):
        cur = post_top + i
        ov = _text_overrides(caps[dr], donor_acct, a["name"])
        g.stamp_row(cur, caps[dr], overrides=ov)
        if dr == u["subtotal_donor_row"]:
            tot = cur
            for cl in u["sum_cols"]:
                ws[f"{cl}{cur}"] = (f"=SUM({cl}{data_top}:{cl}{de})" if n > 0 else 0)
        elif dr == u["tie_donor_row"]:
            tie = cur
            ws.cell(cur, _col(u["tie_label_col"])).value = u["tie_label"].format(name=a["name"])
            for d2col, field in u["tie_cols"].items():
                lead_col = lead["base_col"] if field == "base" else lead["end_col"]
                ws[f"{d2col}{cur}"] = f"={ref}{lead_col}{a['row']}"
        elif dr == u["check_donor_row"]:
            for cl, tmpl in u["check_formulas"].items():
                ws[f"{cl}{cur}"] = tmpl.format(tot=tot, tie=tie)
        elif dr in sampling:
            for c in range(1, ws.max_column + 1):
                ws.cell(cur, c).value = None      # 샘플링 데이터는 빈 양식 유지
    return post_top + len(post_rows)


def _render_block(g, ws, top, idx, a, recs, b, lead, ref,
                  st_title, st_blank, cap_header, st_data, st_sub, st_tie, st_check) -> int:
    """한 계정 블록 렌더: 제목·blank·헤더·데이터×N·합계·blank·tie·check·blank. 다음 top 반환."""
    cset = b["cols"]
    cc = _col(cset["거래처"])
    ce = _col(cset["기초"])
    cg = _col(cset["기말"])

    # 제목
    g.apply_row_style(top, st_title)
    ws.cell(top, _col(b["title_col"])).value = f"{idx}. {a['name']}"
    r = top + 1
    # blank
    g.apply_row_style(r, st_blank)
    r += 1
    # 헤더(라벨+[배치]+스타일 그대로)
    g.stamp_row(r, cap_header)
    r += 1
    # 데이터
    data_top = r
    n = len(recs)
    for rec in recs:
        g.apply_row_style(r, st_data)
        ws.cell(r, 1).value = None                       # A열 [배치] 마커는 데이터행에 비움
        ws.cell(r, cc).value = _clean(rec.get("거래처명"))
        ws.cell(r, ce).value = _amt(rec.get("전기이월"))
        ws.cell(r, cg).value = _amt(rec.get("잔액"))
        for cl, tmpl in b.get("row_formulas", {}).items():
            ws[f"{cl}{r}"] = tmpl.format(r=r)
        r += 1
    if n == 0:                                            # 부분출력: placeholder 1행(양식만)
        g.apply_row_style(r, st_data)
        ws.cell(r, 1).value = None
        for cl in b.get("row_formulas", {}):
            ws[f"{cl}{r}"] = None
        r += 1
    data_end = data_top + max(n, 1) - 1

    # 합계
    tot = r
    g.apply_row_style(tot, st_sub)
    ws.cell(tot, _col(b["subtotal_label_col"])).value = b["subtotal_label"]
    if b.get("subtotal_merge"):
        g.safe_merge(b["subtotal_merge"].format(r=tot))
    if n > 0:
        for cl in b["sum_cols"]:
            ws[f"{cl}{tot}"] = f"=SUM({cl}{data_top}:{cl}{data_end})"
        for cl, tmpl in b.get("subtotal_formulas", {}).items():
            ws[f"{cl}{tot}"] = tmpl.format(r=tot)
    else:
        for cl in b["sum_cols"]:
            ws[f"{cl}{tot}"] = 0
    r += 1
    # blank
    g.apply_row_style(r, st_blank)
    r += 1
    # tie-out (총괄표 실제 행 참조)
    tie = r
    g.apply_row_style(tie, st_tie)
    ws.cell(tie, _col(b["tie_label_col"])).value = b["tie_label"].format(name=a["name"])
    for d2col, field in b["tie_cols"].items():
        lead_col = lead["base_col"] if field == "base" else lead["end_col"]
        ws[f"{d2col}{tie}"] = f"={ref}{lead_col}{a['row']}"
    r += 1
    # check
    chk = r
    g.apply_row_style(chk, st_check)
    ws.cell(chk, _col(b["check_label_col"])).value = b["check_label"]
    for cl, tmpl in b["check_formulas"].items():
        ws[f"{cl}{chk}"] = tmpl.format(tot=tot, tie=tie)
    r += 1
    # 블록간 blank
    g.apply_row_style(r, st_blank)
    r += 1
    return r
