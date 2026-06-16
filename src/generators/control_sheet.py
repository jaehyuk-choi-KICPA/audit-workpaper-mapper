"""A-1 양식의 'Control Sheet' 시트 채우기.

발송엑셀 Control Sheet의 데이터를 양식의 Control Sheet 시트로 옮긴다.
양식에만 있는 '조회서 번호' 열은 취합엑셀의 조서번호(금융기관명 기준)로 채운다.

- 양식에 존재하는 칸으로만 매핑 (config column_map).
- 발송 대비 양식은 조회서번호가 삽입돼 col4 이후가 한 칸 밀림 → column_map으로 흡수.
- 데이터 행 = 발송 일련번호가 정수인 행 (각주·안내문구 제외).
- base 공용 도구(capture_row_style/clear_region)로 양식(테두리 등) 보존.
"""

import re

import openpyxl

from .base import WorkpaperGenerator


def _cell_has_border(cell) -> bool:
    b = cell.border
    return any(side and side.style for side in (b.left, b.right, b.top, b.bottom))


def _as_int(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v) if v.is_integer() else None
    if isinstance(v, str) and v.strip().isdigit():
        return int(v.strip())
    return None


def _norm_inst(s) -> str:
    """금융기관명 정규화: 괄호 수식어 제거 + 공백 제거 (취합↔발송 매칭 키)."""
    if not s:
        return ""
    s = re.sub(r"\(.*?\)", "", str(s))
    return re.sub(r"\s+", "", s).strip()


_RE_REF = re.compile(r"^\d{0,4}C-?\d+$")


def build_ref_map(confirm_xlsx: str) -> dict:
    """취합엑셀의 '요약' 시트에서 {정규화 금융기관명: (조서번호, 원본 금융기관명)} 맵 생성.

    요약 시트는 전 금융기관의 조서번호↔금융기관명을 한 표로 담은 가장 완전한 근거다.
    조서번호 패턴(예: 3000C-01) 셀을 찾고, 오른쪽 첫 텍스트를 금융기관명으로 본다.
    (요약 시트가 없으면 빈 맵 반환 → 조회서번호는 감사인 영역으로 공란)
    """
    wb = openpyxl.load_workbook(confirm_xlsx, read_only=True, data_only=True)
    sheet = next((s for s in wb.sheetnames if "요약" in s), None)
    ref = {}
    if sheet is None:
        wb.close()
        return ref
    for row in wb[sheet].iter_rows(values_only=True):
        ref_no, name = None, None
        for v in row:
            if v is None:
                continue
            s = str(v).strip()
            if ref_no is None and _RE_REF.match(s):
                ref_no = s
            elif ref_no is not None and name is None and not _RE_REF.match(s):
                name = s
                break
        if ref_no and name:
            k = _norm_inst(name)
            if k and k not in ref:
                ref[k] = (ref_no, name)
    wb.close()
    return ref


def lookup_ref(ref_map: dict, inst) -> tuple:
    """발송 금융기관명으로 조서번호 조회.

    Returns (조서번호 | None, 사유 | None):
      - 정확일치  → (조서번호, None)
      - 부분일치  → (조서번호, '감사인 확인필요 - 기관명 부분일치: 발송 "X" ↔ 취합 "Y"')
      - 미일치    → (None, None)
    """
    n = _norm_inst(inst)
    if not n:
        return (None, None)
    if n in ref_map:
        return (ref_map[n][0], None)
    for k, (ref_no, orig) in ref_map.items():
        if k and (k in n or n in k):
            note = f"감사인 확인필요 - 기관명 부분일치: 발송 '{inst}' ↔ 취합 '{orig}'"
            return (ref_no, note)
    return (None, None)


class ControlSheetGenerator(WorkpaperGenerator):
    """발송 Control Sheet → 양식 Control Sheet 시트로 이전."""

    def fill(self, sang_path: str, ref_map: dict, params: dict = None) -> None:
        self.open_template()
        cfg = self.config
        ws = self.ws
        params = params or {}
        colmap = {int(k): int(v) for k, v in cfg["column_map"].items()}
        s_serial = cfg["serial_col_src"]
        s_inst = cfg["inst_col_src"]
        ref_col = cfg["ref_col_tmpl"]
        note_label_col = cfg.get("note_label_col", 1)
        start = cfg["data_start_row"]
        end = cfg["data_end_row"]

        # 회사명: 제목의 자리표시자(OO)를 params 회사명으로 치환
        company = params.get("회사명")
        if company and cfg.get("title_cell"):
            cell = ws[cfg["title_cell"]]
            token = cfg.get("company_token", "OO")
            if isinstance(cell.value, str) and token in cell.value:
                cell.value = cell.value.replace(token, company, 1)
            elif cell.value:
                cell.value = f"{company} {cell.value}"

        # 발송 Control Sheet 읽기
        wb_s = openpyxl.load_workbook(sang_path, read_only=True, data_only=True)
        sheet = "Control Sheet" if "Control Sheet" in wb_s.sheetnames else wb_s.sheetnames[0]
        sang = list(wb_s[sheet].iter_rows(values_only=True))
        wb_s.close()

        def g(row, col1based):
            i = col1based - 1
            return row[i] if i < len(row) else None

        data_rows = [row for row in sang if _as_int(g(row, s_serial)) is not None]

        # ★ 양식 보존: 고정 격자 템플릿이므로 값만 비우고 격자(테두리)는 보존.
        #   격자는 불연속일 수 있으므로 행별로 테두리 유무를 보고, 없는 행에만 도너 복사.
        self.clear_values(start, end)
        grid_cols = list(colmap.values())
        grid_donor = self.capture_row_style(cfg["grid_donor_row"])

        r = start
        notes = []  # 부분일치 확인필요 사유 (검은 글씨, 표 하단에 기재)
        for row in data_rows:
            if r > end:
                break  # 표 용량 초과 (안전) — 초과분은 감사인 영역
            # 이 행에 격자가 없으면 도너로 부여(있으면 템플릿 격자 보존)
            if not all(_cell_has_border(ws.cell(row=r, column=c)) for c in grid_cols):
                self.apply_row_style(r, grid_donor)
            for scol, tcol in colmap.items():
                ws.cell(row=r, column=tcol).value = g(row, scol)
            inst = g(row, s_inst)
            ref_no, note = lookup_ref(ref_map, inst)
            ws.cell(row=r, column=ref_col).value = ref_no   # 정확·부분 모두 검은 글씨로 매핑
            if note:
                notes.append(note)
            r += 1

        # 부분일치 사유 → 표 하단(마지막 데이터 행 다음 줄부터) 검은 글씨 블록
        if notes:
            r += 1
            for note in notes:
                if r > end:
                    break
                ws.cell(row=r, column=note_label_col).value = f"※ {note}"
                r += 1

        # 헤더 라벨 교정(주소↔우편번호 스왑 반영) — 헤더 행은 clear 대상 아님
        hr = cfg.get("header_row", 3)
        for col, label in (cfg.get("header_labels") or {}).items():
            ws.cell(row=hr, column=int(col)).value = label

        # 컬럼 너비 보정(주소 등 텍스트 열이 내용에 맞게)
        af = cfg.get("autofit")
        if af:
            self.fit_column_widths(af["cols"], start, end,
                                   min_width=af.get("min_width", 8),
                                   max_width=af.get("max_width", 100),
                                   header_row=hr)
