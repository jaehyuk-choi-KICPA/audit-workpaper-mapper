"""P(매출)·Q(매출원가) 총괄표 생성기.

이 두 조서는 회사마다 계정 구성이 크게 달라(매출 종류·상품매출원가 유무) 고정행 리필로는
부족하다. 따라서:
  - Q(매출원가): 고정 격자(미완성공사·당기총공사비용·공사매출원가)를 보존하며 당기총공사비용을
    제조원가명세서 총계로 채우고, **상품매출원가가 있으면 공사매출원가 아래에 행을 삽입**해 매칭.
  - P(매출): IS 매출 표를 **있는 매출 대분류만 동적 렌더**(회사마다 매출 종류가 달라서),
    BS 교차참조 행(공사선수금·매출채권)은 이름으로 찾아 채운다.

제조원가명세서 행(제조원가=True)과 손익 매출원가 행(제조원가=False)은 별도정산표 파서가
구분해 둔다. 회사 무관 동작을 위해 계정명은 키워드 규칙으로 매칭한다.
"""

import re
from collections import defaultdict
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.utils import column_index_from_string

from .base import shift_formula_rows

_NUMF = '#,##0;[Red]\\(#,##0\\);"-"'
_PCTF = '0%;[Red]\\(0%\\)'


def _norm(s) -> str:
    return re.sub(r"\s+", "", str(s)) if s is not None else ""


def _agg_rule(tb_rows, rule):
    """tb_rows에서 규칙(제조 flag + 포함/제외 키워드)에 맞는 행을 합산. (기초/기말/수정사항)"""
    want_mfg = rule.get("제조")
    inc = [_norm(x) for x in rule.get("포함", [])]
    exc = [_norm(x) for x in rule.get("제외", [])]
    agg = {"기초": 0, "기말": 0, "수정사항": 0}
    hit = False
    for rec in tb_rows:
        if want_mfg is not None and bool(rec.get("제조원가")) != bool(want_mfg):
            continue
        text = _norm(rec["대분류"]) + _norm(rec["계정명"])
        if inc and not any(k in text for k in inc):
            continue
        if exc and any(k in text for k in exc):
            continue
        hit = True
        for k in ("기초", "기말", "수정사항"):
            agg[k] += rec[k] or 0
    return agg, hit


def _copy_row_style(ws, src_r, dst_r, ncols):
    """src_r 행의 셀 스타일을 dst_r 행으로 복사(테두리/서식 보존)."""
    for c in range(1, ncols + 1):
        sc = ws.cell(src_r, c)
        dc = ws.cell(dst_r, c)
        if sc.has_style:
            dc.font = copy(sc.font); dc.border = copy(sc.border)
            dc.fill = copy(sc.fill); dc.alignment = copy(sc.alignment)
            dc.number_format = sc.number_format


def _widen(ws, cols, minw=16):
    for col in cols:
        cur = ws.column_dimensions[col].width or 0
        ws.column_dimensions[col].width = max(cur, minw)


# ─────────────────────────────── Q: 매출원가 ───────────────────────────────

def fill_cogs(template, tb_rows, output, cfg):
    """Q 매출원가 총괄표. 공사원가 롤포워드(기초미완성·당기총공사비용·기말미완성)를 제조원가
    명세서 행으로 채워 공사매출원가가 손익 도급공사매출원가와 일치하게 하고, 상품매출원가가
    있으면 공사매출원가 아래 '상품매출원가'+'매출원가계' 행을 삽입한다.

    cfg: {sheet, name_col, base_col, end_col, adj_col, formulas{col:tmpl},
          name_map{라벨: 규칙}, cogs_anchor, commodity_label, commodity_rule, total_label}
    참고: 타계정 전입/전출액은 (에서/으로)가 상쇄되어 공사매출원가에 영향 없으므로 채우지 않는다.
    """
    sheet = cfg["sheet"]
    nci = cfg.get("name_col", 2)
    bcol, ecol, acol = cfg["base_col"], cfg["end_col"], cfg["adj_col"]
    formulas = cfg.get("formulas", {})

    wb = openpyxl.load_workbook(template)
    ws = wb[sheet]

    def put(r, rec):
        ws[f"{bcol}{r}"] = rec["기초"]; ws[f"{bcol}{r}"].number_format = _NUMF
        ws[f"{ecol}{r}"] = rec["기말"]; ws[f"{ecol}{r}"].number_format = _NUMF
        if rec.get("수정사항"):
            ws[f"{acol}{r}"] = rec["수정사항"]; ws[f"{acol}{r}"].number_format = _NUMF
        for col, tmpl in formulas.items():
            ws[f"{col}{r}"] = tmpl.format(r=r)
            ws[f"{col}{r}"].number_format = _PCTF if col == cfg.get("pct_col") else _NUMF

    # 1) 롤포워드 행(기초미완성·당기총공사비용·기말미완성)을 제조원가명세서 행으로 채움
    name_map = {_norm(k): v for k, v in cfg.get("name_map", {}).items()}
    cogs_anchor = _norm(cfg["cogs_anchor"])
    cogs_row = None
    for r in range(1, ws.max_row + 1):
        nm = _norm(ws.cell(r, nci).value)
        if nm in name_map:
            agg, _ = _agg_rule(tb_rows, name_map[nm])
            put(r, agg)
        if nm.startswith(cogs_anchor) and cogs_row is None:
            cogs_row = r

    # 2) 공사매출원가 아래에 매출원가 종류별(상품·제품·용역 등) 집계를 펼친다. 각 종류는 손익 매출원가
    #    섹션의 하위행(원가상세, 상위=종류)을 가진다: 기말재고 하위행이 있으면 롤포워드(기초+투입−기말),
    #    없으면 구성요소 합(용역수수료 등). 종류마다 매출원가 행 아래 검은(medium) 밑줄로 구분. 공사·도급은
    #    템플릿 고정 블록이라 제외. 최종 '계' = 공사 + 각 종류 매출원가.
    ncols = ws.max_column
    total_label = cfg.get("total_label", "계")
    n_extra = 0

    def _vrow(r, label, rec):
        _copy_row_style(ws, cogs_row, r, ncols)
        ws.cell(r, nci).value = label
        put(r, rec)

    def _frow(r, label, fml):
        _copy_row_style(ws, cogs_row, r, ncols)
        ws.cell(r, nci).value = label
        for col in (bcol, ecol, acol):
            ws[f"{col}{r}"] = fml(col); ws[f"{col}{r}"].number_format = _NUMF
        for col, tmpl in formulas.items():
            ws[f"{col}{r}"] = tmpl.format(r=r)
            ws[f"{col}{r}"].number_format = _PCTF if col == cfg.get("pct_col") else _NUMF

    _right = max([nci] + [column_index_from_string(c) for c in (bcol, ecol, acol, *formulas.keys())])

    def _underline(r):                              # 매출원가 종류 구분 검은(medium) 밑줄
        for ci in range(nci, _right + 1):
            cell = ws.cell(r, ci); bd = cell.border
            cell.border = Border(left=bd.left, right=bd.right, top=bd.top, bottom=Side(style="medium"))

    def _is_end(name):                              # 기말재고(차감 항목) 판정
        n = _norm(name)
        return "기말" in n and "재고" in n

    def _subs(parent):                              # 그 종류의 하위행(정산표 순서)
        return [{"name": r["계정명"], "기초": r["기초"] or 0, "기말": r["기말"] or 0,
                 "수정사항": r["수정사항"] or 0}
                for r in tb_rows if r.get("원가상세") and r.get("상위") == parent]

    # 매출원가 종류(계정명 단위, 정산표 순서, 공사/도급 제외 — 공사 롤포워드는 템플릿 블록)
    types, seen = [], set()
    for r in tb_rows:
        if not r.get("매출원가"):
            continue
        nm, cls = r["계정명"], r["대분류"]
        key = _norm(nm)
        if not key or key in seen or any(k in _norm(cls) + key for k in ("공사", "도급")):
            continue
        if not any(r.get(k) for k in ("기초", "기말", "수정사항")):
            continue
        seen.add(key)
        types.append((nm, cls, {"기초": r["기초"] or 0, "기말": r["기말"] or 0, "수정사항": r["수정사항"] or 0}))

    def _type_rows(parent):                         # 그 종류가 차지할 행수
        subs = _subs(parent)
        if not subs:
            return 1
        adds = [s for s in subs if not _is_end(s["name"])]
        ends = [s for s in subs if _is_end(s["name"])]
        return len(adds) + (1 + len(ends) if ends else 0) + 1

    if cogs_row and types:
        n_rows = sum(_type_rows(cls) for _, cls, _ in types) + 1   # 종류들 + 최종 계
        ws.insert_rows(cogs_row + 1, n_rows)
        shift_formula_rows(ws, cogs_row + 1, n_rows)   # 밀린 보존영역 수식 행참조 +n
        _underline(cogs_row)                        # 공사매출원가 아래 밑줄
        r = cogs_row + 1
        sum_rows = [cogs_row]
        for label, cls, agg in types:
            subs = _subs(cls)
            if not subs:                            # 하위행 없음 → 단일 매출원가 행
                _vrow(r, label, agg); sum_rows.append(r); _underline(r); r += 1
                continue
            adds = [s for s in subs if not _is_end(s["name"])]
            ends = [s for s in subs if _is_end(s["name"])]
            a0 = r
            for s in adds:
                _vrow(r, s["name"], s); r += 1
            if ends:
                _frow(r, "계", lambda c, x=a0, n=len(adds): "=" + "+".join(f"{c}{x+i}" for i in range(n)))
                gye = r; r += 1
                for s in ends:
                    _vrow(r, s["name"], s); r += 1
                _frow(r, label, lambda c, g=gye, e0=gye + 1, ne=len(ends):
                      f"={c}{g}-(" + "+".join(f"{c}{e0+i}" for i in range(ne)) + ")")
            else:                                   # 차감 없음 → 구성요소 합
                _frow(r, label, lambda c, x=a0, n=len(adds): "=" + "+".join(f"{c}{x+i}" for i in range(n)))
            sum_rows.append(r); _underline(r); r += 1
        _frow(r, total_label, lambda c, rows=tuple(sum_rows): "=" + "+".join(f"{c}{x}" for x in rows))
        n_extra = n_rows

    _widen(ws, [bcol, ecol, acol] + list(formulas.keys()))
    Path(output).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return {"rows_inserted": n_extra}


# ─────────────────────────────── EE: 퇴직급여 ───────────────────────────────

def fill_retirement(template, tb_rows, output, cfg):
    """EE 퇴직급여 총괄표. 퇴직급여 비용(판관/제조)을 충당부채 위에 행 삽입해 제조 flag로
    채우고, BS 행(충당부채·운용자산)은 정산표에 있으면 채우고 없으면 비운다(타 회사 잔재 제거).

    cfg: {sheet, name_col, base_col(기초), end_col(기말), adj_col(수정), formulas,
          insert_above, total_anchor, expense_rows[{label, rule}], bs_rows[{label, name}]}
    """
    sheet = cfg["sheet"]
    nci = cfg.get("name_col", 2)
    bcol, ecol, acol = cfg["base_col"], cfg["end_col"], cfg["adj_col"]
    formulas = cfg.get("formulas", {})

    by_class = defaultdict(lambda: {"기초": 0, "기말": 0, "수정사항": 0})
    for rec in tb_rows:
        g = by_class[_norm(rec["대분류"])]
        for k in ("기초", "기말", "수정사항"):
            g[k] += rec[k] or 0

    wb = openpyxl.load_workbook(template)
    ws = wb[sheet]
    ncols = ws.max_column

    ins_anchor = _norm(cfg["insert_above"])
    tot_anchor = _norm(cfg["total_anchor"])
    ins_row = tot_row = None
    for r in range(1, ws.max_row + 1):
        nm = _norm(ws.cell(r, nci).value)
        if ins_row is None and nm.startswith(ins_anchor):
            ins_row = r
        if tot_row is None and nm.startswith(tot_anchor):
            tot_row = r

    def _put(r, rec, label):
        ws.cell(r, nci).value = label
        ws[f"{bcol}{r}"] = rec["기초"]; ws[f"{bcol}{r}"].number_format = _NUMF
        ws[f"{ecol}{r}"] = rec["기말"]; ws[f"{ecol}{r}"].number_format = _NUMF
        ws[f"{acol}{r}"] = rec.get("수정사항") or 0; ws[f"{acol}{r}"].number_format = _NUMF
        for col, tmpl in formulas.items():
            ws[f"{col}{r}"] = tmpl.format(r=r); ws[f"{col}{r}"].number_format = _NUMF

    exp = cfg.get("expense_rows", [])
    n_exp = 0
    if ins_row and exp:
        ws.insert_rows(ins_row, len(exp))
        shift_formula_rows(ws, ins_row, len(exp))   # 밀린 보존영역 수식(PL표·체크행) 행참조 +n
        donor = ins_row + len(exp)            # 밀려난 원래 BS 행 = 도너
        for i in range(len(exp)):
            _copy_row_style(ws, donor, ins_row + i, ncols)
        if tot_row and tot_row >= ins_row:
            tot_row += len(exp)
        for i, spec in enumerate(exp):
            agg, _ = _agg_rule(tb_rows, spec["rule"])
            _put(ins_row + i, agg, spec["label"])
        n_exp = len(exp)

    # BS 행: 정산표에 있으면 채우고, 없으면 값을 비운다(타사 잔재 제거)
    data_start = ins_row if ins_row else 1
    n_bs = 0
    for spec in cfg.get("bs_rows", []):
        lbl = _norm(spec["label"])
        rec = by_class.get(_norm(spec["name"]))
        for r in range(data_start, (tot_row or ws.max_row)):
            if not _norm(ws.cell(r, nci).value).startswith(lbl):
                continue
            if rec and any(rec[k] for k in ("기초", "기말", "수정사항")):
                _put(r, rec, ws.cell(r, nci).value)
                n_bs += 1
            else:    # 없는 항목 → 값 비움(라벨·테두리는 보존), 변동분해/체크열까지 클리어
                from openpyxl.utils import get_column_letter
                for ci in range(column_index_from_string(bcol), ncols + 1):
                    cell = ws.cell(r, ci)
                    if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                        cell.value = None
            break

    # Total 행 SUM 범위 재작성(삽입으로 행이 밀렸으므로) + 라벨(Total→계 등)
    if tot_row and data_start < tot_row:
        if cfg.get("total_label"):
            ws.cell(tot_row, nci).value = cfg["total_label"]
        for col in (bcol, ecol, acol):
            ws[f"{col}{tot_row}"] = f"=SUM({col}{data_start}:{col}{tot_row - 1})"
            ws[f"{col}{tot_row}"].number_format = _NUMF
        for col, tmpl in formulas.items():
            ws[f"{col}{tot_row}"] = tmpl.format(r=tot_row); ws[f"{col}{tot_row}"].number_format = _NUMF

    _widen(ws, [bcol, ecol, acol] + list(formulas.keys()))
    Path(output).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return {"expense_rows": n_exp, "bs_filled": n_bs}


# ─────────────────────────────── P: 매출 ───────────────────────────────

def fill_sales(template, tb_rows, output, cfg):
    """P 매출 총괄표. IS 매출 표를 '있는 매출 대분류'만 동적 렌더하고, BS 교차참조 행
    (공사선수금·매출채권)은 이름으로 찾아 채운다.

    cfg: {sheet, is_header_row, is_start_row, is_total_label, name_col,
          base_col, end_col, adj_col, formulas, is_groups[대분류...],
          bs_rows[{label, name}], pct_col?}
    """
    sheet = cfg["sheet"]
    nci = cfg.get("name_col", 2)
    bcol, ecol, acol = cfg["base_col"], cfg["end_col"], cfg["adj_col"]
    formulas = cfg.get("formulas", {})
    is_start = cfg["is_start_row"]
    total_label = _norm(cfg["is_total_label"])

    by_class = defaultdict(lambda: {"기초": 0, "기말": 0, "수정사항": 0, "계정명": None})
    by_name = {}
    for rec in tb_rows:
        if rec.get("제조원가"):           # 제조원가명세서 행은 매출 아님
            continue
        g = by_class[_norm(rec["대분류"])]
        for k in ("기초", "기말", "수정사항"):
            g[k] += rec[k] or 0
        if g["계정명"] is None:
            g["계정명"] = rec["대분류"]
        by_name[_norm(rec["계정명"])] = {"기초": rec["기초"] or 0, "기말": rec["기말"] or 0,
                                        "수정사항": rec["수정사항"] or 0}

    wb = openpyxl.load_workbook(template)
    ws = wb[sheet]
    ncols = ws.max_column

    def setfmt_formulas(r):
        for col, tmpl in formulas.items():
            ws[f"{col}{r}"] = tmpl.format(r=r)
            ws[f"{col}{r}"].number_format = _PCTF if col == cfg.get("pct_col") else _NUMF

    def put(r, rec, name):
        ws.cell(r, nci).value = name
        ws[f"{bcol}{r}"] = rec["기초"]; ws[f"{bcol}{r}"].number_format = _NUMF
        ws[f"{ecol}{r}"] = rec["기말"]; ws[f"{ecol}{r}"].number_format = _NUMF
        ws[f"{acol}{r}"] = rec.get("수정사항") or 0; ws[f"{acol}{r}"].number_format = _NUMF
        setfmt_formulas(r)

    # 1) 현재 IS 합계 행 위치 탐색
    total_row = None
    for r in range(is_start, ws.max_row + 1):
        if _norm(ws.cell(r, nci).value).startswith(total_label):
            total_row = r
            break
    if total_row is None:
        total_row = is_start + 1

    # 2) 렌더할 매출 계정 = 손익 '매출액' 섹션(매출 flag) 전부. 회사마다 계정명이 달라도 섹션
    #    통째로 편입. 일부 회사는 한 대분류(Ⅰ.매출액)에 여러 계정 → 계정명 단위로 렌더.
    flag = cfg.get("owns_flag", "매출")
    present = [{"name": rec["계정명"], "기초": rec["기초"] or 0, "기말": rec["기말"] or 0,
                "수정사항": rec["수정사항"] or 0}
               for rec in tb_rows
               if rec.get(flag) and any(rec.get(k) for k in ("기초", "기말", "수정사항"))]
    if not present:                          # 폴백: 매출 flag 없으면 is_groups(구버전 대비)
        present = [{"name": by_class[_norm(g)]["계정명"] or g,
                    "기초": by_class[_norm(g)]["기초"], "기말": by_class[_norm(g)]["기말"],
                    "수정사항": by_class[_norm(g)]["수정사항"]}
                   for g in cfg.get("is_groups", [])
                   if _norm(g) in by_class and any(by_class[_norm(g)][k] for k in ("기초", "기말", "수정사항"))]
    needed = max(len(present), 1)
    avail = total_row - is_start            # 합계 위 데이터 행 수
    delta = needed - avail
    if delta > 0:
        ws.insert_rows(is_start, delta)     # 데이터 행 삽입(합계·BS 아래로 밀림)
        shift_formula_rows(ws, is_start, delta)   # 밀린 보존영역 수식 행참조 +n
        for i in range(delta):
            _copy_row_style(ws, is_start + delta, is_start + i, ncols)  # 도너=원래 첫 데이터행
        total_row += delta
    # 데이터 행 비우고 렌더
    for r in range(is_start, total_row):
        for col in (bcol, ecol, acol):
            ws[f"{col}{r}"] = None
        ws.cell(r, nci).value = None
    r = is_start
    for rec in present:
        put(r, rec, rec["name"])
        r += 1
    # 3) 합계 행: SUM 범위 재작성(삽입으로 범위가 어긋날 수 있음)
    ws.cell(total_row, nci).value = cfg.get("is_total_label", "합계")
    for col in (bcol, ecol, acol):
        ws[f"{col}{total_row}"] = f"=SUM({col}{is_start}:{col}{total_row - 1})"
        ws[f"{col}{total_row}"].number_format = _NUMF
    setfmt_formulas(total_row)

    # 4) BS 교차참조 행: 이름으로 찾아 채움(공사선수금·매출채권 등)
    n_bs = 0
    for spec in cfg.get("bs_rows", []):
        lbl = _norm(spec["label"])
        nm = _norm(spec["name"])
        rec = by_name.get(nm) or by_class.get(nm)
        if not rec:
            continue
        for rr in range(total_row + 1, ws.max_row + 1):
            if _norm(ws.cell(rr, nci).value).startswith(lbl):
                ws[f"{bcol}{rr}"] = rec["기초"]; ws[f"{bcol}{rr}"].number_format = _NUMF
                ws[f"{ecol}{rr}"] = rec["기말"]; ws[f"{ecol}{rr}"].number_format = _NUMF
                ws[f"{acol}{rr}"] = rec.get("수정사항") or 0; ws[f"{acol}{rr}"].number_format = _NUMF
                setfmt_formulas(rr)
                n_bs += 1
                break

    _widen(ws, [bcol, ecol, acol] + list(formulas.keys()))
    Path(output).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return {"n_sales": len(present), "n_bs": n_bs}
