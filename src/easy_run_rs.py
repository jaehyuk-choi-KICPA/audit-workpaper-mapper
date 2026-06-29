# -*- coding: utf-8 -*-
"""R(판매비와관리비)·S(기타손익) 조서 생성 — 쉬운 실행기 (비전문가용 대화형).

**정산표 조서생성의 후속 단계**(자기완결). 정산표로 R·S 총괄표를 직접 만들고(조서생성의 R·S 항목만
결합 — 단독 실행 가능), 그 총괄표에 **연동**해 보조시트를 동적 구성한다:
  · R110 증감분석 : 1)전기대비증감 표를 총괄표 판관비 계정과 동기화 + ABOVE 비고 X마커·증감분석 블록.
  · R 주석검토    : "15.부가가치 계산" 고정계정을 총괄표에 =ROUND(!/1000,0) 연동(천원).
  · S200 증감분석 : 섹션형 표를 총괄표와 동기화 + F열 수식·Threshold·결론 블록.

Threshold = 별도정산표의 PM(수행중요성) × 0.5.

사용:
  1) `입력자료/`에 정산표(필수)를 넣는다.
  2) `양식자료/`에 R·S 템플릿(R_4000_template_판매비와관리비.xlsx, S_4000_template_기타손익.xlsx)을 둔다.
  3) `R-S생성.bat` 더블클릭 (또는 python src/easy_run_rs.py)
  4) 회사명·기준일·작성자·검토자 입력 → `출력조서/{회사}/`에 R·S 생성(완성본 보존 graft).
"""
import glob
import io
import os
import re
import sys
import warnings
import zipfile
from pathlib import Path

warnings.filterwarnings("ignore", message=".*[Hh]eader or footer.*")
sys.path.insert(0, str(Path(__file__).resolve().parent))
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stdin.reconfigure(encoding="utf-8")
except Exception:
    pass

import openpyxl
from pipeline import build_lead_all
from generators.sheet_surgery import graft_sheet
from generators import rs_detail

# R·S 총괄표 생성용 레지스트리(조서생성의 R·S 항목만 따옴 — 단독 실행 가능하게 결합)
_R = {"code": "R", "config": "r.yaml", "template": "R_4000_template_판매비와관리비.xlsx"}
_S = {"code": "S", "config": "s.yaml", "template": "S_4000_template_기타손익.xlsx"}
R_LEAD = "R_100총괄표"
R_SECOND = "R110_계정별 Nature 파악 및 증감분석"
R_NOTES = "주석검토"
S_LEAD = "S100_총괄표"


from workspace import select_workspace, template_dir, config_dir, prompt_company_info

# 양식·config = 회사 무관 공유 자산. 입력·변환·출력은 회사별 워크스페이스에서 받는다.
TEMPLATE_DIR = template_dir()
CONFIG_DIR = config_dir()
INPUT_DIR = PARSED_DIR = OUTPUT_DIR = None  # main()에서 워크스페이스로 설정


def _line(m=""):
    print(m, flush=True)


def _ask(label, required=True, example=""):
    hint = f" (예: {example})" if example else ""
    while True:
        v = input(f"  {label}{hint}: ").strip()
        if v or not required:
            return v
        _line("  ▷ 값을 입력해 주세요.")


def _find(pattern):
    c = [h for h in glob.glob(str(INPUT_DIR / "**" / pattern), recursive=True) if "~$" not in h]
    return sorted(c)[0] if c else None


def _extract_one(src, sheet, out):
    """완성본에서 단일 시트만 가벼운 임시본으로 추출(도형·외부링크 제거)."""
    buf = io.BytesIO()
    with zipfile.ZipFile(src) as zin, zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zo:
        for n in zin.namelist():
            if n.startswith(("xl/drawings/", "xl/media/")):
                continue
            b = zin.read(n)
            if n.startswith("xl/worksheets/") and n.endswith(".xml"):
                b = re.sub(rb"<drawing\b[^>]*/>", b"", b); b = re.sub(rb"<legacyDrawing\b[^>]*/>", b"", b)
            elif n.endswith(".rels"):
                b = re.sub(rb'<Relationship[^>]*(?:drawing|image|vmlDrawing)"[^>]*/>', b"", b)
            elif n == "[Content_Types].xml":
                b = re.sub(rb"<Override[^>]*(?:drawing|vmlDrawing)[^>]*/>", b"", b)
            zo.writestr(n, b)
    buf.seek(0)
    wb = openpyxl.load_workbook(buf)
    for s in list(wb.sheetnames):
        if s != sheet:
            del wb[s]
    wb[sheet].sheet_state = "visible"
    wb.save(out)


def _edit_sheet(path, sheet, fn, tmp):
    """extract_one → fn(ws) → graft_sheet(완성본 보존)로 한 시트 적용."""
    light = str(Path(tmp) / "light.xlsx")
    _extract_one(path, sheet, light)
    wb = openpyxl.load_workbook(light)
    fn(wb[sheet])
    wb.save(light)
    out = path + ".graft"
    graft_sheet(path, light, sheet, out)
    os.replace(out, path)


def _resolve_sheet(path, prefix):
    """워크북에서 prefix로 시작하는 시트명 반환(비ASCII 방어)."""
    z = zipfile.ZipFile(path)
    wbx = z.read("xl/workbook.xml").decode("utf-8", "ignore"); z.close()
    for m in re.finditer(r'<sheet[^>]*name="([^"]+)"', wbx):
        nm = m.group(1)
        if nm.startswith(prefix):
            return nm.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">")
    return None


def _out_path(out_dir, item):
    return str(out_dir / item["template"].replace("_template_", "_"))


def main():
    _line("=" * 56)
    _line("  R(판관비)·S(기타손익) 조서 생성 — 정산표 조서생성 후속")
    _line("=" * 56)

    global INPUT_DIR, PARSED_DIR, OUTPUT_DIR
    ws = select_workspace(line=_line)
    if ws is None:
        return 2
    INPUT_DIR, PARSED_DIR, OUTPUT_DIR = ws.input_dir, ws.parsed_dir, ws.output_dir
    TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
    _line(f"\n[회사] {ws.company}   (작업폴더: {ws.input_dir.parent})")

    settlement = _find("*정산표*.xls*")
    if not settlement:
        _line(f"\n[안내] 정산표를 찾지 못했습니다(파일명에 '정산표' 포함). 입력자료에 넣어 주세요:\n  {INPUT_DIR}")
        return 2
    present = [it for it in (_R, _S) if (TEMPLATE_DIR / it["template"]).exists()]
    missing = [it["code"] for it in (_R, _S) if not (TEMPLATE_DIR / it["template"]).exists()]
    _line(f"\n[정산표] {Path(settlement).name}")
    _line(f"[양식]   {TEMPLATE_DIR}  (확인 {len(present)}/2)")
    if missing:
        _line(f"         미발견: {', '.join(missing)} — 해당 조서 제외")
    if not present:
        _line("         R·S 템플릿이 없습니다. 양식자료에 완성본을 넣어 주세요.")
        return 2

    _line("\n[회사 정보 입력]")
    company = ws.company
    date, preparer, reviewer = prompt_company_info(ws, input, _line)

    out_dir = ws.output_dir
    # 있으면 그대로 사용(조서생성 산출물 재사용), 없으면 정산표로 직접 생성.
    if all(os.path.exists(_out_path(out_dir, it)) for it in present):
        _line("\n[총괄표] 기존 산출물 재사용 (R·S)")
        _line("        (정산표가 바뀌었으면 '조서생성'을 다시 돌리거나 해당 파일을 지우고 실행)")
    else:
        _line("\n총괄표가 없어 정산표로 생성 중...\n")
        try:
            build_lead_all(
                settlement=settlement, registry=present, config_dir=str(CONFIG_DIR),
                template_root=str(TEMPLATE_DIR), output_dir=str(out_dir),
                parsed_dir=str(ws.parsed_dir),
                params={"회사명": company, "날짜": date, "preparer": preparer, "reviewer": reviewer},
            )
        except PermissionError:
            _line(f"\n[오류] 출력 파일이 열려 있습니다. 닫고 다시 실행해 주세요:\n  {out_dir}")
            return 1
        except Exception as e:
            _line(f"\n[오류] 총괄표 생성 실패: {type(e).__name__}: {e}")
            return 1

    pm = rs_detail.read_pm(settlement)
    amounts = rs_detail.lead_amounts(settlement)
    threshold = round(pm * 0.5) if pm else None

    tmp = str(out_dir / "_tmp_rs")
    Path(tmp).mkdir(parents=True, exist_ok=True)
    notes = []
    notes.append("PM(수행중요성): %s → Threshold(×0.5): %s"
                 % (format(int(pm), ",") if pm else "자동탐지 실패(수동입력 필요)",
                    format(int(threshold), ",") if threshold else "-"))

    # R: R110 + 주석검토
    r_path = _out_path(out_dir, _R)
    if _R in present and os.path.exists(r_path):
        try:
            rwb = openpyxl.load_workbook(r_path, data_only=False)
            lead = rwb[R_LEAD]
            # graft 경유 편집(완성본 보존). lead_ws는 생성 파일에서 읽어 전달.
            r110_stat = {}
            def _do_r110(ws):
                r110_stat.update(rs_detail.fill_r110(ws, lead, R_LEAD, threshold, amounts))
            _edit_sheet(r_path, R_SECOND, _do_r110, tmp)
            notes.append("R110 증감분석: 판관비 %d계정 동기화, ABOVE %d건(X3~)"
                         % (r110_stat.get("accounts", 0), r110_stat.get("above", 0)))
            notes_stat = {}
            def _do_notes(ws):
                notes_stat.update(rs_detail.fill_r_notes(ws, lead, R_LEAD))
            _edit_sheet(r_path, R_NOTES, _do_notes, tmp)
            notes.append("R 주석검토: 부가가치 계정 %d/%d 연동(천원)"
                         % (notes_stat.get("matched", 0), notes_stat.get("total", 0)))
        except Exception as e:
            notes.append(f"R 보조시트 일부 실패: {type(e).__name__}: {e}")

    # S: S200
    s_path = _out_path(out_dir, _S)
    if _S in present and os.path.exists(s_path):
        try:
            swb = openpyxl.load_workbook(s_path, data_only=False)
            lead = swb[S_LEAD]
            s_second = _resolve_sheet(s_path, "S200") or "S200_영업외손익 Test"
            s_stat = {}
            def _do_s200(ws):
                s_stat.update(rs_detail.fill_s200(ws, lead, S_LEAD, threshold, amounts))
            _edit_sheet(s_path, s_second, _do_s200, tmp)
            ab = s_stat.get("above", 0)
            notes.append("S200 증감분석: %d섹션 동기화, ABOVE %d건(%s)"
                         % (s_stat.get("sections", 0), ab,
                            "빨강 X1~" if ab else "특이사항 없음"))
        except Exception as e:
            notes.append(f"S 보조시트 실패: {type(e).__name__}: {e}")

    try:
        import shutil
        shutil.rmtree(tmp, ignore_errors=True)
    except Exception:
        pass

    lines = ["=" * 60, "  R·S 생성 완료", "=" * 60, f"  결과 폴더: {out_dir}"]
    for n in notes:
        lines.append("   - " + n)
    lines.append("  ※ 초안입니다. 증감분석 서술·결론·Threshold는 감사인이 검토·보완하세요.")
    summary = "\n".join(lines)
    _line("\n" + summary)
    try:
        (out_dir / f"{out_dir.name}_RS_실행리포트.txt").write_text(summary, encoding="utf-8")
    except Exception:
        pass
    return 0


if __name__ == "__main__":
    rc = main()
    try:
        input("\n[Enter] 키를 누르면 창이 닫힙니다...")
    except EOFError:
        pass
    sys.exit(rc)
