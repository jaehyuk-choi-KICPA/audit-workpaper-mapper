"""조서 양식 무결성 검증기.

엑셀 수식은 값만 계산하므로 시각적 양식 붕괴(테두리 소실·소계 어긋남·행 밀림)를
잡지 못한다. 이 모듈은 openpyxl로 **구조 불변식**을 검사해 맵핑 후 양식이
무너졌는지 게이트한다. (git 훅=데이터 유출 차단, 이 검증기=양식 무결성)

issue 리스트를 반환한다 (빈 리스트 = 통과).
"""

import re

import openpyxl


def _has_border(cell) -> bool:
    b = cell.border
    return any(side and side.style for side in (b.left, b.right, b.top, b.bottom))


def _disp_len(v) -> int:
    """표시 폭(한글 등 전각 2, 그 외 1)."""
    return sum(2 if ord(ch) > 0x1100 else 1 for ch in str(v)) if v is not None else 0


def check_column_fit(path: str, sheet_name: str, *, cols, data_start: int, data_end: int,
                     header_row: "int | None" = None, tolerance: int = 1) -> list[str]:
    """열 너비가 내용보다 좁아 잘리는지 검사(시각 확인 대체). 내용>너비면 issue."""
    from openpyxl.utils import get_column_letter
    issues = []
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name]
    for c in cols:
        rows = list(range(data_start, data_end + 1)) + ([header_row] if header_row else [])
        maxlen = max((_disp_len(ws.cell(row=r, column=c).value) for r in rows), default=0)
        if maxlen == 0:
            continue
        width = ws.column_dimensions[get_column_letter(c)].width or 8.43
        if width + tolerance < maxlen:
            issues.append(f"[너비] {sheet_name} {get_column_letter(c)}열: 너비 {width:.0f} < 내용 {maxlen} (잘림 의심)")
    wb.close()
    return issues


def check_cells_present(path: str, sheet_name: str, cells: "list[str]") -> list[str]:
    """제목·헤더 등 필수 셀이 비어있지 않은지 검사(제목 날아감 등 탐지)."""
    issues = []
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name]
    for addr in cells:
        v = ws[addr].value
        if v is None or str(v).strip() == "":
            issues.append(f"[필수셀] {sheet_name} {addr} 비어있음(제목·헤더 누락 의심)")
    wb.close()
    return issues


def validate_grid(path: str, sheet_name: str, *, header_row: int, data_start: int,
                  n_rows: int, cols: "list[int]", require_text: bool = True) -> list[str]:
    """채워진 데이터 행의 격자(테두리) 무결성을 검사한다 (시트 무관, 재사용).

    데이터 행에 테두리가 없으면(=격자 붕괴) issue로 보고. 모든 생성기가 데이터 채우기
    후 이 게이트를 거치도록 한다. (A200 검사의 _has_border 로직을 일반화)

    Args:
        header_row: 헤더 행(있으면 테두리 기준 행으로도 참고)
        data_start: 데이터 시작 행
        n_rows:     검사할 데이터 행 수
        cols:       격자가 있어야 할 열들(1-indexed)
        require_text: True면 비어있지 않은 행만 검사(완전 빈 행은 건너뜀)
    """
    issues: list[str] = []
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name]
    for i in range(n_rows):
        r = data_start + i
        row_has_value = any(ws.cell(row=r, column=c).value not in (None, "") for c in cols)
        if require_text and not row_has_value:
            continue
        missing = [c for c in cols if not _has_border(ws.cell(row=r, column=c))]
        if missing:
            issues.append(f"[격자] {sheet_name} R{r}: 테두리 소실 열 {missing} (양식 붕괴 의심)")
    wb.close()
    return issues


def check_outside_col_fill(path: str, sheet_name: str, *,
                           outside_gray_cols: list,
                           sample_rows: "list[int] | None" = None) -> list[str]:
    """컨텐츠 바깥 여백 열의 회색 음영이 보존됐는지 검사.

    양식 원본에서 의도적으로 회색인 여백 열(예: A100의 A·M, A200의 A·X)이
    데이터 채우기 후에도 유지되는지 확인한다. 열 단위(column_dimensions.fill)
    또는 셀 단위 채움 중 하나라도 있으면 통과.
    """
    if not outside_gray_cols:
        return []

    from openpyxl.utils import column_index_from_string
    issues = []
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name]
    rows = sample_rows or [5, 10, 15]

    for col_letter in outside_gray_cols:
        col_idx = column_index_from_string(col_letter)
        # 열 단위 채움 확인
        cd = ws.column_dimensions[col_letter]
        cf = cd.fill
        has_fill = (cf is not None and
                    getattr(cf, "patternType", None) not in (None, "none", ""))
        # 셀 단위 채움 샘플 (열 단위가 없을 때)
        if not has_fill:
            for r in rows:
                c = ws.cell(row=r, column=col_idx)
                sf = c.fill
                if sf is not None and getattr(sf, "patternType", None) not in (None, "none", ""):
                    has_fill = True
                    break
        if not has_fill:
            issues.append(f"[바깥열음영] {sheet_name} {col_letter}열: "
                          f"회색 음영 소실 — 양식 여백 디자인 훼손(템플릿 확인)")
    wb.close()
    return issues


def check_header_fill(path: str, template_path: str, sheet_name: str, *,
                      header_row: int, cols) -> list[str]:
    """생성본 헤더 행의 채움색(fill)이 **템플릿과 동일**한지 검사한다.

    데이터 채우기 중 헤더 색이 (1) 소실되거나 (2) 다른 색(예: 파란색)으로 변질되는
    반복 사고를 프로그래밍으로 잡는 게이트. 헤더는 절대 새 색을 입히지 말고 템플릿 색을
    그대로 보존해야 한다(clear_region 후엔 capture_row/stamp_row로 원래 fill째 재현).
    """
    from openpyxl.utils import get_column_letter

    def _fkey(c):
        f = c.fill
        if f is None or getattr(f, "patternType", None) in (None, "none", ""):
            return None
        fg = getattr(f, "fgColor", None)
        return (f.patternType, getattr(fg, "rgb", None), getattr(fg, "theme", None),
                getattr(fg, "indexed", None), getattr(fg, "tint", None))

    wt = openpyxl.load_workbook(template_path)
    wo = openpyxl.load_workbook(path)
    if sheet_name not in wt.sheetnames or sheet_name not in wo.sheetnames:
        wt.close(); wo.close()
        return []
    tws, ows = wt[sheet_name], wo[sheet_name]
    issues = []
    for ci in cols:
        tk = _fkey(tws.cell(header_row, ci))
        ok = _fkey(ows.cell(header_row, ci))
        if tk != ok:
            col = get_column_letter(ci)
            if tk and not ok:
                issues.append(f"[헤더색] {sheet_name} {col}{header_row}: 헤더 채움색 소실(템플릿엔 있음)")
            else:
                issues.append(f"[헤더색] {sheet_name} {col}{header_row}: "
                              f"헤더 채움색 변질(템플릿과 다름 — 파란색 등 변경 주의)")
    wt.close(); wo.close()
    return issues


def validate_a200_form(path: str, sheet_name: str, config: dict) -> list[str]:
    """생성된 A-200 시트의 양식 무결성을 검사한다.

    검사 항목:
      1. 소계 행(합계) 3개 존재 + B:F 병합
      2. 각 소계 SUM 범위가 바로 위 데이터 블록과 정확히 일치
      3. STEP2 모집단 = 소계 H 합 수식
      4. 데이터 행 테두리 보존(스타일 소실 = 양식 붕괴 탐지)
      5. 표 범위 밖(>회신/결론 열) 값 누출 없음
      6. 데이터 행 차이금액 수식 존재
      7. (*n) 각주 마커 수 = 각주 문구 수
    """
    issues: list[str] = []
    s5 = config["step5"]
    cols = s5["columns"]
    label_col = cols["No."]
    hcol = s5["amount_company_col"]
    ocol = s5["amount_confirm_col"]
    diff_c = s5["diff_col"]
    start = s5["data_start_row"]
    last_col = openpyxl.utils.column_index_from_string(config["step5"]["conclusion_col"])
    # 외화값 열(있으면)은 표의 정당한 확장 → 누출 검사 범위에 포함
    fx_value_col = config["step5"].get("fx_value_col")
    if fx_value_col:
        last_col = max(last_col, openpyxl.utils.column_index_from_string(fx_value_col))

    wb = openpyxl.load_workbook(path)  # 수식 텍스트 검사 → data_only=False
    ws = wb[sheet_name]

    def cell(col_letter, row):
        return ws[f"{col_letter}{row}"]

    # --- 소계 행 탐색 ---
    subtotal_rows = []
    for r in range(start, ws.max_row + 1):
        v = cell(label_col, r).value
        if isinstance(v, str) and v.endswith("합계"):
            subtotal_rows.append(r)
    if len(subtotal_rows) != 3:
        issues.append(f"[소계] 합계 행 3개 기대, {len(subtotal_rows)}개 발견: {subtotal_rows}")

    # --- 소계 병합 + SUM 범위 검사 ---
    merged = {str(m) for m in ws.merged_cells.ranges}
    block_first = start
    for sub in subtotal_rows:
        expect_merge = s5["subtotal_label_merge"].format(r=sub)
        if expect_merge not in merged:
            issues.append(f"[소계] 라벨 병합 누락 R{sub}: {expect_merge}")
        block_last = sub - 1
        hf = cell(hcol, sub).value
        if isinstance(hf, str) and hf.startswith("="):
            if block_last >= block_first:
                want = f"=SUM({hcol}{block_first}:{hcol}{block_last})"
                if hf.replace(" ", "") != want.replace(" ", ""):
                    issues.append(f"[소계] R{sub} 합계 범위 불일치: {hf} (기대 {want})")
            # 블록 내 데이터 행 차이금액 수식 + 테두리 검사
            for dr in range(block_first, block_last + 1):
                if not str(cell(diff_c, dr).value or "").startswith("="):
                    issues.append(f"[차이금액] R{dr} 수식 누락: {cell(diff_c, dr).value!r}")
                if not _has_border(cell(label_col, dr)):
                    issues.append(f"[양식] R{dr} 데이터 행 테두리 소실(양식 붕괴 의심)")
        block_first = sub + 1

    # --- STEP2 모집단 수식 ---
    pop_cell = config["step2"].get("population_cell")
    if pop_cell and subtotal_rows:
        pv = ws[pop_cell].value
        want = "=" + "+".join(f"{hcol}{s}" for s in subtotal_rows)
        if str(pv).replace(" ", "") != want.replace(" ", ""):
            issues.append(f"[STEP2] 모집단 수식 불일치: {pv} (기대 {want})")

    # --- 표 범위 밖 값 누출 ---
    for r in range(start, (subtotal_rows[-1] if subtotal_rows else start) + 1):
        for c in range(last_col + 1, ws.max_column + 1):
            if ws.cell(row=r, column=c).value not in (None, ""):
                issues.append(f"[누출] R{r}C{c} 표 범위 밖 값: {ws.cell(row=r, column=c).value!r}")

    # --- 각주 마커 수 = 문구 수 ---
    conc_c = s5["conclusion_col"]
    markers = set()
    for r in range(start, (subtotal_rows[-1] if subtotal_rows else start) + 1):
        v = cell(conc_c, r).value
        if isinstance(v, str):
            markers.update(re.findall(r"\(\*\d+\)", v))
    note_markers = set()
    for r in range((subtotal_rows[-1] if subtotal_rows else start) + 1, ws.max_row + 1):
        v = cell(label_col, r).value
        if isinstance(v, str):
            note_markers.update(re.findall(r"\(\*\d+\)", v))
    if markers != note_markers:
        issues.append(f"[각주] 마커({sorted(markers)}) ≠ 문구({sorted(note_markers)})")

    wb.close()
    return issues


def validate_detail_blocks(path: str, sheet_name: str, *, expected_blocks: int,
                           lead_sheet: str, tie_label: str = "총괄표",
                           check_label: str = "Check") -> list:
    """D200/CC200 상세 시트 양식 무결성 검사(생성 직후 게이트).

    검사: tie-out 행 수 = 블록 수 = 계정 수, 각 tie-out이 lead_sheet 참조, Check 행 동반,
    합계 SUM이 데이터 범위 참조, 리터럴 오류(#REF!/#NAME?/#VALUE!) 텍스트 없음.
    """
    issues = []
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name]
    tie_norm = re.sub(r"\s+", "", tie_label)
    chk_norm = re.sub(r"\s+", "", check_label)
    lead_ref = lead_sheet if not re.search(r"\s", lead_sheet) else f"'{lead_sheet}'"

    tie_rows, chk_rows, sum_rows = [], [], []
    err_cells = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, min(ws.max_column, 12) + 1):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            vn = re.sub(r"\s+", "", v)
            if tie_norm and vn.startswith(tie_norm) and "→" in v:
                tie_rows.append(r)
            elif vn == chk_norm:
                chk_rows.append(r)
            elif v.startswith("합계") or vn == "합계":
                sum_rows.append(r)
            if any(e in v for e in ("#REF!", "#NAME?", "#VALUE!", "#DIV/0!")):
                err_cells.append(f"{ws.cell(r, c).coordinate}={v}")

    n_tie, n_chk = len(set(tie_rows)), len(set(chk_rows))   # 한 행에 라벨이 여러 열 → 행 단위 dedup
    if n_tie != expected_blocks:
        issues.append(f"[블록수] {sheet_name} tie-out {n_tie}개 ≠ 계정 {expected_blocks}개")
    if n_chk < expected_blocks:
        issues.append(f"[Check] {sheet_name} Check행 {n_chk}개 < 계정 {expected_blocks}개")

    # tie-out 행이 lead_sheet를 참조하는지
    for r in set(tie_rows):
        refs = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)
                if isinstance(ws.cell(r, c).value, str) and ws.cell(r, c).value.startswith("=")]
        if not any(lead_ref + "!" in f for f in refs):
            issues.append(f"[tie] {sheet_name} R{r} 총괄표({lead_sheet}) 참조 누락")

    # 합계 SUM 형식 확인(있으면)
    for r in set(sum_rows):
        has_sum = any(isinstance(ws.cell(r, c).value, str) and "SUM(" in (ws.cell(r, c).value or "")
                      for c in range(1, ws.max_column + 1))
        zero_only = all((ws.cell(r, c).value in (None, "", 0))
                        for c in range(5, 8))   # 수치열(E/F/G)만 — 빈 계정 합계=0은 정상
        if not has_sum and not zero_only:
            issues.append(f"[합계] {sheet_name} R{r} SUM 수식 없음")

    if err_cells:
        issues.append(f"[오류값] {sheet_name} 리터럴 오류 {len(err_cells)}건: {err_cells[:4]}")

    wb.close()
    return issues
