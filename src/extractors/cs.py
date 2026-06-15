"""조회서 발송 Control Sheet 파서.

발송엑셀의 "Control Sheet"에서 구분·일련번호·금융기관명을 추출해
A100 완전성 체크 테이블의 입력 데이터로 사용한다.

발송 Control Sheet 구조 (표준 양식, 실파일 검증):
  - 헤더 행: '일련번호' 컬럼을 포함하는 행
  - 구분:     일련번호 바로 왼쪽 열
              (헤더 라벨은 '금융기관명'이지만 실제 값은 은행/보험/보증기관)
  - 일련번호: 순번 (1, 2, 3, ...) — 조회서 REF(3000C-..)가 아님
  - 금융기관명: 일련번호 오른쪽의 첫 '금융기관명' 헤더 열 (법인명 다음)
  - 데이터: 헤더 다음 행부터, 일련번호 없는 행은 skip
"""

from pathlib import Path

import openpyxl

_DEFAULT_SHEET = "Control Sheet"

_SEQ_HEADER = "일련번호"
_NAME_HEADER = "금융기관명"


def _as_int(value) -> int | None:
    """정수면 int, 정수형 실수(1.0)면 int, 그 외는 None."""
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else None
    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip())
    return None


def _locate_columns(rows: list[tuple]) -> tuple[int, int, int, int]:
    """헤더 행과 구분·일련번호·금융기관명 열 인덱스를 찾는다.

    Returns:
        (header_idx, div_idx, seq_idx, name_idx)

    Raises:
        ValueError: 일련번호/금융기관명 헤더를 찾지 못한 경우
    """
    for header_idx, row in enumerate(rows):
        seq_idx = None
        for c, cell in enumerate(row):
            if cell is not None and str(cell).strip() == _SEQ_HEADER:
                seq_idx = c
                break
        if seq_idx is None:
            continue

        # 금융기관명: 일련번호 오른쪽의 첫 '금융기관명' 헤더
        name_idx = None
        for c in range(seq_idx + 1, len(row)):
            if row[c] is not None and str(row[c]).strip() == _NAME_HEADER:
                name_idx = c
                break
        if name_idx is None:
            continue

        # 구분: 일련번호 바로 왼쪽 열 (라벨과 무관하게 위치로 결정)
        div_idx = seq_idx - 1
        if div_idx < 0:
            raise ValueError("[CS 파서] '일련번호' 왼쪽에 구분 열이 없습니다.")
        return header_idx, div_idx, seq_idx, name_idx

    raise ValueError(
        f"[CS 파서] 헤더 행을 찾을 수 없습니다. "
        f"'{_SEQ_HEADER}'와 '{_NAME_HEADER}' 헤더가 필요합니다."
    )


def parse_cs(path: str, sheet_name: str = _DEFAULT_SHEET) -> list[dict]:
    """Control Sheet에서 조회서 발송 목록을 파싱한다.

    Returns:
        [{"구분", "일련번호", "금융기관명"}, ...]

    Raises:
        ValueError: 시트 없음 / 헤더 행 미발견 / 데이터 없음
    """
    p = Path(path)
    wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"[CS 파서] 시트 '{sheet_name}' 없음. 존재하는 시트: {wb.sheetnames}"
        )

    rows = list(wb[sheet_name].iter_rows(values_only=True))
    wb.close()

    header_idx, c_div, c_seq, c_name = _locate_columns(rows)

    result = []
    for raw in rows[header_idx + 1:]:
        row = list(raw)

        seq = row[c_seq] if c_seq < len(row) else None
        # 발송 기관 행은 일련번호가 정수(1,2,3,...). 헤더 잔재·각주·안내문구
        # (일련번호 칸에 '일련번호'/'2.'/긴 텍스트 등)는 정수가 아니므로 제외.
        seq_int = _as_int(seq)
        if seq_int is None:
            continue

        result.append({
            "구분":       row[c_div]  if c_div  < len(row) else None,
            "일련번호":   seq_int,
            "금융기관명": row[c_name] if c_name < len(row) else None,
        })

    if not result:
        raise ValueError(f"[CS 파서] 데이터 행 없음: {path}")

    return result
