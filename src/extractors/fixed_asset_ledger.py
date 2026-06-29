# -*- coding: utf-8 -*-
"""고정자산 관리대장 → 자산별 이상적 라인 추출 (2양식 자동감지).

회사마다 관리대장 양식이 다르다. 두 실측 양식을 자동 감지해 같은 이상적 출력으로 수렴한다:

**(A) flat 양식 (1행/자산, .xls 흔함)** — 한 헤더행에 모든 필드:
  계정코드|계정과목명|자산코드|자산명|규격|취득일자|…|기초가액|당기증가액|당기감소액|기말잔액|
  전기말상각누계액|당기감가상각비|당기말상각누계액|당기말장부가액. 자산분류=계정과목명, `[…소계]` 스킵.
  (내용연수·상각율 열 없음 → 메모가 내용연수 제공.)

**(B) 기본유형 (2행 페어 + 병합)** — 자산 1건 = 2행:

  [코드] 계정과목명                              ← 섹션 헤더(자산분류)
  자산명 | 경비구분 | 취득일자 | 기초가액 | 당기증감 | 기말잔액 | 전기말상각누계액 | 상각대상금액   ← 헤더行1
         | 양도일자  | 내용연수 | 상각율  | 월수    | 감가상각비| 특별상각비       | 미상각잔액      ← 헤더行2
  옥서스(캐피탈)| 500번대 | 2021-01-01 | 267,272,000 |  | 267,272,000 | 125,283,750 | 267,272,000  ← 자산行1
               |         | 8(내용연수) | 0.125(상각율)| 12 | 33,409,000 |             | 108,579,250  ← 자산行2
  <500번대소계> …                               ← 소계(스킵)
  [자산합계] …                                  ← 합계(스킵)

같은 열이 行1/行2에서 다른 의미를 가지므로(F열=취득일자/내용연수), 2행 헤더를 각각
라벨 매핑해 자산행 페어에서 함께 읽는다. 다른 회사 변형 양식은 후속(기본유형 우선).

이상적 출력(자산 1건):
  {계정과목, 자산명, 취득일자, 기초가액, 당기취득, 당기감소, 기말잔액,
   전기말상각누계, 당기감가상각비, 내용연수, 상각율, 장부가액}
  - 기본유형의 당기증감(net)은 부호로 취득/감소 분리(취득/처분 정밀분해는 분개장 소관).
"""

import re
import warnings
from datetime import date, datetime
from pathlib import Path

import openpyxl

# flat 양식(1행/자산) 헤더 동의어. '전기말상각누계'를 '기말잔액'·'당기말상각누계'보다 먼저.
_FLAT = [
    ("계정과목", ["계정과목명", "계정과목", "계정명"]),   # ⚠️ '계정' 단독 금지(계정코드 오매칭)
    ("자산명", ["자산명", "고정자산명", "자산내역", "품명"]),
    ("취득일자", ["취득일자", "취득일"]),
    ("전기말상각누계", ["전기말상각누계", "전기말감가상각누계", "전기충당금누계", "전기말누계", "기초상각누계"]),
    ("당기감가상각비", ["당기감가상각비", "당기상각비"]),
    ("당기말상각누계", ["당기말상각누계", "당기말감가상각누계", "당기말누계"]),
    ("기초가액", ["기초가액", "기초"]),
    ("당기취득", ["당기증가액", "당기증가", "당기취득"]),
    ("당기감소", ["당기감소액", "당기감소", "당기처분"]),
    ("기말잔액", ["기말잔액", "기말", "취득원가"]),
    ("장부가액", ["당기말장부가액", "장부가액", "미상각잔액"]),
    ("내용연수", ["내용연수"]),
    ("상각율", ["상각율", "상각률"]),
]

# 헤더行1 / 헤더行2 동의어(부분일치). 같은 열이 두 의미를 가져 행별로 따로 매핑.
# ⚠️ '전기말상각누계'를 '기말잔액'보다 먼저(전기'말'이 '기말' 동의어에 오선점되는 것 방지).
_H1 = [
    ("자산명", ["자산명", "자산내역", "품명", "자산번호"]),
    ("경비구분", ["경비구분", "부서"]),
    ("취득일자", ["취득일자", "취득일"]),
    ("전기말상각누계", ["전기말상각누계", "전기말감가상각누계", "기초상각누계", "전기말누계"]),
    ("당기증감", ["당기증감", "증감", "당기증가"]),
    ("기초가액", ["기초가액", "기초", "전기말취득가"]),
    ("기말잔액", ["기말잔액", "기말", "취득원가"]),
    ("상각대상금액", ["상각대상금액", "상각대상"]),
]
_H2 = [
    ("양도일자", ["양도", "폐기", "처분일"]),
    ("내용연수", ["내용연수", "연수"]),
    ("상각율", ["상각율", "상각률", "정율", "정률"]),
    ("월수", ["월수", "상각월수"]),
    ("당기감가상각비", ["감가상각비", "당기상각비", "상각비"]),
    ("특별상각비", ["특별상각"]),
    ("당기말상각누계", ["당기말상각누계", "당기말감가상각누계", "기말상각누계", "당기말누계"]),
    ("장부가액", ["미상각잔액", "장부가액", "장부금액"]),
]
_SECTION_RE = re.compile(r"\[?\s*\d{2,4}\s*\]?\s*(\S.*\S|\S)")  # "[0195] 금융리스자산"
_SKIP_RE = re.compile(r"소\s*계|합\s*계|^\s*[<\[].*[>\]]\s*$|총\s*계")


def _norm(s) -> str:
    return re.sub(r"\s+", "", str(s)).lower() if s is not None else ""


def _amt(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        t = v.replace(",", "").strip()
        if t in ("", "-"):
            return None
        neg = t.startswith("(") and t.endswith(")")
        t = t.strip("()")
        try:
            x = float(t)
            return -x if neg else x
        except ValueError:
            return None
    return None


def _to_date(v):
    if isinstance(v, (datetime, date)):
        return v if isinstance(v, date) and not isinstance(v, datetime) else (v.date() if isinstance(v, datetime) else v)
    if isinstance(v, str):
        m = re.search(r"(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})", v)
        if m:
            try:
                return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            except ValueError:
                return None
    return None


def _field_of(label, table):
    n = _norm(label)
    if not n:
        return None
    for field, kws in table:
        for kw in kws:
            if _norm(kw) in n:
                return field
    return None


def _is_section(row):
    """행이 섹션 헤더(자산분류 '[코드] 명')면 계정과목 문자열 반환, 아니면 None."""
    texts = [c for c in row if isinstance(c, str) and c.strip()]
    nums = [c for c in row if _amt(c) is not None and not isinstance(c, str)]
    if len(texts) != 1 or nums:
        return None
    t = texts[0].strip()
    if _SKIP_RE.search(t):
        return None
    m = re.match(r"\[?\s*\d{2,4}\s*\]\s*(\S.*)$", t)
    if m:
        return m.group(1).strip()
    # 코드 없이 자산분류명만 있는 섹션도 허용(예: '건물', '차량운반구')
    if re.search(r"자산|건물|구축물|차량|비품|기계|시설|공구|기구|소프트|영업권|개발비|토지", t):
        return re.sub(r"^\[?\s*\d*\s*\]?\s*", "", t).strip()
    return None


def _read_rows(path: str):
    """.xlsx(openpyxl) / .xls(xlrd) → 행 튜플 리스트. 관리대장 시트 우선."""
    p = Path(path)
    if p.suffix.lower() == ".xls":
        try:
            import xlrd
        except ImportError:
            warnings.warn(f"[관리대장] .xls 읽기엔 xlrd 필요: {p.name}")
            return []
        try:
            wb = xlrd.open_workbook(str(p))
        except Exception as e:
            warnings.warn(f"[관리대장] .xls 로드 실패 {p.name}: {e}")
            return []
        names = wb.sheet_names()
        sn = next((s for s in names if "관리대장" in s or "고정자산" in s), names[0])
        sh = wb.sheet_by_name(sn)
        return [tuple(sh.row_values(r)) for r in range(sh.nrows)]
    try:
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    except Exception as e:
        warnings.warn(f"[관리대장] 로드 실패 {p.name}: {e}")
        return []
    sheet = next((s for s in wb.sheetnames if "관리대장" in s or "고정자산" in s), wb.sheetnames[0])
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()
    return rows


def parse_fixed_asset_ledger(path: str) -> list[dict]:
    """고정자산 관리대장 → 자산별 라인(flat/기본유형 자동감지). 헤더 없으면 빈 리스트."""
    rows = _read_rows(path)
    if not rows:
        return []
    # flat 양식 감지: 한 행에 자산명 + (당기감가상각비|당기말상각누계) 동시 존재
    # (이 둘은 기본유형에선 헤더行2에만 있어 行1과 구분됨 → 페어 양식 오탐 방지).
    for i, row in enumerate(rows[:40]):
        flat = {_field_of(c, _FLAT) for c in row}
        flat.discard(None)
        if "자산명" in flat and ({"당기감가상각비", "당기말상각누계"} & flat) and len(flat) >= 4:
            return _parse_flat(rows, i)
    return _parse_paired(rows, Path(path).name)


def _parse_flat(rows, h_idx) -> list[dict]:
    """flat 양식(1행/자산) 파싱. h_idx = 헤더행 인덱스."""
    hdr = rows[h_idx]
    col = {c: f for c in range(len(hdr)) if (f := _field_of(hdr[c], _FLAT))}
    name_col = next((c for c, f in col.items() if f == "자산명"), None)
    cls_col = next((c for c, f in col.items() if f == "계정과목"), None)

    def g(row, field):
        for c, f in col.items():
            if f == field and c < len(row):
                return row[c]
        return None

    out, current_class = [], None
    for row in rows[h_idx + 1:]:
        if row is None or all(v in (None, "") for v in row):
            continue
        cls = str(g(row, "계정과목") or "").strip() if cls_col is not None else ""
        name = str(g(row, "자산명") or "").strip()
        if cls and not _SKIP_RE.search(cls):
            current_class = cls
        # 소계/합계 행 스킵(자산명/계정과목 열이 '[…소계]' 등)
        if (name and _SKIP_RE.search(name)) or (cls and _SKIP_RE.search(cls)):
            continue
        if not name:
            continue
        out.append({
            "계정과목": current_class or cls,
            "자산명": name,
            "취득일자": _to_date(g(row, "취득일자")),
            "기초가액": _amt(g(row, "기초가액")) or 0,
            "당기취득": _amt(g(row, "당기취득")) or 0,
            "당기감소": _amt(g(row, "당기감소")) or 0,
            "기말잔액": _amt(g(row, "기말잔액")) or 0,
            "전기말상각누계": _amt(g(row, "전기말상각누계")) or 0,
            "당기감가상각비": _amt(g(row, "당기감가상각비")) or 0,
            "내용연수": _amt(g(row, "내용연수")),
            "상각율": _amt(g(row, "상각율")),
            "장부가액": _amt(g(row, "장부가액")),
        })
    return out


def _parse_paired(rows, fname) -> list[dict]:
    """기본유형(2행 페어) 파싱."""
    # 2행 헤더 탐색: '자산명' 포함 + H1 필드 3개 이상 매칭하는 행(제목行 오탐 방지) = 헤더行1.
    h1_idx = None
    for i, row in enumerate(rows[:40]):
        fields = {_field_of(c, _H1) for c in row}
        fields.discard(None)
        if "자산명" in fields and len(fields) >= 3:
            h1_idx = i
            break
    if h1_idx is None or h1_idx + 1 >= len(rows):
        warnings.warn(f"[관리대장] 헤더 인식 실패: {fname}")
        return []
    hrow1, hrow2 = rows[h1_idx], rows[h1_idx + 1]
    col1 = {c: f for c in range(len(hrow1)) if (f := _field_of(hrow1[c], _H1))}
    col2 = {c: f for c in range(len(hrow2)) if (f := _field_of(hrow2[c], _H2))}
    name_col = next((c for c, f in col1.items() if f == "자산명"), 0)

    def g(row, field, table_cols):
        for c, f in table_cols.items():
            if f == field and c < len(row):
                return row[c]
        return None

    # 헤더 위쪽에서도 섹션 헤더(E6 등) 먼저 반영
    current_class = None
    for row in rows[:h1_idx]:
        sec = _is_section(row)
        if sec:
            current_class = sec

    out = []
    i = h1_idx + 2
    n = len(rows)
    while i < n:
        row = rows[i]
        if row is None or all(v in (None, "") for v in row):
            i += 1
            continue
        sec = _is_section(row)
        if sec:
            current_class = sec
            i += 1
            continue
        name = row[name_col] if name_col < len(row) else None
        name_s = str(name).strip() if name is not None else ""
        if not name_s or _SKIP_RE.search(name_s):
            i += 1   # 소계/합계/빈 자산명 → 스킵(페어 둘째 행은 다음 루프서 자연 스킵)
            continue
        row2 = rows[i + 1] if i + 1 < n else ()
        증감 = _amt(g(row, "당기증감", col1)) or 0
        rec = {
            "계정과목": current_class,
            "자산명": name_s,
            "취득일자": _to_date(g(row, "취득일자", col1)),
            "기초가액": _amt(g(row, "기초가액", col1)) or 0,
            "당기취득": 증감 if 증감 > 0 else 0,
            "당기감소": -증감 if 증감 < 0 else 0,
            "기말잔액": _amt(g(row, "기말잔액", col1)) or 0,
            "전기말상각누계": _amt(g(row, "전기말상각누계", col1)) or 0,
            "당기감가상각비": _amt(g(row2, "당기감가상각비", col2)) or 0,
            "내용연수": _amt(g(row2, "내용연수", col2)),
            "상각율": _amt(g(row2, "상각율", col2)),
            "장부가액": _amt(g(row2, "장부가액", col2)),
        }
        out.append(rec)
        i += 2   # 자산 = 2행 페어
    return out
