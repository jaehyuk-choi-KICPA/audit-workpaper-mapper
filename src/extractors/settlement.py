"""결산보고서 → 이상적 양식 변환 파서.

일부 회사는 거래처원장 대신 **결산보고서**(계정별 명세 시트 모음)로 잔액만 제시한다.
이 파서는 지정된 계정 시트만 골라 이상적 양식 records로 변환한다.

출력 헤더(거래처원장과 동일):
  계정과목 | 적요 | 거래처명 | 전기이월 | 차변/증가 | 대변/감소 | 잔액 | 비고
  - 전기이월·차변/증가·대변/감소 = None (결산보고서는 잔액만 제시)
  - 적요·비고 = 결산보고서의 해당 열에서 채움
  - 잔액 = 금액(또는 미상각잔액)

시트마다 레이아웃이 달라 셀 매핑은 config(_internal/config/settlement_report.yaml)로 둔다.
계정과목은 과목 열(account_col, fill-down) 또는 고정값(account)에서.

결과는 write_ideal_ledger로 거래처원장과 동일한 중간다리 파일이 되어,
하위 생성기(A-200 등)가 입력 소스 구분 없이 그대로 사용한다.
"""

import re
from pathlib import Path

import openpyxl
import yaml


def _norm(v) -> str:
    return re.sub(r"\s+", "", str(v).strip()) if v is not None else ""


_TOTAL_WORDS = {"총계", "합계", "소계", "계"}


def _is_subtotal(v) -> bool:
    """소계/합계 행 표식 판별: '총계'/'합계'/'( 예 금 계 )' 등."""
    n = _norm(v)
    if not n:
        return False
    if n in _TOTAL_WORDS:
        return True
    if re.search(r"계\)$", n):          # (예금계), (현금계)
        return True
    if any(w in n for w in ("총계", "합계", "소계")):
        return True
    return False


def _is_amount(v) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s != "' " else None


def _extract_sheet(ws, entry: dict) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))

    acc_col = entry.get("account_col")          # 1-indexed 과목 열
    fixed_acc = entry.get("account")
    keep = entry.get("keep_accounts")
    keep_norm = {_norm(k) for k in keep} if keep else None
    c_name = entry.get("거래처명")
    c_remark = entry.get("적요")
    c_amt = entry["잔액"]
    c_bigo = entry.get("비고")

    def g(row, col):
        return row[col - 1] if col and (col - 1) < len(row) else None

    cur_acc = fixed_acc
    cur_name = None
    result = []

    for row in rows:
        # 1) 소계/합계 행 skip (과목·거래처·적요 어느 칸이든 표식이면)
        if any(_is_subtotal(g(row, c)) for c in (acc_col, c_name, c_remark) if c):
            continue

        # 2) 과목 fill-down
        if acc_col:
            a = g(row, acc_col)
            if a is not None and str(a).strip():
                cur_acc = _norm(a)

        # 3) 금액 없는 행(헤더/빈행/안내) skip
        amt = g(row, c_amt)
        if not _is_amount(amt):
            continue

        # 4) 거래처명 fill-down
        if c_name:
            nm = g(row, c_name)
            if nm is not None and str(nm).strip():
                cur_name = str(nm).strip()

        # 5) 추출 계정 필터
        if keep_norm is not None and _norm(cur_acc) not in keep_norm:
            continue
        if not cur_acc:
            continue

        result.append({
            "계정과목":  cur_acc,
            "적요":      _clean(g(row, c_remark)) if c_remark else None,
            "거래처명":  cur_name,
            "전기이월":  None,
            "차변/증가": None,
            "대변/감소": None,
            "잔액":      amt,
            "비고":      _clean(g(row, c_bigo)) if c_bigo else None,
        })
    return result


def parse_settlement_report(path: str, config_path: str) -> list[dict]:
    """결산보고서에서 config에 지정된 계정 시트만 추출해 이상적 양식 records 반환.

    Raises:
        ValueError: 파일/시트 열기 실패, config 누락
    """
    cfg = yaml.safe_load(Path(config_path).read_text(encoding="utf-8"))
    entries = cfg.get("sheets", [])

    wb = openpyxl.load_workbook(str(Path(path)), read_only=True, data_only=True)
    names = wb.sheetnames

    out: list[dict] = []
    for entry in entries:
        kw = entry["sheet"]
        excl = entry.get("exclude")
        for name in names:
            if kw not in name:
                continue
            if excl and excl in name:
                continue
            out.extend(_extract_sheet(wb[name], entry))
    wb.close()
    return out
