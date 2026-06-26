"""수정사항집계 파서 — 감사 수정분개표.

정산표의 **수정사항집계(25) 시트**(당기 수정분개)를 entry(분개 묶음) 단위로 구조화한다.
A-0 등 조서 총괄표의 '수정사항' 섹션에 관련 분개를 **그대로 재현**하기 위함이므로,
부호 변환 없이 원본 차변/대변 구조를 보존한다.

시트 구조(헤더 2행):
  행1: # | 계정과목 | 금액 | Effect | Description
  행2:      차변 대변 | 차변 대변 | 손익 이익잉여금
  - entry 시작 = '#' 열에 정수가 있는 행. 다음 정수 전까지 한 entry.
  - 한 줄 = 차변(계정과목 차변열+금액 차변열) 또는 대변(계정과목 대변열+금액 대변열).
  - 설명만 있는 줄은 entry 비고로 모은다. 줄 없는 빈 entry는 스킵.

출력:
  [{"no": int, "lines": [{"side","계정","금액","손익","이익잉여금","설명"}...],
    "notes": [str...]}, ...]
시트/헤더 없으면 빈 리스트.
"""

import re
import warnings
from pathlib import Path

import openpyxl

from ._headers import normalize


def _pick_sheet(names: list) -> "str | None":
    """수정사항집계(NN) 시트 중 **당기연도(가장 높은 NN)** 를 고른다.

    회사마다 감사연도가 달라 시트명이 (25)/(24)/(23) 등으로 바뀐다. 별도정산표가 당기
    (Unaudited)에 앵커하는 것과 일관되게, 연도 접미사가 가장 큰 것(=당기)을 쓴다.
    """
    cands = []
    for n in names:
        m = re.match(r"수정사항집계\((\d+)\)", str(n).replace(" ", ""))
        if m:
            cands.append((int(m.group(1)), n))
    if cands:
        return max(cands)[1]
    for n in names:                       # 접미사 없는 변형 대비
        if str(n).replace(" ", "").startswith("수정사항집계"):
            return n
    return None


def _is_amount(v) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _as_int(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float) and v.is_integer():
        return int(v)
    s = str(v).strip() if v is not None else ""
    return int(s) if s.isdigit() else None


def _find_header(rows: list[tuple]) -> "tuple[int, dict] | None":
    """(데이터 시작행 인덱스, {필드: 열인덱스}) 반환. 0-indexed.

    행1(계정과목/금액/Effect/Description) + 행2(차변/대변/손익/이익잉여금) 구조에 앵커.
    """
    for i in range(len(rows) - 1):
        n1 = [normalize(c) for c in rows[i]]
        n2 = [normalize(c) for c in rows[i + 1]]
        col_no = next((j for j, n in enumerate(n1) if n == "#"), None)
        # 부분일치(회사마다 '금액'↔'수정금액', '계정과목'↔'계정과목명' 등 변형).
        col_계정 = next((j for j, n in enumerate(n1) if "계정과목" in n), None)
        col_금액 = next((j for j, n in enumerate(n1) if "금액" in n), None)
        col_effect = next((j for j, n in enumerate(n1) if n.lower().startswith("effect")), None)
        # Description 열 = **라벨 1차**(영문 Description + 한글 동의어), **두 헤더행 모두** 스캔.
        # 라벨을 1차로 두는 이유: Effect가 병합셀로 나타나면(read_only에서 비-앵커 칸=None) 위치
        # 오프셋(Effect+2)이 어긋날 수 있어 이 표에선 라벨이 더 안정적이다. 라벨이 어느 헤더행에
        # 오든(일부 회사=서브헤더 행) 커버하려고 n1·n2를 함께 본다. 위치는 라벨 자체가 없는 변형 폴백.
        def _label_desc(*hdrs):
            for h in hdrs:
                j = next((k for k, n in enumerate(h)
                          if "description" in n.lower()
                          or any(s in n for s in ("적요", "수정사유", "사유", "내용"))), None)
                if j is not None:
                    return j
            return None
        col_desc = _label_desc(n1, n2)
        if col_desc is None and col_effect is not None:
            col_desc = col_effect + 2          # 폴백: 라벨 없는 변형만(손익·이익잉여금 뒤)
        if None in (col_계정, col_금액):
            continue
        # 행2가 차변/대변 서브헤더인지 확인
        if not (normalize(rows[i + 1][col_계정] if col_계정 < len(rows[i + 1]) else "") == "차변"):
            continue
        cm = {
            "no": col_no,
            "차변계정": col_계정, "대변계정": col_계정 + 1,
            "금액차변": col_금액, "금액대변": col_금액 + 1,
            "손익": col_effect, "이익잉여금": (col_effect + 1) if col_effect is not None else None,
            "설명": col_desc,
        }
        return i + 2, cm   # 데이터는 서브헤더(행2) 다음부터
    return None


def parse_adjustments(path: str, sheet_name: str = None) -> list[dict]:
    """수정사항집계 시트를 entry 리스트로 반환. sheet_name 미지정 시 당기연도 시트 자동 선택."""
    p = Path(path)
    wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    sheet = sheet_name if (sheet_name and sheet_name in wb.sheetnames) else _pick_sheet(wb.sheetnames)
    if sheet is None:
        wb.close()
        warnings.warn(f"[수정사항 파서] 수정사항집계(연도) 시트 없음. 존재: {wb.sheetnames}")
        return []
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()

    found = _find_header(rows)
    if found is None:
        warnings.warn("[수정사항 파서] 헤더(계정과목/금액 + 차변/대변)를 찾지 못함.")
        return []
    start, cm = found

    def g(row, key):
        j = cm.get(key)
        return row[j] if j is not None and j < len(row) else None

    entries: list[dict] = []
    cur = None
    for row in rows[start:]:
        # 표1 끝: '계'(총계) 행에서 중단(이후 '(2) 미수정왜곡표시' 등 다른 표의 헤더 흡수 방지).
        # '계'는 '#'(no) 열에 온다.
        if normalize(g(row, "no")) == "계":
            break
        no = _as_int(g(row, "no"))
        if no is not None:                       # 새 entry 시작
            if cur and (cur["lines"] or cur["notes"]):
                entries.append(cur)
            cur = {"no": no, "lines": [], "notes": []}
        if cur is None:
            continue
        손익 = g(row, "손익"); 이잉 = g(row, "이익잉여금"); 설명 = g(row, "설명")
        dr_acc = g(row, "차변계정"); cr_acc = g(row, "대변계정")
        added = False
        if dr_acc is not None and str(dr_acc).strip():
            cur["lines"].append({
                "side": "차변", "계정": str(dr_acc).strip(),
                "금액": g(row, "금액차변") if _is_amount(g(row, "금액차변")) else None,
                "손익": 손익 if _is_amount(손익) else None,
                "이익잉여금": 이잉 if _is_amount(이잉) else None,
                "설명": str(설명).strip() if 설명 is not None and str(설명).strip() else None,
            })
            added = True
        if cr_acc is not None and str(cr_acc).strip():
            cur["lines"].append({
                "side": "대변", "계정": str(cr_acc).strip(),
                "금액": g(row, "금액대변") if _is_amount(g(row, "금액대변")) else None,
                "손익": 손익 if _is_amount(손익) else None,
                "이익잉여금": 이잉 if _is_amount(이잉) else None,
                "설명": str(설명).strip() if 설명 is not None and str(설명).strip() else None,
            })
            added = True
        if not added and 설명 is not None and str(설명).strip():
            cur["notes"].append(str(설명).strip())

    if cur and (cur["lines"] or cur["notes"]):
        entries.append(cur)
    return entries
