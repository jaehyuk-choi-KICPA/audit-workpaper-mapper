"""금융기관조회서 취합엑셀 INSURANCE 시트 파서.

보험사 회신 1번 항목(보험거래 내용)에서 **해약환급금이 있는 적립식 보험**만
장기금융상품으로 추출한다. (해약환급금 0인 손해·차량보험 등은 제외)

분류·그룹화 규칙 (실파일 검증):
  - 해약환급금_금액(col29) > 0 → 장기금융상품
  - (금융기관명, 보험의종류) 동일한 정책을 1행으로 묶음
    · 증권번호: 쉼표 연결
    · 조회서금액: 해약환급금_금액 합산

⚠️ INSURANCE 시트는 6개 섹션(1.보험 2.대출 3.지급보증 …)으로 구성되며
   1번 섹션만 사용한다. col29에는 ''(빈문자열)·None이 섞여 있으므로
   금액 변환은 반드시 안전 변환(_to_amount)을 거친다. (int('') 크래시 방지)
"""

import re
from pathlib import Path

import openpyxl

from ._sections import find_section_bounds
from ._headers import map_headers

# 헤더 동의어 → 표준 컬럼명.
# 주의: '해약환급금'(col15 텍스트)은 매핑하지 않는다. 숫자열인
#       '해약환급금_금액'(col29)만 사용한다.
_SYNONYMS: dict[str, str] = {
    "조서번호":     "조서번호",
    "금융기관명":   "금융기관명",
    "보험의 종류":  "보험의종류",
    "보험의종류":   "보험의종류",
    "증권번호":     "증권번호",
    # 해약환급금: 리치 변형은 분리된 숫자열(_금액, 예: 별도 col), 단순 변형은 단일열(예: 통합 col).
    # 둘 다 인식하고, 추출 시 _금액(파란 숫자)을 우선한다.
    "해약환급금_금액": "해약환급금_금액",
    "해약환급금":     "해약환급금",
}

# 필수: 금융기관명·보험의종류 + (해약환급금_금액 또는 해약환급금 중 하나)
_REQUIRED_BASE = {"금융기관명", "보험의종류"}

# 부분일치(fuzzy) 폴백 키워드 — 정확매칭으로 못 찾은 필드만 보강(폴백 전용, 회귀 위험 0).
# 주의: '해약환급금_금액'(파란 숫자열)을 '해약환급금'(텍스트열)보다 먼저 두어 우선 매칭.
_SUB_KEYWORDS: list[tuple] = [
    ("금융기관명",     ["금융기관", "기관명", "금융기관명"]),
    ("보험의종류",     ["보험의종류", "보험종류", "종류", "상품명"]),
    ("증권번호",       ["증권번호", "증권", "증서", "계약번호"]),
    ("해약환급금_금액", ["해약환급금_금액", "해지환급금_금액"]),
    ("해약환급금",     ["해약환급금", "해지환급금", "환급금"]),
]

_DEFAULT_SHEET = "INSURANCE"


def _map_headers(row: tuple) -> dict[str, int]:
    """행에서 {표준컬럼명: 열인덱스} 매핑. 필수 컬럼 없으면 빈 dict.

    1차 정확 + 2차 부분일치(공용 헬퍼). 금융기관명·보험의종류는 필수이고,
    해약환급금_금액 또는 해약환급금 중 하나는 반드시 있어야 한다.
    """
    mapping = map_headers(row, synonyms=_SYNONYMS, required=_REQUIRED_BASE,
                          sub_keywords=_SUB_KEYWORDS, parser_key="insurance")
    if mapping is None:
        return {}
    has_amount = "해약환급금_금액" in mapping or "해약환급금" in mapping
    return mapping if has_amount else {}


def _to_amount(value) -> float:
    """금액 셀을 안전하게 숫자로 변환. None/''/비숫자 → 0 (절대 예외 없음)."""
    if isinstance(value, bool):
        return 0
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        s = re.sub(r"[^\d.\-]", "", value)  # 'KRW 1,234' → '1234'
        try:
            return float(s) if s else 0
        except ValueError:
            return 0
    return 0


def parse_insurance(path: str, sheet_name: str = _DEFAULT_SHEET) -> list[dict]:
    """INSURANCE 1번 섹션에서 장기금융상품(해약환급금 보유)을 추출·그룹화한다.

    Returns:
        [{"금융기관명", "보험의종류", "증권번호", "조회서금액", "건수"}, ...]
        (금융기관명, 보험의종류) 기준 그룹. 조회서금액 = 해약환급금 합산.

    참고: 시트 없음/1번 섹션·헤더 미발견(=보험거래 없는 회사 등)은 정상 케이스로 보고
    예외 대신 빈 리스트를 반환한다 (한 파서의 빈 결과가 전체 A-1 생성을 막지 않도록).
    """
    p = Path(path)
    wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []

    rows = list(wb[sheet_name].iter_rows(values_only=True))
    wb.close()

    bounds = find_section_bounds(rows, "1.")
    if bounds is None:
        return []
    sec_idx, sec_end = bounds

    header_idx, col_map = None, None
    for i in range(sec_idx + 1, min(sec_idx + 6, sec_end)):
        cm = _map_headers(rows[i])
        if cm:
            header_idx, col_map = i, cm
            break
    if col_map is None:
        return []

    c_inst = col_map["금융기관명"]
    c_kind = col_map["보험의종류"]
    c_amt = col_map.get("해약환급금_금액", col_map.get("해약환급금"))  # 파란 숫자열 우선
    c_pol = col_map.get("증권번호")
    c_ref = col_map.get("조서번호")

    def g(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    # (금융기관명, 보험의종류) → 그룹 누적 (등장 순서 유지)
    groups: dict[tuple, dict] = {}
    for i in range(header_idx + 1, sec_end):
        row = rows[i]
        inst = g(row, c_inst)
        if inst is None or str(inst).strip() == "":
            continue

        amount = _to_amount(g(row, c_amt))
        if amount <= 0:
            continue  # 해약환급금 없는 보험(손해·차량 등) → 장기금융상품 아님

        kind = g(row, c_kind)
        key = (str(inst).strip(), str(kind).strip() if kind else "")
        policy = g(row, c_pol)

        if key not in groups:
            groups[key] = {
                "금융기관명": key[0],
                "보험의종류": key[1],
                "증권번호목록": [],
                "조회서금액": 0,
                "건수": 0,
                "조서번호": None,
            }
        grp = groups[key]
        grp["조회서금액"] += amount
        grp["건수"] += 1
        if policy is not None and str(policy).strip() not in ("", "' "):
            grp["증권번호목록"].append(str(policy).strip())
        ref = g(row, c_ref)
        if grp["조서번호"] is None and ref is not None and str(ref).strip():
            grp["조서번호"] = str(ref).strip()

    result = []
    for grp in groups.values():
        result.append({
            "금융기관명": grp["금융기관명"],
            "보험의종류": grp["보험의종류"],
            "증권번호":   ", ".join(grp["증권번호목록"]),
            "조회서금액": grp["조회서금액"],
            "건수":       grp["건수"],
            "조서번호":   grp["조서번호"],
        })
    return result
