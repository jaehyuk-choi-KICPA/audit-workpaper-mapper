"""금융기관조회서 취합엑셀 BANK 시트 파서.

은행 회신 1번 항목(금융상품의 내용)에서 계좌별 잔액을 추출한다.
금융상품의 종류로 현금성자산 / 퇴직연금을 분류한다:
  - 퇴직연금 → A300 (기타조회내역)
  - 그 외     → A200 (현금성자산)

시트 구조 (회사 무관, 표준 양식):
  - 섹션 마커: A열(또는 인접) 텍스트가 `숫자.` / `숫자-숫자.` 로 시작
    1번 = 금융상품, 2-1/2-2 = 대출, 9번 = 담보 ...
  - 1번 섹션: [마커행] → [+2행 헤더] → [+3행~ 데이터] → [다음 마커 전까지]
"""

import re
from pathlib import Path

import openpyxl

from ._sections import find_section_bounds

# 헤더 동의어 → 표준 컬럼명
_SYNONYMS: dict[str, str] = {
    "조서번호":      "조서번호",
    "금융기관명":    "금융기관명",
    "금융상품의 종류": "금융상품종류",
    "금융상품의종류": "금융상품종류",
    "계좌번호":      "계좌번호",
    "금액":          "금액",
    "연이자율":      "연이자율",
    "만기일":        "만기일",
    "금액_통화":     "통화",
    "금액_금액":     "외화금액",
}

# 1번 섹션 헤더로 인정하기 위한 최소 필수 컬럼
_REQUIRED = {"금융기관명", "금융상품종류", "계좌번호", "금액"}

# 만기일 sentinel — 날짜가 아닌 값은 수시입출금으로 간주
_MATURITY_NONE = {"00000000", "00010101", "0", ""}

# 퇴직연금 분류 키워드 (금융상품종류 컬럼값)
_PENSION_KEYWORD = "퇴직연금"

_DEFAULT_SHEET = "BANK"


def _map_headers(row: tuple) -> dict[str, int]:
    """행에서 {표준컬럼명: 열인덱스} 매핑. 필수 컬럼 없으면 빈 dict."""
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(row):
        if cell is None:
            continue
        std = _SYNONYMS.get(str(cell).strip())
        if std and std not in mapping:
            mapping[std] = idx
    return mapping if _REQUIRED.issubset(mapping.keys()) else {}


def _format_maturity(value) -> str:
    """만기일 셀값을 표준화. 무만기 sentinel → '수시입출금'."""
    if value is None:
        return "수시입출금"
    s = str(value).strip()
    if s in _MATURITY_NONE:
        return "수시입출금"
    # YYYYMMDD 형태면 YYYY-MM-DD 로
    if re.fullmatch(r"\d{8}", s):
        return f"{s[:4]}-{s[4:6]}-{s[6:]}"
    return s


def _classify(product_kind) -> str:
    """금융상품종류 → 계정분류."""
    if product_kind and _PENSION_KEYWORD in str(product_kind):
        return "퇴직연금"
    return "현금성자산"


def _to_record(row: tuple, cm: dict[str, int]) -> dict:
    def g(col: str):
        idx = cm.get(col)
        return row[idx] if idx is not None and idx < len(row) else None

    kind = g("금융상품종류")
    return {
        "조서번호":     g("조서번호"),
        "금융기관명":   g("금융기관명"),
        "금융상품종류": kind,
        "계좌번호":     g("계좌번호"),
        "금액":         g("금액"),
        "연이자율":     g("연이자율"),
        "만기":         _format_maturity(g("만기일")),
        "통화":         g("통화"),
        "외화금액":     g("외화금액"),
        "계정분류":     _classify(kind),
    }


def parse_bank(path: str, sheet_name: str = _DEFAULT_SHEET) -> list[dict]:
    """BANK 시트 1번 항목을 파싱해 계좌별 표준 행 리스트를 반환한다.

    Returns:
        [{"조서번호","금융기관명","금융상품종류","계좌번호","금액",
          "연이자율","만기","통화","외화금액","계정분류"}, ...]
        계정분류 = "현금성자산" | "퇴직연금"

    Raises:
        ValueError: 시트 없음 / 1번 섹션·헤더 미발견 / 데이터 없음
    """
    p = Path(path)
    wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"[BANK 파서] 시트 '{sheet_name}' 없음. 존재: {wb.sheetnames}")

    rows = list(wb[sheet_name].iter_rows(values_only=True))
    wb.close()

    # 1번 섹션 경계 [start, end) 탐색
    bounds = find_section_bounds(rows, "1.")
    if bounds is None:
        raise ValueError("[BANK 파서] 1번 섹션(금융상품) 마커를 찾을 수 없습니다.")
    sec1_idx, sec1_end = bounds

    # 마커 다음 행들 중 헤더 행 탐색 (필수 컬럼 모두 포함하는 첫 행)
    header_idx, col_map = None, None
    for i in range(sec1_idx + 1, min(sec1_idx + 6, sec1_end)):
        cm = _map_headers(rows[i])
        if cm:
            header_idx, col_map = i, cm
            break
    if col_map is None:
        raise ValueError("[BANK 파서] 1번 섹션 헤더를 찾을 수 없습니다.")

    # 데이터: 헤더 다음 ~ 섹션 끝(다음 마커 전)
    result = []
    for i in range(header_idx + 1, sec1_end):
        row = rows[i]
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        # 금융기관명 없는 행(주석·합계 등) skip
        inst_idx = col_map["금융기관명"]
        inst = row[inst_idx] if inst_idx < len(row) else None
        if inst is None or str(inst).strip() == "":
            continue
        result.append(_to_record(row, col_map))

    if not result:
        raise ValueError(f"[BANK 파서] 1번 섹션 데이터 행 없음: {path}")

    return result


# ---------------------------------------------------------------------------
# 2-2번(대출거래) · 9번(담보제공) 섹션 파서 — A-300용
#
# 금액은 검은색 텍스트열이 아니라 파란색 숫자열('_금액' 접미)을 사용한다.
# "해당 없음"·"담보보증내역없음" 등 실거래 없는 행은 제외한다.
# ---------------------------------------------------------------------------

_LOAN_SYNONYMS = {
    "조서번호":            "조서번호",
    "금융기관명":          "금융기관명",
    "대출 종류":           "대출종류",
    "대출종류":            "대출종류",
    "금액_약정한도액_금액": "약정한도액",
    "금액_대출금액_금액":   "대출금액",
    "대출일":              "대출일",
    "최종만기일":          "최종만기일",
    "이자_연이율":         "연이율",
    "이자_최종이자지급일":  "최종이자지급일",
    "상환방법":            "상환방법",
    "담보 보증 및 관련약정": "담보보증",
}
_LOAN_REQUIRED = {"금융기관명", "대출종류", "약정한도액", "대출금액"}

_COLLATERAL_SYNONYMS = {
    "조서번호":          "조서번호",
    "금융기관명":        "금융기관명",
    "구분":              "구분",
    "담보보증의 내용":   "담보보증내용",
    "소유자(제공자)":    "소유자",
    "감정금액_금액":     "감정금액",
    "설정금액_금액":     "설정금액",
    "설정순위":          "설정순위",
    "선순위 설정금액_금액": "선순위설정금액",
}
_COLLATERAL_REQUIRED = {"금융기관명", "구분", "감정금액", "설정금액"}

_NONE_TOKENS = {"", "' ", "해당 없음", "해당없음", "담보보증내역없음"}


def _map_by(row: tuple, synonyms: dict, required: set) -> dict:
    mapping = {}
    for idx, cell in enumerate(row):
        if cell is None:
            continue
        std = synonyms.get(str(cell).strip())
        if std and std not in mapping:
            mapping[std] = idx
    return mapping if required.issubset(mapping.keys()) else {}


def _section_rows(path, sheet_name, prefix, synonyms, required):
    p = Path(path)
    wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"[BANK 파서] 시트 '{sheet_name}' 없음")
    rows = list(wb[sheet_name].iter_rows(values_only=True))
    wb.close()

    bounds = find_section_bounds(rows, prefix)
    if bounds is None:
        return []
    sec_idx, sec_end = bounds

    header_idx, col_map = None, None
    for i in range(sec_idx + 1, min(sec_idx + 6, sec_end)):
        cm = _map_by(rows[i], synonyms, required)
        if cm:
            header_idx, col_map = i, cm
            break
    if col_map is None:
        return []

    def g(row, key):
        idx = col_map.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    out = []
    for i in range(header_idx + 1, sec_end):
        row = rows[i]
        inst = g(row, "금융기관명")
        if inst is None or str(inst).strip() == "":
            continue
        out.append({k: g(row, k) for k in col_map})
    return out


def _is_none_token(v) -> bool:
    return v is None or str(v).strip() in _NONE_TOKENS


def parse_bank_loans(path: str, sheet_name: str = _DEFAULT_SHEET) -> list[dict]:
    """BANK 2-2번(대출거래) — 실대출만. 금액은 파란 숫자열 사용."""
    rows = _section_rows(path, sheet_name, "2-2", _LOAN_SYNONYMS, _LOAN_REQUIRED)
    return [r for r in rows if not _is_none_token(r.get("대출종류"))]


def parse_bank_collateral(path: str, sheet_name: str = _DEFAULT_SHEET) -> list[dict]:
    """BANK 9번(담보제공) — 실제 담보만(감정/설정금액>0). 금액은 파란 숫자열 사용."""
    rows = _section_rows(path, sheet_name, "9.", _COLLATERAL_SYNONYMS, _COLLATERAL_REQUIRED)
    def has_collateral(r):
        from .insurance import _to_amount as _amt
        return _amt(r.get("감정금액")) > 0 or _amt(r.get("설정금액")) > 0
    return [r for r in rows if has_collateral(r)]
