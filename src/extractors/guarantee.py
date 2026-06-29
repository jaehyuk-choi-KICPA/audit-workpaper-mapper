# -*- coding: utf-8 -*-
"""금융기관조회서 취합엑셀 GUARANTEE 시트 1번(당사가 제공한 보증) 박스 추출.

1번 섹션(보증(보험)의 내용)을 **그대로** 떠와 A-1 FN 시트에 붙인다.
정산표/조회서에 GUARANTEE가 있을 때만(데이터 존재 시) 반영하고, 없으면 None.
값은 회신값 그대로 보존(재계산 없음 — '그대로 갖다 붙이기').
"""
from pathlib import Path

import openpyxl

from ._sections import find_section_bounds

_DEFAULT_SHEET = "GUARANTEE"


def parse_guarantee(path: str, sheet_name: str = _DEFAULT_SHEET):
    """GUARANTEE 1번 섹션을 {title, header, rows}로 반환(없으면 None).

    header/rows = 헤더 첫~끝 비어있지 않은 열 범위의 셀값(그대로). rows는 선금/지급/합계 등.
    """
    p = Path(path)
    try:
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    except Exception:
        return None
    if sheet_name not in wb.sheetnames:
        wb.close()
        return None
    rows = list(wb[sheet_name].iter_rows(values_only=True))
    wb.close()

    b = find_section_bounds(rows, "1.")
    if b is None:
        return None
    si, se = b
    title = (str(rows[si][1]).strip() if len(rows[si]) > 1 and rows[si][1]
             else "1. 조회기준일 현재 당사가 제공한 보증(보험)의 내용")

    # 헤더 행: 조서번호 + 금융기관명 포함
    hi = None
    for i in range(si + 1, min(si + 6, se)):
        vals = [str(v).strip() for v in rows[i] if v not in (None, "")]
        if any("조서번호" in v for v in vals) and any("금융기관" in v for v in vals):
            hi = i
            break
    if hi is None:
        return None
    hdr = rows[hi]
    cols = [c for c, v in enumerate(hdr) if v not in (None, "")]
    if not cols:
        return None
    c0, c1 = min(cols), max(cols)

    # 종류·금액 열(데이터 행 판별용)
    kind_c = next((c for c, v in enumerate(hdr) if v and "종류" in str(v)), None)
    amt_c = next((c for c, v in enumerate(hdr) if v and ("보증금액" in str(v) or "보험가입" in str(v))), None)

    def slab(r):
        return [(r[c] if c < len(r) else None) for c in range(c0, c1 + 1)]

    data = []
    for i in range(hi + 1, se):
        r = rows[i]
        kind = r[kind_c] if (kind_c is not None and kind_c < len(r)) else None
        amt = r[amt_c] if (amt_c is not None and amt_c < len(r)) else None
        if (kind in (None, "") and amt in (None, "")):
            continue                       # 빈/주석 행 skip
        data.append(slab(r))
    if not data:
        return None
    return {"title": title, "header": slab(hdr), "rows": data}
