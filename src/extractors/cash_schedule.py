# -*- coding: utf-8 -*-
"""잔액 스냅샷 소스(현금성자산 등 계좌관리 엑셀) → 이상적 양식 변환기.

거래처원장과 달리 일부 회사는 은행계좌를 **별도 엑셀**로 관리한다(시점 잔액 스냅샷):
  제목/회사명 행 + 헤더행(상단 아닐 수 있음) + 소계행(예: '보통예금 합계') + 데이터행.
  컬럼: 개설일·계정과목·거래처명·계좌번호·용도·잔액·이자율·만기일 (전기이월/증감 없음).

→ 이상적 양식(계정과목|적요|거래처명|전기이월|차변/증가|대변/감소|잔액|비고)으로 수렴시킨다.
   전기이월/증감은 공란(스냅샷이라 없음). 적요=용도, 비고=계좌번호. A-1/A-0은 잔액만 사용.

견고성: 헤더/데이터 미발견은 예외 대신 **빈 리스트**(부분 출력 보장). 계정과목은 행 컬럼에서
도출하되, 컬럼이 없으면 직전 '○○ 합계' 소계행 라벨로 포워드필(섹션 = 계정과목).

포맷 호환:
  - C-1차: `_headers.map_headers`(정확+부분일치)로 거래처명·잔액(필수) 등 열 인식.
  - C-2차(폴백): 필수 열을 못 찾고 resolver가 주어지면(키 있을 때) 마스킹 LLM로 보정.
"""

from pathlib import Path

import openpyxl

from ._headers import map_headers, normalize

# 헤더 동의어 → 표준 필드 (용도→적요, 계좌번호는 비고로 보냄)
_SYNONYMS = {
    "계정과목": "계정과목", "계정": "계정과목", "과목": "계정과목",
    "거래처명": "거래처명", "거래처": "거래처명", "예금주": "거래처명",
    "잔액": "잔액", "기말잔액": "잔액", "장부가액": "잔액",
    "계좌번호": "계좌번호",
    "용도": "적요", "적요": "적요",
}

_SUB_KEYWORDS = [
    ("계정과목", ["계정과목", "계정", "과목"]),
    ("거래처명", ["거래처", "은행", "금융기관", "예금주", "계좌명", "기관명"]),
    ("잔액",     ["잔액", "기말", "장부가", "금액"]),
    ("계좌번호", ["계좌", "번호"]),
    ("적요",     ["용도", "적요", "내용"]),
]

# 거래처명·잔액만 있으면 잔액 소스로 인정(계정과목·계좌번호·적요는 선택)
_REQUIRED = {"거래처명", "잔액"}

_IDEAL_FIELDS = ["계정과목", "적요", "거래처명", "전기이월", "차변/증가", "대변/감소", "잔액", "비고"]


def _has_total_marker(row) -> bool:
    """행 어느 셀에든 '합계'가 들어가면 소계/총계행으로 본다(데이터행 아님)."""
    for c in row:
        if c is not None and "합계" in normalize(c):
            return True
    return False


def _section_label(row) -> "str | None":
    """'보 통 예 금 합 계' 같은 소계행에서 계정과목 라벨 추출('합계' 제거)."""
    for c in row:
        if c is not None:
            n = normalize(c)
            if "합계" in n:
                lab = n.replace("합계", "").strip()
                return lab or None
    return None


def _find_header(rows, *, resolver=None):
    """헤더행 인덱스 + 컬럼맵을 찾는다. 못 찾으면 (None, None).

    상단에 제목/회사명 행이 있을 수 있으므로 첫 15행을 훑는다.
    1차(오프라인): C-1차로 필수열을 가진 첫 행을 찾는다(대부분 여기서 끝 — LLM 호출 0).
    2차(폴백): 1차가 끝까지 실패하고 resolver가 있을 때만, **가장 헤더다운 후보 1행**에만
               마스킹 LLM 보정을 1회 적용한다(검색 중 매 행 호출 금지 — 오탐·불필요 egress 방지).
    """
    scan = min(15, len(rows))
    cands = []  # (부분매칭수, 비어있지않은셀수, i, 부분맵)
    for i in range(scan):
        cm = map_headers(rows[i], synonyms=_SYNONYMS, required=_REQUIRED,
                         sub_keywords=_SUB_KEYWORDS, parser_key="cash")
        if cm is not None:
            return i, cm  # 1차 성공 — 오프라인
        partial = map_headers(rows[i], synonyms=_SYNONYMS, required=set(),
                              sub_keywords=_SUB_KEYWORDS, parser_key="cash") or {}
        nonempty = sum(1 for c in rows[i] if c not in (None, ""))
        cands.append((len(partial), nonempty, i, partial))

    if resolver is None or not cands:
        return None, None

    # 2차 — 가장 헤더다운 후보(부분매칭↑, 그다음 비어있지 않은 셀↑) 1행만 LLM 보정
    _, _, i, cm = max(cands, key=lambda t: (t[0], t[1]))
    cm = dict(cm)
    missing = list(_REQUIRED - cm.keys())
    samples = [r for r in rows[i + 1:i + 14] if any(v not in (None, "") for v in r)]
    for field, idx in (resolver(rows[i], samples, missing) or {}).items():
        if field in (_REQUIRED - cm.keys()):
            cm[field] = idx
    return (i, cm) if _REQUIRED.issubset(cm.keys()) else (None, None)


def parse_cash_schedule(path: str, *, resolver=None) -> list[dict]:
    """현금성자산 등 잔액 스냅샷 엑셀을 이상적 양식 records로 변환.

    Args:
        path:     원본 엑셀(.xlsx/.xlsm)
        resolver: 선택 — (header_cells, sample_rows, missing_fields)->{field:idx}.
                  C-1차로 필수열을 못 찾을 때만 호출(C-2차 마스킹 LLM 폴백).

    Returns:
        이상적 양식 records. 시트/헤더/데이터 미발견은 [] (부분 출력 보장).
    """
    p = Path(path)
    if p.suffix.lower() not in (".xlsx", ".xlsm"):
        return []

    wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    result: list[dict] = []
    for sn in wb.sheetnames:
        rows = [list(r) for r in wb[sn].iter_rows(values_only=True)]
        hdr_i, cm = _find_header(rows, resolver=resolver)
        if cm is None:
            continue

        c_acct = cm.get("계정과목")
        c_name = cm.get("거래처명")
        c_bal  = cm.get("잔액")
        c_memo = cm.get("적요")
        c_acc  = cm.get("계좌번호")

        def g(row, idx):
            return row[idx] if idx is not None and idx < len(row) else None

        section = None  # 직전 '○○ 합계' 섹션 라벨(계정과목 컬럼 없을 때 포워드필)
        for r in rows[hdr_i + 1:]:
            if all(v is None or str(v).strip() == "" for v in r):
                continue
            if _has_total_marker(r):
                lab = _section_label(r)
                if lab:
                    section = lab
                continue  # 소계/총계행은 데이터 아님(중복합산 방지)

            bal = g(r, c_bal)
            if bal is None or str(bal).strip() == "":
                continue  # 잔액 없는 행은 데이터 아님

            acct = g(r, c_acct)
            acct = str(acct).strip() if acct not in (None, "") else section
            name = g(r, c_name)
            result.append({
                "계정과목":  acct,
                "적요":      g(r, c_memo),
                "거래처명":  (str(name).strip() if name not in (None, "") else None),
                "전기이월":  None,
                "차변/증가": None,
                "대변/감소": None,
                "잔액":      bal,
                "비고":      (str(g(r, c_acc)).strip() if g(r, c_acc) not in (None, "") else None),
            })
    wb.close()
    return result
