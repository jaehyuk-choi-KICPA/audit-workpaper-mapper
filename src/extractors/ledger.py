"""거래처잔액 (계정별원장) 변형 파서.

회사마다 ERP 출력 형태가 다르므로 시트명 패턴으로 형식을 감지한 뒤
이상적 헤더로 변환한다.

지원 형식 (시트명 패턴 → 없으면 계정과목 컬럼으로 자동 감지):
  num_name_code  시트명 `N_계정명(코드)`  예: `0_보통예금(10300)`, `9_장기금융상품(17600)`
                 헤더: 코드|거래처명|전기(월)이월|증가|감소|잔액  (데이터 col0~)
  code_name      시트명 `(코드)계정명`     예: `(0103)보통예금`
                 헤더: 코드|거래처명|전기(월)이월|차변|대변|잔액   (데이터 col2~)
  account_col    단일/통합 시트(계정별 분리 안 됨, 예 '명세서') — 계정과목이 **행 컬럼**으로 존재.
                 헤더: 거래처코드|거래처|코드|계정과목명|전기(월)이월|차변|대변|잔액
                 사실상 이미 변환된 '거래처원장 총괄잔액' 출력. 시트명 패턴이 안 맞을 때 폴백.

출력 이상적 헤더:
  계정과목 | 적요 | 거래처명 | 전기이월 | 차변/증가 | 대변/감소 | 잔액 | 비고
  (거래처원장은 적요·비고가 없어 공란. 결산보고서 변환 시 적요·비고가 채워짐)

★ 계정과목 결정 원칙 (절대 규칙)
  계정과목은 **오직 시트명에서만** 도출한다. 원본 ERP 출력은 계정별로 시트가
  분리되어 있고(시트명에 계정과목 포함), 데이터 영역에는 계정과목 컬럼이 없다.
  따라서 행 단위 계정과목 컬럼을 신뢰하면 안 된다.
  (수동으로 여러 시트를 한 시트로 합치는 과정에서 계정과목 라벨이 어긋나는
   오류가 실제로 발생했음 → 시트명을 권위 있는 출처로 고정해 원천 차단)
"""

import re
from pathlib import Path

import openpyxl

try:
    import xlrd
    _HAS_XLRD = True
except ImportError:
    _HAS_XLRD = False


# ---------------------------------------------------------------------------
# 헤더 동의어 → 표준 컬럼명
# ---------------------------------------------------------------------------

from ._headers import normalize as _normalize, map_headers


_SYNONYMS: dict[str, str] = {
    "코드":         "코드",
    "전기(월)이월": "전기이월",
    "전기이월":     "전기이월",
    "거래처명":     "거래처명",
    "거래처":       "거래처명",   # '거래처'(이름)를 fuzzy가 '거래처코드'로 오인하지 않게 exact로 고정
    "증가":         "차변/증가",
    "차변":         "차변/증가",
    "감소":         "대변/감소",
    "대변":         "대변/감소",
    "잔액":         "잔액",
}

# 다중시트 형식에선 계정과목을 시트명에서만 도출한다(의도적). 단일/통합 시트(account_col)
# 형식에선 계정과목이 행 컬럼으로 존재하므로 아래 _SYNONYMS_ACCT로 그 컬럼을 인식한다.
_SYNONYMS_ACCT: dict[str, str] = {
    **_SYNONYMS,
    "계정과목":   "계정과목",
    "계정과목명": "계정과목",
    "계정":       "계정과목",
}

_REQUIRED = {"코드", "거래처명", "전기이월", "차변/증가", "대변/감소", "잔액"}

# account_col 형식은 계정과목 컬럼이 추가로 필수.
_REQUIRED_ACCT = _REQUIRED | {"계정과목"}

# 부분일치(fuzzy) 폴백용 키워드 — 정확매칭으로 못 찾은 필수 컬럼만 보강한다.
# 헤더 라벨이 ERP마다 조금씩 달라도(예: '기말잔액', '거 래 처', '당기증가') 흡수.
# 순서 = 구체적(specific) 우선. 정확매칭이 1차이므로 회귀 위험은 없다(폴백 전용).
_SUB_KEYWORDS: list[tuple] = [
    ("코드",      ["코드"]),
    ("거래처명",  ["거래처", "거래처명", "상호", "계좌명", "예금주"]),
    ("전기이월",  ["전기이월", "기초", "이월"]),
    ("차변/증가", ["차변", "증가", "입금"]),
    ("대변/감소", ["대변", "감소", "출금"]),
    ("잔액",      ["잔액", "기말"]),
]

# account_col 폴백용: 위 키워드 + 계정과목. 계정과목은 '거래처코드'·'코드'보다 뒤에 와도
# 무방하나(정확매칭이 코드를 먼저 점유), 구체적 라벨 우선 위해 별도 키워드를 둔다.
_SUB_KEYWORDS_ACCT: list[tuple] = _SUB_KEYWORDS + [
    ("계정과목", ["계정과목", "계정", "과목"]),
]

_TOTAL_MARKERS = {"합계"}  # _normalize 후 비교하므로 '합 계'도 매칭됨


# ---------------------------------------------------------------------------
# 형식 감지 및 계정과목 추출
# ---------------------------------------------------------------------------

# N_계정명(코드)  예: 9_장기금융상품(17600)
_RE_NUM_NAME_CODE = re.compile(r"^\d+_(.+)\(\d+\)$")
# (코드)계정명    예: (0103)보통예금
_RE_CODE_NAME = re.compile(r"^\(\d+\)(.+)$")

_FORMAT_RE = {
    "num_name_code": _RE_NUM_NAME_CODE,
    "code_name": _RE_CODE_NAME,
}


def _detect_format(sheet_names: list[str]) -> str:
    """시트명 패턴으로 'num_name_code' | 'code_name' | 'unknown' 반환."""
    a = sum(1 for n in sheet_names if _RE_NUM_NAME_CODE.match(n))
    b = sum(1 for n in sheet_names if _RE_CODE_NAME.match(n))
    if a > 0 and a >= b:
        return "num_name_code"
    if b > 0:
        return "code_name"
    return "unknown"


def _account_name(sheet_name: str, fmt: str) -> "str | None":
    """시트명에서 계정과목을 추출. 형식 패턴에 맞지 않으면 None (= 계정 시트 아님)."""
    rx = _FORMAT_RE.get(fmt)
    if rx is None:
        return None
    m = rx.match(sheet_name)
    return m.group(1).strip() if m else None


# ---------------------------------------------------------------------------
# 헤더 매핑
# ---------------------------------------------------------------------------

def _try_map_headers(header_row: list) -> "dict[str, int] | None":
    """헤더 행 시도. 필수 컬럼 미충족이면 None 반환 (예외 없음).

    1차: 정확 동의어 매칭(기존, 검증됨). 2차: 못 찾은 필수 컬럼만 부분일치(fuzzy) 폴백.
    """
    return map_headers(header_row, synonyms=_SYNONYMS, required=_REQUIRED,
                       sub_keywords=_SUB_KEYWORDS, parser_key="ledger")


# ---------------------------------------------------------------------------
# 합계행 판별
# ---------------------------------------------------------------------------

def _is_total(row: list, c_code: int, c_name: int) -> bool:
    for idx in (c_code, c_name):
        if idx < len(row) and row[idx] is not None:
            if _normalize(row[idx]) in _TOTAL_MARKERS:
                return True
    return False


# ---------------------------------------------------------------------------
# 행 → 표준 dict  (계정과목은 인자로 받은 시트명 기반 값을 그대로 사용)
# ---------------------------------------------------------------------------

def _to_record(row: list, account: str, col_map: dict[str, int]) -> dict:
    def g(col: str):
        idx = col_map.get(col)
        return row[idx] if idx is not None and idx < len(row) else None

    # 코드는 데이터 행 탐색용으로만 읽고 출력 헤더에는 포함하지 않는다.
    # 거래처원장에는 적요·비고가 없으므로 공란으로 둔다(결산보고서에서만 채워짐).
    return {
        "계정과목":  account,
        "적요":      None,
        "거래처명":  g("거래처명"),
        "전기이월":  g("전기이월"),
        "차변/증가": g("차변/증가"),
        "대변/감소": g("대변/감소"),
        "잔액":      g("잔액"),
        "비고":      None,
    }


# ---------------------------------------------------------------------------
# 헤더 행 탐색 (첫 12행 내에서 필수 컬럼을 모두 가진 첫 행)
# ---------------------------------------------------------------------------

def _find_header(get_row, nrows: int) -> "tuple[int, dict[str, int]] | None":
    for i in range(min(12, nrows)):
        cm = _try_map_headers(get_row(i))
        if cm is not None:
            return i, cm
    return None


# ---------------------------------------------------------------------------
# 단일 시트 파싱 (openpyxl / xlrd 공통 로직)
# ---------------------------------------------------------------------------

def _parse_sheet(get_row, nrows: int, account: str, sheet_name: str) -> list[dict]:
    found = _find_header(get_row, nrows)
    if found is None:
        raise ValueError(
            f"[거래처잔액 파서] 시트 '{sheet_name}': 헤더 행 미발견 (첫 12행 탐색)"
        )
    header_idx, col_map = found
    c_code = col_map["코드"]
    c_name = col_map["거래처명"]

    result = []
    for r in range(header_idx + 1, nrows):
        row = get_row(r)
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        if _is_total(row, c_code, c_name):
            continue
        # 거래처명·코드 모두 비어있으면 데이터 행 아님
        code = row[c_code] if c_code < len(row) else None
        name = row[c_name] if c_name < len(row) else None
        if (code is None or str(code).strip() == "") and \
           (name is None or str(name).strip() == ""):
            continue
        result.append(_to_record(row, account, col_map))
    return result


def _parse_account_col(sheets: list) -> list[dict]:
    """단일/통합 시트(계정과목이 행 컬럼) 폴백 파싱. 시트명 패턴이 안 맞을 때만 호출된다.

    계정과목을 **시트명이 아니라 행의 계정과목 컬럼**에서 도출한다. 이 형식은 ERP의
    '거래처원장 총괄잔액' 기계 출력(수기 병합본이 아님)이라 컬럼을 신뢰할 수 있다 —
    다중시트 형식의 '시트명만' 규칙(수기 병합 오류 방지)의 정당한 예외.

    sheets: [(시트명, get_row, nrows), ...]. 헤더가 매핑되는 시트만 사용(표지 등 무시).
    """
    out: list[dict] = []
    for name, get_row, nrows in sheets:
        found = None
        for i in range(min(12, nrows)):
            cm = map_headers(get_row(i), synonyms=_SYNONYMS_ACCT, required=_REQUIRED_ACCT,
                             sub_keywords=_SUB_KEYWORDS_ACCT, parser_key="ledger")
            if cm is not None:
                found = (i, cm)
                break
        if found is None:
            continue
        header_idx, col_map = found
        c_acct = col_map["계정과목"]
        c_code = col_map["코드"]
        c_name = col_map["거래처명"]
        for r in range(header_idx + 1, nrows):
            row = get_row(r)
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            if _is_total(row, c_code, c_name):
                continue
            acct = row[c_acct] if c_acct < len(row) else None
            if acct is None or str(acct).strip() == "":   # 계정과목 없는 행=소계/공백
                continue
            out.append(_to_record(row, str(acct).strip(), col_map))
    return out


# ---------------------------------------------------------------------------
# 검증 게이트
# ---------------------------------------------------------------------------

def _validate_result(rows: list[dict], path: str) -> None:
    if not rows:
        raise ValueError(f"[거래처잔액 파서] 파싱 결과 행 없음: {path}")

    errors = []
    for r in rows:
        prev = r["전기이월"] or 0
        dr   = r["차변/증가"] or 0
        cr   = r["대변/감소"] or 0
        bal  = r["잔액"]
        if bal is None:
            continue
        try:
            # 부호 무관: 자산형(기초+차변-대변=잔액)·부채/자본/수익형(기초+대변-차변=잔액) 중
            # 하나만 맞으면 정상. 총괄잔액(account_col)은 계정유형이 섞여 있어 필수.
            p, d, c, b = float(prev), float(dr), float(cr), float(bal)
            diff = min(abs(p + d - c - b), abs(p + c - d - b))
        except (TypeError, ValueError):
            continue
        if diff > 1:
            errors.append(
                f"  거래처 '{r['거래처명']}' ({r['계정과목']}): "
                f"전기이월{prev}±차변{dr}∓대변{cr}≠잔액{bal} (차이={diff:.0f})"
            )

    if errors:
        import warnings
        warnings.warn(
            f"[거래처잔액 파서] 잔액 불일치 {len(errors)}건 (검토 필요):\n"
            + "\n".join(errors[:10])
        )


# ---------------------------------------------------------------------------
# 공개 API
# ---------------------------------------------------------------------------

def parse_ledger(path: str) -> list[dict]:
    """거래처잔액 원장(회사제시 원본)을 파싱해 표준 행 리스트로 반환한다.

    계정과목은 **시트명에서만** 도출한다 (형식 패턴에 맞는 시트만 파싱).

    Returns:
        [{"계정과목", "코드", "거래처명", "전기이월", "차변/증가", "대변/감소", "잔액"}, ...]

    Raises:
        ValueError : 형식 미인식 / 헤더 미발견 / 파싱 결과 없음 / 미지원 확장자
        ImportError: .xls 처리 시 xlrd 미설치
    """
    p = Path(path)
    ext = p.suffix.lower()
    rows: list[dict] = []

    # ① 시트 접근자 [(시트명, get_row, nrows)]를 형식 무관하게 한 번에 수집.
    sheets: list = []
    if ext == ".xls":
        if not _HAS_XLRD:
            raise ImportError(".xls 처리에는 xlrd가 필요합니다: pip install xlrd")
        wb = xlrd.open_workbook(str(p))
        names = wb.sheet_names()
        for name in names:
            sh = wb.sheet_by_name(name)
            sheets.append((name, sh.row_values, sh.nrows))
    elif ext in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
        names = wb.sheetnames
        for name in names:
            data = list(wb[name].iter_rows(values_only=True))
            # 기본인자로 data 바인딩(루프 변수 늦은바인딩 방지).
            sheets.append((name, lambda i, d=data: list(d[i]), len(data)))
        wb.close()
    else:
        raise ValueError(f"지원하지 않는 파일 형식: {ext}")

    # ② 시트명 패턴(다중시트, 계정=시트명) 우선. 없으면 계정과목 컬럼(account_col) 폴백.
    fmt = _detect_format(names)
    if fmt != "unknown":
        for name, get_row, nrows in sheets:
            account = _account_name(name, fmt)
            if account is None:
                continue  # 형식 패턴에 맞지 않는 시트(표지·요약 등) 건너뜀
            rows.extend(_parse_sheet(get_row, nrows, account, name))
    else:
        rows = _parse_account_col(sheets)   # 단일/통합 시트(계정과목이 행 컬럼)
        if not rows:
            raise ValueError(f"[거래처잔액 파서] 시트명 형식·계정과목 컬럼 모두 미인식: {names[:5]}")

    _validate_result(rows, str(p))
    return rows


# ---------------------------------------------------------------------------
# 이상적 양식(중간다리) 입출력
#
#   원본(.xls/.xlsx, 계정별 다중시트) → parse_ledger → records
#   records → write_ideal_ledger → 단일 시트 표준 파일 (회계사 검수용 + 하위단계 입력)
#   단일 시트 표준 파일 → read_ideal_ledger → records (하위 생성기가 소비)
#
#   계정과목 컬럼은 parse_ledger가 시트명에서 생성한 값이므로 분류가 정확하다.
#   (수기 편집 산물이 아니므로, 표준 파일을 읽을 때는 행별 계정과목 컬럼을 신뢰한다)
# ---------------------------------------------------------------------------

_IDEAL_HEADER = ["계정과목", "적요", "거래처명", "전기이월", "차변/증가", "대변/감소", "잔액", "비고"]

_DEFAULT_IDEAL_SHEET = "거래처원장"


def write_ideal_ledger(rows: list[dict], output_path: str,
                       sheet_title: str = _DEFAULT_IDEAL_SHEET) -> Path:
    """parse_ledger / parse_settlement_report 결과를 단일 시트 '이상적 양식'으로 생성한다.

    헤더: 계정과목 | 적요 | 거래처명 | 전기이월 | 차변/증가 | 대변/감소 | 잔액 | 비고
    계정과목은 시트명(원장) 또는 과목 컬럼(결산보고서)에서 도출된 값이다.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(_IDEAL_HEADER)
    for r in rows:
        ws.append([r.get(k) for k in _IDEAL_HEADER])

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    wb.close()
    return out


def read_ideal_ledger(path: str) -> list[dict]:
    """이상적 양식(단일 시트, 행별 계정과목 컬럼)을 읽어 records로 반환한다.

    이 파일은 write_ideal_ledger로 코드 생성된 표준 형식이므로
    행별 계정과목 컬럼을 신뢰한다. (회계사 검수 시 수정 가능 = human-in-the-loop)
    """
    wb = openpyxl.load_workbook(str(Path(path)), read_only=True, data_only=True)
    ws = wb.active
    data = list(ws.iter_rows(values_only=True))
    wb.close()

    if not data:
        raise ValueError(f"[이상적 양식 리더] 빈 파일: {path}")

    # 헤더 위치 매핑 (공백 정규화)
    header = data[0]
    pos = {}
    for idx, cell in enumerate(header):
        key = _normalize(cell)
        if key in _IDEAL_HEADER and key not in pos:
            pos[key] = idx
    missing = set(_IDEAL_HEADER) - pos.keys()
    if missing:
        raise ValueError(f"[이상적 양식 리더] 헤더 누락 {missing}: {[str(c) for c in header]}")

    result = []
    for row in data[1:]:
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        rec = {k: (row[pos[k]] if pos[k] < len(row) else None) for k in _IDEAL_HEADER}
        # 계정과목·거래처명 모두 비면 데이터 행 아님
        if not (rec["계정과목"] or rec["거래처명"]):
            continue
        result.append(rec)
    return result
