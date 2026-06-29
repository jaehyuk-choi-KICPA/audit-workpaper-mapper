"""자본변동표(CE 시트) 파서 — 이익잉여금처분계산서 전기 칼럼·연도헤더 소스.

별도정산표 파일의 'CE' 시트(자본변동표)에서 **이익잉여금 열**의 전기/당기 흐름과
기수·기간 라벨을 추출한다. GG 자본조서 이익잉여금처분계산서의 **전기 칼럼**과 **제N기 연도헤더**가
이 값을 쓴다(당기 칼럼은 조서 내부수식=총괄표 링크 유지 — generators.capital 참고).

견고성: 시트/열/앵커가 없으면 예외 대신 None을 반환한다(자본조서 나머지는 정상 생성).
헤더 라벨('이익잉여금')은 _x000D_·개행이 섞여 들어와 정규화 후 매칭한다(다중 헤더행 스캔).
기간 앵커는 '(전기초)/(전기말)/(당기초)/(당기말)' 마커로 잡아 회사마다 행수가 달라도 안정적이다.
"""

import re
import warnings
from pathlib import Path

import openpyxl

_DEFAULT_SHEET = "CE"


def _n(s) -> str:
    """헤더/라벨 정규화 — _x000D_(CR 인코딩)·개행·공백 제거."""
    if s is None:
        return ""
    return re.sub(r"\s+", "", str(s).replace("_x000D_", "").replace("\r", ""))


def _num(v):
    return v if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def _parse_term(label):
    """'제 19(당) 기 2025년 01월 01일부터 2025년 12월 31일까지' → (기수, (시작, 종료))."""
    s = str(label or "")
    m = re.search(r"제\s*(\d+)", s)
    dates = re.findall(r"(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일", s)
    term = m.group(1) if m else None
    rng = None
    if len(dates) >= 2:
        (y1, m1, d1), (y2, m2, d2) = dates[0], dates[1]
        rng = (f"{y1}년 {int(m1)}월 {int(d1)}일", f"{y2}년 {int(m2)}월 {int(d2)}일")
    return term, rng


def parse_equity_changes(path: str, sheet_name: str = _DEFAULT_SHEET) -> "dict | None":
    """자본변동표(CE)에서 이익잉여금 흐름·기수 라벨을 추출한다.

    Returns:
        {"전기": {전기이월,당기순이익,연차배당,이익준비금적립,임의적립금이입,차기이월,기수,기간},
         "당기": {...동일...}}  (단위: 원, 그대로)
        시트/열/앵커 없으면 None.
    """
    p = Path(path)
    wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        warnings.warn(f"[CE 파서] 시트 '{sheet_name}' 없음.")
        return None
    rows = list(wb[sheet_name].iter_rows(values_only=True))
    wb.close()

    # 이익잉여금 열(다중 헤더행 스캔 — _x000D_·개행 정규화 후 매칭)
    col_re = None
    for row in rows[:12]:
        for j, c in enumerate(row):
            if "이익잉여금" in _n(c):
                col_re = j
                break
        if col_re is not None:
            break
    if col_re is None:
        warnings.warn("[CE 파서] 이익잉여금 열을 찾지 못함.")
        return None

    # 과목(라벨) 열 = 기간 앵커('(전기초)' 등)가 있는 열(이익잉여금 열 왼쪽, 보통 B)
    label_col = None
    for j in range(col_re):
        for r in range(len(rows)):
            cell = rows[r][j] if j < len(rows[r]) else None
            if cell and re.search(r"\((전기초|전기말|당기초|당기말)\)", str(cell)):
                label_col = j
                break
        if label_col is not None:
            break
    if label_col is None:
        warnings.warn("[CE 파서] 기간 앵커(전기초/전기말/당기초/당기말) 열을 찾지 못함.")
        return None

    def lab(r):
        return str(rows[r][label_col]) if label_col < len(rows[r]) and rows[r][label_col] is not None else ""

    def val(r):
        return _num(rows[r][col_re]) if col_re < len(rows[r]) else None

    # 기수/기간 라벨(제 N (당/전) 기 YYYY…) 두 줄
    term_labels = []
    for row in rows[:10]:
        for c in row:
            if c and re.search(r"제\s*\d+.*기.*\d{4}\s*년", str(c)):
                term_labels.append(str(c))
                break
    cur_term = _parse_term(term_labels[0]) if len(term_labels) >= 1 else (None, None)
    prev_term = _parse_term(term_labels[1]) if len(term_labels) >= 2 else (None, None)

    # 기간 앵커 행
    anchor = {}
    for r in range(len(rows)):
        nl = _n(lab(r))
        for key in ("전기초", "전기말", "당기초", "당기말"):
            if f"({key})" in nl:
                anchor[key] = r

    def block(start_key, end_key):
        if start_key not in anchor or end_key not in anchor:
            return None
        s, e = anchor[start_key], anchor[end_key]
        b = {"전기이월": val(s) or 0.0, "차기이월": val(e) or 0.0,
             "연차배당": 0.0, "당기순이익": 0.0, "이익준비금적립": 0.0, "임의적립금이입": 0.0}
        for r in range(s + 1, e):                      # 기초~기말 사이 증감행
            nm = _n(lab(r))
            v = val(r) or 0.0
            if "배당" in nm:
                b["연차배당"] += v
            elif "순이익" in nm or "순손익" in nm or "순손실" in nm:
                b["당기순이익"] += v
            elif "이익준비금" in nm:
                b["이익준비금적립"] += v
            elif "임의적립금" in nm or "이입" in nm:
                b["임의적립금이입"] += v
        return b

    prev = block("전기초", "전기말")
    cur = block("당기초", "당기말")
    if prev is None and cur is None:
        warnings.warn("[CE 파서] 전기/당기 블록을 구성하지 못함.")
        return None
    return {
        "전기": {**(prev or {}), "기수": prev_term[0], "기간": prev_term[1]},
        "당기": {**(cur or {}), "기수": cur_term[0], "기간": cur_term[1]},
    }
