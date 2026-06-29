# -*- coding: utf-8 -*-
"""금융기관조회서에서 '해약환급금' 보유 보험을 시트·위치 불문하고 장기금융상품으로 추출.

표준 케이스는 INSURANCE 1번 섹션(parse_insurance)이지만, 회사에 따라 해약환급금이
**다른 시트/다른 섹션**에 등장한다(실측: INSURANCE 시트가 아예 없고
GUARANTEE 보증 박스에 해약환급금이 있는 변형도 있음). 위치가 고정이 아니므로
시트 전체를 **라벨로 스캔**해 '해약환급금' 열을 찾고, 금액>0 행을 장기금융으로 끌어온다.

설계 원칙:
  · 라벨 우선(위치 하드코딩 금지) — '금융기관명'+'해약환급금' 라벨이 같은 헤더행에 있을 때만 인식.
  · 다음 섹션 마커(예: '4.')에서 데이터 수집 중단(섹션 침범 방지).
  · parse_insurance와 동일 스키마로 그룹을 만들 수 있게 per-row를 반환.
  · 못 찾으면 빈 리스트(예외 금지 — 부분실패 허용 원칙).
"""

import re
from pathlib import Path

import openpyxl

from ._headers import map_headers
from .insurance import parse_insurance

_SYN = {
    "금융기관명": "금융기관명", "금융기관": "금융기관명",
    "보험의 종류": "보험의종류", "보험의종류": "보험의종류",
    "상품의 종류": "보험의종류", "상품의종류": "보험의종류",
    "증권번호": "증권번호", "조서번호": "조서번호",
    "해약환급금_금액": "해약환급금_금액",
    "해약환급금": "해약환급금", "해지환급금": "해약환급금",
}
_REQ = {"금융기관명"}  # 해약환급금 컬럼은 매핑 후 별도 확인
_SUB = [
    ("금융기관명",     ["금융기관", "기관명"]),
    ("보험의종류",     ["상품의종류", "보험의종류", "상품종류", "보험종류", "종류"]),
    ("증권번호",       ["증권번호", "증권", "계약번호", "증서"]),
    ("해약환급금_금액", ["해약환급금_금액", "해지환급금_금액"]),
    ("해약환급금",     ["해약환급금", "해지환급금", "환급금"]),
]

_SECTION_RE = re.compile(r"^\s*\d+\.")


def _amt(v) -> float:
    if isinstance(v, bool):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = re.sub(r"[^\d.\-]", "", v)
        try:
            return float(s) if s else 0.0
        except ValueError:
            return 0.0
    return 0.0


def parse_surrender_value(path: str, *, exclude_sheets: tuple = ()) -> list[dict]:
    """모든 시트를 스캔해 '해약환급금'>0 보험을 장기금융 per-row로 반환.

    Returns: [{금융기관명, 보험의종류, 증권번호, 해약환급금, 조서번호}]
    (exclude_sheets: parse_insurance가 이미 처리한 INSURANCE 등을 빼 중복 방지)
    """
    try:
        wb = openpyxl.load_workbook(str(Path(path)), read_only=True, data_only=True)
    except Exception:
        return []
    out: list[dict] = []
    try:
        for sn in wb.sheetnames:
            if sn in exclude_sheets:
                continue
            rows = list(wb[sn].iter_rows(values_only=True))
            n = len(rows)
            i = 0
            while i < n:
                row = rows[i]
                # 헤더 후보: 이 행에 '해약환급금' 라벨이 있나? ('원금만 포함' 같은 주석 셀은 제외)
                is_hdr_candidate = any(
                    v is not None and "해약환급금" in str(v) and "원금" not in str(v)
                    for v in row)
                if is_hdr_candidate:
                    cm = map_headers(row, synonyms=_SYN, required=_REQ,
                                     sub_keywords=_SUB, parser_key="surrender")
                    amt_key = ("해약환급금_금액" if (cm and "해약환급금_금액" in cm)
                               else "해약환급금")
                    if cm and "금융기관명" in cm and amt_key in cm:
                        c_inst, c_amt = cm["금융기관명"], cm[amt_key]
                        c_kind = cm.get("보험의종류")
                        c_pol = cm.get("증권번호")
                        c_ref = cm.get("조서번호")

                        def g(r, idx):
                            return r[idx] if (idx is not None and idx < len(r)) else None

                        j = i + 1
                        while j < n:
                            rj = rows[j]
                            first = next((str(v).strip() for v in rj if v not in (None, "")), "")
                            if _SECTION_RE.match(first):
                                break  # 다음 섹션 마커 → 중단
                            inst = g(rj, c_inst)
                            amount = _amt(g(rj, c_amt))
                            if (inst not in (None, "") and str(inst).strip() not in ("", "' ")
                                    and amount > 0):
                                out.append({
                                    "금융기관명": str(inst).strip(),
                                    "보험의종류": str(g(rj, c_kind) or "").strip(),
                                    "증권번호": str(g(rj, c_pol) or "").strip(),
                                    "해약환급금": amount,
                                    "조서번호": (str(g(rj, c_ref)).strip()
                                              if g(rj, c_ref) not in (None, "") else None),
                                })
                            j += 1
                        i = j
                        continue
                i += 1
    finally:
        wb.close()
    return out


def parse_longterm_groups(path: str) -> list[dict]:
    """장기금융상품 그룹 = parse_insurance(INSURANCE) + 해약환급금 스캔(기타 시트), 중복 제거.

    parse_insurance와 동일 스키마를 반환하므로 기존 소비처(A200·총괄표 장기금융)에 그대로 쓸 수 있다.
    Returns: [{금융기관명, 보험의종류, 증권번호, 조회서금액, 건수, 조서번호}]
    """
    try:
        base = parse_insurance(path)
    except Exception:
        base = []
    # INSURANCE는 parse_insurance가 권위 처리 → 스캔에서 제외(중복 방지). 나머지 시트만 스캔.
    extra_rows = parse_surrender_value(path, exclude_sheets=("INSURANCE",))

    groups: dict[tuple, dict] = {}
    for r in extra_rows:
        key = (r["금융기관명"], r["보험의종류"])
        grp = groups.setdefault(key, {
            "금융기관명": r["금융기관명"], "보험의종류": r["보험의종류"],
            "증권번호목록": [], "조회서금액": 0, "건수": 0, "조서번호": None})
        grp["조회서금액"] += r["해약환급금"]
        grp["건수"] += 1
        if r["증권번호"]:
            grp["증권번호목록"].append(r["증권번호"])
        if grp["조서번호"] is None and r.get("조서번호"):
            grp["조서번호"] = r["조서번호"]

    result = list(base)
    for grp in groups.values():
        result.append({
            "금융기관명": grp["금융기관명"],
            "보험의종류": grp["보험의종류"],
            "증권번호": ", ".join(grp["증권번호목록"]),
            "조회서금액": grp["조회서금액"],
            "건수": grp["건수"],
            "조서번호": grp["조서번호"],
        })
    return result
