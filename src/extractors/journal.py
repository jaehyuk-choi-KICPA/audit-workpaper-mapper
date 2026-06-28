"""분개장(journal) → 이상적 양식 거래라인 추출.

회사마다 분개장 양식이 크게 다르다(실측 형태):
  A. 단일계정+차대: 월/일|번호|계정과목|적요|거래처|차변|대변      (한 행 = 한 라인)
  B. 이중계정(좌차/우대): [2단헤더 구분{차변,대변}] 월일|번호|계정|금액|계정|금액
                           (한 전표행 = 차변라인 + 대변라인 2개)
  C. 코드형: 년/월/일|번호|금액|계정과목|계정과목|금액|적요|코드|거래처명

헤더 위치·라벨이 회사마다 달라 **다중행 헤더 스캔 + 동의어(부분일치) 매핑(C-1차)**으로 열을 잡고,
필수열(계정과목·금액)을 못 잡으면 **빈 결과**를 반환(전체 흐름 안 막음). 라벨이 아예 다른 변형은
LLM(C-2차, format_adapt)으로 헤더만 좁게 질의하는 여지를 남긴다(여기선 훅만, 라이브 미연결).

이상적 출력 라인(거래 1건):
  {날짜, 번호, 계정과목, 거래처, 적요, 차변, 대변}
  - 차변/대변 중 한쪽만 금액(다른쪽 0/None). 이중계정 양식은 전표행을 두 라인으로 분해.
  - 계정과목의 [코드] 접두는 보존(예: '[13300]선급비용').
"""

import re
import warnings
from pathlib import Path

import openpyxl

# 필드별 동의어(부분일치). 구체적 라벨이 먼저 오도록.
_SYN = [
    ("날짜",   ["년/월/일", "년월일", "월/일", "일자", "거래일", "전표일자", "date"]),
    ("번호",   ["전표번호", "번호", "전표no", "no", "전표"]),
    ("계정과목", ["계정과목", "계정", "과목", "account"]),
    ("적요",   ["적요", "내용", "비고", "description", "메모"]),
    ("거래처",  ["거래처명", "거래처", "상대처", "거래상대", "vendor"]),
    ("차변",   ["차변", "출금", "debit"]),
    ("대변",   ["대변", "입금", "credit"]),
    ("금액",   ["금액", "amount"]),
    ("코드",   ["거래처코드", "코드", "code"]),
]
_HEADER_KW = {"계정", "과목", "차변", "대변", "금액", "월", "일자", "번호", "적요", "거래처", "구분"}


def _norm(s) -> str:
    return re.sub(r"\s+", "", str(s)).lower() if s is not None else ""


def _amt(v):
    """숫자 금액으로 해석(쉼표·괄호 음수 허용). 아니면 None."""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        t = v.replace(",", "").strip()
        if t in ("", "-"):
            return None
        neg = t.startswith("(") and t.endswith(")")
        t = t.strip("()")
        try:
            x = float(t)
            return -x if neg else x
        except ValueError:
            return None
    return None


def _field_of(label: str):
    """정규화 라벨 → 표준필드(부분일치, 구체적 우선). 없으면 None."""
    n = _norm(label)
    if not n:
        return None
    for field, kws in _SYN:
        for kw in kws:
            if _norm(kw) in n:
                return field
    return None


def _find_header(rows, scan=14):
    """헤더 행 인덱스와 열별 필드맵을 찾는다. 2단 헤더(상/하 행 라벨 병합)도 지원.

    Returns: (data_start_idx, col_fields) where col_fields = {col_idx: 표준필드}.
             못 찾으면 None.
    """
    best = None  # (점수, hdr_idx, col_fields, data_start)
    n = min(scan, len(rows))
    for i in range(n):
        row = rows[i]
        # 이 행 + 다음 행을 합친 라벨로 열별 필드 추정(2단 헤더 대비)
        nxt = rows[i + 1] if i + 1 < len(rows) else ()
        col_fields = {}
        for c in range(max(len(row), len(nxt))):
            top = row[c] if c < len(row) else None
            bot = nxt[c] if c < len(nxt) else None
            f = _field_of(top) or _field_of(bot)
            if f:
                col_fields[c] = f
        # 헤더다움 점수: 서로 다른 필드 종류 수 + 계정/금액 존재 가중
        fields = set(col_fields.values())
        if not fields:
            continue
        score = len(fields) + (2 if "계정과목" in fields else 0) \
            + (1 if ({"차변", "대변"} & fields) or "금액" in fields else 0)
        # 2단 헤더면 데이터는 i+2부터, 아니면 i+1부터(다음 행이 데이터처럼 보이면 i+1)
        two_stage = bool(_field_of_row_has_group(row))
        data_start = i + (2 if two_stage else 1)
        if best is None or score > best[0]:
            best = (score, i, col_fields, data_start)
    if best is None or best[0] < 2:
        return None
    return best[3], best[2]


def _field_of_row_has_group(row):
    """행이 그룹헤더(구분/차변/대변만)인지 — 2단 헤더 상단 판별용."""
    labels = [_norm(c) for c in row if c is not None and _norm(c)]
    if not labels:
        return False
    groupish = sum(1 for l in labels if l in ("구분", "차변", "대변", "차", "대"))
    return groupish >= 1 and len(labels) <= 4


def parse_journal(path: str) -> list[dict]:
    """분개장에서 거래라인을 추출. 시트/헤더 없으면 빈 리스트(예외 대신)."""
    p = Path(path)
    if p.suffix.lower() == ".xls":
        warnings.warn(f"[분개장] 구형 .xls는 미지원(.xlsx로 변환 필요): {p.name}")
        return []
    try:
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    except Exception as e:
        warnings.warn(f"[분개장] 로드 실패 {p.name}: {e}")
        return []
    # 분개장 시트 선택: '분개장' 포함 시트 우선, 없으면 첫 시트
    sheet = next((s for s in wb.sheetnames if "분개" in s), wb.sheetnames[0])
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []

    found = _find_header(rows)
    if found is None:
        warnings.warn(f"[분개장] 헤더 인식 실패: {p.name}")
        return []
    data_start, col_fields = found

    # ── 데이터 타입 프로파일링 ── 라벨이 부정확한 변형(일부 양식: '계정과목' 라벨인데 값이 숫자) 보정.
    # 표본 데이터 행에서 각 열의 숫자비율을 재고, 라벨과 데이터가 어긋나면 데이터를 신뢰한다.
    sample = [r for r in rows[data_start:data_start + 40] if r and any(v not in (None, "") for v in r)]
    maxc = max((len(r) for r in sample), default=0)

    def numeric_ratio(c):
        vals = [r[c] for r in sample if c < len(r) and r[c] not in (None, "")]
        if not vals:
            return 0.0
        return sum(1 for v in vals if _amt(v) is not None) / len(vals)

    def text_acctish(c):
        vals = [r[c] for r in sample if c < len(r) and r[c] not in (None, "")]
        if not vals:
            return 0.0
        return sum(1 for v in vals if _amt(v) is None and re.search(r"[가-힣A-Za-z]", str(v))) / len(vals)

    def cols(field):
        return [c for c, f in sorted(col_fields.items()) if f == field]

    # 계정과목 라벨열 중 실제 텍스트(계정)인 것만 account, 숫자인 것은 amount로 재분류
    acct_cols, reclassed_amt = [], []
    for c in cols("계정과목"):
        if numeric_ratio(c) > 0.6:
            reclassed_amt.append(c)
        else:
            acct_cols.append(c)
    debit_cols = [c for c in cols("차변") if numeric_ratio(c) > 0.4]
    credit_cols = [c for c in cols("대변") if numeric_ratio(c) > 0.4]
    amt_cols = sorted(set([c for c in cols("금액") + reclassed_amt if numeric_ratio(c) > 0.4]))
    date_c = (cols("날짜") or [None])[0]
    no_c = (cols("번호") or [None])[0]
    # 적요/거래처 라벨열이 실제 텍스트인지 확인(라벨-데이터 어긋남 보정)
    desc_c = next((c for c in cols("적요") if text_acctish(c) > 0.3), None)
    vend_c = next((c for c in cols("거래처") if text_acctish(c) > 0.3), None)

    def g(row, c):
        return row[c] if c is not None and c < len(row) else None

    # 단일계정인데 차/대 라벨이 없고 계정 좌우로 금액열이 있으면 좌=차변·우=대변(좌측금액형)
    if len(acct_cols) == 1 and not debit_cols and not credit_cols and len(amt_cols) >= 2:
        ac = acct_cols[0]
        left = [c for c in amt_cols if c < ac]
        right = [c for c in amt_cols if c > ac]
        if left and right:
            debit_cols = [max(left)]
            credit_cols = [min(right)]
            amt_cols = []

    _HDR_WORDS = {"계정과목", "계정", "과목", "년/월/일", "월/일", "일자", "번호", "차변", "대변", "금액"}
    out = []
    last_date = last_no = None
    for row in rows[data_start:]:
        if row is None or all(v in (None, "") for v in row):
            continue
        # 반복 헤더 행 스킵(계정/날짜 열 값이 헤더 라벨이면)
        if _norm(g(row, acct_cols[0] if acct_cols else None)) in _HDR_WORDS \
           or _norm(g(row, date_c)) in _HDR_WORDS:
            continue
        date = g(row, date_c) or last_date
        no = g(row, no_c) or last_no
        if g(row, date_c):
            last_date = g(row, date_c)
        if g(row, no_c):
            last_no = g(row, no_c)
        desc = g(row, desc_c)
        vend = g(row, vend_c)

        def emit(acct, dr, cr):
            acct = str(acct).strip() if acct is not None else ""
            if not acct and not dr and not cr:
                return
            if not acct or re.search(r"합\s*계|소\s*계|^계$", acct):
                return
            out.append({
                "날짜": date, "번호": no, "계정과목": acct,
                "거래처": str(vend).strip() if vend else None,
                "적요": str(desc).strip() if desc else None,
                "차변": dr or 0, "대변": cr or 0,
            })

        if len(acct_cols) >= 2:
            # 이중계정: 계정[0]=차변 라인, 계정[1]=대변 라인. 각 계정의 금액은 **거리상 최근접**
            # 금액열(좌/우 무관 — 유형A=계정|금액 우측, 유형B=금액|계정 좌측)을 중복없이 배정.
            pool = sorted(set(amt_cols + debit_cols + credit_cols))
            used = set()
            paired = []
            for ac in acct_cols[:2]:
                cand = sorted([c for c in pool if c not in used], key=lambda c: abs(c - ac))
                amc = cand[0] if cand else None
                if amc is not None:
                    used.add(amc)
                paired.append(_amt(g(row, amc)) if amc is not None else None)
            emit(g(row, acct_cols[0]), paired[0], 0)
            emit(g(row, acct_cols[1]), 0, paired[1])
        else:
            ac = acct_cols[0] if acct_cols else None
            dr = _amt(g(row, debit_cols[0])) if debit_cols else None
            cr = _amt(g(row, credit_cols[0])) if credit_cols else None
            if dr is None and cr is None and amt_cols:   # 금액 단일열 양식
                dr = _amt(g(row, amt_cols[0]))
            emit(g(row, ac), dr, cr)
    return out


def _nearest_amt(row, acct_col, amt_cols):
    """계정과목 열의 오른쪽에서 가장 가까운 금액열 값(이중계정 라인 금액)."""
    cands = sorted([c for c in amt_cols if c > acct_col])
    for c in cands:
        v = _amt(row[c]) if c < len(row) else None
        if v is not None:
            return v
    # 오른쪽에 없으면 전체에서
    for c in sorted(amt_cols):
        v = _amt(row[c]) if c < len(row) else None
        if v is not None:
            return v
    return None
