"""별도정산표(trial balance) 파서.

정산표 파일의 **별도정산표 시트만** 사용한다. 이 시트는 계정과목별 1행으로
기초(전기말 수정후) · 기말(당기 회사제시) · 당기수정(차변/대변) · 수정후 잔액을 담는다.
(당기수정 차변/대변은 별도정산표가 이미 '수정사항집계' 시트를 SUMIF로 물고 있어 부호가
정산표 규약대로 계산돼 있다 → 코드는 그 값을 그대로 읽는다.)

출력(이상적 양식 records):
  계정명 | 대분류 | 기초 | 기말 | 수정차변 | 수정대변 | 수정사항 | 수정후
  - 계정명  = 회사사용 계정과목(별도정산표 B열) → 총괄표 '회사사용 계정과목'
  - 대분류  = 공시용 계정과목(별도정산표 C열) → 총괄표 '대분류'(계정→조서 라우팅 키)
  - 수정사항 = 수정차변 - 수정대변 (편의 필드; 부호는 정산표 규약 그대로)

헤더 라벨이 회사/연도마다 조금씩 달라(날짜가 헤더에 들어감) 위치 탐지는 **안정적 구조 라벨**
('공시용'·'차변'·'대변'·'Audited'·'Unaudited')에 앵커한다. 셀값이 없거나 시트가 없으면
예외 대신 빈 리스트를 반환한다(한 파서의 빈 결과가 전체 생성을 막지 않도록).
"""

import warnings
from pathlib import Path

import openpyxl

from ._headers import normalize
from .settlement import _is_subtotal

_DEFAULT_SHEET = "별도정산표"


def _is_amount(v) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _find_header(rows: list[tuple]) -> "tuple[int, dict] | None":
    """헤더 행 인덱스와 {필드: 열인덱스} 매핑을 찾는다(0-indexed).

    별도정산표는 전기/당기 두 블록이 있어 '차변'/'대변'·'Unaudited'·'Audited'가 여러 개 나온다.
    **당기 수정후 = 마지막 Audited 열**에 앵커해 역추적한다(회사 무관 — 초도감사처럼 전기가
    Unaudited여도, Unaudited가 여러 개여도 안전):
      수정후 = 마지막 Audited / 당기 차변·대변 = 수정후 직전의 마지막 차변·대변 /
      기말(회사제시) = 당기 차변 직전의 마지막 Unaudited / 기초 = 기말 바로 왼쪽 열(전기말;
      Audited이거나 초도감사면 Unaudited) / 대분류 = '공시용' / 계정명 = 대분류 바로 왼쪽.
    (과거: '첫 Unaudited'를 기말로 잡아 초도감사 회사에서 전전기 열을 잡고 기초 Audited 못 찾아 실패.)
    """
    def has(n, key):
        return key.lower() in n.lower()

    for i, row in enumerate(rows):
        norms = [normalize(c) for c in row]
        col_대분류 = next((j for j, n in enumerate(norms) if "공시용" in n), None)
        if col_대분류 is None:
            continue
        audited = [j for j, n in enumerate(norms) if has(n, "audited") and not has(n, "unaudited")]
        unaud = [j for j, n in enumerate(norms) if has(n, "unaudited")]
        if not audited or not unaud:
            continue
        col_수정후 = max(audited)                          # 마지막 Audited = 당기 수정후
        cha = [j for j, n in enumerate(norms) if n == "차변" and j < col_수정후]
        dae = [j for j, n in enumerate(norms) if n == "대변" and j < col_수정후]
        col_차변 = max(cha) if cha else None
        col_대변 = max(dae) if dae else None
        lim = col_차변 if col_차변 is not None else col_수정후
        u_before = [j for j in unaud if j < lim]
        col_기말 = max(u_before) if u_before else None    # 당기 차변 직전 마지막 Unaudited
        col_기초 = col_기말 - 1 if col_기말 else None       # 당기 기초 = 기말 바로 왼쪽(전기말)
        col_계정명 = col_대분류 - 1 if col_대분류 > 0 else None
        cm = {
            "계정명": col_계정명, "대분류": col_대분류,
            "기초": col_기초, "기말": col_기말,
            "수정차변": col_차변, "수정대변": col_대변, "수정후": col_수정후,
        }
        if all(v is not None for v in cm.values()):
            return i, cm
    return None


def parse_trial_balance(path: str, sheet_name: str = _DEFAULT_SHEET) -> list[dict]:
    """별도정산표에서 계정과목별 행을 추출한다.

    Returns:
        [{"계정명","대분류","기초","기말","수정차변","수정대변","수정사항","수정후"}, ...]
        (시트/헤더 없으면 빈 리스트)
    """
    p = Path(path)
    wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        warnings.warn(f"[별도정산표 파서] 시트 '{sheet_name}' 없음. 존재: {wb.sheetnames}")
        return []
    rows = list(wb[sheet_name].iter_rows(values_only=True))
    wb.close()

    found = _find_header(rows)
    if found is None:
        warnings.warn("[별도정산표 파서] 헤더 행(공시용/차변/대변)을 찾지 못함.")
        return []
    header_idx, cm = found

    def g(row, key):
        j = cm[key]
        return row[j] if j is not None and j < len(row) else None

    result = []
    # 섹션 추적 → 상세행에 매출/매출원가/판관비/유형자산/무형자산 플래그(회사마다 계정명이 달라도 섹션
    # 통째로 조서에 편입: P=매출, Q=매출원가, R=판관비, G=유형·무형자산). 헤더('Ⅰ.매출액','Ⅱ.매출원가',
    # 'Ⅲ.매출총이익','Ⅳ.판매비와관리비','(2)유형자산','(3)무형자산','(4)기타비유동자산')는 계정명(B)에만
    # 있고 대분류(C)는 비어 있다. (제조원가명세서 플래그와 동일 패턴.)
    판관비_section = 매출_section = 매출원가_section = False
    유형자산_section = 무형자산_section = False
    재고_section = 자본_section = 영업외수익_section = 영업외비용_section = False
    유동부채_section = 비유동부채_section = False
    for row in rows[header_idx + 1:]:
        대분류 = g(row, "대분류")
        계정명 = g(row, "계정명")
        _nm = str(계정명 or "")
        nn = _nm.replace(" ", "")
        _hdr = (대분류 is None or str(대분류).strip() == "")
        # 섹션 헤더 전환(헤더는 대분류 비어 있음)
        if "판매비와관리비" in nn or "판매비및관리비" in nn:
            판관비_section, 매출_section, 매출원가_section = True, False, False
            continue
        if _hdr and nn.endswith("매출액"):
            매출_section, 매출원가_section, 판관비_section = True, False, False
            continue
        if _hdr and nn.endswith("매출원가"):
            매출원가_section, 매출_section, 판관비_section = True, False, False
            continue
        if "매출총이익" in nn or "매출총손익" in nn:
            매출_section = 매출원가_section = False
        if 판관비_section and any(k in nn for k in ("영업손익", "영업이익", "영업외", "영업비용")):
            판관비_section = False
        # 비유동자산: (2)유형자산~(3)무형자산~(4)기타비유동자산. G 조서가 유·무형 통째로 렌더.
        if _hdr and "무형자산" in nn:
            무형자산_section, 유형자산_section = True, False
            continue
        if _hdr and "유형자산" in nn:
            유형자산_section, 무형자산_section = True, False
            continue
        if _hdr and any(k in nn for k in ("기타비유동", "투자자산", "유동부채", "부채총계", "자본")):
            유형자산_section = 무형자산_section = False
        # 재고자산: (2)재고자산~(2)유형자산/비유동자산. E 조서가 통째로 렌더.
        if _hdr and "재고자산" in nn and "평가" not in nn:
            재고_section = True
            continue
        if 재고_section and _hdr and any(k in nn for k in ("비유동자산", "투자자산", "유형자산")):
            재고_section = False
        # 자본: '자본'~'자본총계'(Ⅰ.자본금/잉여금/기타포괄손익/이익잉여금 포함). GG 소유(분개·완전성).
        if _hdr and ("자본총계" in nn or "부채와자본총계" in nn):
            자본_section = False
        elif _hdr and "자본" in nn:
            자본_section = True
        # 부채: 'Ⅰ.유동부채'/'Ⅱ.비유동부채'~자본/부채총계 직전(매입채무·차입금·미지급금·충당부채 등).
        # CC가 다른 부채조서(AA/BBDD/EE)가 안 가져간 '잔여 부채'를 유동/비유동 구분해 흡수(완전성·분개).
        # '비유동부채'가 '유동부채'를 부분포함하므로 비유동을 먼저 본다.
        if _hdr and ("자본" in nn or "부채총계" in nn or "부채와자본총계" in nn):
            유동부채_section = 비유동부채_section = False
        elif _hdr and "비유동부채" in nn:
            비유동부채_section, 유동부채_section = True, False
        elif _hdr and "유동부채" in nn:
            유동부채_section, 비유동부채_section = True, False
        # 영업외수익/비용: Ⅵ.영업외수익~Ⅶ.영업외손실(비용)~법인세차감전. S가 통째로 렌더(2섹션).
        if _hdr and ("영업외손실" in nn or "영업외비용" in nn):
            영업외비용_section, 영업외수익_section = True, False
        elif _hdr and "영업외수익" in nn:
            영업외수익_section, 영업외비용_section = True, False
        if _hdr and any(k in nn for k in ("법인세", "당기순이익", "당기순손익", "총포괄", "기타포괄손익")):
            영업외수익_section = 영업외비용_section = False
        # 상세행만: 대분류(공시용) 있고, 소계/합계/섹션머리 아님
        if 대분류 is None or str(대분류).strip() == "":
            continue
        if _is_subtotal(대분류) or _is_subtotal(계정명):
            continue
        기초 = g(row, "기초"); 기말 = g(row, "기말")
        수정차변 = g(row, "수정차변"); 수정대변 = g(row, "수정대변")
        수정후 = g(row, "수정후")
        dr = 수정차변 if _is_amount(수정차변) else 0
        cr = 수정대변 if _is_amount(수정대변) else 0
        # 수정사항 = 수정후(M) - 기말(I). 별도정산표는 자산은 M=I+K-L, 부채/자본은 M=I+L-K로
        # 부호를 반영해 M을 계산하므로, M-I가 계정유형 무관하게 부호까지 정확하다(K-L은 자산에만 맞음).
        adj = None
        if _is_amount(수정후) and _is_amount(기말):
            adj = round(수정후) - round(기말)
        else:
            adj = dr - cr
        result.append({
            "계정명": str(계정명).strip() if 계정명 is not None else None,
            "대분류": str(대분류).strip(),
            "판관비": 판관비_section,
            "매출": 매출_section,
            "매출원가": 매출원가_section,
            "유형자산": 유형자산_section,
            "무형자산": 무형자산_section,
            "재고": 재고_section,
            "자본": 자본_section,
            "유동부채": 유동부채_section,
            "비유동부채": 비유동부채_section,
            "부채": 유동부채_section or 비유동부채_section,
            "영업외수익": 영업외수익_section,
            "영업외비용": 영업외비용_section,
            "기초": 기초 if _is_amount(기초) else None,
            "기말": 기말 if _is_amount(기말) else None,
            "수정차변": dr,
            "수정대변": cr,
            "수정사항": adj,
            "수정후": 수정후 if _is_amount(수정후) else None,
        })

    # ---- 제조원가명세서 섹션(별도정산표 하단) ----
    # 대분류(C열) 없이 계정명(B)만 있고 컬럼은 동일(H전기/I당기/M수정후). Q 매출원가·R-2 인건비 노무비
    # 소스. 대분류=계정명으로 부여해 동일 엔진/소유맵으로 라우팅(예: 당기총공사비용, 급여_제조).
    mfg_idx = None
    name_col = cm["계정명"]
    for i, row in enumerate(rows):
        v = row[name_col] if name_col < len(row) else None
        if v is not None and "제조원가명세서" in str(v):
            mfg_idx = i
            break
    if mfg_idx is not None:
        for row in rows[mfg_idx + 1:]:
            계정명 = g(row, "계정명")
            if 계정명 is None or str(계정명).strip() == "" or _is_subtotal(계정명):
                continue
            기말 = g(row, "기말"); 수정후 = g(row, "수정후"); 기초 = g(row, "기초")
            if not (_is_amount(기말) or _is_amount(수정후)):
                continue
            adj = (round(수정후) - round(기말)) if (_is_amount(수정후) and _is_amount(기말)) else 0
            nm = str(계정명).strip()
            result.append({
                "계정명": nm, "대분류": nm, "제조원가": True,
                "기초": 기초 if _is_amount(기초) else None,
                "기말": 기말 if _is_amount(기말) else None,
                "수정차변": 0, "수정대변": 0, "수정사항": adj,
                "수정후": 수정후 if _is_amount(수정후) else None,
            })

    # ---- 매출원가 집계요소(손익계산서 매출원가 섹션, 대분류 None 하위행) ----
    # 각 매출원가 종류(상품/제품/용역매출원가)의 하위행을 '상위'(그 종류)와 함께 캡처해 Q가 종류별로
    # 집계를 펼치게 한다. 상품/제품=기초재고+투입−기말재고 롤포워드, 용역=구성요소(용역수수료 등) 합.
    # 'Ⅱ.매출원가'~'Ⅲ.매출총이익' 구간만, 대분류 None 행만(대분류 있는 종류행은 본문서 매출원가=True).
    in_mc = False
    mc_parent = None
    for row in rows[header_idx + 1:]:
        계정명 = g(row, "계정명"); 대분류 = g(row, "대분류")
        nn = str(계정명 or "").replace(" ", "")
        has_cls = 대분류 is not None and str(대분류).strip() != ""
        if (not has_cls) and nn.endswith("매출원가"):           # 'Ⅱ. 매출원가' 섹션 헤더
            in_mc, mc_parent = True, None
            continue
        if "매출총이익" in nn or "매출총손익" in nn or ((not has_cls) and "판매비와관리비" in nn):
            in_mc = False
        if not in_mc:
            continue
        if has_cls:                                            # 매출원가 종류 행(상품/제품/용역매출원가)
            mc_parent = str(대분류).strip()
            continue
        if not nn or _is_subtotal(계정명) or "원가율" in nn:    # 소계·원가율 등 스킵
            continue
        기초 = g(row, "기초"); 기말 = g(row, "기말"); 수정후 = g(row, "수정후")
        if not (_is_amount(기말) or _is_amount(수정후)):
            continue
        adj = (round(수정후) - round(기말)) if (_is_amount(수정후) and _is_amount(기말)) else 0
        nm = str(계정명).strip()
        result.append({
            "계정명": nm, "대분류": nm, "원가상세": True, "상위": mc_parent,
            "기초": 기초 if _is_amount(기초) else None,
            "기말": 기말 if _is_amount(기말) else None,
            "수정차변": 0, "수정대변": 0, "수정사항": adj,
            "수정후": 수정후 if _is_amount(수정후) else None,
        })
    return result
