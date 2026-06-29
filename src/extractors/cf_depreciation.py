# -*- coding: utf-8 -*-
"""CF(정산표)_25 시트 → 자산별 당기 감가상각비 추출.

유무형자산 총괄표(G100)의 감가상각비(J열)·주석 변동표는 **자산분류별 당기 감가상각비**가
필요하다. 정산표의 `CF(정산표)_25` 시트(현금흐름 정산표)는 각 감가상각누계액 계정 옆
'계정분석/비용의 가산' 칼럼에 자산별 감가상각비를 이미 분해해 둔다(실측):

  B(계정과목)            … I(계정분석)   J(비용의 가산)
  감가상각누계액_건물          감가상각비      436,608,170
  감가상각누계액_차량운반구    감가상각비      300,522,933
  소프트웨어                  무형고정자산상각  98,828,915

→ {자산계정: 감가상각비} 반환(계정명은 '감가상각누계액_' 접두 제거해 건물/차량운반구/…로).
시트/칼럼 없으면 빈 dict(전체 흐름 안 막음).
"""

import re
import warnings
from pathlib import Path

import openpyxl


def _norm(s) -> str:
    return re.sub(r"\s+", "", str(s)) if s is not None else ""


def _amt(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        t = v.replace(",", "").strip().strip("()")
        try:
            return float(t)
        except ValueError:
            return None
    return None


def _clean_acct(name: str) -> str:
    """'감가상각누계액_건물' / '건물 감가상각누계액' → '건물'."""
    t = re.sub(r"감가상각누계액|상각누계액", "", str(name))
    return t.strip(" _\t").strip()


def parse_cf_depreciation(settlement: str) -> dict:
    """정산표 CF(정산표)_25 → {자산계정: 당기감가상각비(float)}."""
    try:
        wb = openpyxl.load_workbook(str(settlement), data_only=True, read_only=True)
    except Exception as e:
        warnings.warn(f"[CF감가상각비] 로드 실패: {e}")
        return {}
    # 'CF(정산표)' 우선, 없으면 'CF'+'정산표' 포함 시트
    sheet = next((s for s in wb.sheetnames if "cf(정산표)" in s.lower().replace(" ", "")), None) \
        or next((s for s in wb.sheetnames if "cf" in s.lower() and "정산표" in s), None)
    if sheet is None:
        wb.close()
        return {}
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()

    # 헤더 행: '비용의 가산' 셀이 있는 행. 그 칼럼 = 금액열, 왼쪽 = 계정분석(라벨)열.
    amt_col = label_col = acct_col = None
    for row in rows[:15]:
        for c, v in enumerate(row):
            n = _norm(v)
            if "비용의가산" in n:
                amt_col, label_col = c, c - 1
            if n == "계정과목":
                acct_col = c
        if amt_col is not None:
            break
    if amt_col is None:
        return {}
    if acct_col is None:
        acct_col = 1  # 통상 B열

    out: dict[str, float] = {}
    for row in rows:
        if amt_col >= len(row):
            continue
        label = _norm(row[label_col]) if 0 <= label_col < len(row) else ""
        if "상각" not in label:
            continue
        amt = _amt(row[amt_col])
        if not amt:
            continue
        acct = _clean_acct(row[acct_col] if acct_col < len(row) else "")
        if not acct:
            continue
        out[acct] = out.get(acct, 0.0) + amt
    return out
