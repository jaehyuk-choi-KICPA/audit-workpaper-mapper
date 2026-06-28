"""금융기관조회서 취합엑셀 INVESTMENT 시트 파서.

증권사 회신에서 두 섹션을 추출한다:
  - **1번(요약)**: 금융상품 금액 + **예수금**(증권사 예수금 = 현금성자산 하위 항목).
  - **2번(상세명세)**: 종목별 **평가액**(시가). 단기금융상품의 회신금액(=평가액 합)으로 쓴다.

A-200에서:
  - 예수금(1번 K열) → 현금성자산 섹션의 '예수금' 행.
  - 단기금융상품 회신금액(O) = 2번 평가액의 **계좌(금융기관)별 합**. 회사계상액(거래처원장)과
    다르면 (*) '단기금융상품 미평가 의심' 각주.

시트 구조: BANK/INSURANCE와 동일한 섹션 패턴 (마커 col1, 번호 섹션).
1번 헤더(예): 조서번호 | … | 금융기관명 | … | 계좌번호 | 금융상품의 종류 | 금융상품 금액 | 예수금 | …
2번 헤더(예): … | 금융기관명 | … | 계좌번호 | 종목명 | 액면금액 | 기준가 | 평가액 | …
"""

import re
from pathlib import Path

import openpyxl

from ._sections import find_section_bounds

# 1번(요약) 섹션 헤더 동의어
_SYNONYMS: dict[str, str] = {
    "조서번호":       "조서번호",
    "금융기관명":     "금융기관명",
    "계좌번호":       "계좌번호",
    "금융상품의 종류": "금융상품종류",
    "금융상품의종류":  "금융상품종류",
    "금융상품 금액":   "금액",
    "금융상품금액":    "금액",
    "예수금":         "예수금",
}
_REQUIRED = {"금융기관명", "금융상품종류", "금액"}

# 2번(상세명세) 섹션 헤더 동의어 — 평가액(시가) 추출용
_EVAL_SYNONYMS: dict[str, str] = {
    "금융기관명": "금융기관명",
    "계좌번호":   "계좌번호",
    "종목명":     "종목명",
    "평가액":     "평가액",
    "평가금액":   "평가액",
}
_EVAL_REQUIRED = {"금융기관명", "평가액"}

_DEFAULT_SHEET = "INVESTMENT"


def _map_headers(row: tuple, synonyms: dict = None, required: set = None) -> dict[str, int]:
    synonyms = synonyms or _SYNONYMS
    required = required or _REQUIRED
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(row):
        if cell is None:
            continue
        std = synonyms.get(str(cell).strip())
        if std and std not in mapping:
            mapping[std] = idx
    return mapping if required.issubset(mapping.keys()) else {}


# ★ 예외조항: INVESTMENT 종목명이 '확정기여형(DC)'이면 **외부적립 퇴직급여**라
#   현금성/단기금융 어디에도 넣지 않는다(퇴직급여 조서 소관). 평가액·예수금 모두 제외.
_DC_RE = re.compile(r"확정기여|외부적립|\(DC\)|DC형|디씨형")


def _is_dc(s) -> bool:
    return bool(s) and bool(_DC_RE.search(str(s)))


def _dc_accounts(rows) -> set:
    """2번(상세명세) 종목명이 DC인 계좌번호 집합 — 그 계좌는 1번 예수금도 제외 대상."""
    b = find_section_bounds(rows, "2.")
    if b is None:
        return set()
    si, se = b
    hi = cm = None
    for i in range(si + 1, min(si + 6, se)):
        m = _map_headers(rows[i], _EVAL_SYNONYMS, _EVAL_REQUIRED)
        if m:
            hi, cm = i, m
            break
    if cm is None or "종목명" not in cm:
        return set()
    accs, last_acc = set(), None
    ai, ji = cm.get("계좌번호"), cm["종목명"]
    for i in range(hi + 1, se):
        row = rows[i]
        if ai is not None and ai < len(row) and row[ai] not in (None, ""):
            last_acc = str(row[ai]).strip()
        if ji < len(row) and _is_dc(row[ji]) and last_acc:
            accs.add(last_acc)
    return accs


def _to_amount(value) -> float:
    if isinstance(value, bool):
        return 0
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        s = re.sub(r"[^\d.\-]", "", value)
        try:
            return float(s) if s else 0
        except ValueError:
            return 0
    return 0


def _load_rows(path: str, sheet_name: str):
    p = Path(path)
    wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return None
    rows = list(wb[sheet_name].iter_rows(values_only=True))
    wb.close()
    return rows


def parse_invest(path: str, sheet_name: str = _DEFAULT_SHEET) -> list[dict]:
    """INVESTMENT 1번(요약) 섹션의 계좌별 행을 반환한다.

    Returns:
        [{"조서번호","금융기관명","계좌번호","금융상품종류","금액","예수금"}, ...]
        (시트 없으면 빈 리스트 — 단기금융 미보유 회사 대응)
        예수금 = 증권사 예수금(현금성자산 하위) — 없으면 0.
    """
    rows = _load_rows(path, sheet_name)
    if rows is None:
        return []

    bounds = find_section_bounds(rows, "1.")
    if bounds is None:
        return []
    sec_idx, sec_end = bounds

    header_idx, col_map = None, None
    for i in range(sec_idx + 1, min(sec_idx + 6, sec_end)):
        cm = _map_headers(rows[i])
        if cm:
            header_idx, col_map = i, cm
            break
    if col_map is None:
        return []

    def g(row, key):
        idx = col_map.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    dc_acc = _dc_accounts(rows)        # DC(외부적립퇴직급여) 계좌 → 예수금도 제외
    result = []
    for i in range(header_idx + 1, sec_end):
        row = rows[i]
        inst = g(row, "금융기관명")
        if inst is None or str(inst).strip() == "":
            continue
        acc = g(row, "계좌번호")
        if (acc is not None and str(acc).strip() in dc_acc) or _is_dc(g(row, "금융상품종류")):
            continue                   # 확정기여형(DC) 계좌 → 현금성/예수금 제외
        result.append({
            "조서번호":     (str(g(row, "조서번호")).strip() if g(row, "조서번호") else None),
            "금융기관명":   str(inst).strip(),
            "계좌번호":     acc,
            "금융상품종류": g(row, "금융상품종류"),
            "금액":         _to_amount(g(row, "금액")),
            "예수금":       _to_amount(g(row, "예수금")),
        })
    return result


def parse_invest_eval(path: str, sheet_name: str = _DEFAULT_SHEET) -> list[dict]:
    """INVESTMENT 2번(상세명세) 섹션의 종목 평가액을 **계좌(금융기관)별로 합산**해 반환.

    Returns:
        [{"금융기관명","계좌번호","평가액"}, ...]  (평가액 = 계좌별 종목 평가액 합)
        2번 섹션/헤더 없으면 빈 리스트. 단기금융상품 회신금액(O)의 소스.
    """
    rows = _load_rows(path, sheet_name)
    if rows is None:
        return []

    bounds = find_section_bounds(rows, "2.")
    if bounds is None:
        return []
    sec_idx, sec_end = bounds

    header_idx, col_map = None, None
    for i in range(sec_idx + 1, min(sec_idx + 6, sec_end)):
        cm = _map_headers(rows[i], _EVAL_SYNONYMS, _EVAL_REQUIRED)
        if cm:
            header_idx, col_map = i, cm
            break
    if col_map is None:
        return []

    def g(row, key):
        idx = col_map.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    # 계좌(금융기관)별 평가액 합. 금융기관명/계좌번호는 위에서 내려쓰기(병합셀 대비).
    from collections import OrderedDict
    groups: "OrderedDict[tuple, dict]" = OrderedDict()
    last_inst = last_acc = None
    for i in range(header_idx + 1, sec_end):
        row = rows[i]
        inst = g(row, "금융기관명")
        acc = g(row, "계좌번호")
        if inst is not None and str(inst).strip():
            last_inst = str(inst).strip()
        if acc is not None and str(acc).strip():
            last_acc = str(acc).strip()
        if _is_dc(g(row, "종목명")):       # 확정기여형(DC) = 외부적립퇴직급여 → 단기금융 제외
            continue
        ev = _to_amount(g(row, "평가액"))
        if not ev or last_inst is None:
            continue
        key = (last_inst, last_acc)
        if key not in groups:
            groups[key] = {"금융기관명": last_inst, "계좌번호": last_acc, "평가액": 0}
        groups[key]["평가액"] += ev
    return list(groups.values())
