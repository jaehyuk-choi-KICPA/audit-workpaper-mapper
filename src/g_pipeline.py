# -*- coding: utf-8 -*-
"""G 유무형자산 조서 통합 생성(build_g) — 다중시트 연쇄 graft.

정산표(별도정산표·CF·분개장) + 고정자산 관리대장 + 참고자료 메모를 입력으로,
완성본 G 조서의 5개 시트를 채워 **연쇄 graft**(컨트롤·매크로 무손실)로 한 워크북에 이식한다:
  ① 고정자산 관리대장_25 (회사제시 입력 + 감사인 재계산)
  ② G100 총괄표 (엔진 기초/기말/수정후 + 이동컬럼[취득/처분/감가상각비] + 회계정책 박스)
  ③ G300 감가상각비 검토 (모집단·재계산표 동적, G100·관리대장 참조)
  ④ G200 취득/처분 Test (샘플선정·TARGET TEST 분개장 매핑 + 처분 폼 신설)
부분 실패 허용(한 시트가 안 돼도 나머지 생성). 주석은 후속.
"""

import re
import shutil
import warnings
from pathlib import Path

import yaml


def _norm(s):
    return re.sub(r"\s+", "", str(s)) if s is not None else ""


def build_g(*, settlement, register, ref_dir, template, output,
            params=None, parsed_dir=None, config_dir=None, progress=None):
    """G 조서 통합 생성 — **독립 실행**(총괄표 본문까지 자체 생성, 검증용).

    완성본 G 템플릿에 ① 관리대장_25 ② G100 총괄표(엔진 기초/기말/수정후 + 이동컬럼[취득/처분/감가상각비]
    + 회계정책) ③ G300 ④ G200(취득/처분 Test) ⑤ 주석을 채워 연쇄 graft(컨트롤·매크로 무손실).
    총괄표 본문은 정산표 EXE와 동일 엔진(A0Generator+g.yaml)을 재사용 — 독립 검증 위해 G-0도 자체 생성.
    CF·분개장은 settlement(정산표 파일)에서 읽는다. 항상 출력 생성(부분 실패 허용).
    """
    import sys
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from extractors import (parse_trial_balance, parse_cf_depreciation,
                            parse_fixed_asset_ledger, parse_journal, extract_fixed_asset_movements)
    from extractors.fixed_asset_memo import load_memo
    from cache import load_with_cache
    from generators import A0Generator
    from generators.fixed_asset_register import FixedAssetRegisterGenerator
    from generators.g_movements import fill_g100_movements, fill_g100_policy, fill_note
    from generators.g_depreciation import fill_g300
    from generators.g_sampling import fill_g200
    from generators.sheet_surgery import extract_light, graft_sheet
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string

    params = params or {}
    cfgdir = Path(config_dir or (Path(__file__).resolve().parent.parent / "_internal" / "config"))
    raw = yaml.safe_load(open(cfgdir / "g.yaml", encoding="utf-8"))
    out = Path(output)
    out.parent.mkdir(parents=True, exist_ok=True)
    tmp = out.parent / "_g_tmp"
    tmp.mkdir(exist_ok=True)
    done, warn = [], []

    def _safe(label, fn, default=None):
        try:
            return fn()
        except Exception as e:
            warn.append(f"{label} 실패({type(e).__name__}: {e})")
            return default

    # ── 파싱 (정산표: 총괄표 본문 tb + CF 감가상각비) ──
    tb = _safe("정산표", lambda: parse_trial_balance(settlement), []) or []
    cf = _safe("CF 감가상각비", lambda: parse_cf_depreciation(settlement), {}) or {}
    # 분개장: **이상적 분개장 양식(parse_journal)으로 변환·캐시 후** 거기서 취득/처분 추출(원칙).
    # 변환은 정산표 EXE가 미리 해두면 같은 캐시(변환자료) 재사용 — G는 꺼내 쓰기만.
    parsed = parsed_dir or str(out.parent / "변환자료")
    journal = _safe("분개장→이상적 변환",
                    lambda: load_with_cache(settlement, parse_journal, cache_dir=parsed, tag="journal_v1"), []) or []
    mv = _safe("유무형 취득/처분 추출",
               lambda: extract_fixed_asset_movements(journal, list(cf.keys())), {}) or {}
    ledger = _safe("고정자산 관리대장", lambda: parse_fixed_asset_ledger(register), []) if register else []
    memo = _safe("참고자료 메모", lambda: load_memo(ref_dir), []) if ref_dir else []

    SHEETS = {
        "register": "고정자산 관리대장_25",
        "g100": "G100_총괄표",
        "g300": raw.get("g300", {}).get("sheet", "G300_감가상각비 검토"),
        "g200": raw.get("g200", {}).get("sheet", "G200_취득,처분 Test"),
        "note": raw.get("note", {}).get("sheet", "주석"),
    }

    # ── ① 관리대장_25 ──
    reg_light = str(tmp / "register.xlsx")
    subtotals = {}
    if ledger:
        def _reg():
            extract_light(template, SHEETS["register"], reg_light)
            g = FixedAssetRegisterGenerator(reg_light, str(cfgdir / "g_register.yaml"))
            info = g.fill(ledger, memo, params)
            g.save(reg_light)
            return info
        info = _safe("관리대장_25 생성", _reg)
        if info:
            subtotals = info.get("subtotal_rows", {})
            done.append(SHEETS["register"])

    # ── ② G100 (엔진 총괄표 본문 + 이동컬럼 + 회계정책) ──
    g100_light = str(tmp / "g100.xlsx")
    accum_row, base_row = {}, {}

    def _g100():
        extract_light(template, SHEETS["g100"], g100_light)
        g = A0Generator(g100_light, str(cfgdir / "g.yaml"))
        g.fill(tb, [], params)
        fill_g100_movements(g.ws, raw, mv, cf)
        fill_g100_policy(g.ws, raw, memo)
        g.save(g100_light)
        return True
    if _safe("G100 생성", _g100):
        done.append(SHEETS["g100"])
        # 누계행/본체행 스캔(G300/G200 좌표용)
        wb = openpyxl.load_workbook(g100_light)
        ws = wb[SHEETS["g100"]]
        nci = column_index_from_string(raw.get("movement_name_col", "C"))
        ds = raw.get("body", {}).get("data_start_row", 11)
        for r in range(ds, ds + 60):
            v = ws.cell(r, nci).value
            if not v:
                continue
            nm = str(v).strip()
            if _norm(raw.get("body", {}).get("body_end_anchor", "계정과목")) in _norm(nm):
                break
            if "감가상각누계액" in nm:
                accum_row[_norm(re.sub("감가상각누계액", "", nm).strip(" _"))] = r
            else:
                base_row[_norm(nm)] = r

    # ── dep_list (G300용) ──
    tangible, intangible = [], []
    for a in cf:
        na = _norm(a)
        rec = {"계정": a, "회사감가비": cf[a], "register_row": subtotals.get(a) or subtotals.get(na)}
        if na in accum_row:
            rec["g100_row"] = accum_row[na]
            tangible.append(rec)
        else:
            rec["g100_row"] = base_row.get(na)
            intangible.append(rec)
    intang_names = [r["계정"] for r in intangible]

    # ── ③ G300 ──
    g300_light = str(tmp / "g300.xlsx")

    def _g300():
        extract_light(template, SHEETS["g300"], g300_light)
        wb = openpyxl.load_workbook(g300_light)
        ws = wb[SHEETS["g300"]]
        fill_g300(ws, raw, tangible, intangible)
        wb.save(g300_light)
        return True
    g300_ok = (tangible or intangible) and _safe("G300 생성", _g300)
    if g300_ok:
        done.append(SHEETS["g300"])

    # ── ④ G200 ──
    g200_light = str(tmp / "g200.xlsx")

    def _g200():
        extract_light(template, SHEETS["g200"], g200_light)
        wb = openpyxl.load_workbook(g200_light)
        ws = wb[SHEETS["g200"]]
        fill_g200(ws, raw, mv, intangible=intang_names)   # 취득·처분 표 모두 채움(처분 폼 선구성)
        wb.save(g200_light)
        return True
    g200_ok = mv and _safe("G200 생성", _g200)
    if g200_ok:
        done.append(SHEETS["g200"])

    # ── ⑤ 주석 (헬퍼 테이블을 G100 참조로 미러 → 변동표 ROUND 자동) ──
    note_light = str(tmp / "note.xlsx")

    def _note():
        extract_light(template, SHEETS["note"], note_light)
        wb = openpyxl.load_workbook(note_light)
        ws = wb[SHEETS["note"]]
        fill_note(ws, raw, base_row, accum_row)
        wb.save(note_light)
        return True
    note_ok = (base_row or accum_row) and _safe("주석 생성", _note)
    if note_ok:
        done.append(SHEETS["note"])

    # ── 연쇄 graft (완성본 → 시트별 무손실 이식) ──
    cur = str(template)
    seq = [("register", reg_light), ("g100", g100_light), ("g300", g300_light),
           ("g200", g200_light), ("note", note_light)]
    seq = [(k, p) for k, p in seq if SHEETS[k] in done]
    for i, (k, p) in enumerate(seq):
        nxt = str(out) if i == len(seq) - 1 else str(tmp / f"chain{i}.xlsx")
        if _safe(f"{SHEETS[k]} graft", lambda c=cur, pp=p, kk=k, nn=nxt: graft_sheet(c, pp, SHEETS[kk], nn) or nn):
            cur = nxt
        if progress:
            progress(SHEETS[k], True)
    if cur != str(out):
        shutil.copy(template, out)   # graft 하나도 못 했으면 템플릿 복사(항상 출력)
    return done, warn
