# -*- coding: utf-8 -*-
"""A-0(현금및현금성자산, 장단기금융상품) 개별시트 채움 — 쉬운 실행기 (비전문가용 대화형).

**자체완결**: 정산표로 A-0 총괄표를 직접 만들고(조서생성의 A-0 항목만 결합 — 단독 검증 가능),
그 총괄표에 **연동**해 개별시트를 동적 구성한다:
  · 총괄표 5)장기금융 : 조회서 보험사 ↔ 거래처원장 잔액(이름매칭) → 잔액/해약환급금/차액
  · 총괄표 4)단기금융 : INVESTMENT 평가액(확정기여형 DC=외부적립퇴직급여 제외) — 있을 때만(prune)
  · A050 주석         : 총괄표 계정 수정후/기초를 =ROUND(!/1000,0) 참조(이름매칭)
  · A020 실사         : 기말환율표(참고자료) + 외화 원금×환율 가이드
  · 보험가입현황      : 조회서 INSURANCE 해약환급금 per-policy 매핑

사용:
  1) `입력자료/`에 정산표(필수)·금융기관조회서·거래처원장(잔액)을 넣는다.
  2) `양식자료/`에 A-0 템플릿(A-0_4000_template_현금및현금성자산.xlsx)을 둔다.
  3) `A-0생성.bat` 더블클릭 (또는 python src/easy_run_a0.py)
  4) 회사명·기준일·작성자·검토자 입력 → `출력조서/{회사}/`에 A-0 생성(완성본 보존 graft).
"""
import glob
import io
import os
import re
import sys
import warnings
import zipfile
from pathlib import Path

warnings.filterwarnings("ignore", message=".*[Hh]eader or footer.*")
sys.path.insert(0, str(Path(__file__).resolve().parent))
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stdin.reconfigure(encoding="utf-8")
except Exception:
    pass

import openpyxl
from pipeline import build_lead_all
from extractors import load_fx_rates
from generators.sheet_surgery import graft_sheet
from generators import a0_detail

# A-0 총괄표 생성용 레지스트리(조서생성의 A-0 항목만 따옴 — 단독 실행 가능하게 결합)
_A0 = {"code": "A-0", "config": "a0.yaml", "template": "A-0_4000_template_현금및현금성자산.xlsx"}
TOTAL_SHEET = "4000_A000 총괄표"
A050_SHEET = "A050_주석검토"
A020_SHEET = "A020_금융자산 실사"
HOLD_SHEET = "장기금융상품_보험가입현황"


from workspace import select_workspace, template_dir, config_dir, prompt_company_info

# 양식·config = 회사 무관 공유 자산. 입력·변환·참고·출력은 회사별 워크스페이스에서 받는다.
TEMPLATE_DIR = template_dir()
CONFIG_DIR = config_dir()
INPUT_DIR = PARSED_DIR = OUTPUT_DIR = REF_DIR = None  # main()에서 워크스페이스로 설정


def _line(m=""):
    print(m, flush=True)


def _ask(label, required=True, example=""):
    hint = f" (예: {example})" if example else ""
    while True:
        v = input(f"  {label}{hint}: ").strip()
        if v or not required:
            return v
        _line("  ▷ 값을 입력해 주세요.")


def _find(pattern):
    c = [h for h in glob.glob(str(INPUT_DIR / "**" / pattern), recursive=True) if "~$" not in h]
    return sorted(c)[0] if c else None


def _extract_one(src, sheet, out):
    """완성본에서 단일 시트만 가벼운 임시본으로 추출(도형·외부링크 제거)."""
    buf = io.BytesIO()
    with zipfile.ZipFile(src) as zin, zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zo:
        for n in zin.namelist():
            if n.startswith(("xl/drawings/", "xl/media/")):
                continue
            b = zin.read(n)
            if n.startswith("xl/worksheets/") and n.endswith(".xml"):
                b = re.sub(rb"<drawing\b[^>]*/>", b"", b); b = re.sub(rb"<legacyDrawing\b[^>]*/>", b"", b)
            elif n.endswith(".rels"):
                b = re.sub(rb'<Relationship[^>]*(?:drawing|image|vmlDrawing)"[^>]*/>', b"", b)
            elif n == "[Content_Types].xml":
                b = re.sub(rb"<Override[^>]*(?:drawing|vmlDrawing)[^>]*/>", b"", b)
            zo.writestr(n, b)
    buf.seek(0)
    wb = openpyxl.load_workbook(buf)
    for s in list(wb.sheetnames):
        if s != sheet:
            del wb[s]
    wb[sheet].sheet_state = "visible"
    wb.save(out)


def _edit_sheet(path, sheet, fn, tmp):
    """extract_one → fn(ws) → graft_sheet(완성본 보존)로 한 시트 적용."""
    light = str(Path(tmp) / "light.xlsx")
    _extract_one(path, sheet, light)
    wb = openpyxl.load_workbook(light)
    fn(wb[sheet])
    wb.save(light)
    out = path + ".graft"
    graft_sheet(path, light, sheet, out)
    os.replace(out, path)


def _inject_cols(path, sheet, widths):
    """graft가 빈 템플릿 cols를 제거하므로, 시트에 <cols> 너비를 XML로 재주입."""
    import xml.etree.ElementTree as ET
    M = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    RID = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
    z = zipfile.ZipFile(path)
    wbx = ET.fromstring(z.read("xl/workbook.xml"))
    rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
    r2t = {r.get("Id"): r.get("Target") for r in rels}
    target = None
    for s in wbx.iter(M + "sheet"):
        if s.get("name") == sheet:
            target = "xl/" + r2t[s.get(RID)].lstrip("/")
    z.close()
    if not target:
        return
    cols = "<cols>" + "".join(
        '<col min="%d" max="%d" width="%d" customWidth="1"/>' % (c, c, w) for c, w in sorted(widths.items())
    ) + "</cols>"
    tmp = path + ".w"
    with zipfile.ZipFile(path) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zo:
        for it in zin.infolist():
            b = zin.read(it.filename)
            if it.filename == target:
                b = re.sub(rb"<cols>.*?</cols>", b"", b, flags=re.DOTALL)
                b = b.replace(b"<sheetData", cols.encode() + b"<sheetData", 1)
            zo.writestr(it, b)
    os.replace(tmp, path)


def _find_a0_outputs():
    """출력조서/**/에서 조서생성이 만든 A-0 파일을 찾는다(총괄표 시트 보유 확인)."""
    cands = [h for h in glob.glob(str(OUTPUT_DIR / "**" / "*.xls*"), recursive=True)
             if "~$" not in h and re.search(r"A-?0[_\b]", Path(h).name) and "현금" in Path(h).name]
    good = []
    for h in cands:
        try:
            z = zipfile.ZipFile(h)
            wbx = z.read("xl/workbook.xml").decode("utf-8", "ignore"); z.close()
            if "총괄표" in wbx:
                good.append(h)
        except Exception:
            pass
    return sorted(set(good))


def main():
    _line("=" * 56)
    _line("  A-0 현금및현금성자산 조서 생성 (총괄표 + 개별시트 · 단독 실행)")
    _line("=" * 56)

    global INPUT_DIR, PARSED_DIR, OUTPUT_DIR, REF_DIR
    ws = select_workspace(line=_line)
    if ws is None:
        return 2
    INPUT_DIR, PARSED_DIR, REF_DIR, OUTPUT_DIR = (
        ws.input_dir, ws.parsed_dir, ws.ref_dir, ws.output_dir)
    TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
    _line(f"\n[회사] {ws.company}   (작업폴더: {ws.input_dir.parent})")

    settlement = _find("*정산표*.xls*")
    if not settlement:
        _line(f"\n[안내] 정산표를 찾지 못했습니다(파일명에 '정산표' 포함). 입력자료에 넣어 주세요:\n  {INPUT_DIR}")
        return 2
    tpl = TEMPLATE_DIR / _A0["template"]
    if not tpl.exists():
        _line(f"\n[안내] A-0 템플릿이 없습니다: {tpl}")
        return 2
    confirm = _find("*조회서*기말감사*.xls*") or _find("필수자료2*") or _find("*금융기관조회서*.xls*")
    ledger = _find("필수자료3*") or _find("*잔액*.xls*") or _find("*거래처원장*.xls*")
    _line(f"\n[정산표]   {Path(settlement).name}")
    _line(f"[조회서]   {Path(confirm).name if confirm else '없음 — 장기금융/보험가입현황 생략'}")
    _line(f"[잔액원장] {Path(ledger).name if ledger else '없음 — 장기금융 잔액 매칭 생략'}")

    _line("\n[회사 정보 입력]")
    company = ws.company
    date, preparer, reviewer = prompt_company_info(ws, input, _line)

    out_dir = ws.output_dir
    a0_path = str(out_dir / _A0["template"].replace("_template_", "_"))
    # 있으면 그대로 사용(조서생성 산출물 재사용), 없으면 정산표로 직접 생성.
    if os.path.exists(a0_path):
        _line(f"\n[총괄표] 기존 산출물 재사용: {Path(a0_path).name}")
        _line("        (정산표가 바뀌었으면 '조서생성'을 다시 돌리거나 이 파일을 지우고 실행)")
    else:
        _line("\n총괄표가 없어 정산표로 생성 중...\n")
        try:
            build_lead_all(
                settlement=settlement, registry=[_A0], config_dir=str(CONFIG_DIR),
                template_root=str(TEMPLATE_DIR), output_dir=str(out_dir),
                parsed_dir=str(ws.parsed_dir),
                params={"회사명": company, "날짜": date, "preparer": preparer, "reviewer": reviewer},
            )
        except PermissionError:
            _line(f"\n[오류] 출력 파일이 열려 있습니다. 닫고 다시 실행해 주세요:\n  {out_dir}")
            return 1
        except Exception as e:
            _line(f"\n[오류] 총괄표 생성 실패: {type(e).__name__}: {e}")
            return 1
        if not os.path.exists(a0_path):
            _line("\n[오류] 총괄표 산출물이 없습니다."); return 1
    _line("개별시트 채우는 중...")

    tmp = str(out_dir / "_tmp_a0")
    Path(tmp).mkdir(parents=True, exist_ok=True)
    fx = load_fx_rates(str(REF_DIR)) or {}
    policies = a0_detail.parse_insurance_holdings(confirm) if confirm else []
    recon = a0_detail.build_longterm_recon(confirm, ledger) if (confirm and ledger) else []
    recon_st = a0_detail.build_shortterm_recon(confirm, ledger) if (confirm and ledger) else []
    notes = []

    # 총괄표 연동값 읽기용(A050이 참조) — 같은 파일의 총괄표 ws
    try:
        twb = openpyxl.load_workbook(a0_path, data_only=False)
        total_ws = twb[TOTAL_SHEET]

        st_placed = {"n": 0}

        def _do_total(ws):
            a0_detail.cleanup_total_marker(ws)
            a0_detail.fill_total_longterm(ws, recon)        # 5) 장기금융 평가대사(A-1/조회서 기반)
            st_placed["n"] = a0_detail.fill_total_shortterm(ws, recon_st)  # 4) 단기금융(존재+표 있을 때)
        _edit_sheet(a0_path, TOTAL_SHEET, _do_total, tmp)
        notes.append("총괄표 5)장기금융: 조회서 보험사 %d건 평가대사" % len(recon))
        if not recon_st:
            notes.append("총괄표 4)단기금융: 없음(조회서 INVESTMENT 평가액 0 → prune)")
        elif st_placed["n"]:
            notes.append("총괄표 4)단기금융: %d건 평가대사" % st_placed["n"])
        else:
            notes.append("총괄표 4)단기금융: %d건 감지(%s 등) — 단 양식에 4)평가대사 표가 없어 미배치(수동/표추가 필요)"
                         % (len(recon_st), str(recon_st[0]["거래처"])[:12]))

        _edit_sheet(a0_path, A050_SHEET, lambda ws: a0_detail.fill_a050_notes(ws, total_ws, TOTAL_SHEET), tmp)
        notes.append("A050 주석: 총괄표 ROUND 연동")
        _edit_sheet(a0_path, A020_SHEET, lambda ws: a0_detail.fill_a020_fx(ws, fx), tmp)
        notes.append("A020 실사: 기말환율표(%d통화)" % len(fx))
        _edit_sheet(a0_path, HOLD_SHEET, lambda ws: a0_detail.fill_insurance_holdings(ws, policies), tmp)
        _inject_cols(a0_path, HOLD_SHEET, a0_detail.HOLDINGS_COL_WIDTHS)
        notes.append("보험가입현황: 조회서 %d건 매핑" % len(policies))
    except Exception as e:
        notes.append(f"개별시트 일부 실패: {type(e).__name__}: {e}")

    # 정리
    try:
        import shutil
        shutil.rmtree(tmp, ignore_errors=True)
    except Exception:
        pass

    lines = ["=" * 60, "  A-0 생성 완료", "=" * 60, f"  결과: {a0_path}"]
    for n in notes:
        lines.append("   - " + n)
    lines.append("  ※ 초안입니다. 보험가입현황은 조회서 회신분 — 회사 보험명세로 보완하세요.")
    summary = "\n".join(lines)
    _line("\n" + summary)
    try:
        (out_dir / f"{out_dir.name}_A0_실행리포트.txt").write_text(summary, encoding="utf-8")
    except Exception:
        pass
    return 0


if __name__ == "__main__":
    rc = main()
    try:
        input("\n[Enter] 키를 누르면 창이 닫힙니다...")
    except EOFError:
        pass
    sys.exit(rc)
