"""완성본 조서(.xlsx/.xlsm)에서 한 시트만 떼어 깨끗한 작업 템플릿(.xlsx)을 만든다.

회계법인 완성본은 깨진 drawing(Target=NULL)·외부링크·VBA·타시트가 많아 openpyxl/Excel COM이
힘들어한다. 깨진 drawing을 무력화한 뒤 대상 시트의 값+스타일+병합+행/열치수만 새 워크북에 복사한다.
(셀복사 방식이 가장 안전 — 외부링크·VBA·drawing 전부 떨어져 나감.)
"""
import os
import tempfile
import zipfile
from copy import copy

import openpyxl

_EMPTY_RELS = b'<?xml version="1.0"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"></Relationships>'
_EMPTY_DR = b'<?xml version="1.0"?><xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"></xdr:wsDr>'


def clean_sheet(src: str, sheet: str, out: str) -> str:
    """src의 sheet를 깨끗한 단일시트 .xlsx(out)으로 만든다."""
    fixed = os.path.join(tempfile.gettempdir(), "tp_fixed.xlsx")
    zin = zipfile.ZipFile(src)
    zout = zipfile.ZipFile(fixed, "w", zipfile.ZIP_DEFLATED)
    for it in zin.infolist():
        data = zin.read(it.filename)
        f = it.filename
        if f.startswith("xl/drawings/_rels/") and f.endswith(".rels") and b"NULL" in data:
            data = _EMPTY_RELS
        elif f.startswith("xl/drawings/drawing") and f.endswith(".xml"):
            data = _EMPTY_DR
        zout.writestr(it, data)
    zin.close(); zout.close()

    sw = openpyxl.load_workbook(fixed)
    sws = sw[sheet]
    nw = openpyxl.Workbook()
    nws = nw.active
    nws.title = sheet
    for r in range(1, sws.max_row + 1):
        for c in range(1, sws.max_column + 1):
            sc = sws.cell(r, c)
            nc = nws.cell(r, c)
            nc.value = sc.value
            if sc.has_style:
                nc.font = copy(sc.font); nc.border = copy(sc.border)
                nc.fill = copy(sc.fill); nc.alignment = copy(sc.alignment)
                nc.number_format = sc.number_format
    for m in sws.merged_cells.ranges:
        nws.merge_cells(str(m))
    for k, d in sws.column_dimensions.items():
        nws.column_dimensions[k].width = d.width
    for k, d in sws.row_dimensions.items():
        nws.row_dimensions[k].height = d.height
    os.makedirs(os.path.dirname(out) or ".", exist_ok=True)
    nw.save(out)
    return out
